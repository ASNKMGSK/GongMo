"""Golden-set 빌드 — md 1파일 = 1 청크 (item × sample).

청크 구조 (4 섹션):
  1. 파싱 원문 — Layer1 가 파싱한 transcript (turn 단위, speaker 분리)
  2. 점수 — GT 점수
  3. 이유 — GT 사유 (비고)
  4. 근거 — GT STT 발췌

Layer1 의 rule_pre_verdict score / rationale 등은 메타가 아니라 시스템 1차 판정이므로
golden_set 청크에는 포함하지 않음.

총 산출: 17 평가항목 × 14 학습셋 = 238 md 파일.
파일명: `{NN}_{slug}_{sample_id}.md`
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[4]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from v2.layer1.run_layer1 import run_layer1

GT_SRC = r"C:\Users\META M\Desktop\qa테스트 정답\STT QA 정답표_재채점 및 근거 작성(STT 원문 및 상담유형 추가).xlsx"
GT_TMP = r"C:\Users\META M\Desktop\_gt_inspect.xlsx"
LEARN_DIR = Path(r"C:\Users\META M\Desktop\qa 샘플\학습셋")
OUT_DIR = Path(__file__).parent

ITEM_NUMBERS = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
SLUGS = {
    1: "initial_greeting", 2: "closing_greeting", 4: "empathy",
    5: "hold_notice", 6: "polite_language", 7: "cushion_words",
    8: "needs_identification", 9: "customer_info_verification",
    10: "explanation_clarity", 11: "top_down_answer",
    12: "problem_solving_attitude", 13: "additional_guidance",
    14: "follow_up", 15: "correct_information",
    16: "mandatory_notice", 17: "pii_verification", 18: "pii_protection",
}
# 항목 번호 → Layer1 agent_turn_assignments 그룹 키 매핑
# (CLAUDE.md QA Pipeline 8 평가 노드 매핑 기준)
ITEM_TO_GROUP = {
    1: "greeting", 2: "greeting",
    4: "understanding", 5: "understanding",
    6: "courtesy", 7: "courtesy",
    8: "mandatory", 9: "mandatory",
    10: "scope", 11: "scope",
    12: "proactiveness", 13: "proactiveness", 14: "proactiveness",
    15: "work_accuracy", 16: "work_accuracy",
    17: "incorrect_check", 18: "incorrect_check",
}


def _bucket(score: int | None, max_score: int | None) -> str:
    if score is None or max_score is None:
        return "unknown"
    if score == max_score:
        return "full"
    if score == 0:
        return "zero"
    return "partial"


def _clean_excerpt(text: str) -> str:
    """GT 엑셀 STT 발췌 셀 정화.

    엑셀 작성자가 여러 발췌 사이에 ``-------`` / ``=======`` 같은 dash 구분자를 넣어둔
    경우 가시성에 도움이 안 되고 RAG 임베딩에도 노이즈로 작용. 다음을 수행:
      - dash/equal/star 만으로 이루어진 줄 제거
      - 연속 빈 줄 1줄로 압축
      - 양 끝 trim
    """
    if not text:
        return ""
    out_lines: list[str] = []
    for raw in text.splitlines():
        line = raw.rstrip()
        stripped = line.strip()
        # dash/equal/star/underscore 만으로 이루어진 구분 줄은 제거
        if stripped and re.fullmatch(r"[\-=*_~·•·\s]{3,}", stripped):
            continue
        out_lines.append(line)
    # 연속 빈 줄 1줄로 압축
    compact: list[str] = []
    prev_blank = False
    for line in out_lines:
        is_blank = not line.strip()
        if is_blank and prev_blank:
            continue
        compact.append(line)
        prev_blank = is_blank
    return "\n".join(compact).strip()


def _turns_to_text(turns: list[dict]) -> str:
    """파싱된 turns 리스트 → 'turn_id [speaker]: text' 라인 join."""
    out = []
    for t in turns or []:
        tid = t.get("turn_id")
        spk = t.get("speaker", "?")
        txt = (t.get("text") or "").strip()
        out.append(f"[{tid}] {spk}: {txt}")
    return "\n".join(out)


def main() -> None:
    sample_parsed: dict[str, dict] = {}
    print("== Layer1 파싱 (14 samples) ==")
    for fn in sorted(LEARN_DIR.glob("*.json")):
        sid = fn.name.split("_")[0]
        with open(fn, encoding="utf-8") as f:
            d = json.load(f)
        parsed = run_layer1(d["transcript"])
        sample_parsed[sid] = {"label": d.get("label", ""), "parsed": parsed}
    print(f"  {len(sample_parsed)} samples parsed.")

    shutil.copyfile(GT_SRC, GT_TMP)
    wb = openpyxl.load_workbook(GT_TMP, data_only=True)
    sample_to_sheet = {}
    for sid in sample_parsed:
        matches = [s for s in wb.sheetnames if s.endswith(f"_{sid}")]
        if matches:
            sample_to_sheet[sid] = matches[0]

    items_data = {
        n: {"item_name": "", "category": "", "max_score": None, "chunks": []}
        for n in ITEM_NUMBERS
    }
    for sid in sorted(sample_parsed.keys()):
        if sid not in sample_to_sheet:
            continue
        ws = wb[sample_to_sheet[sid]]
        cur_idx = 0
        cur_category = None
        for r in range(6, 50):
            cat = ws.cell(r, 1).value
            item_name = ws.cell(r, 2).value
            max_score = ws.cell(r, 4).value
            score = ws.cell(r, 5).value
            note = ws.cell(r, 6).value
            stt = ws.cell(r, 7).value
            if cat and isinstance(cat, str) and "총점" in cat:
                break
            if cat:
                cur_category = str(cat).strip()
            if item_name is None:
                continue
            if cur_idx >= len(ITEM_NUMBERS):
                break
            n = ITEM_NUMBERS[cur_idx]
            if not items_data[n]["item_name"]:
                items_data[n]["item_name"] = re.sub(r"\n.*", "", str(item_name).strip())
                items_data[n]["category"] = re.sub(r"\n.*", "", (cur_category or "")).strip()
                items_data[n]["max_score"] = (
                    int(max_score) if isinstance(max_score, (int, float)) else None
                )
            items_data[n]["chunks"].append({
                "sample_id": sid,
                "score": int(score) if isinstance(score, (int, float)) else None,
                "note": _clean_excerpt(str(note)) if note else "",
                "evidence_excerpt": _clean_excerpt(str(stt)) if stt else "",
            })
            cur_idx += 1

    os.remove(GT_TMP)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for old in list(OUT_DIR.glob("*.md")) + list(OUT_DIR.glob("*.json")):
        if old.name.startswith("_"):
            continue
        old.unlink()

    written = 0
    for n in ITEM_NUMBERS:
        d = items_data[n]
        slug = SLUGS[n]

        group = ITEM_TO_GROUP.get(n, "")
        for c in d["chunks"]:
            sid = c["sample_id"]
            sm = sample_parsed.get(sid, {})
            pp = sm.get("parsed", {})
            ata = pp.get("agent_turn_assignments") or {}
            assignment = ata.get(group, {}) if group else {}
            assigned_text = (assignment or {}).get("text", "") or ""
            assigned_turn_ids = (assignment or {}).get("turn_ids", [])
            split_desc = (assignment or {}).get("description", "")
            bucket = _bucket(c["score"], d["max_score"])

            md: list[str] = []
            md.append(f"# [item_{n:02d}] {d['item_name']} — sample {sid}")
            md.append("")
            md.append(
                f"_대분류: {d['category']} · 배점 {d['max_score']}점 · "
                f"평가그룹: {group} · sample_label: {sm.get('label', '')}_"
            )
            md.append("")

            md.append("## 파싱 원문 (Layer1 → 평가그룹별 분할 문맥)")
            md.append("")
            if split_desc:
                md.append(f"_{split_desc} · 할당 turn 수: {len(assigned_turn_ids)}_")
                md.append("")
            md.append("```")
            md.append(assigned_text if assigned_text else "-")
            md.append("```")
            md.append("")

            md.append("## 점수")
            md.append("")
            md.append(f"**{c['score']}점** / {d['max_score']}점 ({bucket})")
            md.append("")

            md.append("## 이유")
            md.append("")
            md.append(c["note"] if c["note"] else "-")
            md.append("")

            md.append("## 근거")
            md.append("")
            md.append("```")
            md.append(c["evidence_excerpt"] if c["evidence_excerpt"] else "-")
            md.append("```")
            md.append("")

            fn = f"{n:02d}_{slug}_{sid}.md"
            (OUT_DIR / fn).write_text("\n".join(md), encoding="utf-8")
            written += 1

    print(f"=== {written} md 파일 빌드 완료 ===")
    print(f"출력 경로: {OUT_DIR}")


if __name__ == "__main__":
    main()
