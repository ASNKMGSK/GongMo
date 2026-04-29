# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""HITL human_reviews → xlsx export.

`export_reviews_to_xlsx()` 는 human_reviews 테이블을 필터링 조회한 뒤 openpyxl 로
xlsx 파일을 생성해 `~/Desktop/QA평가결과/HITL_비교/<timestamp>_reviews.xlsx`
(또는 env ``QA_HITL_EXPORT_ROOT`` override) 에 저장한다.

빈 결과여도 header 만 있는 0-row xlsx 가 생성된다.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any


logger = logging.getLogger(__name__)


ITEM_NAMES: dict[int, str] = {
    1: "첫인사",
    2: "끝인사",
    3: "호응 및 공감",
    4: "호응 및 공감",
    5: "대기 멘트",
    6: "정중한 표현",
    7: "쿠션어 활용",
    8: "문의 파악 및 재확인(복창)",
    9: "고객정보 확인",
    10: "설명의 명확성",
    11: "두괄식 답변",
    12: "문제 해결 의지",
    13: "부연 설명 및 추가 안내",
    14: "사후 안내",
    15: "정확한 안내",
    16: "필수 안내 이행",
    17: "정보 확인 절차",
    18: "정보 보호 준수",
}


EXPORT_HEADERS: list[str] = [
    "consultation_id",
    "item_number",
    "item_name",
    "ai_score",
    "human_score",
    "delta",
    "ai_confidence",
    "force_t3",
    "status",
    "reviewer_id",
    "reviewer_role",
    "ai_judgment",
    "human_note",
    "created_at",
    "confirmed_at",
]


def _export_root() -> Path:
    override = os.environ.get("QA_HITL_EXPORT_ROOT")
    if override:
        return Path(override)
    return Path.home() / "Desktop" / "QA평가결과" / "HITL_비교"


def _build_row(rec: dict[str, Any]) -> list[Any]:
    ai = rec.get("ai_score")
    human = rec.get("human_score")
    if ai is not None and human is not None:
        try:
            delta: Any = abs(float(ai) - float(human))
        except (TypeError, ValueError):
            delta = ""
    else:
        delta = ""

    force_t3_val = rec.get("force_t3")
    force_t3_yn = "Y" if (force_t3_val == 1 or force_t3_val is True) else "N"

    item_num = rec.get("item_number")
    try:
        item_name = ITEM_NAMES.get(int(item_num), "") if item_num is not None else ""
    except (TypeError, ValueError):
        item_name = ""

    return [
        rec.get("consultation_id", ""),
        item_num if item_num is not None else "",
        item_name,
        ai if ai is not None else "",
        human if human is not None else "",
        delta,
        rec.get("ai_confidence") if rec.get("ai_confidence") is not None else "",
        force_t3_yn,
        rec.get("status", ""),
        rec.get("reviewer_id") or "",
        rec.get("reviewer_role") or "",
        rec.get("ai_judgment") or "",
        rec.get("human_note") or "",
        rec.get("created_at", ""),
        rec.get("confirmed_at") or "",
    ]


def _load_rows(status: str | None, consultation_id: str | None) -> list[dict[str, Any]]:
    """human_reviews 를 필터 기준으로 읽어 dict 리스트 반환."""
    from v2.hitl import db as _hitl_db

    _hitl_db.init_db()

    import sqlite3  # noqa: F401  # db.get_conn 이 이미 sqlite3 기반

    sql = "SELECT * FROM human_reviews"
    clauses: list[str] = []
    params: list[Any] = []
    if status and status.lower() != "all":
        clauses.append("status = ?")
        params.append(status)
    if consultation_id:
        clauses.append("consultation_id = ?")
        params.append(consultation_id)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY consultation_id, item_number"

    with _hitl_db.get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]


def export_reviews_to_xlsx(
    *,
    status: str | None = None,
    consultation_id: str | None = None,
) -> tuple[Path, int]:
    """human_reviews 를 xlsx 로 저장. (path, row_count) 반환.

    Parameters
    ----------
    status : "pending" | "confirmed" | "rejected" | "all" | None
        None 또는 "all" 이면 전체.
    consultation_id : 단일 상담 필터.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl 필요") from exc

    rows = _load_rows(status=status, consultation_id=consultation_id)

    root = _export_root()
    root.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = root / f"{ts}_reviews.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "검토 큐"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    ws.append(EXPORT_HEADERS)
    for col_idx, _ in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for rec in rows:
        ws.append(_build_row(rec))

    # 열폭 대충 기본값 이상으로
    widths = [14, 10, 22, 9, 11, 8, 12, 8, 10, 14, 12, 40, 40, 22, 22]
    for col_idx, w in enumerate(widths, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = w

    wb.save(str(path))
    logger.info("export_reviews_to_xlsx: saved %s (rows=%d)", path, len(rows))
    return path, len(rows)
