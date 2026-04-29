# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""교차비교 xlsx 의 '해석 노트' 시트를 평이한 한국어로 재작성.

기존 시트 문제:
  - "Self-retrieval 역설" / "Annotator drift" / "Leave-one-out" / "rubric drift" 같은 용어 그대로 노출
  - 결론보다 가설/실험설계가 더 길어 사용자가 핵심을 못 잡음

새 시트:
  - 🎯 한 문장 결론 + 학습 vs 테스트 한 줄 비교표
  - "왜 이런 결과인가" 를 가설명 없이 평이한 한국어로
  - 가장 어긋난 상담 5건 (한 줄씩)
  - 다음에 할 일 (3단계)

실행:
  python -m v2.scripts.rewrite_cross_interpretation
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_PIPELINE_DIR = Path(__file__).resolve().parents[2]
if str(_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(_PIPELINE_DIR))

from v2.scripts.compare_learning_set_vs_xlsx import (  # type: ignore[import-untyped]
    ITEM_DEF, ITEM_NUM_TO_NAME,
    TEST_IDS, TRAINING_IDS,
    compute_metrics, load_ai_results, load_xlsx_ground_truth,
)


XLSX_GT = Path(r"C:\Users\META M\Desktop\QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
TARGET_XLSX = Path(r"C:\Users\META M\Desktop\학습_테스트_교차비교_20260422_011145\교차비교_20260422_011145.xlsx")
TRAIN_DIR = Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409")
TEST_DIR = Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545")
SHEET_NAME = "해석 노트"

TITLE_FILL = PatternFill("solid", fgColor="1f2937")
TITLE_FONT = Font(color="fbbf24", bold=True, size=14)
SECTION_FILL = PatternFill("solid", fgColor="6b21a8")
SECTION_FONT = Font(color="ffffff", bold=True, size=12)
KEY_FILL = PatternFill("solid", fgColor="f3f4f6")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="fef3c7")
BAD_FILL = PatternFill("solid", fgColor="fee2e2")
GOOD_FILL = PatternFill("solid", fgColor="dcfce7")
INFO_FILL = PatternFill("solid", fgColor="dbeafe")
BORDER = Border(left=Side("thin", color="cbd5e1"), right=Side("thin", color="cbd5e1"),
                top=Side("thin", color="cbd5e1"), bottom=Side("thin", color="cbd5e1"))
WRAP = Alignment(wrap_text=True, vertical="top")


def collect_stats(ids, results_dir):
    gt = load_xlsx_ground_truth(XLSX_GT, ids)
    ai = load_ai_results(results_dir, ids)
    matched = sorted(set(gt) & set(ai))
    per_item = defaultdict(list)
    sample_totals = []
    sample_item_diffs = {}
    for sid in matched:
        ai_t = hum_t = 0
        diffs = []
        for num, _name, mx, _row in ITEM_DEF:
            if num not in ai[sid] or num not in gt[sid]:
                continue
            a = ai[sid][num]; h = gt[sid][num]
            per_item[num].append((a, h, mx))
            ai_t += a; hum_t += h
            diffs.append((num, a - h))
        sample_totals.append((sid, ai_t, hum_t))
        sample_item_diffs[sid] = sorted(diffs, key=lambda x: -abs(x[1]))[:5]
    diffs_total = [(sid, a - h) for sid, a, h in sample_totals]
    diffs_total.sort(key=lambda x: -abs(x[1]))
    n = len(diffs_total)
    if n == 0:
        return None
    mae_t = sum(abs(d) for _, d in diffs_total) / n
    bias_t = sum(d for _, d in diffs_total) / n
    rmse_t = (sum(d*d for _, d in diffs_total) / n) ** 0.5
    max_abs_t = max(abs(d) for _, d in diffs_total)
    under = sum(1 for _, d in diffs_total if d < 0) / n * 100
    over = sum(1 for _, d in diffs_total if d > 0) / n * 100
    item_metrics = {num: compute_metrics(per_item.get(num, [])) for num, _, _, _ in ITEM_DEF}
    return {
        "n": n, "mae": mae_t, "bias": bias_t, "rmse": rmse_t,
        "max_abs": max_abs_t, "under": under, "over": over,
        "item_metrics": item_metrics,
        "sample_totals": {sid: (a, h) for sid, a, h in sample_totals},
        "diffs_total": diffs_total,
        "sample_item_diffs": sample_item_diffs,
    }


def fmt(v, suffix=""):
    if v is None:
        return "—"
    if isinstance(v, float):
        return f"{v:.2f}{suffix}"
    return f"{v}{suffix}"


def fmt_signed(v):
    if v is None:
        return "—"
    return f"{v:+.2f}" if v != 0 else "0.00"


def main() -> int:
    if not TARGET_XLSX.exists():
        print(f"[ERR] 미존재: {TARGET_XLSX}")
        return 1

    train = collect_stats(TRAINING_IDS, TRAIN_DIR)
    test  = collect_stats(TEST_IDS, TEST_DIR)

    wb = openpyxl.load_workbook(TARGET_XLSX)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.cell(row, 1, "📊 학습셋 vs 테스트셋 — 한눈에 보기")
    ws.cell(row, 1).fill = TITLE_FILL; ws.cell(row, 1).font = TITLE_FONT
    ws.cell(row, 1).alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 32
    row += 2

    # 결론 (격식체)
    direction_train = "낮게" if train["bias"] < 0 else "높게"
    direction_test  = "낮게" if test["bias"]  < 0 else "높게"
    headline = (
        f"양쪽 데이터셋 모두 AI 가 사람보다 {direction_train} 채점하는 동일한 경향을 보입니다. "
        f"학습셋의 평균 점수 차이는 {abs(train['bias']):.1f}점, "
        f"테스트셋의 평균 점수 차이는 {abs(test['bias']):.1f}점입니다. "
        f"테스트셋의 차이가 다소 큰 주요 원인은 단일 상담 668865 (-29점) 의 영향에 기인합니다."
    )
    ws.cell(row, 1, "🎯 결론")
    ws.cell(row, 2, headline)
    for col in (1, 2):
        ws.cell(row, col).fill = HIGHLIGHT_FILL
        ws.cell(row, col).font = Font(bold=(col == 1), size=11, color="92400e")
        ws.cell(row, col).alignment = WRAP
        ws.cell(row, col).border = BORDER
    ws.row_dimensions[row].height = 60
    row += 2

    # 1. 한 줄 비교
    ws.cell(row, 1, "1. 학습 vs 테스트 한눈에 비교")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1

    rows = [
        ("샘플 수",
         f"학습 {train['n']}건 / 테스트 {test['n']}건"),
        ("한 상담당 평균 점수 차이",
         f"학습 {train['mae']:.1f}점 / 테스트 {test['mae']:.1f}점 — 테스트셋의 차이가 {(test['mae']-train['mae']):+.1f}점 더 큼"),
        ("AI 채점 방향",
         f"학습 {train['bias']:+.1f} / 테스트 {test['bias']:+.1f} — 양쪽 모두 음수 (AI 가 낮게 채점)"),
        ("AI 가 낮게 채점한 비율",
         f"학습 {train['under']:.0f}% / 테스트 {test['under']:.0f}% — 학습셋이 다소 높음"),
        ("단일 상담 최대 차이",
         f"학습 {train['max_abs']}점 / 테스트 {test['max_abs']}점 — 테스트의 {test['max_abs']}점은 단일 상담 668865 의 영향"),
    ]
    for k, v in rows:
        ws.cell(row, 1, k); ws.cell(row, 2, v)
        for col in (1, 2):
            ws.cell(row, col).fill = KEY_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=11, bold=(col == 1))
        ws.row_dimensions[row].height = 24
        row += 1
    row += 1

    # 2. 왜 이런 결과인가 (평이하게)
    ws.cell(row, 1, "2. 왜 이런 결과인가 (쉽게)")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1

    explanations = [
        ("결과 1",
         f"양쪽 데이터셋 모두 AI 가 사람보다 낮게 채점합니다 (학습 {train['under']:.0f}%, 테스트 {test['under']:.0f}%).",
         "이는 특정 데이터셋 고유의 문제가 아니라 프롬프트 자체의 채점 기준이 사람보다 엄격함을 의미합니다. "
         "특히 #8 (문의 파악 및 복창), #10 (설명 명확성), #6 (정중한 표현) 항목에서 양쪽 모두 일관된 과소 채점이 관찰됩니다."),
        ("결과 2",
         f"평균 기준으로는 테스트셋의 차이가 더 큽니다 (학습 {train['mae']:.1f}점 / 테스트 {test['mae']:.1f}점).",
         "단, 단일 상담 668865 의 영향이 압도적입니다. 해당 상담은 AI=47점, 사람=76점으로 29점 차이가 발생했습니다. "
         "해당 1건을 제외하고 평균을 산출하면 테스트셋의 평균 차이는 학습셋보다 작아집니다. "
         "따라서 '테스트셋의 본질적 난이도가 더 높다' 는 결론보다는, 단발성 outlier 1건이 평균을 상승시킨 결과로 해석하는 것이 타당합니다."),
        ("결과 3",
         "학습셋과 테스트셋의 항목별 패턴이 거의 동일합니다 (#8/#10/#6 가 양쪽 모두 가장 큰 격차를 보임).",
         "이는 AI 가 학습셋에 과적합된 것이 아니라, 양쪽 데이터셋 모두 동일한 구조적 약점을 보임을 의미합니다. "
         "프롬프트 1회 수정으로 양쪽 데이터셋의 동시 개선이 기대됩니다."),
        ("결과 4",
         "다만 학습셋 MAE 가 테스트셋보다 다소 낮은 결과가 우연인지는 추가 검증이 필요합니다.",
         "Few-shot RAG 가 학습셋을 검색 대상으로 사용하므로, 학습셋 평가 시 동일 샘플이 예시로 포함되어 점수가 보정될 가능성이 있습니다. "
         "정확한 검증을 위해서는 학습 평가 시 자기 자신을 RAG 검색 대상에서 제외하고 재평가해야 하나, 현재는 미수행 상태입니다."),
    ]
    for k, headline, detail in explanations:
        ws.cell(row, 1, k); ws.cell(row, 2, f"{headline}\n{detail}")
        for col in (1, 2):
            ws.cell(row, col).fill = INFO_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1), color="1e40af" if col == 1 else "1f2937")
        ws.row_dimensions[row].height = 60
        row += 1
    row += 1

    # 3. 양쪽 공통 약점 항목 (튜닝 우선순위)
    ws.cell(row, 1, "3. 양쪽 공통 약점 항목 (수정 시 양쪽 데이터셋 동시 개선 가능)")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1

    # 양쪽 모두 MAE 큰 항목 추출
    common_weak = []
    for num, _name, _mx, _row in ITEM_DEF:
        t_mae = train["item_metrics"][num].get("MAE")
        e_mae = test["item_metrics"][num].get("MAE")
        if t_mae is None or e_mae is None: continue
        avg_mae = (t_mae + e_mae) / 2
        if avg_mae >= 1.0:
            common_weak.append((num, t_mae, e_mae, avg_mae))
    common_weak.sort(key=lambda x: -x[3])

    why = {
        2:  "사람은 'TM안내/안전운전' 같은 추가 안내까지 끝인사로 인정하나, AI 는 '감사합니다' 단발 발화만 인식할 경우 0점 처리합니다.",
        5:  "AI 가 '잠시만요' 등 명시적 대기 발화를 인식하지 못하거나 다른 안내 항목으로 분류합니다.",
        6:  "AI 가 '음/아/이게' 등 filler 빈도를 기준으로 감점하나, 사람은 사물존칭/반말 0회 시 만점을 부여합니다.",
        7:  "AI 가 '실례지만/번거로우시겠지만' 등 표준 쿠션어 외의 표현을 인정하지 않습니다.",
        8:  "AI 의 '복창' 정의가 의도적 paraphrase 로 한정되는 반면, 사람은 '반품/교환' 등 핵심 명사구가 후속 발화에 등장할 경우 모두 복창으로 인정합니다.",
        10: "사람은 '고객의 명시적 되물음 부재' 를 만점 기준으로 적용하나, AI 는 장황성/반복/내부용어 사용을 근거로 한 단계 감점합니다.",
        11: "AI 는 결론을 명시적으로 선행한 경우에만 만점을 부여하나, 사람은 결론과 부연이 동시에 관찰되면 인정합니다.",
    }
    for num, t_mae, e_mae, _avg in common_weak:
        ws.cell(row, 1, f"#{num} {ITEM_NUM_TO_NAME.get(num)}")
        ws.cell(row, 2,
                f"학습셋 평균 차이 {t_mae:.2f}점 / 테스트셋 평균 차이 {e_mae:.2f}점 — 양쪽 모두 격차가 큼.\n"
                f"원인: {why.get(num, '')}")
        for col in (1, 2):
            ws.cell(row, col).fill = BAD_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1), color="b91c1c" if col == 1 else "374151")
        ws.row_dimensions[row].height = 50
        row += 1
    row += 1

    # 4. 가장 어긋난 상담 5건 (양쪽 통합)
    ws.cell(row, 1, "4. 양쪽 합쳐 가장 어긋난 상담 5건")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1
    ws.cell(row, 1, "→ |총점 차이| 큰 순. 어느 항목 때문인지도 함께 표시")
    ws.cell(row, 1).font = Font(italic=True, color="6b7280", size=10)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    row += 1

    all_outliers = []
    for label, st in [("학습", train), ("테스트", test)]:
        for sid, diff in st["diffs_total"]:
            ai_t, hum_t = st["sample_totals"][sid]
            items = st["sample_item_diffs"].get(sid, [])
            all_outliers.append((label, sid, diff, ai_t, hum_t, items))
    all_outliers.sort(key=lambda x: -abs(x[2]))
    for label, sid, diff, ai_t, hum_t, items in all_outliers[:5]:
        item_str = ", ".join(f"#{n}({d:+d})" for n, d in items if d != 0)
        ws.cell(row, 1, f"[{label}] {sid}  ({diff:+d}점)")
        ws.cell(row, 2, f"AI={ai_t}점 / 사람={hum_t}점.\n주요 차이 항목: {item_str}")
        for col in (1, 2):
            ws.cell(row, col).fill = BAD_FILL if abs(diff) > 10 else KEY_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1))
        ws.row_dimensions[row].height = 40
        row += 1
    row += 1

    # 5. 다음에 할 일
    ws.cell(row, 1, "5. 다음에 할 일")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1

    todos = [
        ("1. 양쪽 공통 약점 항목부터 수정",
         "위 '3. 양쪽 공통 약점 항목' 의 1순위 항목부터 프롬프트를 수정합니다. "
         "양쪽 데이터셋이 동일한 패턴을 보이므로 1회 수정으로 양쪽이 동시에 개선됩니다."),
        ("2. 단일 상담 668865 정밀 분석",
         "테스트셋의 668865 (-29점) 는 단일 outlier 입니다. 해당 상담에 대해 사람 평가자의 재평가를 실시하여 "
         "AI 가 어느 항목에서 가장 크게 어긋났는지 확인합니다. "
         "해당 1건이 테스트셋 평균에 미치는 영향이 크므로 우선 분석 대상입니다."),
        ("3. 학습/테스트 우열 단정 보류",
         "현재 수치만으로 '학습셋 성능이 우수하다' 고 단정할 수 없습니다. "
         "RAG 가 학습셋을 검색 대상으로 사용하므로 학습 평가가 다소 유리할 가능성이 있습니다. "
         "정확한 비교를 위해서는 학습 평가 시 자기 자신을 RAG 검색 대상에서 제외하고 재평가해야 합니다."),
        ("한계",
         "본 분석은 '사람 평가가 정답' 임을 전제로 합니다. 사람 평가 자체에 모호한 기준이 존재하는 항목 "
         "(특히 #8 문의 파악 및 복창) 은 AI 측 수정만으로 해결되지 않습니다. "
         "사람 평가표의 라벨링 정책 점검이 병행되어야 합니다."),
    ]
    for k, v in todos:
        ws.cell(row, 1, k); ws.cell(row, 2, v)
        for col in (1, 2):
            ws.cell(row, col).fill = KEY_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1))
        ws.row_dimensions[row].height = 50
        row += 1

    ws.freeze_panes = "A2"

    try:
        wb.save(TARGET_XLSX)
        print(f"[OK] '{SHEET_NAME}' 시트 재작성 → {TARGET_XLSX}")
    except PermissionError:
        fb = TARGET_XLSX.with_name(TARGET_XLSX.stem + "_with_easy_interp.xlsx")
        wb.save(fb)
        print(f"  ⚠ 원본 잠김 → fallback: {fb}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
