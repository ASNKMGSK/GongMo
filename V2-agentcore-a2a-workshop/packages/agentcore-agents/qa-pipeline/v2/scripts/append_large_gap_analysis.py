# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""기존 비교분석 xlsx 에 '큰차이 심층분석' 시트 1개를 추가.

대상 파일:
  - C:/.../학습셋_비교분석_<ts>/비교분석_<ts>.xlsx
  - C:/.../테스트셋_비교분석_<ts>/비교분석_<ts>.xlsx

분석 대상: |AI_score - GT_score| >= 3 인 (sample, item) 쌍 모두.

새 시트 컬럼:
  sample_id | # | 항목명 | 만점 | AI | 사람 | Δ | AI 판정 요약 | AI 근거 | AI 감점 사유 |
  사람 비고 (xlsx F열) | 패턴 가설 | 파일

실행:
  python -m v2.scripts.append_large_gap_analysis
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_PIPELINE_DIR = Path(__file__).resolve().parents[2]
if str(_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(_PIPELINE_DIR))

from v2.scripts.compare_learning_set_vs_xlsx import (  # type: ignore[import-untyped]
    ITEM_DEF, ITEM_NUM_TO_MAX, ITEM_NUM_TO_NAME,
    TEST_IDS, TRAINING_IDS,
)


XLSX_GT = Path(r"C:\Users\META M\Desktop\QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
TRAIN_DIR = Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409")
TEST_DIR = Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545")
SAMPLES_TRAIN = Path(r"C:\Users\META M\Desktop\qa 샘플\학습셋")
SAMPLES_TEST = Path(r"C:\Users\META M\Desktop\qa 샘플\테스트셋")

LARGE_DIFF_THRESHOLD = 3
SHEET_NAME = "큰차이 심층분석"


# ───────────────────────────────────────────────────────────
# 데이터 로더
# ───────────────────────────────────────────────────────────


def load_gt_with_notes(xlsx_path: Path, ids: list[str]) -> dict[str, dict[int, dict[str, Any]]]:
    """{sid: {item_num: {score, note}}}"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[str, dict[int, dict[str, Any]]] = {}
    allowed = set(ids)
    for sn in wb.sheetnames:
        tail = sn.strip().split("_")[-1]
        if not (tail.isdigit() and len(tail) == 6 and tail in allowed):
            continue
        ws = wb[sn]
        sample: dict[int, dict[str, Any]] = {}
        for num, _name, _mx, row in ITEM_DEF:
            score = ws.cell(row, 5).value
            note = ws.cell(row, 6).value
            if score is None:
                continue
            try:
                score_int = int(score)
            except (TypeError, ValueError):
                continue
            sample[num] = {
                "score": score_int,
                "note": (str(note).strip() if isinstance(note, str) else "") or "",
            }
        out[tail] = sample
    wb.close()
    return out


def load_ai_full(results_dir: Path, ids: list[str]) -> dict[str, dict[int, dict[str, Any]]]:
    """{sid: {item_num: {score, evidence, judgment, deductions}}}"""
    out: dict[str, dict[int, dict[str, Any]]] = {}
    allowed = set(ids)
    for jp in sorted(results_dir.glob("*_result.json")):
        sid = jp.stem.replace("_result", "")
        if sid not in allowed:
            continue
        try:
            data = json.loads(jp.read_text(encoding="utf-8"))
        except Exception:
            continue
        sample: dict[int, dict[str, Any]] = {}
        for ev in data.get("evaluations", []) or []:
            inner = ev.get("evaluation") if isinstance(ev.get("evaluation"), dict) else ev
            if not isinstance(inner, dict):
                continue
            num = inner.get("item_number")
            if not isinstance(num, int):
                continue
            score = inner.get("score")
            if not isinstance(score, (int, float)):
                continue
            ev_lines: list[str] = []
            for it in inner.get("evidence") or []:
                if isinstance(it, str):
                    if it.strip():
                        ev_lines.append(it.strip())
                elif isinstance(it, dict):
                    speaker = (it.get("speaker") or it.get("role") or "").strip()
                    turn = it.get("turn") if it.get("turn") is not None else it.get("turn_id")
                    text = (it.get("text") or it.get("quote") or "").strip()
                    if not text:
                        continue
                    prefix = ""
                    if turn is not None:
                        prefix += f"[T{turn}] "
                    if speaker:
                        prefix += f"{speaker}: "
                    ev_lines.append(prefix + text)
            ded_lines: list[str] = []
            for d in inner.get("deductions") or []:
                if isinstance(d, dict):
                    reason = d.get("reason") or d.get("rule") or d.get("description") or ""
                    pts = d.get("points_lost", d.get("points", 0))
                    if reason:
                        ded_lines.append(f"-{pts}점: {reason}")
            sample[int(num)] = {
                "score": int(score),
                "evidence": ev_lines,
                "judgment": (inner.get("judgment") or inner.get("summary") or "").strip(),
                "deductions": ded_lines,
            }
        out[sid] = sample
    return out


def load_transcript_filename_map(samples_dir: Path) -> dict[str, str]:
    """sid → 원본 transcript 파일명 (참조용)."""
    out: dict[str, str] = {}
    for jp in samples_dir.glob("*.json"):
        sid = jp.stem.split("_")[0]
        if sid.isdigit() and len(sid) == 6:
            out[sid] = jp.name
    return out


# ───────────────────────────────────────────────────────────
# 패턴 가설
# ───────────────────────────────────────────────────────────


def infer_hypothesis(item_num: int, ai_score: int, gt_score: int,
                     ai_evidence: list[str], gt_note: str,
                     ai_judgment: str) -> str:
    """rule 기반 가설 — 차이 패턴 + 항목별 휴리스틱."""
    diff = ai_score - gt_score
    direction = "AI 과소" if diff < 0 else "AI 과대"
    pieces: list[str] = []

    # 항목별 알려진 패턴
    item_pattern = {
        2:  "끝인사: 사람은 'TM안내/안전운전' 부분 안내까지 인정, AI 는 '감사합니다' 단일 발화만 보면 0점 처리 경향",
        4:  "호응 및 공감: 사람은 명시적 공감 발화 없으면 0점, AI 는 '네/예'만으로도 인정 경향 (과대)",
        5:  "대기 멘트: 사람은 '잠시만요/확인 도와드리겠습니다' 등 명시적 대기 발화 인정, AI 는 발화 누락 또는 다른 안내로 분류",
        7:  "쿠션어: 사람은 '실례지만/번거로우시겠지만' 등 표준 쿠션어 기준 엄격, AI 는 '혹시/우선' 등 약한 쿠션어 미인정 (과소)",
        8:  "문의 파악·복창: 가장 빈번한 불일치. 사람은 '~인 거 맞으세요?' 같은 부분 복창도 인정, AI 는 핵심 명사구 누락 시 0점 처리",
        10: "설명의 명확성: 10점 만점이라 차이도 큼. 사람은 부분 정확성 7점도 줌, AI 는 두괄식+근거 모두 만족해야 만점",
        11: "두괄식 답변: 사람은 결론 먼저 + 후속 부연이 보이면 인정, AI 는 결론 위치 strict 판정",
        14: "사후 안내: '추가 문의 사항' 멘트 인정 범위 차이. AI 는 표준 클로징 패턴만 인정",
    }
    if item_num in item_pattern:
        pieces.append(item_pattern[item_num])

    # AI 근거 vs 사람 비고 충돌 진단
    if ai_score == 0 and gt_score > 0:
        if not ai_evidence:
            pieces.append("AI 가 evidence 를 못 찾음 → 화자/구간 매칭 실패 의심 (Layer 1 turn assignment 또는 evidence_refiner 필터링)")
        elif "키워드" in ai_judgment or "누락" in ai_judgment:
            pieces.append("AI 가 키워드 기반 판정으로 0점 → 사람은 표현 변형도 인정")
    elif ai_score > 0 and gt_score == 0:
        pieces.append("AI 가 적극 인정 → 사람 기준이 더 엄격 (rubric drift, 사람이 보수적)")
    elif diff <= -3:
        pieces.append(f"AI 가 사람 대비 {abs(diff)}점 과소 → 프롬프트 rubric 의 '0점 처리' 조건이 너무 광범위")
    elif diff >= 3:
        pieces.append(f"AI 가 사람 대비 {diff}점 과대 → 인정 임계가 너무 관대")

    # GT note 가 풍부한지
    if gt_note and len(gt_note) > 50:
        pieces.append("(사람 비고 풍부 — 세부 근거 명시되어 있어 비교 가능)")
    elif not gt_note:
        pieces.append("(사람 비고 없음 — 점수만 부여, 근거 추정 어려움)")

    return " · ".join(pieces) or f"{direction} {abs(diff)}점"


# ───────────────────────────────────────────────────────────
# 시트 작성
# ───────────────────────────────────────────────────────────


_HEADERS = [
    ("sample_id", 14),
    ("#", 5),
    ("항목명", 22),
    ("만점", 6),
    ("AI", 6),
    ("사람", 6),
    ("Δ", 6),
    ("AI 판정 요약", 50),
    ("AI 근거", 70),
    ("AI 감점 사유", 50),
    ("사람 비고 (xlsx F열)", 70),
    ("패턴 가설 / 차이 원인", 70),
    ("transcript 파일", 60),
]

_HEADER_FILL = PatternFill("solid", fgColor="1f2937")
_HEADER_FONT = Font(color="fbbf24", bold=True, size=10)
_DELTA_NEG_FILL = PatternFill("solid", fgColor="dbeafe")  # AI 과소
_DELTA_POS_FILL = PatternFill("solid", fgColor="fee2e2")  # AI 과대
_WRAP = Alignment(wrap_text=True, vertical="top")


def write_sheet(xlsx_path: Path, results_dir: Path, samples_dir: Path,
                ids: list[str], dataset_label: str) -> int:
    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 1)  # 요약 다음에 위치

    # 헤더
    for col, (name, width) in enumerate(_HEADERS, start=1):
        c = ws.cell(1, col, name)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # 데이터 로드
    gt = load_gt_with_notes(XLSX_GT, ids)
    ai = load_ai_full(results_dir, ids)
    fname_map = load_transcript_filename_map(samples_dir)

    rows: list[dict[str, Any]] = []
    for sid in sorted(set(gt) & set(ai)):
        for num, _name, mx, _row in ITEM_DEF:
            g = gt[sid].get(num); a = ai[sid].get(num)
            if not g or not a:
                continue
            diff = a["score"] - g["score"]
            if abs(diff) < LARGE_DIFF_THRESHOLD:
                continue
            rows.append({
                "sid": sid, "num": num, "name": ITEM_NUM_TO_NAME[num], "max": mx,
                "ai": a["score"], "human": g["score"], "diff": diff,
                "ai_judgment": a["judgment"], "ai_evidence": a["evidence"],
                "ai_deductions": a["deductions"], "gt_note": g["note"],
                "fname": fname_map.get(sid, ""),
            })

    # |diff| 큰 순 → 같으면 sid 순
    rows.sort(key=lambda r: (-abs(r["diff"]), r["sid"], r["num"]))

    for i, r in enumerate(rows, start=2):
        hypo = infer_hypothesis(r["num"], r["ai"], r["human"],
                                r["ai_evidence"], r["gt_note"], r["ai_judgment"])
        ev_text = "\n".join(f"• {ln}" for ln in r["ai_evidence"]) if r["ai_evidence"] else "(없음)"
        ded_text = "\n".join(r["ai_deductions"]) if r["ai_deductions"] else "(없음)"

        cells = [
            r["sid"], r["num"], r["name"], r["max"],
            r["ai"], r["human"], (f"+{r['diff']}" if r["diff"] > 0 else str(r["diff"])),
            r["ai_judgment"] or "(없음)",
            ev_text, ded_text, r["gt_note"] or "(없음)",
            hypo, r["fname"],
        ]
        fill = _DELTA_POS_FILL if r["diff"] > 0 else _DELTA_NEG_FILL
        for col, val in enumerate(cells, start=1):
            c = ws.cell(i, col, val)
            c.alignment = _WRAP
            if col == 7:
                c.fill = fill
                c.font = Font(bold=True, size=11,
                              color="b91c1c" if r["diff"] > 0 else "1d4ed8")
        ws.row_dimensions[i].height = max(60, min(220, 14 + 14 * (
            len(r["ai_evidence"]) + len((r["gt_note"] or "").split("\n")) + len(r["ai_deductions"])
        )))

    # 안내 행 — 데이터 위 1줄 (제목 행 사이는 freeze 라 못 끼움. 시트 마지막에 메모)
    last = len(rows) + 3
    ws.cell(last, 1, f"[{dataset_label}] |Δ|≥{LARGE_DIFF_THRESHOLD} 항목 {len(rows)}건 | "
                     f"AI 과소(파랑) {sum(1 for r in rows if r['diff']<0)}건 / "
                     f"AI 과대(빨강) {sum(1 for r in rows if r['diff']>0)}건")
    ws.cell(last, 1).font = Font(italic=True, color="64748b", size=10)
    ws.merge_cells(start_row=last, end_row=last, start_column=1, end_column=len(_HEADERS))

    try:
        wb.save(xlsx_path)
        return len(rows)
    except PermissionError:
        # Excel 등이 파일을 열고 있으면 동일 디렉토리에 _with_gap.xlsx 로 폴백
        fallback = xlsx_path.with_name(xlsx_path.stem + "_with_gap.xlsx")
        wb.save(fallback)
        print(f"  ⚠ 원본 파일 잠김 → 별도 파일 저장: {fallback}")
        return len(rows)


def main() -> int:
    targets = [
        (Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409\비교분석_20260422_011144.xlsx"),
         TRAIN_DIR, SAMPLES_TRAIN, TRAINING_IDS, "학습셋"),
        (Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545\비교분석_20260422_011144.xlsx"),
         TEST_DIR, SAMPLES_TEST, TEST_IDS, "테스트셋"),
    ]
    for xlsx_path, results_dir, samples_dir, ids, label in targets:
        if not xlsx_path.exists():
            print(f"[SKIP] xlsx 없음: {xlsx_path}")
            continue
        n = write_sheet(xlsx_path, results_dir, samples_dir, ids, label)
        print(f"[{label}] '{SHEET_NAME}' 시트 {n} 행 → {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
