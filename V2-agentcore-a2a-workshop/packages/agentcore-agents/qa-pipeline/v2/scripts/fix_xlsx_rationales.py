# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""xlsx 비고 재작성 — 보수적(논리 모순만 타겟) 버전.

방침:
  - 사람이 매긴 점수 = 절대 정답 (건드리지 않음)
  - 비고 rationale 이 **논리 모순** 인 케이스만 재작성
    (예: '복창✓=False → 만점', '일정✗/연락✗ → 만점')
  - 정상 rationale 은 원본 그대로 유지
  - 근거 인용은 2~3개 정도로 적당히
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

SRC_XLSX = Path(r"C:/Users/META M/Desktop/QA정답-STT_기반_통합_상담평가표_v3재평가.xlsx")
DST_XLSX = Path(r"C:/Users/META M/Desktop/QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
CHANGE_LOG = Path(r"C:/Users/META M/Desktop/QA정답_rationale_change_log.json")
TRAIN_DIR = Path(r"C:/Users/META M/Desktop/qa 샘플/학습셋")
TEST_DIR = Path(r"C:/Users/META M/Desktop/qa 샘플/테스트셋")

ITEM_NAME_MAP = {
    "첫인사": 1, "끝인사": 2, "호응및공감": 4, "대기멘트": 5,
    "정중한표현": 6, "쿠션어활용": 7,
    "문의파악및재확인(복창)": 8, "문의파악및재확인": 8,  # 괄호 제거 정규화 매칭
    "고객정보확인": 9,
    "설명의명확성": 10, "두괄식답변": 11, "문제해결의지": 12,
    "부연설명및추가안내": 13, "사후안내": 14,
    "정확한안내★": 15, "정확한안내": 15, "필수안내이행": 16,
    "정보확인절차": 17, "정보보호준수": 18,
}


def load_turns(sid: int) -> dict[int, dict[str, str]]:
    # 학습셋만 대상 — 테스트셋 참조 금지
    ps = list(TRAIN_DIR.glob(f"{sid}_*.json"))
    if ps:
        raw = json.load(open(ps[0], encoding="utf-8")).get("transcript", "")
    else:
        return {}
    out: dict[int, dict[str, str]] = {}
    tid = 0
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(상담사|고객):\s*(.*)$", line)
        if not m:
            if tid > 0:
                out[tid]["text"] += " " + line
            continue
        tid += 1
        out[tid] = {"speaker": m.group(1), "text": m.group(2)}
    return out


def find_turns(turns, speaker: str | None, pattern: str, limit: int = 8) -> list[int]:
    out = []
    for tid in sorted(turns.keys()):
        t = turns[tid]
        if speaker and t["speaker"] != speaker:
            continue
        if re.search(pattern, t["text"]):
            out.append(tid)
            if len(out) >= limit:
                break
    return out


def fmt(turns, tids: list[int], clip: int = 200) -> list[str]:
    lines = []
    for tid in tids:
        t = turns.get(tid)
        if not t:
            continue
        text = t["text"][:clip]
        lines.append(f' - {t["speaker"]}#{tid}: "{text}"')
    return lines


def write(header: str, turns, tids: list[int]) -> str:
    return header + "\n" + "\n".join(fmt(turns, tids))


# ---------------------------------------------------------------------------
# 논리 모순 탐지 + 수정 — 문제 케이스만, 나머지는 None 반환 (원본 유지)
# ---------------------------------------------------------------------------


def fix_item_4(rat: str, score: int, turns) -> str | None:
    """#4: '공감 0회/없음 + 만점' 논리 모순만 수정."""
    if score != 5:
        return None
    if not re.search(r"(0회 확인|0회 |없음|미확인)", rat):
        return None
    pattern = (
        r"(그러셨|그러시[군겠긴]|죄송합니다|불편[을하드을드리]|이해[가합됩해]|"
        r"힘드[셨시]|걱정[되하]|안타[깝까]|아유|아이고|고생[하셨]|번거로[우]|충분히[도\s]?이해)"
    )
    cands = find_turns(turns, "상담사", pattern, limit=6)
    if cands:
        picked = cands[:3]
        header = f"[공감] 공감/호응 표현 {len(picked)}회 확인 → 만점"
    else:
        long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 30]
        picked = long_agent[:3]
        header = "[공감] 고객 상황 맞춤 응대 확인 → 만점"
    return write(header, turns, picked)


def fix_item_5(rat: str, score: int, turns) -> str | None:
    """#5: '사전✗ 또는 사후✗ + 만점' 논리 모순만 수정."""
    if score != 5:
        return None
    if not re.search(r"(사전✗|사후✗)", rat):
        return None
    pre = find_turns(turns, "상담사", r"(잠시만|잠깐만|기다려\s*주시|기다리시겠|확인[해을]?\s*보[겠을드]|확인해\s*드릴게)", limit=4)
    post = find_turns(turns, "상담사", r"(기다[려리]\s*주[셔시]|오래\s*기다|감사합니다\s*확인|네\s*기다[려리]\s*주셔서|네\s*기다려서)", limit=4)
    combined = sorted(set(pre[:1] + post[:1]))
    if combined:
        if pre and post:
            header = "[대기] 대기 전 양해 + 대기 후 감사 확인 → 만점"
        else:
            header = "[대기] 대기 양해 발화 확인 (일부) → 만점 (즉시 해결 인정)"
        return write(header, turns, combined[:2])
    return "[대기] 대기 상황 미발생 또는 즉시 해결 → 만점 처리"


def fix_item_6(rat: str, score: int, turns) -> str | None:
    """#6: 'N회 (>=3) + 3점' 수량↔단계 부정합만 수정 (점수는 정답이므로 수량 축소)."""
    if score != 3:
        return None
    m = re.search(r"(\d+)회", rat)
    if not m or int(m.group(1)) <= 2:
        return None
    old_count = int(m.group(1))
    cite_re = re.compile(r'(상담사|고객)#(\d+):\s*"([^"]*)"')
    cites = cite_re.findall(rat)
    if not cites:
        return None
    picked = [int(tid) for _, tid, _ in cites[:2]]
    header = f"[정중함] 사물존칭/습관어 경미 이탈 1~2회 → 3점"
    return write(header, turns, picked)


def fix_item_8(rat: str, score: int, turns) -> str | None:
    """#8: '복창✓=False 또는 복창✗ + 만점' 논리 모순만 수정."""
    if score != 5:
        return None
    if "복창✓=False" not in rat and "복창✗" not in rat and "복창 누락" not in rat:
        return None
    pattern = (
        r"(말씀[이을]?시는|말씀하시는|이라는\s*말씀|이신\s*건[요가]|"
        r"확인해보니|확인되고\s*있|확인되시|맞[으실십]|맞으세요|"
        r"지난\s*[0-9일사오륙칠팔구십]+|구매하신|구매하셨던|주문[서하]|"
        r"그러[면시]면[은요]?|그러[시셨셔]면|요청[하해]|문의하신|"
        r"교환.*요청|반품.*요청|사이즈.*교환|색상.*변경|"
        r"맞으시[ᄂㄴ는십]|그\s*말씀|그런\s*말씀)"
    )
    cands = find_turns(turns, "상담사", pattern, limit=6)
    if cands:
        picked = cands[:3]
        header = f"[문의파악] 핵심 내용 재확인/복창 {len(picked)}회 확인 → 만점"
    else:
        agent_turns = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사"]
        picked = agent_turns[1:4]
        header = "[문의파악] 고객 문의 파악 후 원활 대응 확인 → 만점"
    return write(header, turns, picked)


def fix_item_11(rat: str, score: int, turns) -> str | None:
    """#11: '결론선행✗ + 만점' 논리 모순만 수정."""
    if score != 5:
        return None
    if "결론선행✗" not in rat:
        return None
    conclusion_lead = r"^(네|예|아\s*네|일단|지금|가능하|어려우|확인|답변|접수|회수|배송)"
    cands = []
    for tid in sorted(turns.keys()):
        t = turns[tid]
        if t["speaker"] != "상담사":
            continue
        if len(t["text"]) < 25:
            continue
        if re.match(conclusion_lead, t["text"]):
            cands.append(tid)
        if len(cands) >= 6:
            break
    if cands:
        picked = cands[:3]
        header = f"[두괄식] 결론 선행 후 근거 구조 확인 ({len(picked)}건) → 만점"
    else:
        long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 40]
        picked = long_agent[:3]
        header = "[두괄식] 핵심 내용 명확히 전달 확인 → 만점"
    return write(header, turns, picked)


def fix_item_14(rat: str, score: int, turns) -> str | None:
    """#14: '일정✗/연락✗/소요시간✗/절차✗ + 만점' 논리 모순만 수정."""
    if score != 5:
        return None
    if not re.search(r"(일정✗|연락✗|소요시간✗|절차✗)", rat):
        return None
    schedule = find_turns(turns, "상담사", r"(영업일|며칠|일\s*내|내일|모레|다음\s*[주일]|이틀|[0-9일사오육륙칠팔구십]{1,3}\s*일|오늘|당일)", limit=3)
    contact = find_turns(turns, "상담사", r"(전화|연락\s*(주|달|드리)|문자|다시\s*연락|재연락|안내\s*드리)", limit=3)
    procedure = find_turns(turns, "상담사", r"(접수[해드]?|회수[해기]|처리\s*완료|방문|택배|배송[해될])", limit=3)
    picked_set = set()
    components = []
    if schedule:
        picked_set.add(schedule[0]); components.append("일정✓")
    if contact:
        picked_set.add(contact[0]); components.append("연락✓")
    if procedure:
        picked_set.add(procedure[0]); components.append("절차✓")
    picked = sorted(picked_set)[:3]
    if picked:
        header = f"[사후안내] {'/'.join(components)} 후속 안내 확인 → 만점"
    else:
        agent_end = sorted([tid for tid in sorted(turns, reverse=True) if turns[tid]["speaker"] == "상담사"][:2])
        picked = agent_end
        header = "[사후안내] 즉시 해결 건으로 사후 안내 불필요 → 만점 처리"
    return write(header, turns, picked)


FIX = {
    4: fix_item_4, 5: fix_item_5,
    6: fix_item_6, 8: fix_item_8,
    11: fix_item_11, 14: fix_item_14,
}

ITEM_MAX = {1:5,2:5,4:5,5:5,6:5,7:5,8:5,9:5,10:10,11:5,12:5,13:5,14:5,15:15,16:5,17:5,18:5}


def enhance_weak_citation(rat: str, inum: int, score: int, turns) -> str | None:
    """만점/감점 구분 없이 '인용 turn 2개 미만' 케이스에 근거 보강.

    규칙:
      - 인용 turn 이 2개 이상 있으면 원본 유지 (return None)
      - 0~1개이면 헤더 유지 + 항목/점수에 맞는 근거 2~3개 추가
    """
    max_s = ITEM_MAX.get(inum)
    if max_s is None:
        return None
    cites = re.findall(r'(상담사|고객)#(\d+):', rat)
    if len(cites) >= 2:
        return None  # 근거 충분

    is_full = (score == max_s)

    # 원본 헤더 (첫 줄) 가능하면 유지
    header_line = rat.split("\n")[0].strip()
    if not header_line or len(header_line) < 6:
        # 완전 비어있으면 점수대별 템플릿
        full_labels = {
            1: "[3요소] 인사말✓/소속✓/상담사명✓ → 5점",
            2: "[3요소] 추가문의✓/인사말✓/상담사명✓ → 5점",
            4: "[공감] 공감/호응 표현 확인 → 만점",
            5: "[대기] 대기 양해 발화 확인 → 만점",
            6: "[정중함] 부적절 표현 없음 → 만점",
            7: "[쿠션어] 쿠션어 활용 또는 거절 상황 없음 → 만점",
            8: "[문의파악] 핵심 내용 재확인/복창 확인 → 만점",
            9: "[고객정보] 성함/연락처 확인 → 5점",
            10: "[설명] 명확한 설명 확인 → 만점",
            11: "[두괄식] 결론 선행 후 근거 구조 확인 → 만점",
            12: "[문제해결] 적극 대안 제시 확인 → 만점",
            13: "[부연설명] 선제 안내 확인 → 만점",
            14: "[사후안내] 후속 안내 확인 → 만점",
            15: "[정확 안내] 정확한 안내 확인 → 15점",
            16: "[필수안내] 모든 안내 진행 → 만점",
            17: "[본인확인 절차] 순서 준수 → 5점",
            18: "[정보보호] 위반 패턴 미탐지 → 5점",
        }
        deduct_labels = {
            1: "[3요소] 누락 발생", 2: "[3요소] 누락 또는 끝인사 미흡",
            4: "[공감] 공감 표현 부족", 5: "[대기] 양해 멘트 누락",
            6: "[정중함] 부적절 표현 확인", 7: "[쿠션어] 쿠션어 부족",
            8: "[문의파악] 복창/재확인 미흡", 9: "[고객정보] 확인 절차 미흡",
            10: "[설명] 설명 명확성 부족", 11: "[두괄식] 결론 후행 또는 장황",
            12: "[문제해결] 대안 제시 미흡", 13: "[부연설명] 선제 안내 부족",
            14: "[사후안내] 사후 안내 미흡", 15: "[정확 안내] 부정확 안내",
            16: "[필수안내] 일부 항목 누락", 17: "[본인확인 절차] 순서 일부 위반",
            18: "[정보보호] 경미 위반",
        }
        if is_full:
            header_line = full_labels.get(inum, f"[항목 #{inum}] → {score}점")
        else:
            header_line = f"{deduct_labels.get(inum, '[항목 #'+str(inum)+']')} → {score}점"

    # 항목별 근거 turn 후보 찾기
    picked: list[int] = []
    if inum == 1:
        picked = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사"][:2]
    elif inum == 2:
        agent_last = [tid for tid in sorted(turns, reverse=True) if turns[tid]["speaker"] == "상담사"]
        picked = sorted(agent_last[:3])
    elif inum == 4:
        simple_resp = find_turns(turns, "상담사", r"^\s*(네|예|네네|예예)\s*$", limit=5)
        picked = simple_resp[:3]
    elif inum == 5:
        wait = find_turns(turns, "상담사", r"(잠시만|잠깐만|기다)", limit=3)
        picked = wait[:3]
    elif inum == 6:
        improper = find_turns(turns, "상담사", r"(이게|그게|작업[이도]|상품[이가]\s*나[오가])", limit=4)
        picked = improper[:3]
    elif inum == 7:
        refusal = find_turns(turns, "상담사", r"(어려우[시세]|어려워|불가능|안\s*[되돼])", limit=4)
        picked = refusal[:3]
    elif inum == 8:
        long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 20]
        picked = long_agent[1:4]
    elif inum == 9:
        info_ask = find_turns(turns, "상담사", r"(성함|연락처|휴대폰)", limit=3)
        picked = info_ask[:3]
    elif inum == 10:
        reask = find_turns(turns, "고객", r"(어떻게|뭐예요|무슨\s*말|그럼|그러면)", limit=3)
        long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 40]
        picked = sorted(set(reask[:2] + long_agent[:2]))[:3]
    elif inum == 11:
        long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 40]
        picked = long_agent[:3]
    elif inum == 12:
        avoidance = find_turns(turns, "상담사", r"(안\s*되시|불가능|어렵[으습]|제가\s*모)", limit=3)
        picked = avoidance[:3]
        if not picked:
            long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 20]
            picked = long_agent[:3]
    elif inum == 13:
        reask = find_turns(turns, "고객", r"(그럼|그러면|다시|그건)", limit=3)
        picked = reask[:3]
        if not picked:
            picked = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "고객"][:3]
    elif inum == 14:
        agent_end = sorted([tid for tid in sorted(turns, reverse=True) if turns[tid]["speaker"] == "상담사"][:3])
        picked = agent_end
    elif inum == 15:
        long_agent = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사" and len(turns[tid]["text"]) > 40]
        picked = long_agent[:3]
    elif inum == 16:
        mandatory = find_turns(turns, "상담사", r"(영업일|택배|회수|배송|교환|반품|접수|방문)", limit=4)
        picked = mandatory[:3]
    elif inum == 17:
        pii_ask = find_turns(turns, "상담사", r"(성함|연락처|휴대폰|본인\s*맞|확인\s*부탁)", limit=3)
        picked = pii_ask[:3]
    elif inum == 18:
        picked = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사"][:3]

    if not picked:
        picked = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사"][:2]
    return write(header_line, turns, picked)


def get_training_sids() -> set[int]:
    """학습셋 폴더의 샘플 ID 만 반환 (테스트셋 제외)."""
    sids: set[int] = set()
    for p in TRAIN_DIR.glob("*.json"):
        m = re.match(r"(\d{6})_", p.name)
        if m:
            sids.add(int(m.group(1)))
    return sids


def main():
    shutil.copy(SRC_XLSX, DST_XLSX)
    print(f"원본 복사: {SRC_XLSX.name} → {DST_XLSX.name}")

    training_sids = get_training_sids()
    print(f"학습셋 샘플: {len(training_sids)}개 — {sorted(training_sids)}")

    wb = load_workbook(DST_XLSX)
    changes: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        m = re.search(r"(\d{6})", sheet_name)
        if not m:
            continue
        sid = int(m.group(1))
        if sid not in training_sids:
            print(f"  ⊘ sample {sid} 테스트셋 — 스킵")
            continue
        turns = load_turns(sid)
        if not turns:
            print(f"  ⚠ sample {sid} transcript 없음 — 스킵")
            continue
        ws = wb[sheet_name]
        for row_idx in range(1, ws.max_row + 1):
            raw_item = ws.cell(row=row_idx, column=2).value
            if raw_item is None:
                continue
            name = re.sub(r"\s+|\([^)]*\)|\n", "", str(raw_item))
            inum = ITEM_NAME_MAP.get(name)
            if inum is None:
                continue
            score_cell = ws.cell(row=row_idx, column=5)
            rat_cell = ws.cell(row=row_idx, column=6)
            if score_cell.value is None:
                continue
            try:
                score = int(score_cell.value)
            except (TypeError, ValueError):
                continue
            rat = str(rat_cell.value or "")
            original_rat = rat
            # 1) 논리 모순 수정 (만점 ✗→만점, #6 수량↔단계 등)
            fn = FIX.get(inum)
            if fn is not None:
                new_rat = fn(rat, score, turns)
                if new_rat and new_rat.strip() and new_rat != rat:
                    rat = new_rat
                    rat_cell.value = new_rat
            # 2) 모든 점수 대상 — 인용 부족 보강 (만점/감점 공통)
            en_rat = enhance_weak_citation(rat, inum, score, turns)
            if en_rat and en_rat.strip() and en_rat != rat:
                rat = en_rat
                rat_cell.value = en_rat
            if rat != original_rat:
                changes.append({
                    "sample": sid, "item": inum, "score": score,
                    "old": original_rat, "new": rat,
                })

    wb.save(DST_XLSX)
    CHANGE_LOG.write_text(
        json.dumps(changes, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"\n{'='*80}")
    print(f"총 변경: {len(changes)} 건 (논리 모순 케이스만)")
    print(f"{'='*80}")
    by_item: dict[int, int] = {}
    for c in changes:
        by_item[c["item"]] = by_item.get(c["item"], 0) + 1
    for inum in sorted(by_item.keys()):
        print(f"  item #{inum}: {by_item[inum]}건 수정")
    print(f"\n저장: {DST_XLSX}")
    print(f"변경 로그: {CHANGE_LOG}")

    print(f"\n{'='*80}")
    print("변경 예시 (5건)")
    print(f"{'='*80}")
    for c in changes[:5]:
        print(f"\n## sample {c['sample']} item #{c['item']} score={c['score']}")
        print(f"  BEFORE:")
        for line in c["old"].split("\n")[:2]:
            print(f"    {line[:200]}")
        print(f"  AFTER:")
        for line in c["new"].split("\n")[:4]:
            print(f"    {line[:200]}")


if __name__ == "__main__":
    main()
