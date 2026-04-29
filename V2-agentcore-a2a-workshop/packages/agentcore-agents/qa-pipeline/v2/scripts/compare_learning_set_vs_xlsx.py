# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""학습셋 배치 결과(JSON) vs 사람 정답 xlsx 비교 리포트.

- 입력 A: 배치 결과 폴더 (각 sample_id_result.json)
- 입력 B: QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx
- 출력: 동일 폴더에 비교분석_<ts>.xlsx + summary.html

지표 (MAE/RMSE/Bias/MAPE/Accuracy 만 사용 — Pearson/κ 등 상관계수 미사용):
  MAE       : mean(|AI - Human|)
  RMSE      : sqrt(mean((AI - Human)^2))
  Bias      : mean(AI - Human)
  MAPE      : mean(|AI - Human| / max_score) * 100  (%)
  Accuracy  : (|AI - Human| <= 0 인 비율) * 100
"""

from __future__ import annotations

import json
import math
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# 항목 번호 → (이름, 배점, xlsx row)
# xlsx 레이아웃: 각 sample 시트의 R6~R22 에 17 개 항목 순서대로
# ⚠ #15(정확한 안내 ★ 15점), #16(필수 안내 이행 5점) = 업무 정확도(20) 대분류 — 비교에서 제외
ITEM_DEF: list[tuple[int, str, int, int]] = [
    (1,  "첫인사",                  5,  6),
    (2,  "끝인사",                  5,  7),
    (4,  "호응 및 공감",            5,  8),
    (5,  "대기 멘트",               5,  9),
    (6,  "정중한 표현",             5, 10),
    (7,  "쿠션어 활용",             5, 11),
    (8,  "문의 파악 및 재확인(복창)", 5, 12),
    (9,  "고객정보 확인",           5, 13),
    (10, "설명의 명확성",          10, 14),
    (11, "두괄식 답변",              5, 15),
    (12, "문제 해결 의지",           5, 16),
    (13, "부연 설명 및 추가 안내",    5, 17),
    (14, "사후 안내",                5, 18),
    # (15, "정확한 안내 ★",           15, 19),  # 제외: 업무 정확도
    # (16, "필수 안내 이행",           5, 20),   # 제외: 업무 정확도
    (17, "정보 확인 절차",           5, 21),
    (18, "정보 보호 준수",           5, 22),
]
EXCLUDED_ITEMS = {15, 16}  # 업무 정확도 대분류

ITEM_NUM_TO_NAME = {n: name for n, name, _, _ in ITEM_DEF}
ITEM_NUM_TO_MAX = {n: mx for n, _, mx, _ in ITEM_DEF}
ITEM_NUM_TO_ROW = {n: row for n, _, _, row in ITEM_DEF}

# 학습셋 14건 / 테스트셋 9건
TRAINING_IDS = [
    "668437", "668451", "668464", "668481", "668488", "668507", "668526",
    "668542", "668605", "668610", "668675", "668697", "668736", "668771",
]
TEST_IDS = [
    "668797", "668847", "668853", "668865", "668899",
    "668916", "668927", "668941", "668963",
]


def _pick_ids_for_dir(results_dir: Path) -> list[str]:
    """결과 폴더 이름 / 파일 내역 으로 샘플ID 집합 자동 결정."""
    dn = results_dir.name
    if "테스트" in dn or "test" in dn.lower():
        return TEST_IDS
    if "학습" in dn or "train" in dn.lower() or "learning" in dn.lower():
        return TRAINING_IDS
    # fallback: 존재하는 결과 json 기반
    present = {p.stem.replace("_result", "") for p in results_dir.glob("*_result.json")}
    return sorted(present)


def load_xlsx_ground_truth(xlsx_path: Path, allowed_ids: list[str]) -> dict[str, dict[int, int]]:
    """{sample_id: {item_num: human_score}} 반환."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    gt: dict[str, dict[int, int]] = {}
    allowed = set(allowed_ids)
    for sn in wb.sheetnames:
        # 시트 이름 마지막 6자리 = sample_id
        tail = sn.strip().split("_")[-1]
        if not tail.isdigit() or len(tail) != 6:
            continue
        if tail not in allowed:
            continue
        ws = wb[sn]
        sample = {}
        for num, name, maxs, row in ITEM_DEF:
            v = ws.cell(row, 5).value  # E열 = 점수
            if v is None:
                continue
            try:
                sample[num] = int(v)
            except (TypeError, ValueError):
                continue
        gt[tail] = sample
    return gt


def load_ai_results(results_dir: Path, allowed_ids: list[str]) -> dict[str, dict[int, int]]:
    """{sample_id: {item_num: ai_score}} 반환."""
    ai: dict[str, dict[int, int]] = {}
    allowed = set(allowed_ids)
    for jp in sorted(results_dir.glob("*_result.json")):
        sid = jp.stem.replace("_result", "")
        if sid not in allowed:
            continue
        try:
            data = json.loads(jp.read_text(encoding="utf-8"))
        except Exception:
            continue
        evals = data.get("evaluations", []) or []
        sample: dict[int, int] = {}
        for e in evals:
            if not isinstance(e, dict):
                continue
            # 결과 구조: 외피 {status, agent_id, evaluation: {item_number, score, ...}}
            inner = e.get("evaluation") if isinstance(e.get("evaluation"), dict) else e
            num = inner.get("item_number")
            score = inner.get("score")
            if num in ITEM_NUM_TO_MAX and isinstance(score, (int, float)):
                sample[int(num)] = int(score)
        ai[sid] = sample
    return ai


def compute_metrics(pairs: list[tuple[int, int, int]]) -> dict[str, Any]:
    """pairs = [(ai, human, max_score), ...]. 측정지표 반환.

    MAE/RMSE/Bias/MAPE: 표준 오차 지표
    MaxAbs: 최대 절대 오차 (단일 worst case)
    Over%: AI > Human 비율 (과대 채점)
    Under%: AI < Human 비율 (과소 채점)
    """
    n = len(pairs)
    if n == 0:
        return {"n": 0, "MAE": None, "RMSE": None, "Bias": None, "MAPE": None,
                "MaxAbs": None, "Over%": None, "Under%": None}
    diffs = [a - h for a, h, _ in pairs]
    abs_diffs = [abs(d) for d in diffs]
    mae = sum(abs_diffs) / n
    rmse = math.sqrt(sum(d * d for d in diffs) / n)
    bias = sum(diffs) / n
    mape = sum(abs(a - h) / m for a, h, m in pairs) / n * 100.0
    max_abs = max(abs_diffs)
    over = sum(1 for d in diffs if d > 0) / n * 100.0
    under = sum(1 for d in diffs if d < 0) / n * 100.0
    return {
        "n": n,
        "MAE": round(mae, 3),
        "RMSE": round(rmse, 3),
        "Bias": round(bias, 3),
        "MAPE": round(mape, 2),
        "MaxAbs": max_abs,
        "Over%": round(over, 2),
        "Under%": round(under, 2),
    }


def build_report(results_dir: Path, xlsx_path: Path) -> Path:
    allowed_ids = _pick_ids_for_dir(results_dir)
    gt = load_xlsx_ground_truth(xlsx_path, allowed_ids)
    ai = load_ai_results(results_dir, allowed_ids)

    matched_ids = sorted(set(gt.keys()) & set(ai.keys()))
    print(f"xlsx 정답 샘플: {len(gt)} / AI 결과 샘플: {len(ai)} / 매칭: {len(matched_ids)}")

    # === 수집 ===
    # 대화별: {sid: {item: (ai, human)}}
    per_sample_pairs: dict[str, list[tuple[int, int, int, int]]] = {}
    # 항목별: {item_num: [(ai, human, max), ...]}
    per_item_pairs: dict[int, list[tuple[int, int, int]]] = defaultdict(list)
    # 전체
    all_pairs: list[tuple[int, int, int]] = []
    # 총점 (AI vs Human) 샘플별
    total_rows: list[tuple[str, int, int, int]] = []  # (sid, ai_total, human_total, max_total)

    for sid in matched_ids:
        a_map = ai[sid]
        h_map = gt[sid]
        sample_pairs: list[tuple[int, int, int, int]] = []
        ai_total = 0
        human_total = 0
        max_total = 0
        for num, _name, maxs, _row in ITEM_DEF:
            if num not in a_map or num not in h_map:
                continue
            a = a_map[num]
            h = h_map[num]
            sample_pairs.append((num, a, h, maxs))
            per_item_pairs[num].append((a, h, maxs))
            all_pairs.append((a, h, maxs))
            ai_total += a
            human_total += h
            max_total += maxs
        per_sample_pairs[sid] = sample_pairs
        total_rows.append((sid, ai_total, human_total, max_total))

    # === xlsx 리포트 생성 ===
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    def _write_headers(ws, headers: list[str]) -> None:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align

    def _diff_fill(delta: int) -> PatternFill | None:
        if delta == 0:
            return good_fill
        if abs(delta) <= 2:
            return warn_fill
        return bad_fill

    # ──────────────────── Sheet 1: Overview ────────────────────
    ws = out_wb.create_sheet("요약")
    overall = compute_metrics(all_pairs)
    total_pairs = [(a, h, m) for _, a, h, m in total_rows]
    total_metrics = compute_metrics(total_pairs)

    ws["A1"] = "학습셋 AI vs 사람 평가 비교 리포트 (업무정확도 #15/#16 제외)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"] = f"매칭 샘플 수: {len(matched_ids)} / 항목 수: {len(ITEM_DEF)} / 총 쌍: {len(all_pairs)}"
    ws["A4"] = f"제외 항목: #15 정확한 안내 ★ (15점) / #16 필수 안내 이행 (5점) — 업무 정확도 대분류 전체"
    ws["A4"].font = Font(italic=True, color="666666")

    ws["A5"] = "■ 총점 지표 (샘플별 총점 기준)"
    ws["A5"].font = Font(bold=True)
    headers_6 = ["n", "MAE", "RMSE", "Bias", "MAPE(%)", "MaxAbs", "Over%", "Under%"]
    for c, h in enumerate(headers_6, 1):
        cell = ws.cell(6, c, h); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
    for c, k in enumerate(["n", "MAE", "RMSE", "Bias", "MAPE", "MaxAbs", "Over%", "Under%"], 1):
        ws.cell(7, c, total_metrics[k])

    ws["A9"] = "■ 지표 해석"
    ws["A9"].font = Font(bold=True)
    ws["A10"] = "MAE: 평균 절대 오차 (낮을수록 좋음). 점수 단위."
    ws["A11"] = "RMSE: 제곱평균 오차 (큰 오차에 민감)."
    ws["A12"] = "Bias: AI - Human 평균 (양수=AI 과대, 음수=AI 과소)."
    ws["A13"] = "MAPE: 총점 대비 오차율."
    ws["A14"] = "MaxAbs: 단일 샘플 최대 절대 총점 오차 (worst case)."
    ws["A15"] = "Over% / Under%: AI 과대 / 과소 채점 샘플 비율."

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # ──────────────────── Sheet 2: 항목별 지표 ────────────────────
    ws = out_wb.create_sheet("항목별 지표")
    _write_headers(ws, ["item#", "항목명", "배점", "n", "MAE", "RMSE", "Bias", "MAPE(%)", "MaxAbs", "Over%", "Under%"])
    r = 2
    for num, name, maxs, _row in ITEM_DEF:
        pairs = per_item_pairs.get(num, [])
        m = compute_metrics(pairs)
        ws.cell(r, 1, num); ws.cell(r, 2, name); ws.cell(r, 3, maxs)
        ws.cell(r, 4, m["n"]); ws.cell(r, 5, m["MAE"]); ws.cell(r, 6, m["RMSE"])
        ws.cell(r, 7, m["Bias"]); ws.cell(r, 8, m["MAPE"])
        ws.cell(r, 9, m["MaxAbs"]); ws.cell(r, 10, m["Over%"]); ws.cell(r, 11, m["Under%"])
        if m["MAE"] is not None:
            fill = good_fill if m["MAE"] <= 0.5 else warn_fill if m["MAE"] <= 1.5 else bad_fill
            ws.cell(r, 5).fill = fill
        r += 1
    for c, w in enumerate([7, 24, 8, 6, 10, 10, 10, 10, 9, 9, 9], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ──────────────────── Sheet 3: 대화별 지표 ────────────────────
    ws = out_wb.create_sheet("대화별 지표")
    _write_headers(ws, [
        "sample_id", "n", "MAE", "RMSE", "Bias", "MAPE(%)", "MaxAbs", "Over%", "Under%",
        "AI총점", "사람총점", "총점Δ", "불일치 항목 (|Δ|≥2)", "경미 불일치 (|Δ|=1)",
    ])
    r = 2
    for sid in matched_ids:
        pairs_for_sid = [(a, h, mx) for _, a, h, mx in per_sample_pairs[sid]]
        m = compute_metrics(pairs_for_sid)
        ai_t = sum(a for a, _, _ in pairs_for_sid)
        hu_t = sum(h for _, h, _ in pairs_for_sid)
        delta = ai_t - hu_t

        # 항목별 Δ 문자열 구성: "#8(-5), #10(-3), ..." 절대값 내림차순
        item_deltas: list[tuple[int, int]] = [
            (num, a - h) for num, a, h, _mx in per_sample_pairs[sid]
        ]
        major = sorted(
            [(n_, d_) for n_, d_ in item_deltas if abs(d_) >= 2],
            key=lambda x: -abs(x[1]),
        )
        minor = sorted(
            [(n_, d_) for n_, d_ in item_deltas if abs(d_) == 1],
            key=lambda x: -abs(x[1]),
        )
        major_str = ", ".join(f"#{n_}({d_:+d})" for n_, d_ in major) or "-"
        minor_str = ", ".join(f"#{n_}({d_:+d})" for n_, d_ in minor) or "-"

        ws.cell(r, 1, sid); ws.cell(r, 2, m["n"])
        ws.cell(r, 3, m["MAE"]); ws.cell(r, 4, m["RMSE"]); ws.cell(r, 5, m["Bias"])
        ws.cell(r, 6, m["MAPE"]); ws.cell(r, 7, m["MaxAbs"])
        ws.cell(r, 8, m["Over%"]); ws.cell(r, 9, m["Under%"])
        ws.cell(r, 10, ai_t); ws.cell(r, 11, hu_t); ws.cell(r, 12, delta)
        ws.cell(r, 13, major_str); ws.cell(r, 14, minor_str)
        ws.cell(r, 13).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(r, 14).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        fill = _diff_fill(delta)
        if fill:
            ws.cell(r, 12).fill = fill
        r += 1
    for c, w in enumerate([12, 6, 9, 9, 9, 10, 8, 8, 8, 9, 10, 9, 45, 35], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ──────────────────── Sheet 4: 대화 × 항목 매트릭스 (AI) ────────────────────
    # 그리고 사람 매트릭스, Δ 매트릭스 — 한 시트에 stacked
    ws = out_wb.create_sheet("대화x항목 매트릭스")
    # 헤더
    headers = ["sample_id", "종류"] + [f"#{n}\n{ITEM_NUM_TO_NAME[n]}" for n, *_ in ITEM_DEF] + ["총점"]
    _write_headers(ws, headers)
    r = 2
    for sid in matched_ids:
        pairs = per_sample_pairs[sid]
        ai_total = sum(a for _, a, _, _ in pairs)
        hu_total = sum(h for _, _, h, _ in pairs)
        # AI row
        ws.cell(r, 1, sid); ws.cell(r, 2, "AI")
        for i, (num, a, h, _mx) in enumerate(pairs, 3):
            ws.cell(r, i, a)
        ws.cell(r, len(headers), ai_total)
        # Human row
        ws.cell(r + 1, 1, sid); ws.cell(r + 1, 2, "Human")
        for i, (num, a, h, _mx) in enumerate(pairs, 3):
            ws.cell(r + 1, i, h)
        ws.cell(r + 1, len(headers), hu_total)
        # Delta row (AI - Human)
        ws.cell(r + 2, 1, sid); ws.cell(r + 2, 2, "Δ")
        for i, (num, a, h, _mx) in enumerate(pairs, 3):
            d = a - h
            ws.cell(r + 2, i, d)
            f = _diff_fill(d)
            if f:
                ws.cell(r + 2, i).fill = f
        delta_total = ai_total - hu_total
        ws.cell(r + 2, len(headers), delta_total)
        f = _diff_fill(delta_total)
        if f:
            ws.cell(r + 2, len(headers)).fill = f
        r += 3
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    for c in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.row_dimensions[1].height = 36

    # ──────────────────── Sheet 5: 불일치 Top ────────────────────
    ws = out_wb.create_sheet("불일치 TOP")
    _write_headers(ws, ["sample_id", "item#", "항목명", "AI", "Human", "Δ (AI-H)", "max"])
    flat: list[tuple[str, int, str, int, int, int, int]] = []
    for sid in matched_ids:
        for num, a, h, mx in per_sample_pairs[sid]:
            flat.append((sid, num, ITEM_NUM_TO_NAME[num], a, h, a - h, mx))
    flat.sort(key=lambda x: -abs(x[5]))  # 절대값 내림차순
    r = 2
    for sid, num, name, a, h, d, mx in flat[:60]:
        ws.cell(r, 1, sid); ws.cell(r, 2, num); ws.cell(r, 3, name)
        ws.cell(r, 4, a); ws.cell(r, 5, h); ws.cell(r, 6, d); ws.cell(r, 7, mx)
        f = _diff_fill(d)
        if f:
            ws.cell(r, 6).fill = f
        r += 1
    for c, w in enumerate([12, 8, 26, 8, 8, 10, 8], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ──────────────────── Sheet 5.5: 해석 노트 ────────────────────
    _write_interpretation_sheet(
        out_wb, matched_ids, overall, total_metrics,
        per_item_pairs, per_sample_pairs, total_rows, hdr_fill, hdr_font, hdr_align,
    )

    # ──────────────────── Sheet 6: 지표 설명 ────────────────────
    ws = out_wb.create_sheet("지표 설명")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 28

    guide = [
        ("항목",        "설명",                                                                          "해석 기준"),
        ("n",          "비교에 사용된 쌍(pair) 개수. 항목-레벨이면 '샘플 × 항목' 수, 대화별이면 항목 수.",    "—"),
        ("MAE",        "Mean Absolute Error. 평균 |AI - 사람| 점수 오차.",                              "0 이상. 낮을수록 좋음. 5점 척도에서 0.5 이하=우수, 1.5 초과=개선필요."),
        ("RMSE",       "Root Mean Squared Error. 오차 제곱 평균의 제곱근. 큰 오차에 민감한 지표.",             "MAE 보다 크면 'outlier(튀는 오차)' 가 존재한다는 신호."),
        ("Bias",       "AI - 사람 평균. 부호 포함 평균 오차. 과대/과소 채점 방향성.",                        "0 근처=균형. 음수=AI 과소 채점, 양수=AI 과대 채점."),
        ("MAPE(%)",    "Mean Absolute Percentage Error. |AI - 사람| / 항목 배점 평균 × 100.",              "배점이 다른 항목끼리 비교할 때 유용. 10% 이하=양호, 30% 이상=심각."),
        ("MaxAbs",     "단일 쌍 최대 절대 오차. 가장 심한 불일치 한 건의 크기.",                             "5점 항목에서 MaxAbs=5 면 '정반대 판정' 이 최소 1건 있음."),
        ("Over%",      "n 대비 AI > 사람 비율 (%). AI 과대 채점 비율.",                                   "Under% 과 합하면 전체 불일치율 (= 100 - 일치율)."),
        ("Under%",     "n 대비 AI < 사람 비율 (%). AI 과소 채점 비율.",                                    "Over% 대비 높으면 AI 가 체계적으로 엄격함을 의미."),
        ("총점Δ",      "샘플 총점 기준 AI - 사람. 대화별 시트 전용.",                                        "|Δ| ≤ 2 ok, 3~10 warn, 그 이상 bad."),
        ("불일치 항목", "|Δ| ≥ 2 인 항목만 절대값 내림차순으로 `#item(±Δ)` 형식 나열.",                        "튜닝 대상 직접 특정용. 각 샘플에서 어느 항목이 문제인지 스냅샷."),
        ("경미 불일치", "|Δ| = 1 인 항목 나열. 튜닝 노이즈 바운더리 참고용.",                                  "여기에 집중하면 rubric drift 위험."),
    ]
    for ridx, row in enumerate(guide, 1):
        for cidx, v in enumerate(row, 1):
            c = ws.cell(ridx, cidx, v)
            if ridx == 1:
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
            else:
                c.alignment = Alignment(vertical="top", wrap_text=True)

    # 주석: 집계 방식
    ws.cell(len(guide) + 2, 1, "■ 집계 레벨 구분").font = Font(bold=True)
    ws.cell(len(guide) + 3, 2, "항목-레벨: 모든 (샘플 × 항목) 쌍을 한 덩어리로 집계. 채점 정확도의 원자 단위 지표.")
    ws.cell(len(guide) + 4, 2, "총점-레벨: 샘플별로 AI/사람 총점을 먼저 합친 뒤, 총점들끼리 오차 지표. 업무상 최종 등급 영향도 평가용.")

    # ──────────────────── 저장 ────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = results_dir / f"비교분석_{ts}.xlsx"

    # 기존 비교분석_*.xlsx / .html 삭제 (최신만 유지)
    for old in list(results_dir.glob("비교분석_*.xlsx")) + list(results_dir.glob("비교분석_*.html")):
        try:
            old.unlink()
        except Exception:
            pass

    out_wb.save(out_path)

    # HTML 요약도 같이
    html_path = results_dir / f"비교분석_{ts}.html"
    _write_html(
        html_path,
        matched_ids=matched_ids,
        overall=overall,
        total_metrics=total_metrics,
        per_item_pairs=per_item_pairs,
        per_sample_pairs=per_sample_pairs,
    )

    print(f"[완료] xlsx: {out_path}")
    print(f"[완료] html: {html_path}")
    return out_path


def _write_interpretation_sheet(
    wb,
    matched_ids: list[str],
    overall: dict,
    total_metrics: dict,
    per_item_pairs: dict[int, list[tuple[int, int, int]]],
    per_sample_pairs: dict[str, list[tuple[int, int, int, int]]],
    total_rows: list[tuple[str, int, int, int]],
    hdr_fill: PatternFill,
    hdr_font: Font,
    hdr_align: Alignment,
) -> None:
    """데이터에서 자동 도출된 해석 노트 시트."""
    ws = wb.create_sheet("해석 노트")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110

    def _sec(r: int, title: str) -> int:
        c = ws.cell(r, 1, title); c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        c.alignment = hdr_align
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        return r + 1

    def _kv(r: int, k: str, v: str) -> int:
        ws.cell(r, 1, k).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(r, 2, v).alignment = Alignment(vertical="top", wrap_text=True)
        return r + 1

    # 항목별 지표 계산 (MAE / Under% / Over%)
    item_metrics = []
    for num, name, maxs, _row in ITEM_DEF:
        pairs = per_item_pairs.get(num, [])
        if not pairs:
            continue
        m = compute_metrics(pairs)
        item_metrics.append((num, name, maxs, m))

    # 대화별 총점Δ 정렬
    sample_deltas = sorted(
        [(sid, ai_t - hu_t, ai_t, hu_t) for sid, ai_t, hu_t, _max in total_rows],
        key=lambda x: abs(x[1]), reverse=True,
    )

    ws["A1"] = "자동 생성 해석 노트 (AI 평가 vs 사람 정답)"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:B1")

    r = 3
    r = _sec(r, "1. 요약 지표 (총점 기준)")
    r = _kv(r, "샘플 수", f"{len(matched_ids)} 건 (업무정확도 #15/#16 제외)")
    r = _kv(r, "총점 MAE", f"{total_metrics['MAE']} (샘플당 평균 AI-사람 총점 차이)")
    r = _kv(r, "총점 Bias", f"{total_metrics['Bias']} (음수=AI 과소, 양수=AI 과대)")
    r = _kv(r, "총점 Under%", f"{total_metrics['Under%']}% (AI < 사람 샘플 비율)")
    r = _kv(r, "총점 Over%", f"{total_metrics['Over%']}% (AI > 사람 샘플 비율)")
    r = _kv(r, "총점 MaxAbs", f"{total_metrics['MaxAbs']} (최악 샘플 총점 괴리)")
    r += 1

    # 해석 텍스트 - Bias 방향성 (총점 기준)
    bias_v = total_metrics["Bias"] or 0
    if bias_v < -1:
        bias_interp = f"AI 가 사람보다 체계적으로 **엄격** (샘플당 평균 {abs(bias_v):.2f}점 적게 부여). 프롬프트의 감점 트리거 재검토 필요."
    elif bias_v > 1:
        bias_interp = f"AI 가 사람보다 체계적으로 **관대** (샘플당 평균 {abs(bias_v):.2f}점 과대 부여). 만점 조건 재검토 필요."
    else:
        bias_interp = "AI 와 사람이 전체적으로 균형. 개별 항목 편차 확인 필요."

    r = _sec(r, "2. 방향성 해석")
    r = _kv(r, "전체 채점 경향", bias_interp)

    # Over vs Under balance (총점 기준)
    over_pct = total_metrics["Over%"] or 0
    under_pct = total_metrics["Under%"] or 0
    if under_pct > over_pct * 2:
        balance = f"Under:Over = {under_pct:.1f}% : {over_pct:.1f}% — AI 과소가 압도적. rubric 해석이 사람보다 보수적."
    elif over_pct > under_pct * 2:
        balance = f"Over:Under = {over_pct:.1f}% : {under_pct:.1f}% — AI 과대가 압도적. 만점 기준이 느슨."
    else:
        balance = f"Over/Under 균형 (Over {over_pct:.1f}%, Under {under_pct:.1f}%) — 방향성 편향 적음."
    r = _kv(r, "과대/과소 균형", balance)
    r += 1

    # 항목별 worst (MAE 기준)
    r = _sec(r, "3. 튜닝 우선순위 항목 (MAE 내림차순 상위 5)")
    worst_items = sorted(item_metrics, key=lambda x: x[3]["MAE"] or 0, reverse=True)[:5]
    for num, name, maxs, m in worst_items:
        tag = ""
        if (m["Under%"] or 0) > 50:
            tag = " [AI 심하게 엄격]"
        elif (m["Over%"] or 0) > 30:
            tag = " [AI 느슨]"
        r = _kv(
            r, f"#{num} {name}",
            f"MAE={m['MAE']}, Bias={m['Bias']}, Under%={m['Under%']}, Over%={m['Over%']}, MaxAbs={m['MaxAbs']}{tag}",
        )
    r += 1

    # 항목별 best
    r = _sec(r, "4. 정확도 우수 항목 (MAE 오름차순 상위 5)")
    best_items = sorted(item_metrics, key=lambda x: x[3]["MAE"] or 0)[:5]
    for num, name, maxs, m in best_items:
        r = _kv(
            r, f"#{num} {name}",
            f"MAE={m['MAE']}, Bias={m['Bias']}, Accuracy 역할 잘 수행 중.",
        )
    r += 1

    # 대화별 outlier
    r = _sec(r, "5. 주요 outlier 대화 (|총점Δ| 내림차순 상위 5)")
    for sid, delta, ai_t, hu_t in sample_deltas[:5]:
        # 해당 sid 의 주요 불일치 항목
        pairs = per_sample_pairs.get(sid, [])
        item_ds = sorted(
            [(num, a - h) for num, a, h, _mx in pairs if abs(a - h) >= 2],
            key=lambda x: -abs(x[1]),
        )
        major_str = ", ".join(f"#{n}({d:+d})" for n, d in item_ds[:5]) or "-"
        r = _kv(r, f"{sid} (Δ{delta:+d})", f"AI={ai_t}, Human={hu_t}. 주요 불일치: {major_str}")
    r += 1

    # 공통 rubric drift - Under% 50% 이상 항목
    drift_items = [(num, name, m) for num, name, _mx, m in item_metrics if (m["Under%"] or 0) >= 50]
    if drift_items:
        r = _sec(r, "6. 구조적 AI 과소 채점 항목 (Under% ≥ 50%)")
        r = _kv(
            r, "해석",
            "이 항목들은 단일 샘플 문제가 아닌 rubric 해석 차이. 프롬프트의 감점 트리거를 재검토해야 함 "
            "(관대 조항 추가 또는 기준 재정의).",
        )
        for num, name, m in drift_items:
            r = _kv(
                r, f"#{num} {name}",
                f"Under%={m['Under%']}, MAE={m['MAE']}, Bias={m['Bias']} — AI 가 {abs(m['Bias']):.2f}점 체계적 과소.",
            )
        r += 1

    # 실행 제안
    r = _sec(r, "7. 실행 제안")
    if worst_items:
        top_item = worst_items[0]
        r = _kv(
            r, "① 최우선 튜닝",
            f"#{top_item[0]} {top_item[1]} (MAE {top_item[3]['MAE']}) — 해당 프롬프트의 판정 기준을 사람 평가 샘플과 비교 검토.",
        )
    r = _kv(
        r, "② 튜닝 검증 방법",
        "프롬프트 수정 후 동일 샘플로 재평가 → MAE/Bias 가 개선되는지 확인. "
        "다른 항목 MAE 가 악화되는 '풍선 효과' 주의.",
    )
    r = _kv(
        r, "③ outlier 샘플 수동 검토",
        f"|Δ|≥10 인 대화는 rubric 해석 충돌 가능성. 최소 상위 3건은 사람 평가자와 재검토 권장.",
    )
    r = _kv(
        r, "④ 한계 인정",
        "본 지표는 '사람 평가'를 정답으로 간주. 사람 평가 자체에 drift 가 있으면 지표 신뢰도 제한. "
        "두 셋(학습/테스트) 비교에서 공통으로 나타나는 패턴만 구조적 문제로 간주해야 함.",
    )


def _write_html(
    html_path: Path,
    matched_ids: list[str],
    overall: dict[str, Any],
    total_metrics: dict[str, Any],
    per_item_pairs: dict[int, list[tuple[int, int, int]]],
    per_sample_pairs: dict[str, list[tuple[int, int, int, int]]],
) -> None:
    def _fmt(v):
        return "-" if v is None else v

    html = ["""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8"/>
<title>학습셋 AI vs 사람 비교 리포트</title>
<style>
body{font-family:'Segoe UI','Malgun Gothic',sans-serif;max-width:1400px;margin:24px auto;padding:0 16px;color:#222}
h1{border-bottom:3px solid #305496;padding-bottom:6px}
h2{margin-top:32px;border-left:5px solid #305496;padding-left:10px}
table{border-collapse:collapse;width:100%;margin:12px 0;font-size:13px}
th{background:#305496;color:#fff;padding:8px 6px;text-align:center;border:1px solid #1f3864}
td{padding:6px;border:1px solid #bbb;text-align:center}
tr:nth-child(even) td{background:#f2f6fc}
.ok{background:#C6EFCE!important}
.warn{background:#FFEB9C!important}
.bad{background:#FFC7CE!important}
.muted{color:#666;font-size:12px}
.pill{display:inline-block;padding:2px 8px;border-radius:10px;background:#305496;color:#fff;font-size:12px;margin-right:4px}
</style></head><body>"""]
    html.append(f"<h1>학습셋 AI vs 사람 비교 리포트</h1>")
    html.append(f"<p class='muted'>생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} · 매칭 샘플: {len(matched_ids)} 건 · <b>업무 정확도(#15/#16) 제외</b></p>")
    html.append("<h2>총점 지표</h2>")
    html.append("<table><tr><th>n</th><th>MAE</th><th>RMSE</th><th>Bias</th><th>MAPE(%)</th><th>MaxAbs</th><th>Over%</th><th>Under%</th></tr>")
    mm = total_metrics
    html.append(
        f"<tr>"
        f"<td>{_fmt(mm['n'])}</td><td>{_fmt(mm['MAE'])}</td>"
        f"<td>{_fmt(mm['RMSE'])}</td><td>{_fmt(mm['Bias'])}</td>"
        f"<td>{_fmt(mm['MAPE'])}</td><td>{_fmt(mm['MaxAbs'])}</td>"
        f"<td>{_fmt(mm['Over%'])}</td><td>{_fmt(mm['Under%'])}</td></tr>"
    )
    html.append("</table>")

    html.append("<h2>항목별 지표</h2>")
    html.append("<table><tr><th>item#</th><th>항목명</th><th>배점</th><th>n</th><th>MAE</th><th>RMSE</th><th>Bias</th><th>MAPE(%)</th><th>MaxAbs</th><th>Over%</th><th>Under%</th></tr>")
    for num, name, maxs, _row in ITEM_DEF:
        pairs = per_item_pairs.get(num, [])
        m = compute_metrics(pairs)
        cls = ""
        if m["MAE"] is not None:
            cls = "ok" if m["MAE"] <= 0.5 else "warn" if m["MAE"] <= 1.5 else "bad"
        html.append(
            f"<tr><td>#{num}</td><td>{name}</td><td>{maxs}</td><td>{_fmt(m['n'])}</td>"
            f"<td class='{cls}'>{_fmt(m['MAE'])}</td>"
            f"<td>{_fmt(m['RMSE'])}</td><td>{_fmt(m['Bias'])}</td><td>{_fmt(m['MAPE'])}</td>"
            f"<td>{_fmt(m['MaxAbs'])}</td><td>{_fmt(m['Over%'])}</td><td>{_fmt(m['Under%'])}</td></tr>"
        )
    html.append("</table>")

    html.append("<h2>대화별 지표 · 불일치 항목</h2>")
    html.append("<table><tr><th>sample_id</th><th>n</th><th>MAE</th><th>RMSE</th><th>Bias</th><th>MAPE(%)</th><th>MaxAbs</th><th>Over%</th><th>Under%</th><th>AI총점</th><th>사람총점</th><th>Δ</th><th style='text-align:left'>불일치 항목 (|Δ|≥2)</th><th style='text-align:left'>경미 (|Δ|=1)</th></tr>")
    for sid in matched_ids:
        pairs = [(a, h, m) for _, a, h, m in per_sample_pairs[sid]]
        mm = compute_metrics(pairs)
        ai_t = sum(a for a, _, _ in pairs)
        hu_t = sum(h for _, h, _ in pairs)
        d = ai_t - hu_t
        cls = "ok" if d == 0 else "warn" if abs(d) <= 2 else "bad"
        item_deltas = [(num, a - h) for num, a, h, _mx in per_sample_pairs[sid]]
        major = sorted([(n_, d_) for n_, d_ in item_deltas if abs(d_) >= 2], key=lambda x: -abs(x[1]))
        minor = sorted([(n_, d_) for n_, d_ in item_deltas if abs(d_) == 1], key=lambda x: -abs(x[1]))
        major_str = ", ".join(f"#{n_}({d_:+d})" for n_, d_ in major) or "-"
        minor_str = ", ".join(f"#{n_}({d_:+d})" for n_, d_ in minor) or "-"
        html.append(
            f"<tr><td>{sid}</td><td>{_fmt(mm['n'])}</td><td>{_fmt(mm['MAE'])}</td>"
            f"<td>{_fmt(mm['RMSE'])}</td><td>{_fmt(mm['Bias'])}</td><td>{_fmt(mm['MAPE'])}</td>"
            f"<td>{_fmt(mm['MaxAbs'])}</td><td>{_fmt(mm['Over%'])}</td><td>{_fmt(mm['Under%'])}</td>"
            f"<td>{ai_t}</td><td>{hu_t}</td><td class='{cls}'>{d:+d}</td>"
            f"<td style='text-align:left'>{major_str}</td>"
            f"<td style='text-align:left'>{minor_str}</td></tr>"
        )
    html.append("</table>")

    # 해석 노트 (자동 생성 · 총점 기준)
    html.append("<h2>해석 노트 (자동 생성)</h2>")
    bias_v = total_metrics["Bias"] or 0
    over_pct = total_metrics["Over%"] or 0
    under_pct = total_metrics["Under%"] or 0
    if bias_v < -1:
        bias_interp = f"AI 가 사람보다 체계적으로 <b>엄격</b> (샘플당 평균 {abs(bias_v):.2f}점 적게 부여)."
    elif bias_v > 1:
        bias_interp = f"AI 가 사람보다 체계적으로 <b>관대</b> (샘플당 평균 {abs(bias_v):.2f}점 과대 부여)."
    else:
        bias_interp = "AI 와 사람이 전체적으로 균형. 개별 항목 편차 확인 필요."
    if under_pct > over_pct * 2:
        balance = f"Under {under_pct:.1f}% vs Over {over_pct:.1f}% — <b>AI 과소가 압도적</b>."
    elif over_pct > under_pct * 2:
        balance = f"Over {over_pct:.1f}% vs Under {under_pct:.1f}% — AI 과대가 압도적."
    else:
        balance = f"Over/Under 균형 (Over {over_pct:.1f}%, Under {under_pct:.1f}%)."

    html.append(f"<p><b>채점 경향.</b> {bias_interp}<br/><b>과대/과소 균형.</b> {balance}</p>")

    # 워스트/베스트 항목
    item_m_list = []
    for num, name, maxs, _row in ITEM_DEF:
        pairs = per_item_pairs.get(num, [])
        if pairs:
            m = compute_metrics(pairs)
            item_m_list.append((num, name, maxs, m))
    worst5 = sorted(item_m_list, key=lambda x: x[3]["MAE"] or 0, reverse=True)[:5]
    best5 = sorted(item_m_list, key=lambda x: x[3]["MAE"] or 0)[:5]

    html.append("<h3>튜닝 우선순위 항목 (MAE 상위 5)</h3>")
    html.append("<table><tr><th>item</th><th>항목명</th><th>MAE</th><th>Bias</th><th>Under%</th><th>Over%</th><th>비고</th></tr>")
    for num, name, maxs, m in worst5:
        tag = "AI 심하게 엄격" if (m["Under%"] or 0) > 50 else ("AI 느슨" if (m["Over%"] or 0) > 30 else "")
        html.append(
            f"<tr><td>#{num}</td><td>{name}</td><td class='bad'>{m['MAE']}</td>"
            f"<td>{m['Bias']}</td><td>{m['Under%']}</td><td>{m['Over%']}</td><td>{tag}</td></tr>"
        )
    html.append("</table>")

    html.append("<h3>정확도 우수 항목 (MAE 하위 5)</h3>")
    html.append("<table><tr><th>item</th><th>항목명</th><th>MAE</th><th>Bias</th></tr>")
    for num, name, maxs, m in best5:
        html.append(f"<tr><td>#{num}</td><td>{name}</td><td class='ok'>{m['MAE']}</td><td>{m['Bias']}</td></tr>")
    html.append("</table>")

    # outlier 대화
    sample_deltas = []
    for sid in matched_ids:
        pairs = per_sample_pairs.get(sid, [])
        ai_t = sum(a for _, a, _, _ in pairs)
        hu_t = sum(h for _, _, h, _ in pairs)
        sample_deltas.append((sid, ai_t - hu_t, ai_t, hu_t))
    sample_deltas.sort(key=lambda x: abs(x[1]), reverse=True)

    html.append("<h3>Outlier 대화 (|총점Δ| 상위 5)</h3>")
    html.append("<table><tr><th>sample_id</th><th>AI 총점</th><th>사람 총점</th><th>Δ</th><th>주요 불일치 항목</th></tr>")
    for sid, delta, ai_t, hu_t in sample_deltas[:5]:
        pairs = per_sample_pairs.get(sid, [])
        item_ds = sorted(
            [(num, a - h) for num, a, h, _mx in pairs if abs(a - h) >= 2],
            key=lambda x: -abs(x[1]),
        )
        major_str = ", ".join(f"#{n}({d:+d})" for n, d in item_ds[:6]) or "-"
        cls = "ok" if abs(delta) <= 2 else "warn" if abs(delta) <= 10 else "bad"
        html.append(
            f"<tr><td>{sid}</td><td>{ai_t}</td><td>{hu_t}</td>"
            f"<td class='{cls}'>{delta:+d}</td><td style='text-align:left'>{major_str}</td></tr>"
        )
    html.append("</table>")

    # 구조적 drift
    drift = [(num, name, m) for num, name, _mx, m in item_m_list if (m["Under%"] or 0) >= 50]
    if drift:
        html.append("<h3>구조적 AI 과소 채점 항목 (Under% ≥ 50%)</h3>")
        html.append("<p class='muted'>단일 샘플 문제가 아닌 rubric 해석 차이. 프롬프트 감점 트리거 재검토 필요.</p>")
        html.append("<table><tr><th>item</th><th>항목명</th><th>Under%</th><th>MAE</th><th>Bias</th></tr>")
        for num, name, m in drift:
            html.append(
                f"<tr><td>#{num}</td><td>{name}</td><td class='bad'>{m['Under%']}%</td>"
                f"<td>{m['MAE']}</td><td>{m['Bias']}</td></tr>"
            )
        html.append("</table>")

    html.append("""<h3>실행 제안</h3>
<ol>
<li><b>최우선 튜닝</b>: 위 MAE 상위 항목의 프롬프트 감점 트리거를 사람 평가 샘플과 비교 검토.</li>
<li><b>튜닝 검증</b>: 수정 후 동일 샘플 재평가 → 다른 항목 MAE 악화 '풍선 효과' 주의.</li>
<li><b>Outlier 수동 검토</b>: |Δ|≥10 대화는 사람 평가자와 재검토 권장.</li>
<li><b>한계 인정</b>: 사람 평가가 정답. 사람 평가 drift 가 있으면 본 지표 신뢰도 제한. 두 셋 공통 패턴만 구조적 문제로 간주.</li>
</ol>
""")

    html.append("<h2>지표 설명</h2>")
    html.append("<table><tr><th>항목</th><th style='text-align:left'>설명</th><th>해석 기준</th></tr>")
    guide_html = [
        ("n",        "비교에 사용된 쌍(pair) 개수.",                                               "—"),
        ("MAE",      "Mean Absolute Error. 평균 |AI - 사람| 오차.",                              "낮을수록 좋음. 5점 척도 0.5↓ 우수, 1.5↑ 개선필요."),
        ("RMSE",     "Root Mean Squared Error. 큰 오차에 민감.",                                  "MAE 대비 크면 outlier 존재."),
        ("Bias",     "AI - 사람 평균 (부호 포함).",                                                "음수=AI 과소, 양수=AI 과대."),
        ("MAPE(%)",  "|AI-사람| / 항목 배점 평균 × 100.",                                          "배점이 다른 항목 비교용. 10% 이하 양호."),
        ("MaxAbs",   "단일 쌍 최대 절대오차.",                                                     "5점 항목 MaxAbs=5 → 정반대 판정 1건 존재."),
        ("Over%",    "AI > 사람 비율 (%).",                                                      "Over/Under 합 = 전체 불일치율."),
        ("Under%",   "AI < 사람 비율 (%).",                                                      "Over 대비 높으면 AI 엄격."),
        ("총점Δ",    "샘플 총점 AI - 사람.",                                                      "|Δ| ≤ 2 ok, 3~10 warn, 그 이상 bad."),
        ("불일치 항목", "|Δ|≥2 인 항목 나열. 튜닝 대상 특정용.",                                    "절대값 내림차순."),
        ("경미 불일치", "|Δ|=1 인 항목 나열. 노이즈 바운더리 참고.",                                 "—"),
    ]
    for k, desc, guide in guide_html:
        html.append(f"<tr><td><b>{k}</b></td><td style='text-align:left'>{desc}</td><td>{guide}</td></tr>")
    html.append("</table>")
    html.append("<p class='muted'>집계 레벨 — <b>항목-레벨</b>: 모든 (샘플 × 항목) 쌍 한 덩어리. 원자 단위 정확도. / <b>총점-레벨</b>: 샘플별 총점끼리 비교. 최종 등급 영향도.</p>")
    html.append("<p class='muted'>불일치 항목 표기: <b>#item(±Δ)</b> — 음수=AI 과소, 양수=AI 과대. 절대값 내림차순.</p>")
    html.append("</body></html>")
    html_path.write_text("".join(html), encoding="utf-8")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python compare_learning_set_vs_xlsx.py <results_dir>")
        sys.exit(1)
    results_dir = Path(sys.argv[1])
    if not results_dir.exists():
        print(f"결과 폴더 없음: {results_dir}")
        sys.exit(1)
    xlsx_path = Path(r"C:\Users\META M\Desktop\QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
    if not xlsx_path.exists():
        print(f"정답 xlsx 없음: {xlsx_path}")
        sys.exit(1)
    build_report(results_dir, xlsx_path)
