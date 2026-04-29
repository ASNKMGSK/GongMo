# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""xlsx 비고(사람 평가 rationale + 인용) ↔ 원문 transcript 교차 검증.

목적:
  - 비고의 "상담사#N / 고객#N: \"...\"" 인용이 실제 transcript turn N 과 일치하는가
  - rationale 의 판정 근거(3요소 체크 / 감점 / 감점 0회 등) 가 인용 turn 내용으로 성립하는가
  - 비고의 점수가 사람이 실제로 평가 rubric 에 맞춰 부여했는지 합리성 리뷰

샘플 전수 대신 대표 샘플 3~4개 상세 리포트 + 전체 xlsx 통계 집계.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

XLSX_PATH = r"C:/Users/META M/Desktop/QA정답-STT_기반_통합_상담평가표_v3재평가.xlsx"
TRANSCRIPT_DIR = Path(r"C:/Users/META M/Desktop/qa 샘플/학습셋")

ITEM_NAME_MAP = {
    "첫인사": 1, "끝인사": 2,
    "호응및공감": 4, "대기멘트": 5,
    "정중한표현": 6, "쿠션어활용": 7,
    "문의파악및재확인(복창)": 8, "고객정보확인": 9,
    "설명의명확성": 10, "두괄식답변": 11,
    "문제해결의지": 12, "부연설명및추가안내": 13, "사후안내": 14,
    "정확한안내★": 15, "정확한안내": 15, "필수안내이행": 16,
    "정보확인절차": 17, "정보보호준수": 18,
}

ITEM_MAX = {1:5,2:5,4:5,5:5,6:5,7:5,8:5,9:5,10:10,11:5,12:5,13:5,14:5,15:15,16:5,17:5,18:5}


# ---------------------------------------------------------------------------
# transcript 파싱 — 평문 "화자: 텍스트" 라인 형식
# ---------------------------------------------------------------------------


def load_transcript(sample_id: int) -> dict[int, dict[str, str]]:
    """{turn_id: {"speaker": "상담사"|"고객", "text": "..."}} — turn_id 는 1-based 글로벌.

    transcript 필드는 평문 "상담사: ...\\n고객: ...\\n..." 형식.
    """
    for p in TRANSCRIPT_DIR.glob(f"{sample_id}_*.json"):
        data = json.load(open(p, encoding="utf-8"))
        raw = data.get("transcript", "")
        break
    else:
        return {}
    out: dict[int, dict[str, str]] = {}
    turn_id = 0
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(상담사|고객):\s*(.*)$", line)
        if not m:
            # 화자 prefix 없는 줄은 이전 발화 연장 (없으면 무시)
            if turn_id > 0:
                out[turn_id]["text"] += " " + line
            continue
        turn_id += 1
        out[turn_id] = {"speaker": m.group(1), "text": m.group(2)}
    return out


# ---------------------------------------------------------------------------
# xlsx 파싱
# ---------------------------------------------------------------------------


def parse_xlsx() -> dict[int, dict[int, dict[str, Any]]]:
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    out: dict[int, dict[int, dict[str, Any]]] = {}
    for sheet in wb.sheetnames:
        m = re.search(r"(\d{6})", sheet)
        if not m:
            continue
        sid = int(m.group(1))
        ws = wb[sheet]
        per_item: dict[int, dict[str, Any]] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 6:
                continue
            raw_item = row[1]
            if raw_item is None:
                continue
            name = re.sub(r"\s+|\([^)]*\)|\n", "", str(raw_item))
            inum = ITEM_NAME_MAP.get(name)
            if inum is None:
                continue
            score = row[4]
            rationale = row[5] or ""
            if score is None:
                continue
            try:
                score = int(score)
            except (TypeError, ValueError):
                continue
            per_item[inum] = {
                "score": score,
                "rationale": str(rationale),
                "raw_item_name": str(raw_item),
            }
        out[sid] = per_item
    return out


# ---------------------------------------------------------------------------
# 비고 안의 인용 파싱
# ---------------------------------------------------------------------------

CITE_RE = re.compile(r'-?\s*(상담사|고객)#(\d+)\s*:\s*"([^"]*)"')


def extract_citations(rationale: str) -> list[tuple[str, int, str]]:
    return [(s, int(n), q) for s, n, q in CITE_RE.findall(rationale)]


def verify_citation(turn_map: dict[int, dict[str, str]], spk: str, tid: int, quote: str) -> tuple[bool, str, str]:
    """(matches?, expected_text, mismatch_reason)"""
    actual = turn_map.get(tid)
    if actual is None:
        return False, "", f"turn #{tid} 원문에 없음"
    if actual["speaker"] != spk:
        return False, actual["text"], f"화자 불일치 (원문 {actual['speaker']}, 비고 {spk})"
    # 부분 일치 허용 (비고가 일부만 발췌할 수 있음)
    q_norm = re.sub(r"\s+", "", quote)
    a_norm = re.sub(r"\s+", "", actual["text"])
    if q_norm and q_norm in a_norm:
        return True, actual["text"], ""
    return False, actual["text"], "quote 가 실제 발화에 포함되지 않음"


# ---------------------------------------------------------------------------
# rationale 논리 합리성 검토 — 항목별 heuristic
# ---------------------------------------------------------------------------


def check_rationale_logic(item_num: int, score: int, rationale: str) -> list[str]:
    """비고의 판정 논리가 rubric 과 합리적으로 들어맞는지 간이 체크.

    Returns list of warning messages (empty if OK).
    """
    warnings: list[str] = []
    max_score = ITEM_MAX[item_num]
    tail_match = re.search(r"→\s*(\S+?)(?:\s|$)", rationale)
    claimed = tail_match.group(1) if tail_match else None

    # 1. rationale 끝 "→ 만점 / 3점 / 0점" 이 score 와 맞는지
    if claimed == "만점" and score != max_score:
        warnings.append(f"rationale='만점' but score={score} (max={max_score})")
    if claimed and claimed.endswith("점") and claimed != "만점":
        m = re.match(r"(\d+)점", claimed)
        if m and int(m.group(1)) != score:
            warnings.append(f"rationale='{claimed}' but score={score}")

    # 2. 항목별 구체 패턴
    if item_num == 1:  # 첫인사 3요소
        has_greet = "인사말✓" in rationale
        has_org = "소속✓" in rationale
        has_name = "상담사명✓" in rationale
        missing = [k for k, v in [("인사말", has_greet), ("소속", has_org), ("상담사명", has_name)] if not v]
        if score == 5 and any(("인사말✗" in rationale, "소속✗" in rationale, "상담사명✗" in rationale)):
            warnings.append(f"#1 score=5 but rationale 에 ✗ 요소 있음")
        if score == 3:
            x_count = sum(1 for m in ("인사말✗","소속✗","상담사명✗") if m in rationale)
            if x_count > 1:
                warnings.append(f"#1 score=3 but ✗ 요소 {x_count}개 (규약: 1개일 때 3점)")
        if score == 0:
            x_count = sum(1 for m in ("인사말✗","소속✗","상담사명✗") if m in rationale)
            if x_count < 2:
                warnings.append(f"#1 score=0 but ✗ 요소 {x_count}개 (규약: 2개 이상일 때 0점)")

    if item_num == 2:  # 끝인사 3요소
        x_count = sum(1 for m in ("추가문의✗","인사말✗","상담사명✗") if m in rationale)
        check_count = sum(1 for m in ("추가문의✓","인사말✓","상담사명✓") if m in rationale)
        if score == 5 and x_count > 0:
            warnings.append(f"#2 score=5 but ✗ 요소 {x_count}개")
        if score == 0 and x_count < 2 and check_count > 1:
            warnings.append(f"#2 score=0 but ✗ 요소 {x_count}개 (규약: 2개 이상일 때 0점)")

    if item_num == 8:  # 문의 파악 / 복창
        # "복창✓=False" + score=5 (만점) 은 모순
        if "복창✓=False" in rationale and score == 5:
            warnings.append("#8 복창 안 했다고 표시했는데 만점 (논리 모순)")
        if "복창✓=True" in rationale and score == 0:
            warnings.append("#8 복창 했다고 표시했는데 0점 (논리 모순)")

    if item_num == 11:  # 두괄식
        if "결론선행✗" in rationale and score == 5:
            warnings.append("#11 결론선행 안 했다고 표시했는데 만점 (논리 모순)")
        if "결론선행✓" in rationale and score == 0:
            warnings.append("#11 결론선행 했다고 표시했는데 0점 (논리 모순)")

    if item_num == 4:  # 호응/공감
        if score == 5 and "0회" in rationale:
            warnings.append("#4 공감 0회인데 만점 (공감 1회 이상 필요)")

    if item_num == 5:  # 대기 멘트
        if "대기 상황 없음" not in rationale and "미발생" not in rationale:
            if score == 5 and "사후✗" in rationale and "사전✗" in rationale:
                warnings.append("#5 대기 전후 모두 양해 없는데 만점")

    if item_num == 7:  # 쿠션어
        if score == 5 and "전무" in rationale and "거절/불가" in rationale:
            warnings.append("#7 거절/불가 상황에 쿠션어 전무인데 만점")

    if item_num == 12:  # 문제해결
        if score == 5 and "회피" in rationale:
            warnings.append("#12 업무회피 표현이 있는데 만점")

    if item_num == 15:  # 정확한 안내
        if score >= 10 and "오안내" in rationale and "정정" not in rationale:
            warnings.append("#15 오안내 있는데 정정 언급 없음 + 10점 이상")

    return warnings


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------


def audit_sample(sid: int, items: dict[int, dict[str, Any]]) -> dict[str, Any]:
    turn_map = load_transcript(sid)
    if not turn_map:
        return {"transcript_missing": True}
    total_citations = 0
    cite_ok = 0
    cite_fail = []
    logic_warnings = []
    for inum, rec in items.items():
        cites = extract_citations(rec["rationale"])
        for spk, tid, quote in cites:
            total_citations += 1
            ok, actual, reason = verify_citation(turn_map, spk, tid, quote)
            if ok:
                cite_ok += 1
            else:
                cite_fail.append({
                    "item": inum, "score": rec["score"], "speaker": spk, "turn": tid,
                    "quote": quote, "actual": actual, "reason": reason,
                })
        for w in check_rationale_logic(inum, rec["score"], rec["rationale"]):
            logic_warnings.append({"item": inum, "score": rec["score"], "warning": w,
                                   "rationale": rec["rationale"][:200]})
    return {
        "total_citations": total_citations,
        "cite_ok": cite_ok,
        "cite_fail": cite_fail,
        "logic_warnings": logic_warnings,
    }


def main():
    xlsx = parse_xlsx()
    print(f"\nxlsx 샘플: {len(xlsx)}")

    all_stats = {
        "citation_total": 0, "citation_ok": 0,
        "citation_fail": [], "logic_warnings": [],
    }
    per_sample: dict[int, dict] = {}

    for sid, items in sorted(xlsx.items()):
        r = audit_sample(sid, items)
        if r.get("transcript_missing"):
            print(f"  ⚠ sample {sid}: transcript 없음")
            continue
        per_sample[sid] = r
        all_stats["citation_total"] += r["total_citations"]
        all_stats["citation_ok"] += r["cite_ok"]
        all_stats["citation_fail"].extend([{"sample": sid, **f} for f in r["cite_fail"]])
        all_stats["logic_warnings"].extend([{"sample": sid, **w} for w in r["logic_warnings"]])

    # --- 전체 리포트 ---
    print("\n" + "=" * 80)
    print("[종합] 인용 정확도")
    print("=" * 80)
    print(f"  총 인용: {all_stats['citation_total']}")
    print(f"  ✓ 일치: {all_stats['citation_ok']}")
    print(f"  ❌ 불일치: {len(all_stats['citation_fail'])}")
    print(f"  정확도: {all_stats['citation_ok'] / max(1, all_stats['citation_total']):.1%}")

    # 원인 분석
    reason_counter = Counter(f["reason"] for f in all_stats["citation_fail"])
    print("\n  [불일치 원인]")
    for reason, cnt in reason_counter.most_common():
        print(f"    {cnt:4d}  {reason}")

    # 일부 예시
    print("\n  [불일치 예시 20건]")
    for f in all_stats["citation_fail"][:20]:
        print(f"    sample={f['sample']} item#{f['item']} score={f['score']} {f['speaker']}#{f['turn']}")
        print(f"      비고 인용: {f['quote'][:100]!r}")
        print(f"      실제 원문: {f['actual'][:100]!r}")
        print(f"      사유: {f['reason']}")

    print("\n" + "=" * 80)
    print("[종합] rationale 논리 경고")
    print("=" * 80)
    w_by_item = defaultdict(int)
    for w in all_stats["logic_warnings"]:
        w_by_item[w["item"]] += 1
    print(f"  총 경고: {len(all_stats['logic_warnings'])}")
    print("  [항목별]")
    for inum in sorted(w_by_item.keys()):
        print(f"    item #{inum}: {w_by_item[inum]}건")
    print("\n  [경고 예시 30건]")
    for w in all_stats["logic_warnings"][:30]:
        print(f"    sample={w['sample']} item#{w['item']} score={w['score']}: {w['warning']}")
        print(f"      rationale: {w['rationale']}")

    # --- 대표 샘플 상세 ---
    print("\n" + "=" * 80)
    print("[상세 리포트] 대표 샘플 3개")
    print("=" * 80)
    for sid in [668437, 668605, 668675]:
        if sid not in per_sample:
            continue
        r = per_sample[sid]
        print(f"\n### sample {sid}")
        print(f"  총 인용: {r['total_citations']} (✓{r['cite_ok']} / ❌{len(r['cite_fail'])})")
        print(f"  rationale 논리 경고: {len(r['logic_warnings'])}")
        for w in r["logic_warnings"][:10]:
            print(f"    - item#{w['item']} score={w['score']}: {w['warning']}")
        for f in r["cite_fail"][:5]:
            print(f"    - item#{f['item']} {f['speaker']}#{f['turn']}: {f['reason']}")
            print(f"        quote: {f['quote'][:80]!r}")
            print(f"        actual: {f['actual'][:80]!r}")


if __name__ == "__main__":
    main()
