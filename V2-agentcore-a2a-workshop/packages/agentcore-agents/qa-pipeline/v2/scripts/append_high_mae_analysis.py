# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""기존 비교분석 xlsx 에 'MAE 높은항목 원인 분석' 시트 추가.

대상 항목 (MAE 가 다른 항목 대비 현저히 큰 3개):
  - #6  정중한 표현
  - #8  문의 파악 및 재확인(복창)
  - #10 설명의 명확성

각 항목별로 분포 통계 + **원인 진단** (사람 vs AI 평가 기준 차이) + **튜닝 권고** 노출.
원문 케이스는 별도 시트('큰차이 심층분석') 가 이미 있으므로 본 시트는 진단만.

실행:
  python -m v2.scripts.append_high_mae_analysis
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_PIPELINE_DIR = Path(__file__).resolve().parents[2]
if str(_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(_PIPELINE_DIR))

from v2.scripts.compare_learning_set_vs_xlsx import (  # type: ignore[import-untyped]
    ITEM_DEF, ITEM_NUM_TO_MAX, ITEM_NUM_TO_NAME,
    TEST_IDS, TRAINING_IDS,
)


XLSX_GT = Path(r"C:\Users\META M\Desktop\QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
SHEET_NAME = "MAE 높은항목 원인 분석"

# 항목별 대표 샘플 (사람 평가 vs AI 평가 나란히) — 진단 근거가 가장 명확한 케이스 1~2 건씩
# (sample_id 별로 GT 비고와 AI judgment 를 직접 인용 → 차이의 근본을 보여주기 위함)
REPRESENTATIVE_CASES_BY_DATASET: dict[str, dict[int, list[dict]]] = {
    "학습셋": {
        6: [
            {"sid": "668481", "ai_score": 3, "human_score": 5, "diff": -2,
             "human_view": "사람: '비속어/반말/사물존칭/습관어 0회 → 만점' (5점). 사물존칭/습관어 발생 횟수만 본다.",
             "ai_view":    "AI: \"'진짜 궁금한', '그냥', '이게', '이 그럼' 같은 filler 가 다수 확인됨\" → 3점. 음성 채움말을 '습관어' 로 오분류.",
             "why_gap":    "→ AI 가 본 'filler' (음성 채움말) 와 사람이 본 '습관어' (의미 모호화 표현) 가 다른 신호."},
            {"sid": "668610", "ai_score": 5, "human_score": 3, "diff": +2,
             "human_view": "사람: GT note 가 '사물존칭/습관어 0회 → 3점'. **명시된 룰과 점수가 모순** — 사람이 다른 사유로 -2점.",
             "ai_view":    "AI: 반말/비속어/명령조/사물존칭 모두 부재 → 만점.",
             "why_gap":    "→ GT 자체 불일치 케이스. 학습셋에서 1건 발생."},
        ],
        8: [
            {"sid": "668464", "ai_score": 0, "human_score": 5, "diff": -5,
             "human_view": "사람: '복창 3회 확인 → 만점'. 실제 인용 = '반품 접수 도와드리겠습니다' / '코오롱 스카프 구십 확인되고 있는데 맞으실까요' / '반품 사유 어떻게 되실까요'. **고객 문의 핵심 명사 ('반품', '스카프') 가 등장한 모든 후속 발화를 복창으로 카운트**.",
             "ai_view":    "AI: \"문의 내용을 복창하거나 재확인하지 않고 바로 개인정보 확인 절차로 넘어감\" → 0점. **'처리 안내 발화' 와 '복창' 을 분리 판단**.",
             "why_gap":    "→ '복창' 정의 차이: 사람=핵심 명사 재등장 / AI=의도 기반 paraphrase."},
            {"sid": "668605", "ai_score": 0, "human_score": 5, "diff": -5,
             "human_view": "사람: '복창 3회'. 인용 = '본인 맞으실까요' / '며칠 날 주문하신 어떤 제품에 대해서 문의하시는 건가요' / '주문번호 두 건 확인되시는데요'.",
             "ai_view":    "AI: '재질문하여 문의 내용을 파악하지 못했음. 복창은커녕 기본적인 문의 파악도 실패' → 0점.",
             "why_gap":    "→ 동일한 재질문 발화를 사람은 '복창', AI 는 '문의 파악 실패' 로 정반대 해석."},
        ],
        10: [
            {"sid": "668437", "ai_score": 7, "human_score": 10, "diff": -3,
             "human_view": "사람: '장문 3건, 고객 되물음 최소 → 만점'. 신호 = (1) 설명량 (2) 고객 되물음 횟수만.",
             "ai_view":    "AI: '동일 내용 (영업일 2-3일) 반복 설명, 맞교환 프라스 등 불명확 표현, 부분적으로 장황' → 7점.",
             "why_gap":    "→ 사람은 '고객이 이해함' (되물음 0회) 만으로 만점, AI 는 발화 형식 (장황·반복) 으로 한 단계 감점."},
            {"sid": "668771", "ai_score": 5, "human_score": 10, "diff": -5,
             "human_view": "사람: '장문 5건, 고객 되물음 최소 → 만점'. 정상 설명으로 판정.",
             "ai_view":    "AI: \"'사은품', '물류 쪽' 내부 용어 사용, 설명 중간 끊김\" → 5점. **내부 용어 사용을 5점 트리거로 봄**.",
             "why_gap":    "→ '내부 용어' 는 AI 만의 감점 트리거. 사람은 평가 안 함."},
        ],
    },
    "테스트셋": {
        6: [
            {"sid": "668847", "ai_score": 3, "human_score": 5, "diff": -2,
             "human_view": "사람: '비속어/반말/사물존칭/습관어 0회 → 만점' (5점). 학습셋과 동일한 룰.",
             "ai_view":    "AI: filler 빈도로 -2점 부여 (학습셋과 동일 패턴).",
             "why_gap":    "→ 학습셋과 동일 메커니즘. AI 의 filler 트리거가 잘못된 신호."},
            {"sid": "668963", "ai_score": 5, "human_score": 3, "diff": +2,
             "human_view": "사람: '사물존칭/습관어 0회 → 3점' — **GT 자체 모순 케이스**. 명시된 룰과 점수 불일치.",
             "ai_view":    "AI: 트리거 부재로 만점.",
             "why_gap":    "→ 학습셋의 668610 과 동일 패턴 (GT 불일치 케이스)."},
        ],
        8: [
            {"sid": "668797", "ai_score": 0, "human_score": 5, "diff": -5,
             "human_view": "사람: GT note = '복창/재확인 확인 (복창✓=False) → 만점'. **복창 안 했다고 명시했는데 만점 부여**.",
             "ai_view":    "AI: '복창이나 재확인 절차 전무' → 0점. AI 의 일관된 정의 적용.",
             "why_gap":    "→ GT 자체가 '복창 없어도 만점' 으로 라벨링 — 진짜 채점 기준이 GT note 만으로 추정 불가. AI 가 어떤 기준으로 채점하든 정합 불가능."},
            {"sid": "668865", "ai_score": 0, "human_score": 5, "diff": -5,
             "human_view": "사람: '(복창✓=False) → 만점'. 668797 과 동일 패턴.",
             "ai_view":    "AI: '문의 파악 못하고 바로 주문 건 확인으로 넘어감' → 0점.",
             "why_gap":    "→ 테스트셋 9건 중 8건이 같은 패턴 ('복창=False 인데 만점'). GT 라벨링 정책 자체가 학습셋과 다름."},
        ],
        10: [
            {"sid": "668865", "ai_score": 7, "human_score": 10, "diff": -3,
             "human_view": "사람: '장문 4건, 고객 되물음 최소 → 만점'. 학습셋과 동일 룰.",
             "ai_view":    "AI: 장황·반복 트리거로 7점. 학습셋과 동일 패턴.",
             "why_gap":    "→ 학습셋과 동일 메커니즘. 만점 인정 임계 차이."},
        ],
    },
}

# Backward 호환 — 기존 함수에서 REPRESENTATIVE_CASES 참조 시 학습셋 사용
REPRESENTATIVE_CASES: dict[int, list[dict]] = {
    6: [
        {
            "sid": "668481", "ai_score": 3, "human_score": 5, "diff": -2,
            "human_view": "사람: 비속어/반말/사물존칭/습관어 부적절 표현 없음 → 만점 (5점). 정중함 평가 차원은 '공손성 훼손 여부' 만 봄.",
            "ai_view":    "AI: '진짜 궁금한', '그냥', '이게', '이 그럼' 같은 filler/혼잣말이 다수라며 -2점. 정중함을 '발화 매끄러움' 으로 해석.",
            "why_gap":    "→ 같은 단어 ('습관어') 를 양쪽이 다르게 정의. 사람은 호칭 체계, AI 는 발화 스타일.",
        },
        {
            "sid": "668610", "ai_score": 5, "human_score": 3, "diff": +2,
            "human_view": "사람: 사물존칭/습관어 0회라고 보면서도 3점만 부여 (다른 미세 사유 가능성).",
            "ai_view":    "AI: 반말·비속어·명령조·사물존칭 모두 부재라며 만점 5점.",
            "why_gap":    "→ 반대 방향 케이스. AI 의 만점 임계가 더 관대.",
        },
    ],
    8: [
        {
            "sid": "668464", "ai_score": 0, "human_score": 5, "diff": -5,
            "human_view": "사람: '핵심 내용 재확인/복창 3회 확인 → 만점'. 인용 발화 = '제가 빠르게 확인하고 반품 접수 도와드리겠습니다 ... 성함과 휴대폰 번호 말씀 부탁드립니다' (개인정보 확인 도입 발화). '본인 맞으십니까' 도 복창 카운트.",
            "ai_view":    "AI: '문의 내용을 복창하거나 재확인하지 않고 바로 개인정보 확인 절차로 넘어감. Few-shot 예시 1과 유사. 0점.' → 개인정보 확인을 '복창 아님' 으로 봄.",
            "why_gap":    "→ 사람은 '고객 문의 후 후속 응대 (개인정보·상품 식별·처리 안내)' 전부를 복창으로 인정, AI 는 '핵심 명사구 재발화' 만 복창. 사람의 인정 범위가 압도적으로 넓음.",
        },
        {
            "sid": "668605", "ai_score": 0, "human_score": 5, "diff": -5,
            "human_view": "사람: '복창 3회 확인 → 만점'. 인용 = '*** 고객님 본인 맞으실까요' (본인확인) + '잠시만요 며칠 날 주문하신 어떤 제품에 대해서 문의하시는 건가요' (재질문) + '사 월 사 일에 주문하신 주문번호 두 건 확인되시는데요'.",
            "ai_view":    "AI: '재질문하여 문의 내용을 파악하지 못했음을 보여줌. 복창은커녕 기본적인 문의 파악도 실패. 0점.'",
            "why_gap":    "→ 동일한 재질문 발화를 사람은 '복창 인정', AI 는 '문의 파악 실패 신호' 로 정반대 해석.",
        },
        {
            "sid": "668451", "ai_score": 5, "human_score": 3, "diff": +2,
            "human_view": "사람: GT 비고에 '복창 누락 또는 재질의 발생 (복창✗) → 3점' 명시. **사람이 복창 안 됐다고 인정하면서도 만점 5 가 아닌 3점 부여**.",
            "ai_view":    "AI: '반품 접수 도와드리겠습니다' 를 복창으로 인정하여 만점 5점.",
            "why_gap":    "→ GT 라벨 자체에 '복창✗ 인데 부분점수' 같은 모순 케이스 존재. AI 는 일관된 정의로 5점, 사람은 '복창✗' 라고 표시했지만 3점 부여 — 사람 평가표 일관성 결여가 MAE 의 일부를 설명.",
        },
    ],
    10: [
        {
            "sid": "668437", "ai_score": 7, "human_score": 10, "diff": -3,
            "human_view": "사람: '장문 3건, 고객 되물음 최소 → 만점'. 평가 차원 = (1) 충분한 설명량 (2) 고객이 이해함 = OK. **결과(효과) 중심**.",
            "ai_view":    "AI: '동일 내용 (영업일 기준 2-3일) 반복 설명, 맞교환 프라스 등 불명확한 표현으로 부분적으로 장황. 7점.' **형식(유창성·간결성) 중심**.",
            "why_gap":    "→ 평가 차원이 다름. 고객이 이해 잘한 정상 설명도 AI 는 '장황·반복' 으로 -3 처리. 한 단계 차이 (10↔7) 가 누적.",
        },
        {
            "sid": "668771", "ai_score": 5, "human_score": 10, "diff": -5,
            "human_view": "사람: '장문 5건, 고객 되물음 최소 → 만점'. 인용 발화도 정상 응대 ('소중한 정보 확인 감사드립니다 ... 반품 신청되어 있으신데 이 제품 맞으실까요').",
            "ai_view":    "AI: \"'사은품', '물류 쪽' 등 고객이 이해하기 어려운 내부 용어 사용, 설명이 중간에 끊김. 5점.\"",
            "why_gap":    "→ '내부 용어 사용' 은 AI 만의 감점 트리거 — 사람 평가에서는 의미 없는 항목. 두 단계 차이 누적 (10↔5).",
        },
    ],
}

# ───────────────────────────────────────────────────────────────────────
# 진단 — 학습셋 14건 + 테스트셋 9건 GT note 전수 + AI judgment 직접 비교 후 도출.
# 데이터셋 별로 메커니즘이 다른 항목 (#8) 은 별도 명시.
# ───────────────────────────────────────────────────────────────────────

DIAGNOSES_BY_DATASET: dict[str, dict[int, dict]] = {
    "학습셋": {
        6: {
            "name": "정중한 표현",
            "verdict": "AI 가 사람과 다른 신호 (filler 빈도) 로 감점 → 한 칸 (5↔3) 차이가 누적",
            "reasons": [
                "사람의 채점 룰 (GT note 형식): '비속어/반말/사물존칭/습관어 0회 → 만점' / '사물존칭·습관어 N회 → 3점'. **사물존칭과 습관어 발생 횟수만 본다**",
                "사람의 '습관어' = 의미를 모호하게 하는 발화 (예: 책임 회피성 '~인 것 같은데요')",
                "AI 의 감점 트리거 (모든 -2점 케이스 judgment 공통): \"'음·아·이게·그냥' 같은 filler 가 한 발화 내 3회 이상 연속 사용\" → 3점",
                "AI 가 보는 'filler' 는 STT 가 받아쓴 음성 채움말 (말 사이의 공백음) 이지, 사람이 정의한 '습관어' 가 아님",
                "결과: 정상 통화의 자연스러운 음성 채움말이 -2점 처리되어 14건 중 7건이 AI=3 / 사람=5 패턴",
                "GT 자체 모순 1건 (668610): GT note 가 '사물존칭/습관어 0회' 인데 점수는 3점 → 사람도 채점 기준 외 사유로 감점",
            ],
            "fix": [
                "프롬프트 (item_06) 에서 'filler 3회 연속' 트리거 제거",
                "감점 트리거를 GT note 와 동일하게 (1) 사물존칭 (2) 반말체 (3) 비속어 (4) 명령조 4가지로 한정",
                "'음/아/그' 같은 음성 채움말은 평가 대상 아님으로 명시",
            ],
        },
        8: {
            "name": "문의 파악 및 재확인 (복창)",
            "verdict": "AI 가 '복창' 을 의도 기반 paraphrase 로, 사람은 핵심 명사구 재등장으로 정의 → 정의 차이",
            "reasons": [
                "사람의 채점 룰: '복창 N회 확인 → 만점' (10건) / '복창 누락 (복창✗) → 3점' (4건). **자체 일관됨**",
                "실제 사람이 '복창' 으로 카운트한 발화 패턴 (668464 transcript 검증):",
                "  → '반품 접수 도와드리겠습니다' = 고객 문의의 핵심 명사 ('반품') 재등장 → 복창 1회",
                "  → '코오롱 브랜드의 스카프 구십 확인되고 있는데 맞으실까요' = 상품 재확인 → 복창 2회",
                "  → '반품 접수 도와드릴 텐데요 반품 사유가 어떻게 되실까요' = '반품' 재등장 → 복창 3회",
                "→ 사람 기준 = '고객 문의의 핵심 명사구 (반품·교환·환불 등) 가 상담사 후속 발화 어디든 등장하면 복창 카운트'",
                "AI 의 판정 패턴 (모든 0점 케이스 judgment 공통): \"문의 내용을 복창하거나 재확인하지 않고 바로 개인정보/주문서 확인 절차로 넘어감\"",
                "→ AI 기준 = '고객 문의를 의도적으로 paraphrase 하는 별도 발화가 있어야 복창' (절차 진행 발화는 복창 아님)",
                "결과: AI 는 '반품 접수 도와드리겠습니다' 를 '처리 안내' 로 분류하고 복창 카운트에서 제외 → 9건이 AI=0 / 사람=5",
            ],
            "fix": [
                "프롬프트 (item_08) 의 '복창' 정의를 GT 와 동일하게 재정의:",
                "  → '고객 문의의 핵심 명사구 (반품/교환/환불/A·S/주문/취소 등) 가 상담사 후속 발화 (다음 5턴 이내) 에 1회 이상 등장하면 복창 카운트'",
                "  → 처리 절차 진행 발화 ('도와드리겠습니다', '확인하겠습니다') 도 핵심 명사 포함 시 복창으로 인정",
                "1회=3점 / 2회 이상=5점 으로 단계 명시",
            ],
        },
        10: {
            "name": "설명의 명확성",
            "verdict": "AI 가 '발화 형식' 으로, 사람은 '고객 이해 신호 (되물음 횟수)' 로 채점 → 평가 신호 자체가 다름",
            "reasons": [
                "사람의 채점 룰 (GT note 형식 일관됨):",
                "  → '장문 N건, 고객 되물음 최소 → 만점 10' (8건)",
                "  → '부분 장황 (장문 N건, 고객 되물음 1회) → 7점' (4건)",
                "  → '내부용어/일방 나열 또는 고객 되물음 N회 → 5점' (1건)",
                "→ 사람 기준 = (1) 장문 (충분한 설명량) (2) 고객의 명시적 되물음 횟수. **고객 되물음 0회면 무조건 만점**",
                "AI 의 감점 트리거 (모든 7점 케이스 judgment 공통): '동일 내용 반복', '내부 용어 사용', '장황한 설명', 'filler 사용', '말 꼬임'",
                "→ AI 기준 = 발화 형식 (유창성·간결성·정확성). 고객 되물음은 거의 보지 않음",
                "결과: 사람이 '되물음 최소' 로 만점 준 8건 모두 AI=7점 (장황·반복 트리거). 한 단계 차이 (10↔7) 가 누적",
                "10점 만점 항목이라 한 단계 차이가 5점 항목보다 MAE 절대값 크게 보임",
            ],
            "fix": [
                "프롬프트 (item_10) 만점 조건을 GT 와 동일하게 재정의:",
                "  → '고객의 명시적 되물음 (\"네?\", \"무슨 말씀이세요\", \"다시 한 번 말씀해 주세요\") 0~1회 + 핵심 정보 누락 없음 → 10점'",
                "  → '되물음 1~2회 또는 동일 정보 1회 반복 → 7점'",
                "  → '내부 용어 사용 또는 되물음 3회+ → 5점'",
                "'장황함', '발화 매끄러움' 같은 형식 트리거는 모두 제거 (사람은 평가 안 함)",
            ],
        },
    },
    "테스트셋": {
        6: {
            "name": "정중한 표현",
            "verdict": "학습셋과 동일한 메커니즘 — AI 의 filler 트리거가 잘못된 신호 + GT 자체 모순 1건",
            "reasons": [
                "GT note 형식 학습셋과 동일 — '비속어/반말/사물존칭/습관어 0회 → 만점' / '사물존칭·습관어 N회 → 3점'",
                "AI 행동도 학습셋과 동일 — filler 빈도로 -2 점 부여 (9건 중 7건이 |Δ|=2)",
                "GT 자체 모순 1건 (668963): GT note 가 '사물존칭/습관어 0회' 인데 점수는 3점 → 학습셋의 668610 과 동일한 패턴 (명시되지 않은 다른 사유로 감점)",
                "MAE 1.778 의 거의 전부가 정상 통화의 음성 filler 트리거 오작동에서 발생",
            ],
            "fix": [
                "학습셋과 동일한 프롬프트 수정으로 동시 해결 (filler 트리거 제거)",
            ],
        },
        8: {
            "name": "문의 파악 및 재확인 (복창)",
            "verdict": "GT note 형식이 학습셋과 다름 — '복창✓=False → 만점' 이라는 새로운 라벨링 사용. 진짜 채점 기준이 무엇인지 GT 에서 추정 불가",
            "reasons": [
                "테스트셋 9건의 GT note: 8건이 '복창/재확인 확인 (복창✓=False) → 만점', 1건만 '복창✓=True → 만점'",
                "→ 즉 **8건이 모두 '복창은 안 했지만 만점'** 이라는 라벨링. 학습셋의 '복창✗ = 3점' 룰과 정반대",
                "이 형식 변경의 의미는 GT note 만으로 추정 불가 — 가능성 (a) 평가자가 다른 채점 기준 사용 (b) '복창✓=False' 가 단순 사실 기록이고 실제 만점은 다른 신호 기반",
                "AI 행동: 학습셋과 동일한 '복창 정의 좁음' 로직 적용 → 4건은 0점, 5건은 5점",
                "결과: GT 가 '복창=False 인데 만점' 이므로 AI 가 어떤 기준으로 채점하든 정합 불가능 — MAE 2.222 의 핵심 원인은 GT 라벨링 형식 변경",
            ],
            "fix": [
                "**프롬프트 수정만으로 해결 불가** — GT 작성자에게 'complaint✓=False → 만점' 라벨의 진짜 채점 기준 확인 필요",
                "확인 결과에 따라:",
                "  → '복창 외 다른 신호 (예: 본인확인 절차 완료) 로 만점' 이면 그 신호를 프롬프트에 명시",
                "  → '단순 라벨링 오류' 면 GT 재라벨링",
                "이 항목의 MAE 는 AI 측 책임이 아니므로 트래킹 시 별도 분리 권장",
            ],
        },
        10: {
            "name": "설명의 명확성",
            "verdict": "학습셋과 동일한 메커니즘 — 평가 신호 자체 차이",
            "reasons": [
                "GT note 형식 학습셋과 동일 — '장문 N건, 고객 되물음 최소 → 만점' / '부분 장황 (되물음 N회) → 7점'",
                "AI 행동도 학습셋과 동일 — '장황·반복·내부용어' 로 한 단계 감점 (9건 중 6건이 7점)",
                "MAE 1.556 의 거의 전부가 만점 인정 임계 차이에서 발생",
                "테스트셋이 학습셋보다 MAE 약간 낮은 이유: 테스트셋에 짧고 정돈된 통화가 더 많아 AI 가 만점 인정한 케이스 존재 (668847, 668899, 668963)",
            ],
            "fix": [
                "학습셋과 동일한 프롬프트 수정으로 동시 해결 (만점 조건을 '되물음 0~1회' 로)",
            ],
        },
    },
}

DIAGNOSES = DIAGNOSES_BY_DATASET["학습셋"]  # 기본 (하위 호환)


_HEADER_FILL = PatternFill("solid", fgColor="6b21a8")
_HEADER_FONT = Font(color="ffffff", bold=True, size=11)
_VERDICT_FILL = PatternFill("solid", fgColor="fef3c7")
_REASON_FILL = PatternFill("solid", fgColor="ffffff")
_FIX_FILL = PatternFill("solid", fgColor="dcfce7")
_STAT_FILL = PatternFill("solid", fgColor="e0e7ff")
_BORDER = Border(left=Side("thin", color="cbd5e1"), right=Side("thin", color="cbd5e1"),
                 top=Side("thin", color="cbd5e1"), bottom=Side("thin", color="cbd5e1"))


def load_gt(ids: list[str]) -> dict:
    wb = openpyxl.load_workbook(XLSX_GT, data_only=True)
    out = {}
    aset = set(ids)
    for sn in wb.sheetnames:
        tail = sn.strip().split("_")[-1]
        if not (tail.isdigit() and len(tail) == 6 and tail in aset):
            continue
        ws = wb[sn]
        sample = {}
        for num, _name, _mx, row in ITEM_DEF:
            v = ws.cell(row, 5).value
            if v is None:
                continue
            try:
                sample[num] = int(v)
            except (TypeError, ValueError):
                continue
        out[tail] = sample
    wb.close()
    return out


def load_ai(d: Path, ids: list[str]) -> dict:
    out = {}
    aset = set(ids)
    for jp in sorted(d.glob("*_result.json")):
        sid = jp.stem.replace("_result", "")
        if sid not in aset:
            continue
        try:
            data = json.loads(jp.read_text(encoding="utf-8"))
        except Exception:
            continue
        sample = {}
        for ev in data.get("evaluations") or []:
            inn = ev.get("evaluation") if isinstance(ev.get("evaluation"), dict) else ev
            if not isinstance(inn, dict):
                continue
            num = inn.get("item_number")
            sc = inn.get("score")
            if isinstance(num, int) and isinstance(sc, (int, float)):
                sample[num] = int(sc)
        out[sid] = sample
    return out


def stats_for(item_num: int, gt: dict, ai: dict) -> dict:
    diffs = []
    ai0_h_pos = 0
    ai_pos_h0 = 0
    for sid in sorted(set(gt) & set(ai)):
        g = gt[sid].get(item_num); a = ai[sid].get(item_num)
        if g is None or a is None:
            continue
        d = a - g
        diffs.append(d)
        if a == 0 and g > 0:
            ai0_h_pos += 1
        if a > 0 and g == 0:
            ai_pos_h0 += 1
    n = len(diffs)
    if n == 0:
        return {"n": 0}
    abs_d = [abs(x) for x in diffs]
    return {
        "n": n,
        "mae": round(sum(abs_d) / n, 3),
        "bias": round(sum(diffs) / n, 3),
        "max_abs": max(abs_d),
        "exact": sum(1 for x in diffs if x == 0),
        "abs2": sum(1 for x in diffs if abs(x) == 2),
        "abs3p": sum(1 for x in diffs if abs(x) >= 3),
        "ai0_h_pos": ai0_h_pos,
        "ai_pos_h0": ai_pos_h0,
    }


def write_sheet(xlsx_path: Path, gt: dict, ai: dict, dataset_label: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 2)

    widths = [12, 90]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.cell(row, 1, f"[{dataset_label}] MAE 가 다른 항목 대비 큰 3 개 항목 (#6 / #8 / #10) — 원인 진단")
    ws.cell(row, 1).font = Font(bold=True, size=14, color="1f2937")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    row += 1
    ws.cell(row, 1,
            "각 항목별로 (1) 분포 통계 (2) AI vs 사람 평가 기준 충돌 진단 (3) 튜닝 권고. "
            "원문 케이스는 별도 시트 '큰차이 심층분석' 참조.")
    ws.cell(row, 1).font = Font(italic=True, color="6b7280", size=10)
    ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 30
    row += 2

    diagnoses = DIAGNOSES_BY_DATASET.get(dataset_label, DIAGNOSES)
    for num, diag in diagnoses.items():
        st = stats_for(num, gt, ai)
        if not st.get("n"):
            continue

        # ── 항목 헤더 ──
        ws.cell(row, 1, "항목")
        ws.cell(row, 2, f"#{num}  {diag['name']}  (max {ITEM_NUM_TO_MAX[num]})")
        for col in (1, 2):
            ws.cell(row, col).fill = _HEADER_FILL
            ws.cell(row, col).font = _HEADER_FONT
            ws.cell(row, col).alignment = Alignment(vertical="center", horizontal="left")
            ws.cell(row, col).border = _BORDER
        ws.row_dimensions[row].height = 24
        row += 1

        # ── 분포 통계 ──
        ws.cell(row, 1, "분포")
        ws.cell(row, 2,
                f"n={st['n']}  ·  MAE={st['mae']}  ·  Bias={st['bias']:+.2f}  ·  MaxAbs={st['max_abs']}  "
                f"·  정확일치 {st['exact']}건 / |Δ|=2 {st['abs2']}건 / |Δ|≥3 {st['abs3p']}건  "
                f"·  AI=0&사람>0  {st['ai0_h_pos']}건  ·  AI>0&사람=0  {st['ai_pos_h0']}건")
        for col in (1, 2):
            ws.cell(row, col).fill = _STAT_FILL
            ws.cell(row, col).font = Font(size=10, color="1f2937")
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, col).border = _BORDER
        ws.cell(row, 1).font = Font(bold=True, size=10, color="3730a3")
        ws.row_dimensions[row].height = 36
        row += 1

        # ── 진단 한 줄 ──
        ws.cell(row, 1, "진단")
        ws.cell(row, 2, diag["verdict"])
        for col in (1, 2):
            ws.cell(row, col).fill = _VERDICT_FILL
            ws.cell(row, col).font = Font(bold=True, size=11, color="92400e")
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row, col).border = _BORDER
        ws.cell(row, 1).font = Font(bold=True, size=11, color="92400e")
        ws.row_dimensions[row].height = 26
        row += 1

        # ── 원인 ──
        ws.cell(row, 1, "원인")
        ws.cell(row, 2, "다음 사실들이 전 케이스에서 일관되게 관측됨:")
        for col in (1, 2):
            ws.cell(row, col).fill = _REASON_FILL
            ws.cell(row, col).font = Font(bold=True, size=10, color="1f2937")
            ws.cell(row, col).border = _BORDER
        row += 1
        for r_text in diag["reasons"]:
            ws.cell(row, 1, "")
            ws.cell(row, 2, "  • " + r_text)
            for col in (1, 2):
                ws.cell(row, col).fill = _REASON_FILL
                ws.cell(row, col).font = Font(size=10, color="374151")
                ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row, col).border = _BORDER
            ws.row_dimensions[row].height = max(20, min(80, 14 + 14 * (len(r_text) // 70)))
            row += 1

        # ── 대표 샘플 (사람 vs AI) — 데이터셋별로 분리 ──
        cases_by_ds = REPRESENTATIVE_CASES_BY_DATASET.get(dataset_label, {})
        cases = cases_by_ds.get(num, REPRESENTATIVE_CASES.get(num, []))
        if cases:
            ws.cell(row, 1, "대표 샘플")
            ws.cell(row, 2, "사람 평가 vs AI 평가 — 같은 발화를 어떻게 다르게 봤는지")
            for col in (1, 2):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="ede9fe")
                ws.cell(row, col).font = Font(bold=True, size=10, color="5b21b6")
                ws.cell(row, col).border = _BORDER
            row += 1
            for case in cases:
                # 케이스 헤더
                ws.cell(row, 1, f"  [{case['sid']}]")
                ws.cell(row, 2, f"AI={case['ai_score']}점  /  사람={case['human_score']}점  /  Δ={case['diff']:+d}")
                for col in (1, 2):
                    ws.cell(row, col).fill = PatternFill("solid", fgColor="faf5ff")
                    ws.cell(row, col).font = Font(bold=True, size=10,
                        color="b91c1c" if case['diff'] > 0 else "1d4ed8")
                    ws.cell(row, col).border = _BORDER
                row += 1
                # 사람 view
                ws.cell(row, 1, "  사람:")
                ws.cell(row, 2, case["human_view"])
                for col in (1, 2):
                    ws.cell(row, col).fill = PatternFill("solid", fgColor="f0fdf4")
                    ws.cell(row, col).font = Font(size=10, color="166534",
                        bold=(col == 1))
                    ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row, col).border = _BORDER
                ws.row_dimensions[row].height = max(24, min(120, 14 + 14 * (len(case["human_view"]) // 70)))
                row += 1
                # AI view
                ws.cell(row, 1, "  AI:")
                ws.cell(row, 2, case["ai_view"])
                for col in (1, 2):
                    ws.cell(row, col).fill = PatternFill("solid", fgColor="faf5ff")
                    ws.cell(row, col).font = Font(size=10, color="5b21b6",
                        bold=(col == 1))
                    ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row, col).border = _BORDER
                ws.row_dimensions[row].height = max(24, min(120, 14 + 14 * (len(case["ai_view"]) // 70)))
                row += 1
                # 차이 발생 원인
                ws.cell(row, 1, "  → 차이 원인:")
                ws.cell(row, 2, case["why_gap"])
                for col in (1, 2):
                    ws.cell(row, col).fill = PatternFill("solid", fgColor="fef3c7")
                    ws.cell(row, col).font = Font(size=10, color="92400e",
                        bold=(col == 1), italic=True)
                    ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row, col).border = _BORDER
                ws.row_dimensions[row].height = max(24, min(80, 14 + 14 * (len(case["why_gap"]) // 70)))
                row += 1

        # ── 튜닝 권고 ──
        ws.cell(row, 1, "튜닝 권고")
        ws.cell(row, 2, "")
        for col in (1, 2):
            ws.cell(row, col).fill = _FIX_FILL
            ws.cell(row, col).font = Font(bold=True, size=10, color="166534")
            ws.cell(row, col).border = _BORDER
        row += 1
        for f_text in diag["fix"]:
            ws.cell(row, 1, "")
            ws.cell(row, 2, "  ▸ " + f_text)
            for col in (1, 2):
                ws.cell(row, col).fill = _FIX_FILL
                ws.cell(row, col).font = Font(size=10, color="166534")
                ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row, col).border = _BORDER
            ws.row_dimensions[row].height = max(20, min(80, 14 + 14 * (len(f_text) // 70)))
            row += 1

        row += 1  # 항목 간 간격

    # 마무리 — 우선순위 (데이터셋별)
    ws.cell(row, 1, "우선순위")
    if dataset_label == "테스트셋":
        priority_text = (
            "[1순위] #6 정중한 표현 — 학습셋과 동일하게 filler 트리거 제거로 즉시 해결. "
            "[2순위] #10 설명의 명확성 — 학습셋과 동일하게 만점 조건을 '되물음 0~1회' 로 재정의. "
            "[3순위] #8 문의 파악 — **AI 측 책임 아님**. GT note 8건이 '복창=False 인데 만점' 으로 라벨링되어 "
            "프롬프트로 해결 불가. GT 작성자에게 진짜 채점 기준 확인 필요. 트래킹 시 별도 분리 권장."
        )
    else:
        priority_text = (
            "[1순위] #6 정중한 표현 — 프롬프트의 'filler 3회 연속' 감점 트리거 제거 (1줄 수정으로 큰 효과). "
            "[2순위] #10 설명의 명확성 — 만점 조건을 '고객 되물음 0~1회' 로 단순화. 형식 트리거 (장황/내부용어) 제거. "
            "[3순위] #8 문의 파악 — '복창' 정의를 '핵심 명사 재등장' 으로 재정의 (사람 기준과 일치)."
        )
    ws.cell(row, 2, priority_text)
    for col in (1, 2):
        ws.cell(row, col).fill = PatternFill("solid", fgColor="fbbf24")
        ws.cell(row, col).font = Font(bold=True, size=11, color="78350f")
        ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, col).border = _BORDER
    ws.row_dimensions[row].height = 60

    ws.freeze_panes = "A4"

    try:
        wb.save(xlsx_path)
    except PermissionError:
        fb = xlsx_path.with_name(xlsx_path.stem + "_with_high_mae.xlsx")
        wb.save(fb)
        print(f"  ⚠ 원본 잠김 → fallback 저장: {fb}")


def main() -> int:
    targets = [
        ("학습셋", TRAINING_IDS,
         Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409"),
         Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409\비교분석_20260422_011144.xlsx")),
        ("테스트셋", TEST_IDS,
         Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545"),
         Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545\비교분석_20260422_011144.xlsx")),
    ]
    for label, ids, results_dir, xlsx_path in targets:
        if not xlsx_path.exists():
            print(f"[SKIP] {xlsx_path} 없음")
            continue
        gt = load_gt(ids)
        ai = load_ai(results_dir, ids)
        write_sheet(xlsx_path, gt, ai, label)
        print(f"[{label}] '{SHEET_NAME}' 시트 갱신 → {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
