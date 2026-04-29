# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""Golden-set + Reasoning-index 재구축 (v4 — fixed.xlsx 기반).

입력:
  - fixed.xlsx (수정된 사람 평가 + 검증된 rationale)
  - 학습셋/테스트셋 transcripts
산출:
  - tenants/kolon/golden_set/<NN>_<slug>.json  (17 항목)
  - tenants/kolon/reasoning_index/<NN>_<slug>.json  (17 항목)

생성 규칙:
  - 대화 1건 × 17 항목 = 17 레코드 (variant 제거, 대화당 1개씩)
  - segment_text = xlsx 비고의 인용 turn ± 1 context (항목별 근거 구간)
  - rationale = fixed.xlsx 비고 그대로
  - intent = 파일명/비고에서 추론 (primary 1개)
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

FIXED_XLSX = Path(r"C:/Users/META M/Desktop/QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
TRAIN_DIR = Path(r"C:/Users/META M/Desktop/qa 샘플/학습셋")
TEST_DIR = Path(r"C:/Users/META M/Desktop/qa 샘플/테스트셋")

TENANT_DIR = Path(__file__).parent.parent / "tenants" / "kolon"
GOLDEN_DIR = TENANT_DIR / "golden_set"
REASONING_DIR = TENANT_DIR / "reasoning_index"

ITEM_NAME_MAP = {
    "첫인사": 1, "끝인사": 2, "호응및공감": 4, "대기멘트": 5,
    "정중한표현": 6, "쿠션어활용": 7,
    # 괄호 제거 후 정규화 결과도 매핑 (`re.sub` 이 괄호 내용 제거 → "문의파악및재확인")
    "문의파악및재확인(복창)": 8, "문의파악및재확인": 8,
    "고객정보확인": 9,
    "설명의명확성": 10, "두괄식답변": 11, "문제해결의지": 12,
    "부연설명및추가안내": 13, "사후안내": 14,
    "정확한안내★": 15, "정확한안내": 15, "필수안내이행": 16,
    "정보확인절차": 17, "정보보호준수": 18,
}

# RAG 미사용 항목 — Rule 기반 / 금지어 사전 / compliance-based 로 평가하므로
# golden_set / reasoning_index 에 시딩하지 않는다.
# (#1 첫인사, #2 끝인사: Rule + LLM verify — 고정 구간 평가)
# (#6 정중한 표현: LLM + 금지어 사전)
# (#9 고객정보 확인: structural_only — 마스킹으로 내용 검증 불가)
# (#16 필수 안내 이행: Intent 분류 + 스크립트 매칭)
# (#17 정보 확인 절차, #18 정보 보호 준수: compliance_based + T3 라우팅)
RAG_DISABLED_ITEMS: set[int] = {1, 2, 6, 9, 16, 17, 18}
RAG_ENABLED_ITEMS: set[int] = {4, 5, 7, 8, 10, 11, 12, 13, 14, 15}
ITEM_MAX = {1:5,2:5,4:5,5:5,6:5,7:5,8:5,9:5,10:10,11:5,12:5,13:5,14:5,15:15,16:5,17:5,18:5}
ITEM_KO = {
    1: "첫인사", 2: "끝인사", 4: "호응 및 공감", 5: "대기 멘트",
    6: "정중한 표현", 7: "쿠션어 활용", 8: "문의 파악 및 재확인(복창)", 9: "고객정보 확인",
    10: "설명의 명확성", 11: "두괄식 답변", 12: "문제 해결 의지",
    13: "부연 설명 및 추가 안내", 14: "사후 안내",
    15: "정확한 안내", 16: "필수 안내 이행", 17: "정보 확인 절차", 18: "정보 보호 준수",
}
ITEM_CATEGORY = {
    1: "인사 예절", 2: "인사 예절",
    4: "경청 및 소통", 5: "경청 및 소통",
    6: "언어 표현", 7: "언어 표현",
    8: "니즈 파악", 9: "니즈 파악",
    10: "설명력 및 전달력", 11: "설명력 및 전달력",
    12: "적극성", 13: "적극성", 14: "적극성",
    15: "업무 정확도", 16: "업무 정확도",
    17: "개인정보 보호", 18: "개인정보 보호",
}
ITEM_SLUG = {
    1: "first_greeting", 2: "closing_greeting", 4: "empathy", 5: "hold_notice",
    6: "polite_language", 7: "cushion_words", 8: "needs_identification",
    9: "customer_info_verification", 10: "explanation_clarity", 11: "top_down_answer",
    12: "problem_solving_attitude", 13: "additional_guidance", 14: "follow_up",
    15: "correct_information", 16: "mandatory_notice", 17: "pii_verification",
    18: "privacy_compliance",
}

INTENT_KEYWORDS = [
    ("환불취소", ["교환", "반품", "취소", "환불", "돌려받"]),
    ("주문배송", ["배송", "주문", "도착", "배달", "운송장", "송장"]),
    ("상품문의", ["불량", "사이즈", "상품", "제품", "가격", "AS"]),
    ("결제문의", ["결제", "카드", "이체"]),
    ("변경해지", ["변경", "해지", "해약"]),
    ("장애문의", ["오류", "고장", "문제"]),
    ("가입상담", ["가입", "신청", "등록"]),
]

CITE_RE = re.compile(r'(상담사|고객)#(\d+)\s*:\s*"')

MAX_SEGMENT_LEN = 1400
GAP_MARKER = "..."
GAP_THRESHOLD = 3


def load_turns(sid: int) -> dict[int, dict[str, str]]:
    # 학습셋만 — 테스트셋 전사(JSON)는 Golden-set 시드로 사용 금지 (hold-out 용)
    ps = list(TRAIN_DIR.glob(f"{sid}_*.json"))
    if ps:
        raw = json.load(open(ps[0], encoding="utf-8")).get("transcript", "")
    else:
        return {}
    out: dict[int, dict[str, str]] = {}
    tid = 0
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(상담사|고객):\s*(.*)$", line)
        if not m:
            if tid > 0:
                out[tid]["text"] += " " + line
            continue
        tid += 1
        out[tid] = {"speaker": m.group(1), "text": m.group(2)}
    return out


def infer_intent(sid: int, rationale: str) -> str:
    candidates = []
    ps = list(TRAIN_DIR.glob(f"{sid}_*.json"))
    if ps:
        candidates.append(ps[0].stem.lower())
    haystack = " ".join(candidates) + " " + rationale.lower()
    for label, kws in INTENT_KEYWORDS:
        for kw in kws:
            if kw in haystack:
                return label
    return "일반문의"


def parse_xlsx() -> dict[int, dict[int, dict]]:
    wb = load_workbook(FIXED_XLSX, read_only=True, data_only=True)
    out: dict[int, dict[int, dict]] = {}
    for sheet in wb.sheetnames:
        m = re.search(r"(\d{6})", sheet)
        if not m:
            continue
        sid = int(m.group(1))
        ws = wb[sheet]
        per_item: dict[int, dict] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[1] is None:
                continue
            name = re.sub(r"\s+|\([^)]*\)|\n", "", str(row[1]))
            inum = ITEM_NAME_MAP.get(name)
            if inum is None:
                continue
            score = row[4]
            rationale = row[5] or ""
            if score is None:
                continue
            try:
                score = int(score)
            except (TypeError, ValueError):
                continue
            per_item[inum] = {"score": score, "rationale": str(rationale)}
        out[sid] = per_item
    return out


def extract_cited_turns(rationale: str) -> list[int]:
    ids = []
    for _, n in CITE_RE.findall(rationale):
        try:
            ids.append(int(n))
        except ValueError:
            pass
    return sorted(set(ids))


def build_segment_text(cited: list[int], turns: dict[int, dict], *, context_before=1, context_after=1) -> str:
    if not cited:
        return ""
    expanded: set[int] = set()
    max_tid = max(turns.keys()) if turns else 0
    for tid in cited:
        for off in range(-context_before, context_after + 1):
            nb = tid + off
            if 1 <= nb <= max_tid and nb in turns:
                expanded.add(nb)
    ordered = sorted(expanded)
    lines: list[str] = []
    prev = None
    for tid in ordered:
        if prev is not None and tid - prev > GAP_THRESHOLD:
            lines.append(GAP_MARKER)
        lines.append(f"{turns[tid]['speaker']}: {turns[tid]['text']}")
        prev = tid
    seg = "\n".join(lines)
    if len(seg) > MAX_SEGMENT_LEN:
        seg = seg[:MAX_SEGMENT_LEN] + "..."
    return seg


def fallback_cited_turns(inum: int, turns: dict[int, dict]) -> list[int]:
    agent_tids = [tid for tid in sorted(turns) if turns[tid]["speaker"] == "상담사"]
    if inum == 1:
        return agent_tids[:1]
    if inum == 2:
        return sorted(agent_tids[-3:]) if len(agent_tids) >= 3 else agent_tids
    long_agent = [tid for tid in agent_tids if len(turns[tid]["text"]) > 30]
    return long_agent[:3]


def build_record(sid: int, inum: int, score: int, rationale: str, turns: dict[int, dict], intent: str) -> dict:
    cited = extract_cited_turns(rationale)
    if not cited:
        cited = fallback_cited_turns(inum, turns)
    before, after = 1, 1
    if inum == 15:
        before, after = 2, 2
    if inum == 1:
        before, after = 0, 2
    if inum == 2:
        before, after = 2, 0
    segment = build_segment_text(cited, turns, context_before=before, context_after=after)

    max_s = ITEM_MAX[inum]
    if score == max_s:
        bucket = "full"
    elif score == 0:
        bucket = "zero"
    else:
        bucket = "partial"

    evidence_refs = []
    for tid in cited:
        t = turns.get(tid)
        if t:
            evidence_refs.append({"speaker": t["speaker"], "turn_id": tid})

    return {
        "example_id": f"GS-{inum:02d}-{bucket.upper()}-{sid}",
        "item_number": inum,
        "score": score,
        "score_bucket": bucket,
        "intent": intent,
        "segment_text": segment,
        "rationale": rationale.strip(),
        "rationale_tags": [f"item_{inum}", f"{bucket}_compliance", "aggregated"],
        "evidence_refs": evidence_refs,
        "rater_meta": {
            "rater_type": "senior_consensus",
            "source": "kolon_v4_fixed",
            "sample_id": str(sid),
            "variant": "aggregated",
            "human_score_verified": True,
        },
    }


def build_reasoning_record(sid: int, inum: int, score: int, rationale: str, turns: dict[int, dict]) -> dict:
    cited = extract_cited_turns(rationale)
    quote_example = ""
    if cited:
        t = turns.get(cited[0])
        if t:
            quote_example = f"{t['speaker']}: {t['text'][:200]}"
    max_s = ITEM_MAX[inum]
    if score == max_s:
        bucket = "full"
    elif score == 0:
        bucket = "zero"
    else:
        bucket = "partial"
    return {
        "record_id": f"r_{inum:02d}_{sid}",
        "score": score,
        "rationale": rationale.strip(),
        "quote_example": quote_example,
        "evaluator_id": "kolon_v4_senior",
        "tags": [f"item_{inum}", f"{bucket}_compliance", "aggregated"],
        "stub_seed": False,
        "sample_id": str(sid),
    }


def get_training_sids() -> set[int]:
    sids: set[int] = set()
    for p in TRAIN_DIR.glob("*.json"):
        m = re.match(r"(\d{6})_", p.name)
        if m:
            sids.add(int(m.group(1)))
    return sids


def main():
    xlsx = parse_xlsx()
    training_sids = get_training_sids()
    # 학습셋에 포함된 샘플만 선별 (테스트셋 샘플 제외)
    xlsx = {sid: items for sid, items in xlsx.items() if sid in training_sids}
    print(f"학습셋 샘플: {len(training_sids)}개 — {sorted(training_sids)}")
    print(f"xlsx × 학습셋 매칭: {len(xlsx)}개")

    # 기존 JSON 전면 삭제
    for d in (GOLDEN_DIR, REASONING_DIR):
        if d.exists():
            for p in d.glob("*.json"):
                if p.name.startswith("_"):
                    continue
                p.unlink()
            print(f"기존 JSON 삭제: {d}")
    GOLDEN_DIR.mkdir(exist_ok=True, parents=True)
    REASONING_DIR.mkdir(exist_ok=True, parents=True)

    per_item_gs: dict[int, list[dict]] = {i: [] for i in RAG_ENABLED_ITEMS}
    per_item_rs: dict[int, list[dict]] = {i: [] for i in RAG_ENABLED_ITEMS}

    total_records = 0
    skipped_disabled = 0
    for sid, items in sorted(xlsx.items()):
        turns = load_turns(sid)
        if not turns:
            print(f"  ⚠ sample {sid} transcript 없음 — 스킵")
            continue
        for inum, meta in items.items():
            if inum in RAG_DISABLED_ITEMS:
                skipped_disabled += 1
                continue
            intent = infer_intent(sid, meta["rationale"])
            gs = build_record(sid, inum, meta["score"], meta["rationale"], turns, intent)
            rs = build_reasoning_record(sid, inum, meta["score"], meta["rationale"], turns)
            per_item_gs[inum].append(gs)
            per_item_rs[inum].append(rs)
            total_records += 1

    print(f"\nRAG 비사용 항목 스킵: {skipped_disabled}건 (items={sorted(RAG_DISABLED_ITEMS)})")
    for inum in sorted(RAG_ENABLED_ITEMS):
        slug = ITEM_SLUG[inum]
        gs_path = GOLDEN_DIR / f"{inum:02d}_{slug}.json"
        rs_path = REASONING_DIR / f"{inum:02d}_{slug}.json"

        gs_body = {
            "item_number": inum,
            "item_name": ITEM_KO[inum],
            "category": ITEM_CATEGORY[inum],
            "max_score": ITEM_MAX[inum],
            "allowed_steps": [ITEM_MAX[inum], 3, 0] if ITEM_MAX[inum] == 5 else (
                [10, 7, 5, 0] if inum == 10 else [15, 10, 5, 0]
            ),
            "intents": sorted({ex["intent"] for ex in per_item_gs[inum]}) or ["*"],
            "version": "kolon_v4_fixed",
            "examples": per_item_gs[inum],
        }
        rs_body = {
            "item_number": inum,
            "item_name": ITEM_KO[inum],
            "version": "kolon_v4_fixed",
            "reasoning_records": per_item_rs[inum],
        }
        gs_path.write_text(json.dumps(gs_body, ensure_ascii=False, indent=2), encoding="utf-8")
        rs_path.write_text(json.dumps(rs_body, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n{'='*80}")
    print(f"총 레코드: {total_records}")
    print(f"{'='*80}")
    for inum in sorted(RAG_ENABLED_ITEMS):
        n_gs = len(per_item_gs[inum])
        n_rs = len(per_item_rs[inum])
        print(f"  item #{inum:2d} ({ITEM_KO[inum]:20s}): golden_set {n_gs:3d}개 · reasoning {n_rs:3d}개")
    print(f"\n  (RAG 미사용: items {sorted(RAG_DISABLED_ITEMS)} — 파일 생성 안 함)")

    print(f"\n저장:")
    print(f"  Golden-set:      {GOLDEN_DIR}")
    print(f"  Reasoning-index: {REASONING_DIR}")


if __name__ == "__main__":
    main()
