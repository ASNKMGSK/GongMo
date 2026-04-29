# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""Golden-set 전수 품질 감사 스크립트.

입력:
  - xlsx: C:/Users/META M/Desktop/QA정답-STT_기반_통합_상담평가표_v3재평가.xlsx  (사람 평가 GT)
  - transcripts: C:/Users/META M/Desktop/qa 샘플/학습셋/  (원본 STT 전사)
  - golden_set: v2/tenants/kolon/golden_set/  (AOSS 에 시딩된 레코드)

출력: stdout 리포트
  점검 A — xlsx ↔ golden_set 점수 일치도
  점검 B — rationale 논리 모순 (예: "복창✓=False → 만점")
  점검 C — Evidence 인용 ↔ transcript turn 일치
  점검 D — Self-retrieval 중복 경고
  점검 E — 오프닝 지배 segment 탐지
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

XLSX_PATH = r"C:/Users/META M/Desktop/QA정답-STT_기반_통합_상담평가표_v3재평가.xlsx"
TRANSCRIPT_DIR = Path(r"C:/Users/META M/Desktop/qa 샘플/학습셋")
GOLDEN_DIR = Path(__file__).parent.parent / "tenants" / "kolon" / "golden_set"

# xlsx 평가항목명 → item_number 매핑 (개행 정규화)
ITEM_NAME_MAP = {
    "첫인사": 1,
    "끝인사": 2,
    "호응및공감": 4,
    "대기멘트": 5,
    "정중한표현": 6,
    "쿠션어활용": 7,
    "문의파악및재확인(복창)": 8,
    "고객정보확인": 9,
    "설명의명확성": 10,
    "두괄식답변": 11,
    "문제해결의지": 12,
    "부연설명및추가안내": 13,
    "사후안내": 14,
    "정확한안내★": 15,
    "정확한안내": 15,
    "필수안내이행": 16,
    "정보확인절차": 17,
    "정보보호준수": 18,
}


# ---------------------------------------------------------------------------
# xlsx 파싱
# ---------------------------------------------------------------------------


def parse_xlsx() -> dict[int, dict[int, dict[str, Any]]]:
    """{sample_id: {item_number: {"score": int, "rationale": str, "raw_item_name": str}}}"""
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    out: dict[int, dict[int, dict[str, Any]]] = {}
    for sheet in wb.sheetnames:
        m = re.search(r"(\d{6})", sheet)
        if not m:
            continue
        sample_id = int(m.group(1))
        ws = wb[sheet]
        per_item: dict[int, dict[str, Any]] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 6:
                continue
            raw_item = row[1]
            if raw_item is None:
                continue
            name = re.sub(r"\s+|\([^)]*\)|\n", "", str(raw_item))
            item_num = ITEM_NAME_MAP.get(name)
            if item_num is None:
                continue
            score = row[4]
            rationale = row[5] or ""
            if score is None:
                continue
            try:
                score = int(score)
            except (TypeError, ValueError):
                continue
            per_item[item_num] = {
                "score": score,
                "rationale": str(rationale),
                "raw_item_name": str(raw_item),
            }
        out[sample_id] = per_item
    return out


# ---------------------------------------------------------------------------
# transcript 파싱
# ---------------------------------------------------------------------------


def parse_transcripts() -> dict[int, dict[int, str]]:
    """{sample_id: {turn_id: "speaker: text"}}"""
    out: dict[int, dict[int, str]] = {}
    for path in TRANSCRIPT_DIR.glob("*.json"):
        m = re.match(r"(\d{6})_", path.name)
        if not m:
            continue
        sid = int(m.group(1))
        data = json.load(open(path, encoding="utf-8"))
        turns: dict[int, str] = {}
        # 다양한 스키마 지원 탐색
        turn_list = None
        if isinstance(data, list):
            turn_list = data
        elif isinstance(data, dict):
            for key in ("turns", "utterances", "transcript", "conversation"):
                if key in data and isinstance(data[key], list):
                    turn_list = data[key]
                    break
        if turn_list is None:
            continue
        for t in turn_list:
            if not isinstance(t, dict):
                continue
            tid = t.get("turn_id") or t.get("turn") or t.get("id")
            speaker = t.get("speaker", "")
            text = t.get("text") or t.get("utterance") or t.get("content") or ""
            if tid is None:
                continue
            try:
                tid = int(tid)
            except (TypeError, ValueError):
                continue
            spk_ko = {"agent": "상담사", "customer": "고객"}.get(speaker, speaker)
            turns[tid] = f"{spk_ko}: {text}".strip()
        out[sid] = turns
    return out


# ---------------------------------------------------------------------------
# golden_set 로드
# ---------------------------------------------------------------------------


def load_golden_set() -> dict[int, list[dict]]:
    """{item_number: [example_dict, ...]}"""
    out: dict[int, list[dict]] = {}
    for path in sorted(GOLDEN_DIR.glob("[0-9][0-9]_*.json")):
        data = json.load(open(path, encoding="utf-8"))
        item_num = int(data.get("item_number") or 0)
        out[item_num] = data.get("examples") or []
    return out


# ---------------------------------------------------------------------------
# 점검 A — 점수 일치도
# ---------------------------------------------------------------------------


def audit_A_score_match(xlsx_data, golden):
    print("=" * 80)
    print("[점검 A] xlsx ↔ golden_set 점수 일치도 (aggregated variant)")
    print("=" * 80)
    mismatches = []
    missing = []
    for sample_id, items in xlsx_data.items():
        for item_num, gt in items.items():
            examples = golden.get(item_num, [])
            # aggregated variant — example_id 에 sample_id 포함, T 접미사 없음
            agg = [
                ex for ex in examples
                if str(sample_id) in str(ex.get("example_id", ""))
                and "aggregated" in (ex.get("rationale_tags") or [])
            ]
            if not agg:
                missing.append((sample_id, item_num, gt["score"]))
                continue
            gs_score = agg[0].get("score")
            if gs_score != gt["score"]:
                mismatches.append((sample_id, item_num, gt["score"], gs_score, agg[0].get("example_id")))
    print(f"  총 xlsx 레코드 수: {sum(len(v) for v in xlsx_data.values())}")
    print(f"  ❌ 점수 불일치: {len(mismatches)}")
    for sid, inum, gt_sc, gs_sc, eid in mismatches[:30]:
        print(f"    sample={sid} item={inum}: xlsx={gt_sc} vs golden={gs_sc} ({eid})")
    print(f"  ⚠ aggregated variant 누락: {len(missing)}")
    for sid, inum, sc in missing[:15]:
        print(f"    sample={sid} item={inum} (xlsx 점수={sc})")
    return mismatches, missing


# ---------------------------------------------------------------------------
# 점검 B — rationale 논리 모순
# ---------------------------------------------------------------------------


def audit_B_rationale_logic(golden):
    print()
    print("=" * 80)
    print("[점검 B] rationale 논리 모순 탐지")
    print("=" * 80)
    issues: list[tuple[str, int, int, str]] = []
    # 패턴: "복창✓=False → 만점", "복창✓=True → 0점", "감점 0회 → 3점" 등
    contradict_patterns = [
        (r"복창✓=False[^→]*→\s*만점", "복창 안 했는데 만점"),
        (r"복창✓=True[^→]*→\s*0점", "복창 했는데 0점"),
        (r"결론선행✗[^→]*→\s*만점", "두괄식 안 지켰는데 만점"),
        (r"감점\s*0[회건][^→]*→\s*3점", "감점 0인데 3점"),
        (r"감점\s*0[회건][^→]*→\s*0점", "감점 0인데 0점"),
        (r"복창\s*누락[^→]*→\s*만점", "복창 누락인데 만점"),
    ]
    for item_num, examples in golden.items():
        for ex in examples:
            rat = ex.get("rationale", "") or ""
            score = ex.get("score")
            eid = ex.get("example_id", "")
            for pat, desc in contradict_patterns:
                if re.search(pat, rat):
                    issues.append((eid, item_num, score, f"{desc}: {rat[:120]}"))
                    break
            # score ↔ rationale 끝 단어 불일치 추가 체크
            m_score = re.search(r"→\s*(만점|\d+점)", rat)
            if m_score:
                tail = m_score.group(1)
                if tail == "만점":
                    expected_max = {1:5,2:5,4:5,5:5,6:5,7:5,8:5,9:5,10:10,11:5,12:5,13:5,14:5,15:15,16:5,17:5,18:5}
                    if score != expected_max.get(item_num):
                        issues.append((eid, item_num, score,
                            f"rationale='만점' but score={score} (expected {expected_max.get(item_num)}): {rat[:100]}"))
                else:
                    m_num = re.match(r"(\d+)점", tail)
                    if m_num and score != int(m_num.group(1)):
                        issues.append((eid, item_num, score,
                            f"rationale='{tail}' but score={score}: {rat[:100]}"))
    print(f"  발견된 모순/오류: {len(issues)}")
    for eid, inum, sc, msg in issues[:50]:
        print(f"    [{eid}] item={inum} score={sc} — {msg}")
    if len(issues) > 50:
        print(f"    ... + {len(issues) - 50} more")

    # 패턴별 집계
    type_counter = Counter()
    for eid, inum, sc, msg in issues:
        # 첫 콜론 앞부분을 타입으로
        type_key = msg.split(":")[0][:50]
        type_counter[type_key] += 1
    print("\n  [유형별 집계]")
    for t, c in type_counter.most_common():
        print(f"    {c:4d}  {t}")
    return issues


# ---------------------------------------------------------------------------
# 점검 C — Evidence 인용 ↔ transcript 일치
# ---------------------------------------------------------------------------


def audit_C_evidence_citation(golden, transcripts):
    print()
    print("=" * 80)
    print("[점검 C] rationale 인용 ↔ transcript turn 정확도 (샘플링)")
    print("=" * 80)
    mismatches = []
    # rationale 에 " - 상담사#12: "..." " 같은 인용이 있으면 대조
    cite_re = re.compile(r'- (상담사|고객)#(\d+):\s*"([^"]*)"')
    checked = 0
    for item_num, examples in golden.items():
        for ex in examples:
            eid = ex.get("example_id", "")
            rat = ex.get("rationale", "") or ""
            sample_id = None
            for part in str(eid).split("-"):
                if part.isdigit() and len(part) == 6:
                    sample_id = int(part)
                    break
            if sample_id is None:
                continue
            turns = transcripts.get(sample_id)
            if not turns:
                continue
            for spk, tid, quote in cite_re.findall(rat):
                checked += 1
                tid_i = int(tid)
                actual = turns.get(tid_i, "")
                if quote and quote not in actual:
                    mismatches.append((eid, spk, tid_i, quote[:60], actual[:60]))
    print(f"  검증 인용: {checked}")
    print(f"  ❌ 불일치: {len(mismatches)}")
    for eid, spk, tid, quote, actual in mismatches[:20]:
        print(f"    [{eid}] {spk}#{tid}")
        print(f"       rationale 인용: {quote!r}")
        print(f"       실제 turn:      {actual!r}")
    return mismatches


# ---------------------------------------------------------------------------
# 점검 D — Self-retrieval 중복
# ---------------------------------------------------------------------------


def audit_D_self_retrieval(golden):
    print()
    print("=" * 80)
    print("[점검 D] Self-retrieval 위험 — 샘플별 variant 생성 카운트")
    print("=" * 80)
    sample_item_counter: dict[tuple[int, int], int] = Counter()
    for item_num, examples in golden.items():
        for ex in examples:
            eid = str(ex.get("example_id", ""))
            for part in eid.split("-"):
                if part.isdigit() and len(part) == 6:
                    sample_item_counter[(int(part), item_num)] += 1
                    break
    # 샘플당 item 당 variant 수 분포
    variant_distribution = Counter(sample_item_counter.values())
    print("  [sample × item] variant 수 분포:")
    for n, cnt in sorted(variant_distribution.items()):
        print(f"    variant {n:2d} 개: {cnt:4d} (sample,item) 쌍")
    # 상위 레코드
    print("  [variant 수 TOP 10]")
    for (sid, inum), n in Counter(sample_item_counter).most_common(10):
        print(f"    sample={sid} item={inum}: {n}개 variant")


# ---------------------------------------------------------------------------
# 점검 E — 오프닝 지배 segment 탐지
# ---------------------------------------------------------------------------


def audit_E_opening_dominant(golden):
    print()
    print("=" * 80)
    print("[점검 E] 오프닝 멘트 지배 segment (turn_window 길이 < 오프닝 길이 * 2)")
    print("=" * 80)
    OPENING_MARKER = "반갑습니다 코오롱 고객센터"
    danger = []
    for item_num, examples in golden.items():
        for ex in examples:
            tags = ex.get("rationale_tags") or []
            if "turn_window" not in tags:
                continue
            seg = ex.get("segment_text", "") or ""
            if OPENING_MARKER not in seg:
                continue
            # 오프닝 이후 내용이 얼마나 되나
            idx = seg.find(OPENING_MARKER)
            after = seg[idx:]
            # 오프닝 포함 세그먼트 중 너무 짧은 것
            if len(after) < 250:  # 오프닝 1~2줄 + 간단한 응답 정도
                danger.append((ex.get("example_id"), item_num, len(seg), seg.replace("\n", " | ")[:150]))
    print(f"  오프닝 포함 + 짧은 turn_window: {len(danger)}")
    for eid, inum, slen, preview in danger[:15]:
        print(f"    [{eid}] item={inum} len={slen}")
        print(f"       {preview}")


# ---------------------------------------------------------------------------
# 점검 F — intent 분포
# ---------------------------------------------------------------------------


def audit_F_intent_distribution(golden):
    print()
    print("=" * 80)
    print("[점검 F] intent 분포")
    print("=" * 80)
    per_item: dict[int, Counter] = defaultdict(Counter)
    for item_num, examples in golden.items():
        for ex in examples:
            per_item[item_num][ex.get("intent") or "None"] += 1
    for inum in sorted(per_item.keys()):
        print(f"  item #{inum}: ", dict(per_item[inum].most_common()))


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main():
    print("\n--- golden_set 로드 ---")
    golden = load_golden_set()
    total = sum(len(v) for v in golden.values())
    print(f"  items: {len(golden)}, total examples: {total}")

    print("\n--- xlsx 로드 ---")
    xlsx = parse_xlsx()
    print(f"  samples: {len(xlsx)}, total item scores: {sum(len(v) for v in xlsx.values())}")

    print("\n--- transcripts 로드 ---")
    trans = parse_transcripts()
    print(f"  samples: {len(trans)}, total turns: {sum(len(v) for v in trans.values())}")

    audit_A_score_match(xlsx, golden)
    audit_B_rationale_logic(golden)
    audit_C_evidence_citation(golden, trans)
    audit_D_self_retrieval(golden)
    audit_E_opening_dominant(golden)
    audit_F_intent_distribution(golden)


if __name__ == "__main__":
    main()
