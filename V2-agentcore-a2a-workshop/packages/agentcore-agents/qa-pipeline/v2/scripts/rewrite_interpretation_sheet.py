# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""기존 비교분석 xlsx 의 '해석 노트' 시트를 평이한 한국어로 재작성.

기존 시트 문제:
  - MAE/RMSE/Bias/MAPE/Under%/Over%/체계적 과소/rubric drift 등 용어 그대로 노출
  - 사람이 읽으면 한 번에 파악 안 됨

새 시트:
  - "AI 와 사람 점수가 평균 N점 차이남" 같은 평이한 문장
  - 항목별로 "AI 가 짜게 줬다 / 후하게 줬다 + 왜" 한 줄로
  - 가장 어긋난 상담 5건은 "총점 X점 차이 — 어느 항목 때문" 으로

실행:
  python -m v2.scripts.rewrite_interpretation_sheet
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_PIPELINE_DIR = Path(__file__).resolve().parents[2]
if str(_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(_PIPELINE_DIR))

from v2.scripts.compare_learning_set_vs_xlsx import (  # type: ignore[import-untyped]
    ITEM_DEF, ITEM_NUM_TO_MAX, ITEM_NUM_TO_NAME,
    TEST_IDS, TRAINING_IDS,
    compute_metrics, load_ai_results, load_xlsx_ground_truth,
)


XLSX_GT = Path(r"C:\Users\META M\Desktop\QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
SHEET_NAME = "해석 노트"

# 색
TITLE_FILL = PatternFill("solid", fgColor="1f2937")
TITLE_FONT = Font(color="fbbf24", bold=True, size=14)
SECTION_FILL = PatternFill("solid", fgColor="6b21a8")
SECTION_FONT = Font(color="ffffff", bold=True, size=12)
KEY_FILL = PatternFill("solid", fgColor="f3f4f6")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="fef3c7")
BAD_FILL = PatternFill("solid", fgColor="fee2e2")
GOOD_FILL = PatternFill("solid", fgColor="dcfce7")
BORDER = Border(left=Side("thin", color="cbd5e1"), right=Side("thin", color="cbd5e1"),
                top=Side("thin", color="cbd5e1"), bottom=Side("thin", color="cbd5e1"))
WRAP = Alignment(wrap_text=True, vertical="top")


# 항목별 "왜 차이 났는지" 한 줄 — 우리가 이미 분석한 결과 기반
WHY_GAP: dict[int, str] = {
    1:  "잘 맞음 — 첫인사 3요소 (인사말+소속+상담사명) 같이 평가",
    2:  "AI 가 짜게 — '추가 안내 (TM/안전운전)' 부분을 끝인사로 인정 안 함",
    4:  "엇갈림 — '네/예' 단순 호응을 AI 는 공감으로 인정, 사람은 안 함 (반대 케이스)",
    5:  "AI 가 짜게 — '잠시만요' 같은 명시적 대기 발화를 못 찾음. STT 누락 영향도",
    6:  "AI 가 짜게 — '음/아/이게' filler 빈도로 감점, 사람은 사물존칭/반말 0회면 만점",
    7:  "엇갈림 — '실례지만/번거로우시겠지만' 같은 표준 쿠션어 없으면 AI 가 더 박함",
    8:  "AI 가 가장 짜게 — 사람은 '반품/교환' 단어가 후속 발화에 등장하면 복창 인정, AI 는 의도적 paraphrase 만 인정",
    9:  "잘 맞음 — 본인확인 절차는 양쪽 기준 동일",
    10: "AI 가 짜게 — 사람은 '고객이 안 되묻고 이해함' 만 보고 만점, AI 는 장황·반복·내부용어로 한 단계 감점",
    11: "AI 가 짜게 — 결론을 명시적으로 먼저 말해야 만점, 사람은 결론+부연이 보이면 인정",
    12: "잘 맞음 — 문제 해결 의지 표현 양쪽 기준 거의 일치",
    13: "잘 맞음 — 부연 설명 카운트 양쪽 일치",
    14: "AI 가 짜게 — '추가 문의 사항' 표준 클로징 없으면 감점, 사람은 자연스러운 마무리도 인정",
    17: "잘 맞음 — 본인확인 후 정보 조회 절차 양쪽 동일 판정",
    18: "잘 맞음 — 정보 보호 위반 트리거가 명확해서 양쪽 일치",
}


def fmt_signed(n: float, suffix: str = "점") -> str:
    if n is None:
        return "—"
    if n > 0:
        return f"+{n:.1f}{suffix}"
    return f"{n:.1f}{suffix}"


def write_sheet(xlsx_path: Path, results_dir: Path, ids: list[str], dataset_label: str) -> None:
    gt = load_xlsx_ground_truth(XLSX_GT, ids)
    ai = load_ai_results(results_dir, ids)
    matched = sorted(set(gt) & set(ai))

    # 통계 수집
    per_item_pairs: dict[int, list] = defaultdict(list)
    sample_totals: list[tuple[str, int, int, int]] = []   # (sid, ai, hum, max)
    sample_item_diffs: dict[str, list[tuple[int, int]]] = {}  # sid -> [(item, diff), ...]

    for sid in matched:
        ai_total = hum_total = max_total = 0
        diffs = []
        for num, _name, mx, _row in ITEM_DEF:
            if num not in ai[sid] or num not in gt[sid]:
                continue
            a = ai[sid][num]; h = gt[sid][num]
            per_item_pairs[num].append((a, h, mx))
            ai_total += a; hum_total += h; max_total += mx
            diffs.append((num, a - h))
        sample_totals.append((sid, ai_total, hum_total, max_total))
        sample_item_diffs[sid] = sorted(diffs, key=lambda x: -abs(x[1]))[:5]

    # 항목별 메트릭
    item_stats = {}
    for num, _name, _mx, _row in ITEM_DEF:
        m = compute_metrics(per_item_pairs.get(num, []))
        item_stats[num] = m

    # 총점 차이
    total_diffs = [(sid, ai - hum) for sid, ai, hum, _ in sample_totals]
    total_diffs.sort(key=lambda x: -abs(x[1]))
    total_mae = sum(abs(d) for _, d in total_diffs) / len(total_diffs) if total_diffs else 0
    total_bias = sum(d for _, d in total_diffs) / len(total_diffs) if total_diffs else 0
    total_under = sum(1 for _, d in total_diffs if d < 0) / len(total_diffs) * 100 if total_diffs else 0
    total_over = sum(1 for _, d in total_diffs if d > 0) / len(total_diffs) * 100 if total_diffs else 0
    max_abs_total = max((abs(d) for _, d in total_diffs), default=0)

    # 항목 분류
    items_ai_strict = []   # AI 가 짜게 (Bias < -0.5)
    items_ai_loose = []    # AI 가 후하게 (Bias > +0.5)
    items_aligned = []     # 잘 맞음 (|Bias| <= 0.5 and MAE <= 0.5)
    items_mixed = []       # 엇갈림
    for num, st in item_stats.items():
        mae = st.get("MAE")
        bias = st.get("Bias")
        if mae is None:
            continue
        if mae <= 0.5 and abs(bias) <= 0.5:
            items_aligned.append((num, st))
        elif bias <= -0.5:
            items_ai_strict.append((num, st))
        elif bias >= 0.5:
            items_ai_loose.append((num, st))
        else:
            items_mixed.append((num, st))
    items_ai_strict.sort(key=lambda x: x[1]["MAE"], reverse=True)
    items_ai_loose.sort(key=lambda x: x[1]["MAE"], reverse=True)

    # ── 시트 작성 ───────────────────────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.cell(row, 1, f"📊 [{dataset_label}] 비교 분석 — 한눈에 보기")
    ws.cell(row, 1).fill = TITLE_FILL; ws.cell(row, 1).font = TITLE_FONT
    ws.cell(row, 1).alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 32
    row += 2

    # ─── 1. 결론 한 문장 ───
    direction = "사람보다 짜게 채점" if total_bias < 0 else "사람보다 후하게 채점" if total_bias > 0 else "거의 일치"
    headline = (
        f"AI 는 사람보다 한 상담당 평균 **{abs(total_bias):.1f}점** {direction}하고 있습니다. "
        f"전체 {len(total_diffs)}건 중 {int(total_under/100*len(total_diffs))}건 ({total_under:.0f}%) 에서 AI 가 더 짜게 줬고, "
        f"{int(total_over/100*len(total_diffs))}건 ({total_over:.0f}%) 에서 더 후하게 줬습니다. "
        f"가장 크게 어긋난 상담은 총점 **{max_abs_total}점** 차이."
    )
    ws.cell(row, 1, "🎯 결론")
    ws.cell(row, 2, headline)
    for col in (1, 2):
        ws.cell(row, col).fill = HIGHLIGHT_FILL
        ws.cell(row, col).font = Font(bold=(col == 1), size=11, color="92400e")
        ws.cell(row, col).alignment = WRAP
        ws.cell(row, col).border = BORDER
    ws.row_dimensions[row].height = 60
    row += 2

    # ─── 2. 숫자로 보면 ───
    ws.cell(row, 1, "1. 숫자로 보면")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1

    rows = [
        ("샘플 (상담) 수", f"{len(total_diffs)}건  (업무정확도 #15·#16 항목은 평가 제외)"),
        ("한 상담당 평균 점수 차이", f"{total_mae:.1f}점  ← AI 와 사람 총점이 평균적으로 이 정도 벌어짐"),
        ("AI 의 채점 방향", f"{fmt_signed(total_bias)}  ← {'음수면 AI 가 짜게' if total_bias < 0 else '양수면 AI 가 후하게'} 채점한다는 뜻"),
        ("AI 가 더 짜게 준 상담", f"{total_under:.0f}%   ({int(total_under/100*len(total_diffs))} / {len(total_diffs)} 건)"),
        ("AI 가 더 후하게 준 상담", f"{total_over:.0f}%   ({int(total_over/100*len(total_diffs))} / {len(total_diffs)} 건)"),
        ("가장 크게 어긋난 상담의 점수 차이", f"{max_abs_total}점"),
    ]
    for k, v in rows:
        ws.cell(row, 1, k); ws.cell(row, 2, v)
        for col in (1, 2):
            ws.cell(row, col).fill = KEY_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=11, bold=(col == 1))
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ─── 3. AI 가 짜게 채점한 항목 ───
    ws.cell(row, 1, "2. AI 가 사람보다 짜게 채점한 항목")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1
    ws.cell(row, 1, "→ MAE 큰 순서. 사람보다 평균적으로 점수를 덜 주는 항목들 — 프롬프트 튜닝 우선 대상")
    ws.cell(row, 1).font = Font(italic=True, color="6b7280", size=10)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    row += 1
    for num, st in items_ai_strict[:6]:
        why = WHY_GAP.get(num, "")
        ws.cell(row, 1, f"#{num} {ITEM_NUM_TO_NAME.get(num, '')}")
        ws.cell(row, 2,
                f"한 항목당 평균 {st['MAE']:.2f}점 차이, AI 가 평균 {abs(st['Bias']):.2f}점 적게 부여. "
                f"가장 큰 차이 {st['MaxAbs']}점.\n"
                f"💡 {why}")
        for col in (1, 2):
            ws.cell(row, col).fill = BAD_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1), color="b91c1c" if col == 1 else "374151")
        ws.row_dimensions[row].height = 50
        row += 1
    row += 1

    # ─── 4. AI 가 후하게 채점한 항목 ───
    if items_ai_loose:
        ws.cell(row, 1, "3. AI 가 사람보다 후하게 채점한 항목")
        ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
        ws.row_dimensions[row].height = 22
        row += 1
        ws.cell(row, 1, "→ AI 가 점수를 더 많이 주는 항목 — 사람의 기준이 더 엄격한 영역")
        ws.cell(row, 1).font = Font(italic=True, color="6b7280", size=10)
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
        row += 1
        for num, st in items_ai_loose[:5]:
            why = WHY_GAP.get(num, "")
            ws.cell(row, 1, f"#{num} {ITEM_NUM_TO_NAME.get(num, '')}")
            ws.cell(row, 2,
                    f"한 항목당 평균 {st['MAE']:.2f}점 차이, AI 가 평균 {st['Bias']:+.2f}점 더 부여. "
                    f"가장 큰 차이 {st['MaxAbs']}점.\n💡 {why}")
            for col in (1, 2):
                ws.cell(row, col).fill = HIGHLIGHT_FILL
                ws.cell(row, col).alignment = WRAP
                ws.cell(row, col).border = BORDER
                ws.cell(row, col).font = Font(size=10, bold=(col == 1), color="92400e" if col == 1 else "374151")
            ws.row_dimensions[row].height = 50
            row += 1
        row += 1

    # ─── 5. 잘 맞은 항목 ───
    ws.cell(row, 1, "4. 잘 맞은 항목 (튜닝 불필요)")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1
    aligned_names = [f"#{num} {ITEM_NUM_TO_NAME.get(num)}" for num, _ in items_aligned]
    ws.cell(row, 1, "AI ≈ 사람")
    ws.cell(row, 2, ", ".join(aligned_names) if aligned_names else "(없음)")
    for col in (1, 2):
        ws.cell(row, col).fill = GOOD_FILL
        ws.cell(row, col).alignment = WRAP
        ws.cell(row, col).border = BORDER
        ws.cell(row, col).font = Font(size=10, bold=(col == 1), color="166534" if col == 1 else "374151")
    ws.row_dimensions[row].height = 30
    row += 2

    # ─── 6. 가장 어긋난 상담 5건 ───
    ws.cell(row, 1, "5. 가장 어긋난 상담 5건")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1
    ws.cell(row, 1, "→ |총점 차이| 큰 순. 어느 항목 때문에 차이가 났는지 함께 표시")
    ws.cell(row, 1).font = Font(italic=True, color="6b7280", size=10)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    row += 1
    sample_totals_map = {sid: (ai, hum) for sid, ai, hum, _ in sample_totals}
    for sid, diff in total_diffs[:5]:
        ai_t, hum_t = sample_totals_map[sid]
        item_diffs = sample_item_diffs.get(sid, [])
        item_strs = ", ".join(f"#{n}({d:+d})" for n, d in item_diffs if d != 0)
        ws.cell(row, 1, f"{sid}  ({fmt_signed(diff)})")
        ws.cell(row, 2, f"AI={ai_t}점 / 사람={hum_t}점.\n주요 차이 항목: {item_strs}")
        for col in (1, 2):
            ws.cell(row, col).fill = BAD_FILL if abs(diff) > 10 else KEY_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1))
        ws.row_dimensions[row].height = 40
        row += 1
    row += 1

    # ─── 7. 다음에 할 일 ───
    ws.cell(row, 1, "6. 다음에 할 일")
    ws.cell(row, 1).fill = SECTION_FILL; ws.cell(row, 1).font = SECTION_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=2)
    ws.row_dimensions[row].height = 22
    row += 1
    todo = [
        ("1️⃣ 가장 큰 문제부터",
         f"위 '2. AI 가 짜게 채점한 항목' 의 1순위 ({items_ai_strict[0][0] if items_ai_strict else '-'}번) 부터 프롬프트 수정. "
         "수정 방향은 '큰차이 심층분석' 시트와 'MAE 높은항목 원인 분석' 시트 참고."),
        ("2️⃣ 수정 후 확인",
         "프롬프트 고친 뒤 같은 14건 / 9건 다시 평가 → 이 시트의 1번 '결론' 문장의 평균 차이가 줄어드는지 확인. "
         "다른 항목 점수가 악화되지 않는지도 같이 봐야 함."),
        ("3️⃣ 가장 어긋난 상담 직접 확인",
         "'5. 가장 어긋난 상담 5건' 의 상담 ID 를 정답 xlsx 에서 찾아 사람 평가자 비고를 직접 읽어볼 것. "
         "AI 가 놓친 발화가 무엇인지 파악."),
        ("⚠ 한계",
         "이 모든 분석은 '사람 평가가 정답' 이라고 가정한 결과. 사람 평가 자체에 모호한 기준이 있으면 (특히 #8 문의 파악·복창) "
         "AI 만 고쳐서 해결되지 않음. 양쪽 정렬이 필요."),
    ]
    for k, v in todo:
        ws.cell(row, 1, k); ws.cell(row, 2, v)
        for col in (1, 2):
            ws.cell(row, col).fill = KEY_FILL
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER
            ws.cell(row, col).font = Font(size=10, bold=(col == 1))
        ws.row_dimensions[row].height = 50
        row += 1

    ws.freeze_panes = "A2"

    try:
        wb.save(xlsx_path)
    except PermissionError:
        fb = xlsx_path.with_name(xlsx_path.stem + "_with_easy_interp.xlsx")
        wb.save(fb)
        print(f"  ⚠ 원본 잠김 → fallback: {fb}")


def main() -> int:
    targets = [
        ("학습셋", TRAINING_IDS,
         Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409"),
         Path(r"C:\Users\META M\Desktop\학습셋_비교분석_20260422_000409\비교분석_20260422_011144.xlsx")),
        ("테스트셋", TEST_IDS,
         Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545"),
         Path(r"C:\Users\META M\Desktop\테스트셋_비교분석_20260422_003545\비교분석_20260422_011144.xlsx")),
    ]
    for label, ids, results_dir, xlsx_path in targets:
        if not xlsx_path.exists():
            print(f"[SKIP] {xlsx_path} 없음")
            continue
        write_sheet(xlsx_path, results_dir, ids, label)
        print(f"[{label}] '{SHEET_NAME}' 시트 재작성 → {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
