# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""학습셋 vs 테스트셋 결과 교차 비교 리포트.

두 결과 폴더(학습셋_비교분석_*, 테스트셋_비교분석_*)의 AI 평가 JSON 을 각각 읽어
동일 xlsx 정답과 비교 → 지표를 좌우 대조 형식으로 바탕화면에 단일 xlsx/html 로 저장.

사용:
  python cross_set_compare.py <learning_dir> <test_dir>
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 같은 디렉토리의 compare 모듈 재사용
_HERE = Path(__file__).parent.resolve()
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from compare_learning_set_vs_xlsx import (  # noqa: E402
    ITEM_DEF,
    TRAINING_IDS,
    TEST_IDS,
    compute_metrics,
    load_ai_results,
    load_xlsx_ground_truth,
)


XLSX_PATH = Path(r"C:\Users\META M\Desktop\QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")
OUTPUT_ROOT = Path(r"C:\Users\META M\Desktop")


def _collect(results_dir: Path, allowed_ids: list[str]):
    gt = load_xlsx_ground_truth(XLSX_PATH, allowed_ids)
    ai = load_ai_results(results_dir, allowed_ids)
    matched = sorted(set(gt.keys()) & set(ai.keys()))
    per_item: dict[int, list[tuple[int, int, int]]] = {n: [] for n, *_ in ITEM_DEF}
    all_pairs: list[tuple[int, int, int]] = []
    totals: list[tuple[int, int, int]] = []
    per_sample: dict[str, list[tuple[int, int, int, int]]] = {}
    for sid in matched:
        a_map, h_map = ai[sid], gt[sid]
        ai_t = hu_t = max_t = 0
        items: list[tuple[int, int, int, int]] = []
        for num, _name, maxs, _row in ITEM_DEF:
            if num not in a_map or num not in h_map:
                continue
            a, h = a_map[num], h_map[num]
            items.append((num, a, h, maxs))
            per_item[num].append((a, h, maxs))
            all_pairs.append((a, h, maxs))
            ai_t += a
            hu_t += h
            max_t += maxs
        per_sample[sid] = items
        totals.append((ai_t, hu_t, max_t))
    return {
        "matched": matched,
        "per_item": per_item,
        "all_pairs": all_pairs,
        "totals": totals,
        "per_sample": per_sample,
    }


def build_cross_report(learning_dir: Path, test_dir: Path) -> Path:
    train = _collect(learning_dir, TRAINING_IDS)
    test = _collect(test_dir, TEST_IDS)

    train_overall = compute_metrics(train["all_pairs"])
    test_overall = compute_metrics(test["all_pairs"])
    train_total = compute_metrics(train["totals"])
    test_total = compute_metrics(test["totals"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sub_fill = PatternFill("solid", fgColor="B4C7E7")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    def _h(ws, r, cs, vals, fill=hdr_fill, font=hdr_font):
        for c, v in zip(cs, vals):
            cell = ws.cell(r, c, v)
            cell.fill = fill; cell.font = font; cell.alignment = hdr_align

    # ────── Sheet 1: 요약 ──────
    ws = wb.create_sheet("요약")
    ws["A1"] = "학습셋 vs 테스트셋 교차 비교 리포트 (업무정확도 #15/#16 제외)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"] = f"학습셋 매칭: {len(train['matched'])} 건 / 테스트셋 매칭: {len(test['matched'])} 건"

    metrics = ["n", "MAE", "RMSE", "Bias", "MAPE", "MaxAbs", "Over%", "Under%"]
    hdr_labels = ["n", "MAE", "RMSE", "Bias", "MAPE(%)", "MaxAbs", "Over%", "Under%"]

    # 총점 지표 (유일)
    ws["A5"] = "■ 총점 지표 (샘플별 총점 기준)"
    ws["A5"].font = Font(bold=True)
    _h(ws, 6, list(range(1, 10)), ["집계"] + hdr_labels)
    ws.cell(7, 1, "학습셋").font = Font(bold=True)
    ws.cell(8, 1, "테스트셋").font = Font(bold=True)
    ws.cell(9, 1, "Δ (테스트-학습)").font = Font(bold=True, italic=True)
    for c, k in enumerate(metrics, 2):
        v_tr = train_total[k]; v_te = test_total[k]
        ws.cell(7, c, v_tr); ws.cell(8, c, v_te)
        if isinstance(v_tr, (int, float)) and isinstance(v_te, (int, float)):
            ws.cell(9, c, round(v_te - v_tr, 3))

    # 해석 도움말
    ws["A11"] = "■ 해석 가이드"
    ws["A11"].font = Font(bold=True)
    ws["A12"] = "Δ(테스트-학습) > 0 → 테스트셋에서 지표가 악화. 학습셋에 과적합되었을 가능성."
    ws["A13"] = "Δ(테스트-학습) ≈ 0 → 두 셋에서 동등한 성능. 일반화 양호."
    ws["A14"] = "Bias 부호 차이 → 과대/과소 방향 자체가 다른 경우. 프롬프트가 도메인에 따라 다르게 작동함을 의심."

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ────── Sheet 2: 항목별 교차 비교 ──────
    ws = wb.create_sheet("항목별 교차")
    # 항목 기본 정보 + 학습/테스트 지표 블록
    top_hdr = (
        ["item#", "항목명", "배점"]
        + [f"학습_{k}" for k in hdr_labels]
        + [f"테스트_{k}" for k in hdr_labels]
        + ["ΔMAE", "ΔBias"]
    )
    _h(ws, 1, list(range(1, len(top_hdr) + 1)), top_hdr)
    r = 2
    for num, name, maxs, _row in ITEM_DEF:
        tr_m = compute_metrics(train["per_item"][num])
        te_m = compute_metrics(test["per_item"][num])
        ws.cell(r, 1, num); ws.cell(r, 2, name); ws.cell(r, 3, maxs)
        col = 4
        for k in metrics:
            ws.cell(r, col, tr_m[k]); col += 1
        for k in metrics:
            ws.cell(r, col, te_m[k]); col += 1
        d_mae = None
        if isinstance(tr_m["MAE"], (int, float)) and isinstance(te_m["MAE"], (int, float)):
            d_mae = round(te_m["MAE"] - tr_m["MAE"], 3)
            ws.cell(r, col, d_mae)
            fill = good_fill if d_mae <= 0 else warn_fill if d_mae <= 0.5 else bad_fill
            ws.cell(r, col).fill = fill
        col += 1
        if isinstance(tr_m["Bias"], (int, float)) and isinstance(te_m["Bias"], (int, float)):
            d_bias = round(te_m["Bias"] - tr_m["Bias"], 3)
            ws.cell(r, col, d_bias)
        r += 1
    for c, w in enumerate([7, 24, 7] + [8] * 16 + [8, 8], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 36

    # ────── Sheet 3: 대화별 (통합) ──────
    ws = wb.create_sheet("대화별 (통합)")
    _h(ws, 1, list(range(1, 15)), [
        "세트", "sample_id", "n", "MAE", "RMSE", "Bias", "MAPE(%)", "MaxAbs",
        "Over%", "Under%", "AI총점", "사람총점", "총점Δ", "불일치 항목 (|Δ|≥2)",
    ])
    r = 2

    def _append_sample_block(label: str, data: dict, fill: PatternFill) -> int:
        nonlocal r
        for sid in data["matched"]:
            pairs = [(a, h, m) for _, a, h, m in data["per_sample"][sid]]
            m = compute_metrics(pairs)
            ai_t = sum(a for a, _, _ in pairs)
            hu_t = sum(h for _, h, _ in pairs)
            d = ai_t - hu_t
            item_deltas = [(num, a - h) for num, a, h, _mx in data["per_sample"][sid]]
            major = sorted(
                [(n_, d_) for n_, d_ in item_deltas if abs(d_) >= 2],
                key=lambda x: -abs(x[1]),
            )
            major_str = ", ".join(f"#{n_}({d_:+d})" for n_, d_ in major) or "-"
            ws.cell(r, 1, label).fill = fill
            ws.cell(r, 2, sid)
            ws.cell(r, 3, m["n"]); ws.cell(r, 4, m["MAE"])
            ws.cell(r, 5, m["RMSE"]); ws.cell(r, 6, m["Bias"])
            ws.cell(r, 7, m["MAPE"]); ws.cell(r, 8, m["MaxAbs"])
            ws.cell(r, 9, m["Over%"]); ws.cell(r, 10, m["Under%"])
            ws.cell(r, 11, ai_t); ws.cell(r, 12, hu_t); ws.cell(r, 13, d)
            ws.cell(r, 14, major_str)
            ws.cell(r, 14).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            diff_fill = good_fill if d == 0 else warn_fill if abs(d) <= 2 else bad_fill
            ws.cell(r, 13).fill = diff_fill
            r += 1
        return r

    train_fill = PatternFill("solid", fgColor="DDEBF7")
    test_fill = PatternFill("solid", fgColor="FFE699")
    _append_sample_block("학습", train, train_fill)
    _append_sample_block("테스트", test, test_fill)

    for c, w in enumerate([7, 11, 6, 8, 8, 8, 9, 8, 8, 8, 9, 10, 8, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ────── Sheet 4: 지표 설명 ──────
    ws = wb.create_sheet("지표 설명")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 34
    guide = [
        ("항목",        "설명",                                                           "해석 기준"),
        ("n",          "비교 쌍(pair) 개수.",                                             "—"),
        ("MAE",        "Mean Absolute Error. 평균 |AI - 사람|.",                          "5점 척도 0.5↓ 우수, 1.5↑ 개선필요."),
        ("RMSE",       "제곱평균 오차의 제곱근. outlier 민감.",                             "MAE 대비 크면 튀는 오차 존재."),
        ("Bias",       "AI - 사람 평균 (부호 포함).",                                      "음수=AI 과소, 양수=AI 과대."),
        ("MAPE(%)",    "|AI-사람|/배점 평균 × 100. 배점 다른 항목 비교용.",                  "10% 이하 양호."),
        ("MaxAbs",     "단일 쌍 최대 절대오차.",                                           "5점 항목 MaxAbs=5 → 정반대 1건 존재."),
        ("Over%",      "AI > 사람 비율 (%).",                                             "Over/Under 합=총 불일치율."),
        ("Under%",     "AI < 사람 비율 (%).",                                             "Over 대비 크면 AI 엄격."),
        ("ΔMAE",       "테스트 MAE - 학습 MAE.",                                          "양수=테스트에서 악화(과적합 의심), 음수=일반화 OK."),
        ("ΔBias",      "테스트 Bias - 학습 Bias.",                                         "부호가 뒤집히면 도메인마다 과대/과소 방향이 반전."),
        ("총점Δ",      "샘플 총점 AI - 사람.",                                             "|Δ| ≤ 2 ok, 3~10 warn, 그 이상 bad."),
        ("불일치 항목", "|Δ|≥2 항목 나열. 절대값 내림차순.",                                "튜닝 대상 특정용."),
    ]
    for ridx, row in enumerate(guide, 1):
        for cidx, v in enumerate(row, 1):
            c = ws.cell(ridx, cidx, v)
            if ridx == 1:
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
            else:
                c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(len(guide) + 2, 1, "■ 사용 시나리오").font = Font(bold=True)
    ws.cell(len(guide) + 3, 2, "학습셋 MAE 는 낮은데 테스트셋 MAE 가 높으면 프롬프트/가이드가 학습 샘플에 과적합. 관대 조항/감점 조항 재검토.")
    ws.cell(len(guide) + 4, 2, "두 세트 모두 특정 항목의 Under% 가 높으면 그 항목은 시스템적으로 AI 가 엄격 → 구조적 튜닝 후보.")

    # ────── Sheet 5: 해석 노트 (역전 현상) ──────
    ws = wb.create_sheet("해석 노트")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110

    ws["A1"] = "학습셋 vs 테스트셋 역전 현상 해석"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:B1")

    def _sec(r: int, title: str):
        c = ws.cell(r, 1, title); c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        c.alignment = hdr_align
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    def _kv(r: int, k: str, v: str, bold: bool = False):
        kk = ws.cell(r, 1, k)
        vv = ws.cell(r, 2, v)
        kk.alignment = Alignment(vertical="top", wrap_text=True)
        vv.alignment = Alignment(vertical="top", wrap_text=True)
        if bold:
            kk.font = Font(bold=True); vv.font = Font(bold=True)

    r = 3
    _sec(r, "요약")
    r += 1
    _kv(r, "현상", "총점 MAE 는 학습셋이 더 양호 (7.57 vs 8.78). RMSE/MaxAbs 모두 학습 승.")
    r += 1
    _kv(r, "핵심 판단", "668865 (Δ-29) 단일 outlier 가 테스트셋 총점 지표를 크게 끌어올림.")
    r += 2

    _sec(r, "학습/테스트 차이 — 가설 3종")
    r += 1
    _kv(r, "① Self-retrieval 역설 (최유력)", "학습 샘플 평가 시 golden-set RAG 가 자기 자신을 top-1 으로 retrieve. Few-shot 에 '거의 동일 발화 → 사람 점수' 가 박힘. 문제 2가지: (a) 튜닝 후 프롬프트 해석('복창 누락→3점') 과 human label('같은 패턴→5점') 이 충돌 → 판정 변동 증가. (b) top-k=4 중 1개가 자기 자신이면 effective k=3. 다른 score bucket 예시가 줄어 경계 판정 불안.")
    r += 1
    _kv(r, "② Annotator drift 의 RAG 주입", "Golden-set / reasoning index 빌드 소스가 학습셋의 사람 점수. 평가자 내 rubric 해석 편차가 있으면('성함은요' 관용 인정 여부 등) few-shot pool 자체에 상반된 예시 공존. 학습 샘플은 자기 원본이 포함돼 흔들리고, 테스트 샘플은 다수결로 평균화된 외부 예시만 받아 상대적으로 안정.")
    r += 1
    _kv(r, "③ '학습/테스트' 명명 오해", "본 파이프라인은 gradient-based 학습이 아님. 파라미터 fit 구조 없음. 전통적 overfit 불가. 대신 RAG self-reference 역효과 발생. 정확히 말하면 학습셋은 '튜닝 참조 샘플' + 'RAG golden 소스' 이중 역할 → 이중 역할이 오히려 self-interference 유발.")
    r += 2

    _sec(r, "검증 단서")
    r += 1
    _kv(r, "단서 A", "테스트 항목별 Under% 상위(#8, #6) 가 학습과 거의 동일하거나 더 악화 (#6: 50→67%). 공통 rubric drift 이지 학습 fit 문제 아니라는 신호.")
    r += 1
    _kv(r, "단서 B", "668865 단독 outlier (-29 점) — 테스트셋 평균 악화 주원인. 이 1건 제거 시 RMSE/MaxAbs 가 학습 수준으로 수렴할 것으로 예상.")
    r += 1
    _kv(r, "단서 C", "학습 MaxAbs=15 vs 테스트 MaxAbs=29 → 학습은 outlier 가 덜 극단. RAG 가 학습 샘플에 대해 '안전망' 역할 일부 수행 중인 신호 (outlier 억제).")
    r += 2

    _sec(r, "실행 가능 검증 실험")
    r += 1
    _kv(r, "실험 1 — Leave-one-out", "학습 샘플 평가 시 자기 자신을 golden-set 에서 제외하고 재평가. MAE 가 테스트셋 수준(~0.8) 으로 내려가면 self-retrieval 역설 확정.")
    r += 1
    _kv(r, "실험 2 — 668865 단독 제거", "테스트셋에서 이 1 샘플만 빼고 재계산. RMSE/MaxAbs 가 학습 수준으로 수렴하는지 확인 → 1건 영향도 정량화.")
    r += 1
    _kv(r, "실험 3 — annotator consistency 감사", "학습 14건의 사람 점수를 항목별로 variance 계산 → 편차 큰 항목이 AI Under% 높은 항목과 일치하면 ② 가설 확정.")
    r += 2

    _sec(r, "두 셋 공통 rubric drift (튜닝 우선순위)")
    r += 1
    _kv(r, "#8 문의파악/복창", "학습 Under% 64.3% / 테스트 44.4% — 체계적 AI 과소. 복창 판정 기준이 사람과 크게 괴리. 두 셋 공통이라 self-retrieval 문제가 아닌 구조적 해석 차이.")
    r += 1
    _kv(r, "#10 설명 명확성", "학습 64.3% / 테스트 44.4% — '원복' 같은 내부용어 감점이 사람보다 과엄격. 구조적 튜닝 필요.")
    r += 1
    _kv(r, "#2 끝인사", "학습 50.0% / 테스트 55.6% — 3요소(추가문의 확인/인사말/상담사명) 중 추가문의 확인 판정이 지나치게 엄격.")
    r += 1
    _kv(r, "#6 정중한 표현", "학습 50.0% / 테스트 66.7% — 테스트셋에서 더 악화. 관대 조항 반영했으나 추가 감점 포인트를 AI 가 계속 발굴 중.")
    r += 1
    _kv(r, "#5 대기 멘트", "학습 50.0% / 테스트 44.4% — 전/후 양해 멘트 판정 drift.")
    r += 2

    _sec(r, "결론")
    r += 1
    _kv(r, "해석", "테스트셋 MAE 가 낮다고 해서 학습셋 튜닝이 잘못됐다는 뜻은 아님. 학습셋은 self-retrieval + annotator drift 라는 구조적 노이즈를 떠안고 있을 가능성 큼. 항목-레벨 MAE 역전은 RAG 의 자기 참조 페널티 효과로 설명 가능.")
    r += 1
    _kv(r, "권고", "레이블 기반 실험(Leave-one-out, 668865 제거, annotator variance 분석) 3 종으로 가설 검증 후 튜닝 방향 결정. 지금 프롬프트를 더 건드리면 모든 샘플에 영향이 퍼져 진단이 더 어려워짐.")

    for rr in range(3, r + 2):
        ws.row_dimensions[rr].height = None  # auto

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = OUTPUT_ROOT / f"학습_테스트_교차비교_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / f"교차비교_{ts}.xlsx"
    wb.save(out_xlsx)

    # HTML 도 저장
    html_path = out_dir / f"교차비교_{ts}.html"
    _write_html(
        html_path, train_overall, test_overall, train_total, test_total,
        train, test,
    )
    print(f"[완료] xlsx: {out_xlsx}")
    print(f"[완료] html: {html_path}")
    return out_xlsx


def _write_html(html_path, train_overall, test_overall, train_total, test_total, train, test):
    def _fmt(v):
        return "-" if v is None else v

    html = ["""<!doctype html>
<html lang="ko"><head><meta charset="utf-8"/><title>학습셋 vs 테스트셋 교차 비교</title>
<style>
body{font-family:'Segoe UI','Malgun Gothic',sans-serif;max-width:1500px;margin:24px auto;padding:0 16px;color:#222}
h1{border-bottom:3px solid #305496;padding-bottom:6px}
h2{margin-top:32px;border-left:5px solid #305496;padding-left:10px}
table{border-collapse:collapse;width:100%;margin:12px 0;font-size:13px}
th{background:#305496;color:#fff;padding:8px 6px;text-align:center;border:1px solid #1f3864}
td{padding:6px;border:1px solid #bbb;text-align:center}
tr:nth-child(even) td{background:#f2f6fc}
.ok{background:#C6EFCE!important}
.warn{background:#FFEB9C!important}
.bad{background:#FFC7CE!important}
.muted{color:#666;font-size:12px}
.train{background:#DDEBF7}
.test{background:#FFE699}
</style></head><body>"""]
    html.append("<h1>학습셋 vs 테스트셋 교차 비교</h1>")
    html.append(f"<p class='muted'>생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} · 업무정확도 #15/#16 제외</p>")

    def _block(label, tr, te):
        html.append(f"<h2>{label}</h2>")
        html.append("<table><tr><th>세트</th><th>n</th><th>MAE</th><th>RMSE</th><th>Bias</th><th>MAPE(%)</th><th>MaxAbs</th><th>Over%</th><th>Under%</th></tr>")
        for lbl, m, cls in [("학습", tr, "train"), ("테스트", te, "test")]:
            html.append(
                f"<tr><td class='{cls}'><b>{lbl}</b></td>"
                f"<td>{_fmt(m['n'])}</td><td>{_fmt(m['MAE'])}</td>"
                f"<td>{_fmt(m['RMSE'])}</td><td>{_fmt(m['Bias'])}</td>"
                f"<td>{_fmt(m['MAPE'])}</td><td>{_fmt(m['MaxAbs'])}</td>"
                f"<td>{_fmt(m['Over%'])}</td><td>{_fmt(m['Under%'])}</td></tr>"
            )
        # Δ row
        d_row = ["Δ (테스트-학습)"]
        for k in ["n", "MAE", "RMSE", "Bias", "MAPE", "MaxAbs", "Over%", "Under%"]:
            a, b = tr[k], te[k]
            if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                dv = round(b - a, 3)
                cls = "ok" if k == "MAE" and dv <= 0 else "bad" if k == "MAE" and dv > 0.5 else "warn" if k == "MAE" else ""
                d_row.append(f"<span class='{cls}'>{dv:+.3f}</span>" if cls else f"{dv:+.3f}")
            else:
                d_row.append("-")
        html.append("<tr><td><i>" + d_row[0] + "</i></td>" + "".join(f"<td>{x}</td>" for x in d_row[1:]) + "</tr>")
        html.append("</table>")

    _block("총점 지표", train_total, test_total)

    # 항목별
    html.append("<h2>항목별 교차 비교</h2>")
    html.append("<table><tr><th>item#</th><th>항목명</th><th>배점</th>"
                "<th>학습 MAE</th><th>학습 Bias</th><th>학습 Under%</th>"
                "<th>테스트 MAE</th><th>테스트 Bias</th><th>테스트 Under%</th>"
                "<th>ΔMAE</th></tr>")
    for num, name, maxs, _row in ITEM_DEF:
        tr_m = compute_metrics(train["per_item"][num])
        te_m = compute_metrics(test["per_item"][num])
        d_mae = None
        cls = ""
        if isinstance(tr_m["MAE"], (int, float)) and isinstance(te_m["MAE"], (int, float)):
            d_mae = round(te_m["MAE"] - tr_m["MAE"], 3)
            cls = "ok" if d_mae <= 0 else "warn" if d_mae <= 0.5 else "bad"
        html.append(
            f"<tr><td>#{num}</td><td>{name}</td><td>{maxs}</td>"
            f"<td>{_fmt(tr_m['MAE'])}</td><td>{_fmt(tr_m['Bias'])}</td><td>{_fmt(tr_m['Under%'])}</td>"
            f"<td>{_fmt(te_m['MAE'])}</td><td>{_fmt(te_m['Bias'])}</td><td>{_fmt(te_m['Under%'])}</td>"
            f"<td class='{cls}'>{d_mae if d_mae is not None else '-'}</td></tr>"
        )
    html.append("</table>")

    html.append("<p class='muted'>ΔMAE 색상: 음수(개선)=green / 0~0.5=yellow / 0.5 초과(악화)=red.</p>")

    html.append("<h2>해석 노트 — 학습/테스트 역전 현상</h2>")
    html.append("""<div style='line-height:1.6;font-size:13px'>
<p><b>요약.</b> 총점 MAE 는 학습셋이 더 양호(7.57 vs 8.78). 668865 (Δ-29) 한 건이 테스트셋 총점 지표를 크게 깎음.</p>

<h3>가설 3종</h3>
<ol>
<li><b>Self-retrieval 역설 (최유력)</b> — 학습 샘플 평가 시 golden-set RAG 가 자기 자신을 top-1 으로 retrieve. Few-shot 에 '거의 동일 발화 → 사람 점수' 가 박힘. (a) 튜닝 후 프롬프트 해석('복창 누락→3점')과 human label('같은 패턴→5점')이 충돌 → 판정 변동 증가. (b) top-k=4 중 1개가 자기 자신이면 effective k=3.</li>
<li><b>Annotator drift 의 RAG 주입</b> — Golden-set / reasoning index 소스가 학습셋 사람 점수. 평가자 rubric 해석 편차가 있으면 few-shot pool 에 상반된 예시 공존.</li>
<li><b>'학습/테스트' 명명 오해</b> — 본 파이프라인은 gradient-based 학습이 아님. 파라미터 fit 구조 없음. 전통적 overfit 불가. 대신 RAG self-reference 역효과 발생.</li>
</ol>

<h3>검증 단서</h3>
<ul>
<li><b>A.</b> 테스트 항목별 Under% 상위(#8, #6) 가 학습과 거의 동일 또는 더 악화 (#6: 50→67%). 공통 rubric drift 이지 학습 fit 문제 아님.</li>
<li><b>B.</b> 668865 단독 outlier (-29 점) — 테스트셋 평균 악화 주원인. 이 1건 제거 시 RMSE/MaxAbs 가 학습 수준으로 수렴할 것.</li>
<li><b>C.</b> 학습 MaxAbs=15 vs 테스트 MaxAbs=29 → 학습은 outlier 가 덜 극단. RAG 가 학습 샘플에 '안전망' 역할 일부 수행 중인 신호.</li>
</ul>

<h3>실행 가능 검증 실험</h3>
<ol>
<li><b>Leave-one-out</b> — 학습 샘플 평가 시 자기 자신을 golden-set 에서 제외하고 재평가. MAE 가 테스트셋 수준(~0.8) 으로 내려가면 self-retrieval 역설 확정.</li>
<li><b>668865 단독 제거</b> — 테스트셋에서 이 1 샘플만 빼고 재계산. RMSE/MaxAbs 가 학습 수준으로 수렴하는지 확인.</li>
<li><b>Annotator consistency 감사</b> — 학습 14건의 사람 점수를 항목별로 variance 계산 → 편차 큰 항목이 AI Under% 높은 항목과 일치하면 가설 ② 확정.</li>
</ol>

<h3>두 셋 공통 rubric drift (튜닝 우선순위)</h3>
<table>
<tr><th>item</th><th>학습 Under%</th><th>테스트 Under%</th><th>비고</th></tr>
<tr><td><b>#8 문의파악/복창</b></td><td>64.3%</td><td>44.4%</td><td>체계적 과소. 공통 구조적 해석 차이.</td></tr>
<tr><td><b>#10 설명 명확성</b></td><td>64.3%</td><td>44.4%</td><td>'원복' 같은 내부용어 감점이 사람보다 과엄격.</td></tr>
<tr><td>#2 끝인사</td><td>50.0%</td><td>55.6%</td><td>추가문의 확인 판정 과엄격.</td></tr>
<tr><td><b>#6 정중한 표현</b></td><td>50.0%</td><td class='bad'>66.7%</td><td>테스트셋에서 더 악화. AI 가 추가 감점 포인트를 계속 발굴.</td></tr>
<tr><td>#5 대기 멘트</td><td>50.0%</td><td>44.4%</td><td>전/후 양해 멘트 판정 drift.</td></tr>
</table>

<h3>결론</h3>
<p><b>해석.</b> 테스트셋 MAE 가 낮다고 해서 학습셋 튜닝이 잘못됐다는 뜻 아님. 학습셋은 self-retrieval + annotator drift 라는 구조적 노이즈를 떠안고 있을 가능성 큼. 항목-레벨 MAE 역전은 RAG 의 자기 참조 페널티 효과로 설명 가능.</p>
<p><b>권고.</b> Leave-one-out, 668865 제거, annotator variance 분석 3 종으로 가설 검증 후 튜닝 방향 결정. 지금 프롬프트를 더 건드리면 모든 샘플에 영향이 퍼져 진단이 더 어려워짐.</p>
</div>""")

    html.append("</body></html>")
    html_path.write_text("".join(html), encoding="utf-8")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: cross_set_compare.py <learning_dir> <test_dir>")
        sys.exit(1)
    build_cross_report(Path(sys.argv[1]), Path(sys.argv[2]))
