# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""STT원문 문의유형 구분.xlsx → 사람 추정 인텐트 라벨 추출 (1회성).

출력: stdout 에 (case_id, human_intents, note) tuples + 시트 이름/헤더 dump.
"""
from __future__ import annotations

import sys
from pathlib import Path


def main(xlsx_path: Path) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=False)
    print(f"=== Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        print(f"\n=== Sheet '{sheet_name}' (rows={max_row}, cols={max_col}) ===")
        # 첫 5행 dump (헤더 파악)
        for r_idx in range(1, min(max_row, 6) + 1):
            row_vals = []
            for c_idx in range(1, max_col + 1):
                v = ws.cell(r_idx, c_idx).value
                if v is None:
                    row_vals.append("")
                else:
                    s = str(v).replace("\n", " | ")[:80]
                    row_vals.append(s)
            print(f"  R{r_idx}: {row_vals}")
        # 데이터 행 모두 dump — 헤더 위치 자동 검출 (상담ID 가 들어있는 행)
        header_row_idx = None
        for r_idx in range(1, min(max_row, 10) + 1):
            for c_idx in range(1, max_col + 1):
                v = ws.cell(r_idx, c_idx).value
                if isinstance(v, str) and v.strip() == "상담ID":
                    header_row_idx = r_idx
                    break
            if header_row_idx:
                break
        print(f"\n  header_row_idx = {header_row_idx}")
        if header_row_idx is None:
            continue
        # 헤더 매핑
        headers: dict[int, str] = {}
        for c_idx in range(1, max_col + 1):
            v = ws.cell(header_row_idx, c_idx).value
            if isinstance(v, str) and v.strip():
                headers[c_idx] = v.strip()
        print(f"  headers = {headers}")
        print(f"\n  --- ALL DATA ROWS ---")
        for r_idx in range(header_row_idx + 1, max_row + 1):
            row = {}
            for c_idx, hname in headers.items():
                v = ws.cell(r_idx, c_idx).value
                if v is None:
                    continue
                s = str(v).strip() if not isinstance(v, str) else v.strip()
                if s:
                    row[hname] = s
            if row.get("상담ID"):
                # STT 본문은 너무 길어 일단 짧게
                stt = row.get("STT원문(분류 근거가 되는 핵심 발화)", "")
                if len(stt) > 200:
                    row["STT원문(분류 근거가 되는 핵심 발화)"] = stt[:200] + "..."
                print(f"  R{r_idx}: {row}")
            else:
                # section 라벨 (■ 학습셋, ■ 테스트셋 등)
                non_empty = {k: v for k, v in row.items() if v}
                if non_empty:
                    print(f"  R{r_idx} (section): {non_empty}")
    wb.close()


if __name__ == "__main__":
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\META M\Desktop\STT원문 문의유형 구분.xlsx")
    if not xlsx.exists():
        print(f"ERROR: not found: {xlsx}", file=sys.stderr)
        sys.exit(1)
    main(xlsx)
