# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""KMS xlsx → markdown 변환 (1회성 빌드 스크립트).

사용:
  python _convert_xlsx_to_md.py [xlsx_path]
출력:
  같은 디렉토리의 kms_<intent>.md (7개)
"""
from __future__ import annotations

import os
import sys
from pathlib import Path


KMS_INTENT_TABS = ("회원정보", "환불", "교환", "반품", "수선", "배송", "취소")
HEADER_KEYS = ["문의유형", "세부사항", "조건", "필수 키워드", "필수 안내 사항"]


def main(xlsx_path: Path) -> None:
    import openpyxl

    print(f"Loading {xlsx_path} (size={xlsx_path.stat().st_size:,}B)")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=False)
    print(f"Sheets: {wb.sheetnames}")

    out_dir = Path(__file__).parent
    for tab in KMS_INTENT_TABS:
        if tab not in wb.sheetnames:
            print(f"  [skip] {tab}: 시트 없음")
            continue
        ws = wb[tab]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # 헤더 행 위치
        header_row_idx = None
        for r_idx in range(1, min(max_row, 10) + 1):
            for c_idx in range(1, min(max_col, 10) + 1):
                if ws.cell(r_idx, c_idx).value == "문의유형":
                    header_row_idx = r_idx
                    break
            if header_row_idx:
                break
        if not header_row_idx:
            print(f"  [skip] {tab}: 헤더 미검출")
            continue

        header_cols: dict[str, int] = {}
        for c_idx in range(1, max_col + 1):
            v = ws.cell(header_row_idx, c_idx).value
            if isinstance(v, str) and v.strip() in HEADER_KEYS:
                header_cols[v.strip()] = c_idx

        rows = []
        for r_idx in range(header_row_idx + 1, max_row + 1):
            row = {}
            for key, c_idx in header_cols.items():
                row[key] = ws.cell(r_idx, c_idx).value
            if not (row.get("필수 키워드") or row.get("필수 안내 사항")):
                continue
            rows.append(row)

        # markdown 출력
        md_lines = [f"# KMS — {tab}", ""]
        md_lines.append(f"_탭 출처: {xlsx_path.name} / 시트 `{tab}` / 추출 행 {len(rows)}_")
        md_lines.append("")
        for i, row in enumerate(rows, 1):
            sub = (row.get("세부사항") or "").strip() or "(미지정)"
            cond = (row.get("조건") or "").strip()
            keywords = (row.get("필수 키워드") or "").strip()
            statements = (row.get("필수 안내 사항") or "").strip()
            md_lines.append(f"## [{i}] {sub}")
            if cond:
                md_lines.append(f"- **조건**: {cond}")
            if keywords:
                md_lines.append(f"- **필수 키워드**: {keywords}")
            if statements:
                md_lines.append("- **필수 안내 사항**:")
                # 안내 사항은 multiline 가능 — 들여쓰기로 보존
                for line in statements.split("\n"):
                    line = line.strip()
                    if line:
                        md_lines.append(f"  - {line}")
            md_lines.append("")

        out_path = out_dir / f"kms_{tab}.md"
        out_path.write_text("\n".join(md_lines), encoding="utf-8")
        print(f"  [ok] {out_path.name}: {len(rows)}행, {len('\\n'.join(md_lines)):,}B")

    wb.close()
    print("Done.")


if __name__ == "__main__":
    default_xlsx = Path(os.environ.get("USERPROFILE", os.path.expanduser("~"))) / "Desktop" / "코오롱 업무 정확도 auto_qa_criteria.xlsx"
    xlsx_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx
    if not xlsx_arg.exists():
        print(f"ERROR: xlsx not found: {xlsx_arg}", file=sys.stderr)
        sys.exit(1)
    main(xlsx_arg)
