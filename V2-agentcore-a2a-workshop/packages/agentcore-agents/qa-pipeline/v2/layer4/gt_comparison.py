# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""Layer 4 후속 — GT (Ground Truth) vs AI 점수 비교 노드.

Pipeline state.body 에서 `gt_sample_id` 가 있으면 QA 정답 xlsx 에서 해당 시트를
로드해 항목별 AI 점수와 GT 점수를 비교. 업무 정확도 (#15, #16) 는 비교에서 제외
— 현재 AI 가중 (15점) 이 GT 기준 (10점) 과 다르고, GT 는 수동 검수 기반이라 공정 비교 불가.

입력 (state):
  - gt_sample_id : str | None  (없으면 비교 생략)
  - evaluations  : list[ItemVerdict]  (Layer 2 결과)
  - report       : dict (Layer 4 report_generator 결과, 선택)

출력 (state.gt_comparison):
  {
    "enabled": bool,
    "sample_id": str | None,
    "sheet_name": str | None,
    "excluded_items": [15, 16],
    "compared_item_count": int,
    "ai_total": int,       # 비교 대상 항목 합 (업무정확도 제외)
    "gt_total": int,       # GT 대상 항목 합 (업무정확도 제외)
    "diff": int,           # ai_total - gt_total
    "abs_diff": int,
    "mae": float,          # 평균 절대오차 (항목당)
    "rmse": float,
    "accuracy": float,     # 완전일치 비율 (%)
    "match_count": int,
    "mismatch_count": int,
    "items": [
      {item_number, item_name, max_score, ai_score, gt_score, diff, match, note}
    ],
  }

실패 케이스: xlsx 미존재 / sample_id 매칭 실패 / openpyxl 미설치 → enabled=False, error 필드.
"""

from __future__ import annotations

import logging
import math
import os
from pathlib import Path
from typing import Any


logger = logging.getLogger(__name__)


# GT 비교에서 제외할 항목. AI 가중 / 수동 검수 기준 차이 등 이유.
EXCLUDED_ITEMS: frozenset[int] = frozenset({15, 16})


def _default_xlsx_path() -> str:
    """GT xlsx 경로 — `_gt_loader.resolve_gt_xlsx_path` 단일 헬퍼 사용 (S4 fix).

    이전엔 Windows-only Desktop fallback 만 있어 EC2 Linux 에서 layer4 GT 비교가
    무용했음. _gt_loader 가 OS 별 후보 통합 처리.
    """
    from ._gt_loader import resolve_gt_xlsx_path
    resolved = resolve_gt_xlsx_path()
    if resolved:
        return resolved
    # not-found fallback (호출자가 Path.exists() 로 다시 검사)
    env = os.environ.get("QA_GT_XLSX_PATH")
    if env:
        return env
    return str(Path(r"C:\Users\META M\Desktop") / "QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx")


def _load_gt_items(xlsx_path: str, sample_id: str) -> tuple[list[dict], str | None, str | None]:
    """xlsx 에서 sample_id 매칭 시트 로드 → 17 항목 리스트 + 시트명 + 에러.

    Returns
    -------
    (items, sheet_name, error_message)
        에러 시 items=[], sheet_name=None, error_message=사유.
    """
    try:
        import openpyxl
    except Exception as e:
        return [], None, f"openpyxl import 실패: {e}"

    if not Path(xlsx_path).exists():
        return [], None, f"xlsx 파일 없음: {xlsx_path}"

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        return [], None, f"xlsx 로드 실패: {e}"

    try:
        # ★ 2026-04-30 S4 fix: lenient 매칭을 v2.layer4._gt_loader 단일 헬퍼로 통합.
        # 이전엔 server_v2 `/v2/gt-scores` 와 이 함수에 동일 알고리즘 인라인 → drift 위험.
        from ._gt_loader import match_sheet
        matched, _method = match_sheet(wb.sheetnames, sample_id)
        if not matched:
            return [], None, f"sample_id={sample_id} 시트 없음"
        ws = wb[matched[0]]

        item_numbers = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
        items: list[dict] = []
        cur_idx = 0
        current_category: str | None = None

        # ★ 2026-05-08 perf: 셀 단위 ws.cell(r,c) 호출은 openpyxl read-only 에서
        # 매 호출마다 XML 처음부터 재스트리밍 → 44행 × 6컬 = 264회 호출에 ~150ms+.
        # iter_rows(values_only=True) 1회 호출은 동일 범위를 sub-millisecond 처리.
        # ★ 2026-05-07: column 7 = 사람 검수자가 참조한 STT 원문 발췌 (col index 6).
        rows_iter = ws.iter_rows(min_row=6, max_row=49, min_col=1, max_col=7, values_only=True)
        for row_vals in rows_iter:
            cat_cell = row_vals[0] if len(row_vals) > 0 else None
            item_name = row_vals[1] if len(row_vals) > 1 else None
            max_score = row_vals[3] if len(row_vals) > 3 else None
            score = row_vals[4] if len(row_vals) > 4 else None
            note = row_vals[5] if len(row_vals) > 5 else None
            stt_excerpt = row_vals[6] if len(row_vals) > 6 else None

            if cat_cell and isinstance(cat_cell, str) and "총점" in cat_cell:
                break
            if cat_cell:
                current_category = str(cat_cell).strip()
            if item_name is None:
                continue
            if cur_idx >= len(item_numbers):
                break

            items.append({
                "item_number": item_numbers[cur_idx],
                "category": current_category,
                "item_name": (str(item_name).strip() if isinstance(item_name, str) else str(item_name)),
                "max_score": int(max_score) if isinstance(max_score, (int, float)) else None,
                "score": int(score) if isinstance(score, (int, float)) else None,
                "note": str(note).strip() if isinstance(note, str) else None,
                # 검수자가 참조한 실제 STT 발화 — 프론트가 "T 구간" 의 T 가 무엇인지 볼 수 있게.
                "gt_evidence_excerpt": (
                    str(stt_excerpt).strip() if isinstance(stt_excerpt, str)
                    else (str(stt_excerpt) if stt_excerpt is not None else None)
                ),
            })
            cur_idx += 1

        return items, matched[0], None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _extract_ai_scores(state: dict[str, Any]) -> dict[int, dict[str, Any]]:
    """state.evaluations 에서 item_number → {score, item_name, max_score, evidence, judgment} 매핑."""
    out: dict[int, dict[str, Any]] = {}
    evals = state.get("evaluations") or []
    for ev in evals:
        e = ev.get("evaluation") if isinstance(ev, dict) else None
        if not isinstance(e, dict):
            continue
        try:
            num = int(e.get("item_number"))
        except (TypeError, ValueError):
            continue
        # evidence 정규화 — list[dict|str] → list[str] (한 줄)
        ev_raw = e.get("evidence") or []
        ev_lines: list[str] = []
        for it in ev_raw if isinstance(ev_raw, list) else []:
            if isinstance(it, str):
                s = it.strip()
                if s:
                    ev_lines.append(s)
            elif isinstance(it, dict):
                speaker = (it.get("speaker") or it.get("role") or "").strip()
                turn = it.get("turn") if it.get("turn") is not None else it.get("turn_id")
                text = (it.get("text") or it.get("quote") or "").strip()
                if not text:
                    continue
                prefix = ""
                if turn is not None:
                    prefix += f"[T{turn}] "
                if speaker:
                    prefix += f"{speaker}: "
                ev_lines.append(prefix + text)
        deductions = e.get("deductions") or []
        ded_lines: list[str] = []
        for d in deductions if isinstance(deductions, list) else []:
            if isinstance(d, dict):
                reason = d.get("reason") or d.get("rule") or d.get("description") or ""
                pts = d.get("points_lost", d.get("points", 0))
                if reason:
                    ded_lines.append(f"-{pts}점: {reason}")
        out[num] = {
            "score": e.get("score"),
            "max_score": e.get("max_score"),
            "item_name": e.get("item_name"),
            "evaluation_mode": e.get("evaluation_mode"),
            "judgment": e.get("judgment") or e.get("summary") or "",
            "evidence_lines": ev_lines,
            "deduction_lines": ded_lines,
        }
    return out


async def gt_comparison_node(state: dict[str, Any]) -> dict[str, Any]:
    """Layer 4 후속 노드 — AI vs GT 점수 비교 수행.

    state.gt_sample_id 가 없으면 disabled 로 반환 (파이프라인 무영향).
    """
    sample_id = state.get("gt_sample_id") or state.get("sample_id")
    if not sample_id:
        logger.info("gt_comparison: gt_sample_id 없음 — 비교 생략")
        return {
            "gt_comparison": {
                "enabled": False,
                "reason": "gt_sample_id 미주입 (JSON 에 id 필드 또는 body.gt_sample_id 필요)",
            }
        }

    sample_id_str = str(sample_id)
    xlsx_path = _default_xlsx_path()
    gt_items, sheet_name, err = _load_gt_items(xlsx_path, sample_id_str)
    if err:
        logger.warning("gt_comparison: %s", err)
        return {
            "gt_comparison": {
                "enabled": False,
                "sample_id": sample_id_str,
                "error": err,
                "xlsx_path": xlsx_path,
            }
        }

    ai_scores = _extract_ai_scores(state)
    compared: list[dict[str, Any]] = []
    ai_total = 0
    gt_total = 0
    match_count = 0
    mismatch_count = 0
    squared_diff_sum = 0.0
    abs_diff_sum = 0.0

    for gt in gt_items:
        num = gt.get("item_number")
        if num in EXCLUDED_ITEMS:
            # 비교 목록에는 포함하되 match/score 합산에서 제외 (투명성)
            ai_entry = ai_scores.get(num, {})
            compared.append({
                "item_number": num,
                "item_name": gt.get("item_name"),
                "max_score": gt.get("max_score"),
                "ai_score": ai_entry.get("score"),
                "gt_score": gt.get("score"),
                "diff": None,
                "match": None,
                "excluded": True,
                "note": gt.get("note"),
                "ai_evidence": ai_entry.get("evidence_lines") or [],
                "ai_judgment": ai_entry.get("judgment") or "",
                "ai_deductions": ai_entry.get("deduction_lines") or [],
            })
            continue

        ai_entry = ai_scores.get(num, {})
        ai_score = ai_entry.get("score")
        gt_score = gt.get("score")

        try:
            ai_v = int(ai_score) if ai_score is not None else None
            gt_v = int(gt_score) if gt_score is not None else None
        except (TypeError, ValueError):
            ai_v = None
            gt_v = None

        if ai_v is None or gt_v is None:
            diff = None
            match = None
        else:
            diff = ai_v - gt_v
            match = (diff == 0)
            ai_total += ai_v
            gt_total += gt_v
            abs_diff_sum += abs(diff)
            squared_diff_sum += diff * diff
            if match:
                match_count += 1
            else:
                mismatch_count += 1

        compared.append({
            "item_number": num,
            "item_name": gt.get("item_name"),
            "max_score": gt.get("max_score"),
            "ai_score": ai_v,
            "gt_score": gt_v,
            "diff": diff,
            "match": match,
            "excluded": False,
            "note": gt.get("note"),
            "ai_evidence": ai_entry.get("evidence_lines") or [],
            "ai_judgment": ai_entry.get("judgment") or "",
            "ai_deductions": ai_entry.get("deduction_lines") or [],
        })

    n_compared = match_count + mismatch_count
    mae = (abs_diff_sum / n_compared) if n_compared else 0.0
    rmse = math.sqrt(squared_diff_sum / n_compared) if n_compared else 0.0
    accuracy = (match_count / n_compared * 100.0) if n_compared else 0.0

    result = {
        "enabled": True,
        "sample_id": sample_id_str,
        "sheet_name": sheet_name,
        "xlsx_path": xlsx_path,
        "excluded_items": sorted(EXCLUDED_ITEMS),
        "compared_item_count": n_compared,
        "ai_total": ai_total,
        "gt_total": gt_total,
        "diff": ai_total - gt_total,
        "abs_diff": abs(ai_total - gt_total),
        "mae": round(mae, 3),
        "rmse": round(rmse, 3),
        "accuracy": round(accuracy, 2),
        "match_count": match_count,
        "mismatch_count": mismatch_count,
        "items": compared,
    }

    logger.info(
        "gt_comparison: sample_id=%s compared=%d ai_total=%d gt_total=%d "
        "MAE=%.2f RMSE=%.2f accuracy=%.1f%%",
        sample_id_str, n_compared, ai_total, gt_total, mae, rmse, accuracy,
    )

    return {"gt_comparison": result}
