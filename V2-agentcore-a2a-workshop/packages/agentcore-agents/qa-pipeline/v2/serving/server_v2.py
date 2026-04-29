# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""V2 FastAPI server — `v2/graph_v2.py::build_graph_v2()` 를 HTTP 로 노출.

V1 `server.py` 와 독립적으로 동작. V1 원본 수정 없음.
AgentCore Runtime 호환 엔드포인트 (/ping, /invocations, /health, /readyz) + V2 전용
`/evaluate` (JSON) 를 제공한다. SSE `/evaluate/stream` 은 Phase D1 단계에서는 제외
(CLAUDE.md 지침에 따라 프롬프트 튜닝 배치는 `run_direct_batch_v2.py` 인프로세스 사용).

환경변수:
  PORT  (기본 8081 — V1 8080 과 충돌 방지)
  LOG_LEVEL  (기본 INFO)
"""

from __future__ import annotations

import json
import logging
import os
import sys
import threading
import time
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any


# ★ 로깅 설정을 최상단에서 무조건 실행 — `python -m uvicorn v2.serving.server_v2:app`
# 처럼 main_v2.py 를 거치지 않는 경로에서도 로그 포맷 / Bedrock 훅이 설치되게 한다.
from .logging_setup import configure as _configure_logging  # noqa: E402

_configure_logging()


# qa-pipeline 루트를 path 에
_QA_PIPELINE_DIR = Path(__file__).resolve().parents[2]
if str(_QA_PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(_QA_PIPELINE_DIR))


# ── .env.local 자동 로드 ─────────────────────────────────────────────────────
# `python -m uvicorn v2.serving.server_v2:app` 처럼 _launch_server.ps1 우회 기동시에도
# 로컬 dev env (QA_GT_XLSX_PATH, QA_AOSS_ENDPOINT 등) 주입되게 함.
# EC2 배포에는 영향 없음 (.env.local 은 .gitignore + EC2 에는 파일 없음).
# override=False — 이미 프로세스 env 에 있으면 그 값 유지 (_launch_server.ps1 과 호환).
def _load_env_local() -> None:
    env_file = _QA_PIPELINE_DIR / ".env.local"
    if not env_file.exists():
        return
    try:
        with env_file.open("r", encoding="utf-8") as f:
            for raw_line in f:
                line = raw_line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                name, _, value = line.partition("=")
                name = name.strip()
                value = value.strip().strip('"').strip("'")
                if name and name not in os.environ:
                    os.environ[name] = value
    except OSError:
        pass


_load_env_local()


from datetime import UTC  # noqa: E402
from fastapi import BackgroundTasks, FastAPI, Request  # noqa: E402
from fastapi.middleware.cors import CORSMiddleware  # noqa: E402
from fastapi.responses import JSONResponse, StreamingResponse  # noqa: E402
from v2.graph_v2 import build_graph_v2  # noqa: E402


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Lazy graph 빌드 — startup 에서 실패해도 /ping, /health 는 응답하도록 분리
# ---------------------------------------------------------------------------

_compiled_graph: Any = None
_graph_build_error: str | None = None


# ---------------------------------------------------------------------------
# Interactive discussion gates (V3 ensemble 모드 — auto_start=False 제어용)
# ---------------------------------------------------------------------------
# 프론트가 persona_mode=ensemble + auto_start=False 로 요청 시:
#   1. debate_node 가 discussion_id 생성 후 discussion_started 이벤트 발송.
#   2. run_debate 가 _DISCUSSION_GATES[discussion_id] 의 threading.Event.wait() 로 블록.
#   3. 프론트가 POST /v2/discussion/{discussion_id}/start 호출 → event.set() → 토론 재개.
#   4. next-round 엔드포인트도 동일 구조 (라운드 단위 스텝 진행 — future hook).
#
# dict[str, threading.Event]. 프로세스 수명 동안 유지. 완료된 discussion_id 는
# _discussion_gate_done() 에서 제거 (best-effort — 재설정 방지).
_DISCUSSION_GATES: dict[str, threading.Event] = {}
_DISCUSSION_GATES_LOCK = threading.Lock()


def _get_or_create_discussion_gate(discussion_id: str) -> threading.Event:
    """discussion_id 에 해당하는 threading.Event 를 반환. 없으면 생성.

    gate_factory 로 run_debate 에 주입되며, 프론트의 /start 호출이 .set() 을 실행.
    """
    with _DISCUSSION_GATES_LOCK:
        ev = _DISCUSSION_GATES.get(discussion_id)
        if ev is None:
            ev = threading.Event()
            _DISCUSSION_GATES[discussion_id] = ev
    return ev


def _release_discussion_gate(discussion_id: str) -> bool:
    """gate 가 존재하면 .set() 실행 (run_debate 의 wait() 해제). 반환: 존재 여부."""
    with _DISCUSSION_GATES_LOCK:
        ev = _DISCUSSION_GATES.get(discussion_id)
    if ev is None:
        return False
    ev.set()
    return True


def _cleanup_old_discussion_gates(max_keep: int = 200) -> None:
    """메모리 누수 방지 — 200 개 초과 시 FIFO 로 오래된 것부터 제거."""
    with _DISCUSSION_GATES_LOCK:
        if len(_DISCUSSION_GATES) <= max_keep:
            return
        excess = len(_DISCUSSION_GATES) - max_keep
        for key in list(_DISCUSSION_GATES.keys())[:excess]:
            _DISCUSSION_GATES.pop(key, None)


def _get_graph() -> Any:
    """싱글톤 graph 반환. 실패 시 None."""
    global _compiled_graph, _graph_build_error
    if _compiled_graph is not None:
        return _compiled_graph
    try:
        _compiled_graph = build_graph_v2()
        logger.info("server_v2: graph_v2 compiled")
    except Exception as e:
        _graph_build_error = str(e)
        logger.exception("server_v2: graph_v2 build failed")
    return _compiled_graph


@asynccontextmanager
async def lifespan(_app: FastAPI):
    """Startup 에서 graph 빌드 시도 (실패해도 서버는 기동)."""
    _get_graph()
    yield


app = FastAPI(title="QA Evaluation Pipeline V2", version="2.0.0", lifespan=lifespan)

# 브라우저(Next.js qa-webapp / chatbot-ui 로컬 서빙) 에서 호출 허용
# allow_credentials=True 필요 시 와일드카드 금지 → Next.js dev origin 명시
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)


@app.head("/health")
async def health_head() -> JSONResponse:
    return JSONResponse({"status": "healthy"})


# ---------------------------------------------------------------------------
# 헬스 엔드포인트 (AgentCore Runtime 호환)
# ---------------------------------------------------------------------------


@app.get("/ping")
async def ping() -> JSONResponse:
    """AgentCore Runtime liveness probe."""
    return JSONResponse({"status": "ok"})


@app.get("/health")
async def health() -> JSONResponse:
    """일반 liveness probe."""
    return JSONResponse({"status": "healthy", "service": "qa-pipeline-v2", "version": "2.0.0"})


@app.get("/readyz")
async def readyz() -> JSONResponse:
    """Readiness probe — graph 빌드 성공 여부."""
    graph = _get_graph()
    if graph is None:
        return JSONResponse({"ready": False, "graph_build_error": _graph_build_error}, status_code=503)
    return JSONResponse({"ready": True, "pipeline": "v2"})


def _count_source_docs_at(base: Path, kind: str) -> int:
    """주어진 디렉토리 아래 source doc 수 카운트.

    kind: 'golden' | 'reasoning' | 'knowledge'
    """
    import json as _json

    if not base.is_dir():
        return 0
    try:
        if kind == "golden":
            d = base / "golden_set"
            if not d.is_dir():
                return 0
            n = 0
            for f in d.glob("[0-9][0-9]_*.json"):
                try:
                    data = _json.loads(f.read_text(encoding="utf-8"))
                    n += len(data.get("examples", []))
                except Exception:
                    continue
            return n
        if kind == "reasoning":
            d = base / "reasoning_index"
            if not d.is_dir():
                return 0
            n = 0
            for f in d.glob("[0-9][0-9]_*.json"):
                try:
                    data = _json.loads(f.read_text(encoding="utf-8"))
                    n += len(data.get("reasoning_records", []))
                except Exception:
                    continue
            return n
        if kind == "knowledge":
            # manual.md 가 존재하면 H2 청크 개수 기반 추정.
            # 기존 site-level 은 BusinessKnowledgeRAG 엔진 사용 (정확) — _count_source_docs 하위 경로에서 호출.
            d = base / "business_knowledge"
            if not d.is_dir():
                return 0
            manual = d / "manual.md"
            if not manual.is_file():
                return 0
            try:
                import re as _re

                text = manual.read_text(encoding="utf-8")
                return len(_re.findall(r"^##\s", text, flags=_re.M))
            except Exception:
                return 0
    except Exception:
        pass
    return 0


def _count_source_docs(tenant_id: str, kind: str) -> int:
    """tenant(site) 레거시 카운트 — 기존 API 호환.

    레거시 경로 tenants/{site}/{golden_set|reasoning_index|business_knowledge}/ 기준.
    3단계 scope 카운트는 `_count_source_docs_at` + `_list_tenant_scopes` 참조.
    """
    base = Path(__file__).resolve().parents[1] / "tenants" / tenant_id
    if kind == "knowledge":
        # business_knowledge 는 site-level + 부서별 manual.md 까지 모두 합산해야
        # AOSS 인덱싱(부서 manual 포함) 과 일치 (그 외엔 indexed > source 로 "duplicated" 오탐).
        # meta 주석이 붙은 chunk 만 카운트 — bootstrap_aoss_qa.py 인덱싱과 정확히 동일 기준.
        try:
            from v2.rag.business_knowledge import _CHUNK_RE

            n = 0
            if not base.is_dir():
                return 0
            for manual in base.rglob("business_knowledge/manual.md"):
                try:
                    text = manual.read_text(encoding="utf-8")
                    n += sum(1 for _ in _CHUNK_RE.finditer(text))
                except Exception:
                    continue
            return n
        except Exception:
            return 0
    return _count_source_docs_at(base, kind)


def _list_tenant_scopes() -> list[dict[str, Any]]:
    """filesystem 스캔으로 3단계 scope 리스트 반환 (2026-04-27 단순화).

    `_shared` 메타 폴더 패턴 폐기 → 직하 = 공통, 하위 폴더 = override (실무 표준).

    각 scope:
      {
        "site_id": str,
        "channel": str | None,      # None = site 직하 (사이트 공통)
        "department": str | None,   # None = channel 직하 (채널 공통)
        "is_shared": bool,          # 채널/사이트 직하인지 (이전 _shared 와 의미 동일)
        "label": str,
        "source": {"golden": int, "reasoning": int, "knowledge": int},
        "has_config": bool
      }
    """
    root = Path(__file__).resolve().parents[1] / "tenants"
    scopes: list[dict[str, Any]] = []
    if not root.is_dir():
        return scopes

    KINDS = ("golden", "reasoning", "knowledge")
    # 자원 폴더 이름은 부서명으로 사용 불가 (reserved) — 부서 폴더 스캔 시 제외.
    RESERVED_RESOURCE_DIRS = {
        "golden_set",
        "reasoning_index",
        "business_knowledge",
        "mandatory_scripts",
        "_shared",  # _shared 는 폐기됐지만 마이그레이션 안 끝난 환경 대비
    }

    def _source_of(base: Path) -> dict[str, int]:
        return {k: _count_source_docs_at(base, k) for k in KINDS}

    def _has_any_source(base: Path) -> bool:
        return any((base / sub).is_dir() for sub in ("golden_set", "reasoning_index", "business_knowledge"))

    for site_dir in sorted(d for d in root.iterdir() if d.is_dir()):
        site = site_dir.name

        # 1) site 직하 = 사이트 공통 자원
        if _has_any_source(site_dir):
            scopes.append(
                {
                    "site_id": site,
                    "channel": None,
                    "department": None,
                    "is_shared": True,  # 사이트 공통 (이전 _shared 와 동일 의미)
                    "label": f"{site} (사이트 공통)",
                    "source": _source_of(site_dir),
                    "has_config": (site_dir / "tenant_config.yaml").is_file(),
                }
            )

        # 2) channels (inbound / outbound)
        for ch_dir in sorted(d for d in site_dir.iterdir() if d.is_dir() and d.name in ("inbound", "outbound")):
            channel = ch_dir.name

            # 2a) channel 직하 = 채널 공통 자원
            if _has_any_source(ch_dir):
                scopes.append(
                    {
                        "site_id": site,
                        "channel": channel,
                        "department": None,
                        "is_shared": True,  # 채널 공통
                        "label": f"{site}/{channel} (채널 공통)",
                        "source": _source_of(ch_dir),
                        "has_config": (ch_dir / "tenant_config.yaml").is_file(),
                    }
                )

            # 2b) department 하위 — reserved 자원/메타 폴더 제외 모든 폴더
            for dept_dir in sorted(d for d in ch_dir.iterdir() if d.is_dir() and d.name not in RESERVED_RESOURCE_DIRS):
                scopes.append(
                    {
                        "site_id": site,
                        "channel": channel,
                        "department": dept_dir.name,
                        "is_shared": False,
                        "label": f"{site}/{channel}/{dept_dir.name}",
                        "source": _source_of(dept_dir),
                        "has_config": (dept_dir / "tenant_config.yaml").is_file(),
                    }
                )

    return scopes


@app.get("/v2/rag/scopes")
async def rag_scopes() -> JSONResponse:
    """경량 endpoint — scopes 배열만 반환 (filesystem 스캔 only, AOSS 호출 없음).

    `/v2/rag/status` 는 AOSS index count 호출로 ~3초 걸리는데, 프론트엔드의
    TenantStatusBadge 는 scopes 배열만 필요하므로 이 엔드포인트가 즉시 응답.
    """
    return JSONResponse({"scopes": _list_tenant_scopes()})


@app.get("/v2/tenants/{site_id}/rubric")
async def tenant_rubric(site_id: str, department: str | None = None) -> JSONResponse:
    """tenant rubric 메타 반환 — frontend 동적 노드/배점 표시용.

    응답 구조:
        {
          "site_id": "shinhan",
          "department": "collection",
          "common_categories": [...],   # 7개 공통 (greeting / listening_comm / ...)
          "dept_nodes": [               # 부서특화 (대분류 = node, sub-items 포함)
            {
              "node_id": "coll_accuracy",
              "label": "업무 정확도",
              "max_score": 20,
              "items": [
                {"item_number": 901, "item_name": "...", "max_score": 10, "allowed_steps": [10,7,5,0]},
                ...
              ]
            },
            ...
          ],
          "total_max_score": 100
        }

    department 미지정 시 dept_nodes=[] 로 공통 카테고리만 반환.
    """
    from v2.schemas.enums import CATEGORY_META

    # 공통 카테고리 (8개 base — work_accuracy 포함, 부서특화 활성 시 frontend 가 hidden 처리)
    common_categories = []
    for cat_key, meta in CATEGORY_META.items():
        common_categories.append({
            "category_key": cat_key,
            "label_ko": meta["label_ko"],
            "label_en": meta["label_en"],
            "items": list(meta["items"]),
            "max_score": meta["max_score"],
        })

    # 부서특화 노드 (있는 경우만)
    dept_nodes: list[dict[str, Any]] = []
    try:
        from v2.agents.shinhan_dept.registry import (
            DEPT_NODE_REGISTRY,
            get_dept_nodes_for_tenant,
        )
        active_node_ids = get_dept_nodes_for_tenant(site_id, department)
        for nid in active_node_ids:
            spec = DEPT_NODE_REGISTRY.get(nid)
            if spec is None:
                continue
            dept_nodes.append({
                "node_id": spec["node_id"],
                "team_id": spec["team_id"],
                "label": spec["label_ko"],
                "category_key": spec["category_key"],
                "max_score": spec["max_score"],
                "rubric_focus": spec["rubric_focus"],
                "items": list(spec["items"]),  # 이미 dict 리스트
            })
    except Exception as e:  # pragma: no cover — registry 미로드 시 빈 배열
        logger.warning("tenant_rubric: dept registry 로드 실패: %s", e)

    # work_accuracy 가 부서특화 *_accuracy 로 대체되는지 표시 (frontend 와 정합)
    work_accuracy_replaced = bool(dept_nodes) and any(
        n["node_id"].endswith("_accuracy") for n in dept_nodes
    )

    common_score_total = sum(c["max_score"] for c in common_categories)
    if work_accuracy_replaced:
        wa = next((c for c in common_categories if c["category_key"] == "work_accuracy"), None)
        if wa:
            common_score_total -= wa["max_score"]

    dept_score_total = sum(n["max_score"] for n in dept_nodes)
    return JSONResponse({
        "site_id": site_id,
        "department": department,
        "common_categories": common_categories,
        "dept_nodes": dept_nodes,
        "work_accuracy_replaced": work_accuracy_replaced,
        "common_max_score": common_score_total,
        "dept_max_score": dept_score_total,
        "total_max_score": common_score_total + dept_score_total,
    })


@app.get("/v2/rag/status")
async def rag_status() -> JSONResponse:
    """모든 tenant × 모든 RAG 인덱스 도큐먼트 수 매트릭스 + source 비교."""
    import os as _os

    region = _os.environ.get("AWS_REGION") or "us-east-1"

    # tenant 디렉토리 스캔
    _TENANTS_DIR = Path(__file__).resolve().parents[1] / "tenants"
    tenants: list[str] = []
    if _TENANTS_DIR.exists():
        tenants = sorted([d.name for d in _TENANTS_DIR.iterdir() if d.is_dir()])

    # 3단계 scope 리스트 (2026-04-24) — 프론트 department 드롭다운 + scope 테이블용.
    scopes = _list_tenant_scopes()

    out: dict[str, Any] = {"tenants": tenants, "region": region, "indexes": [], "scopes": scopes}
    try:
        from v2.rag.aoss_store import GOLDEN_INDEX, KNOWLEDGE_INDEX, REASONING_INDEX, AossStore
    except Exception as e:
        return JSONResponse({**out, "error": f"aoss_store import 실패: {e}"})

    INDEX_KINDS = [
        (GOLDEN_INDEX, "golden", "Golden-set (few-shot)"),
        (REASONING_INDEX, "reasoning", "Reasoning (stdev)"),
        (KNOWLEDGE_INDEX, "knowledge", "Business Knowledge"),
    ]

    for idx_name, kind, label in INDEX_KINDS:
        entry: dict[str, Any] = {
            "name": idx_name,
            "kind": kind,
            "label": label,
            "exists": False,
            "by_tenant": {},
            "total_indexed": 0,
            "total_source": 0,
        }
        # source counts (AOSS 무관, 디스크 기준)
        for t in tenants:
            entry["by_tenant"].setdefault(t, {})["source"] = _count_source_docs(t, kind)
            entry["total_source"] += entry["by_tenant"][t]["source"]
        # AOSS counts
        try:
            store = AossStore(idx_name)
            if store.index_exists():
                entry["exists"] = True
                entry["total_indexed"] = store.count_docs()
                for t in tenants:
                    cnt = store.count_docs(tenant_id=t)
                    entry["by_tenant"][t]["indexed"] = cnt
        except Exception as e:
            entry["error"] = str(e)[:200]
        # 미색인 tenant 의 indexed 는 0 으로 보강
        for t in tenants:
            entry["by_tenant"][t].setdefault("indexed", 0)
            src = entry["by_tenant"][t]["source"]
            idx = entry["by_tenant"][t]["indexed"]
            # 상태 분류
            if src == 0 and idx == 0:
                status = "empty"  # 양쪽 다 0 — 데이터 없음
            elif src == idx and src > 0:
                status = "synced"  # 정확히 일치
            elif idx == 0 and src > 0:
                status = "not_built"  # 소스는 있는데 색인 안 됨
            elif idx > src:
                status = "duplicated"  # 중복 누적 가능성
            else:
                status = "stale"  # 색인 < 소스 (보강 필요)
            entry["by_tenant"][t]["status"] = status
        out["indexes"].append(entry)
    return JSONResponse(out)


@app.post("/v2/rag/build")
async def rag_build(request: Request) -> StreamingResponse:
    """bootstrap_aoss_qa.py 실행 + 진행 SSE 스트리밍.

    body 예시 (3단계 멀티테넌트 — 2026-04-24):
      {"site_id": "kolon", "channel": "inbound", "department": "cs", "recreate": false}
      {"tenant": "kolon", "recreate": false}   ← 레거시 (site_id alias)
      {"recreate": true}                         ← 전체 tenant (site/channel/department 모두 생략)
    """
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    # 3단계 멀티테넌트 필드 파싱 — 레거시 `tenant` 는 `site_id` alias.
    site_id = body.get("site_id") or body.get("tenant")  # None = 전체
    channel = body.get("channel")  # None = 전체 채널
    department = body.get("department")  # None = 전체 부서
    recreate = bool(body.get("recreate", False))
    clean_tenant = bool(body.get("clean_tenant", False))

    import asyncio as _asyncio
    import subprocess as _subprocess

    pipeline_dir = Path(__file__).resolve().parents[2]
    script_path = pipeline_dir / "v2" / "scripts" / "bootstrap_aoss_qa.py"

    args = [sys.executable, str(script_path)]
    if site_id:
        args += ["--site", str(site_id)]
    if channel:
        args += ["--channel", str(channel)]
    if department:
        args += ["--department", str(department)]
    if recreate:
        args += ["--recreate"]
    if clean_tenant:
        args += ["--clean-tenant"]

    def sse(event: str, data: dict[str, Any]) -> str:
        return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

    async def stream():
        env = os.environ.copy()
        env.setdefault("QA_RAG_EMBEDDING", "titan")
        env.setdefault("PYTHONIOENCODING", "utf-8")
        # QA_AOSS_ENDPOINT 을 명시적으로 subprocess 에 전달 — 없으면 SSM fallback.
        if env.get("QA_AOSS_ENDPOINT"):
            logger.info("🔧 rag_build subprocess env QA_AOSS_ENDPOINT=%s", env["QA_AOSS_ENDPOINT"])
        else:
            logger.warning("⚠️ rag_build: QA_AOSS_ENDPOINT env 없음 — SSM fallback 사용될 것")
        logger.info("🚀 rag_build subprocess spawn — cmd=%s cwd=%s", args, pipeline_dir)
        yield sse(
            "start",
            {
                "args": args,
                "site_id": site_id or "ALL",
                "channel": channel or "ALL",
                "department": department or "ALL",
                "tenant": site_id or "ALL",  # 레거시 호환
                "recreate": recreate,
            },
        )

        # Windows uvicorn SelectorEventLoop 는 asyncio.create_subprocess_exec 미지원 (NotImplementedError).
        # 동기 subprocess.Popen + 백그라운드 스레드 + asyncio.Queue 로 우회.
        import queue as _queue
        import threading as _threading

        line_q: _queue.Queue[str | None] = _queue.Queue()
        proc_holder: dict[str, Any] = {}
        exc_holder: dict[str, BaseException] = {}

        def _worker() -> None:
            try:
                p = _subprocess.Popen(
                    args,
                    cwd=str(pipeline_dir),
                    stdout=_subprocess.PIPE,
                    stderr=_subprocess.STDOUT,
                    env=env,
                    bufsize=1,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                )
                proc_holder["proc"] = p
                proc_holder["pid"] = p.pid
                assert p.stdout is not None
                for line in p.stdout:
                    stripped = line.rstrip()
                    if stripped:
                        line_q.put(stripped)
                rc = p.wait()
                proc_holder["rc"] = rc
            except BaseException as e:  # noqa: BLE001
                exc_holder["exc"] = e
            finally:
                line_q.put(None)  # sentinel — stream 종료

        t = _threading.Thread(target=_worker, name="rag-bootstrap", daemon=True)
        t.start()

        try:
            line_count = 0
            while True:
                # blocking Queue.get 을 이벤트 루프 blocking 없이 기다리려면 to_thread
                item = await _asyncio.to_thread(line_q.get)
                if item is None:
                    break
                line_count += 1
                logger.info("  [bootstrap] %s", item)
                yield sse("log", {"line": item})
            if "exc" in exc_holder:
                raise exc_holder["exc"]
            rc = proc_holder.get("rc", -1)
            logger.info("✅ rag_build 완료 — rc=%s lines=%d pid=%s", rc, line_count, proc_holder.get("pid"))
            yield sse("done", {"return_code": rc, "ok": rc == 0, "lines": line_count})
        except Exception as e:
            logger.exception("❌ rag_build subprocess 실패: %r", e)
            yield sse("error", {"message": f"{type(e).__name__}: {e}" if str(e) else type(e).__name__})

    return StreamingResponse(stream(), media_type="text/event-stream")


@app.get("/v2/gt-scores")
async def gt_scores(sample_id: str, sheet: str | None = None) -> JSONResponse:
    """QA 정답 xlsx 에서 sample_id 매칭 시트 → 항목별 GT 점수 + 비고 반환.

    파일 경로: 환경변수 QA_GT_XLSX_PATH 또는 기본 Desktop 경로.
    시트명 매칭 (관대): suffix `_{sample_id}` → 단순 포함 → 숫자만 비교 (앞 0 무시) 순.
    sheet 쿼리 파라미터로 사용자 직접 시트 지정 가능 (시트 picker fallback).
    반환: {sample_id, sheet_name, total_score, items[], available_sheets[], match_method, xlsx_path}
      items[i] = {category, item_name, max_score, score, note, item_number}
    """
    import os as _os
    import platform as _platform
    import re as _re

    # 후보 경로 — 환경변수 우선, 없으면 OS 별 기본 경로 순차 탐색.
    # 2026-04-27: 환경별 분기 — Windows 는 Desktop, Linux(EC2) 는 홈/표준 데이터 디렉토리.
    _GT_FILENAMES = (
        "QA정답-STT_기반_통합_상담평가표_v3재평가_fixed.xlsx",
        "QA정답-STT_기반_통합_상담평가표_v3재평가.xlsx",
        "코오롱 업무 정확도 auto_qa_criteria.xlsx",
    )
    candidates: list[str] = []
    env_path = _os.environ.get("QA_GT_XLSX_PATH")
    if env_path:
        candidates.append(env_path)

    if _platform.system() == "Windows":
        # 로컬 개발 — 사용자 Desktop
        _base_dirs = [Path(r"C:\Users\META M\Desktop")]
    else:
        # EC2 / Linux — 표준 위치 후보
        _qa_pipeline_root = Path(__file__).resolve().parent.parent.parent  # qa-pipeline/
        _base_dirs = [
            Path.home() / "qa-data",
            Path("/home/ubuntu/qa-data"),
            Path("/opt/qa-pipeline/data"),
            _qa_pipeline_root / "data" / "gt",
            _qa_pipeline_root / "data",
        ]
    for _b in _base_dirs:
        for _fn in _GT_FILENAMES:
            candidates.append(str(_b / _fn))

    xlsx_path = next((p for p in candidates if p and Path(p).exists()), None)

    if not xlsx_path:
        return JSONResponse(
            {
                "error": "gt_xlsx_not_found",
                "tried": [c for c in candidates if c],
                "hint": "환경변수 QA_GT_XLSX_PATH 로 지정하거나 Desktop 에 파일 배치",
            },
            status_code=404,
        )

    try:
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        return JSONResponse({"error": f"xlsx load failed: {e}"}, status_code=500)

    try:
        sheets = wb.sheetnames
        match_method = "none"
        matched: list[str] = []

        # 0) 사용자가 sheet 쿼리 파라미터로 직접 지정
        if sheet and sheet in sheets:
            matched = [sheet]
            match_method = "explicit"
        else:
            # 1) suffix 매칭 (V2 호환)
            target_suffix = f"_{sample_id}"
            matched = [s for s in sheets if s.endswith(target_suffix)]
            if matched:
                match_method = "suffix"
            else:
                # 2) 단순 포함 매칭
                matched = [s for s in sheets if sample_id in s]
                if matched:
                    match_method = "contains"
                else:
                    # 3) 숫자만 비교 (앞 0 패딩 무시) — sample_id "4" / 시트 "_004" 매칭
                    sid_digits = _re.sub(r"\D", "", sample_id) or sample_id
                    sid_int: int | None = None
                    try:
                        sid_int = int(sid_digits) if sid_digits else None
                    except (ValueError, TypeError):
                        sid_int = None
                    if sid_int is not None:
                        for s in sheets:
                            nums = _re.findall(r"\d+", s)
                            if any(int(n) == sid_int for n in nums if n):
                                matched.append(s)
                        if matched:
                            match_method = "digits"

        if not matched:
            return JSONResponse(
                {
                    "error": "sample_id_not_found",
                    "sample_id": sample_id,
                    "xlsx_path": xlsx_path,
                    "available_sheets": sheets,
                    "hint": "sheet 쿼리 파라미터로 직접 지정 가능: /v2/gt-scores?sample_id=X&sheet=시트명",
                },
                status_code=404,
            )
        ws = wb[matched[0]]

        # v3 평가표 17 항목 순서 (ALLOWED_STEPS key 와 일치)
        item_numbers = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
        items: list[dict[str, Any]] = []
        total_score = 0
        cur_idx = 0
        current_category: str | None = None

        for r in range(6, 50):
            cat_cell = ws.cell(row=r, column=1).value
            item_name = ws.cell(row=r, column=2).value
            max_score = ws.cell(row=r, column=4).value
            score = ws.cell(row=r, column=5).value
            note = ws.cell(row=r, column=6).value

            if cat_cell and isinstance(cat_cell, str) and "총점" in cat_cell:
                break
            if cat_cell:
                current_category = str(cat_cell).strip()
            if item_name is None:
                continue
            if cur_idx >= len(item_numbers):
                break

            items.append(
                {
                    "item_number": item_numbers[cur_idx],
                    "category": current_category,
                    "item_name": str(item_name).strip() if isinstance(item_name, str) else item_name,
                    "max_score": int(max_score) if isinstance(max_score, (int, float)) else max_score,
                    "score": int(score) if isinstance(score, (int, float)) else score,
                    "note": str(note).strip() if isinstance(note, str) else note,
                }
            )
            if isinstance(score, (int, float)):
                total_score += int(score)
            cur_idx += 1

        return JSONResponse(
            {
                "sample_id": sample_id,
                "sheet_name": matched[0],
                "total_score": total_score,
                "items": items,
                "available_sheets": sheets,
                "match_method": match_method,
                "xlsx_path": xlsx_path,
            }
        )
    finally:
        try:
            wb.close()
        except Exception:
            pass


@app.post("/save-xlsx")
async def save_xlsx(request: Request) -> JSONResponse:
    """프론트가 생성한 xlsx 파일을 로컬 폴더에 저장.

    multipart/form-data:
      - file: xlsx 바이너리
      - filename: 저장 파일명
      - subfolder: 하위 폴더명 (예: 'QA평가결과/2026-04-22')

    저장 루트: 환경변수 QA_RESULT_SAVE_ROOT (기본: Desktop/QA평가결과)
    """
    import os as _os

    try:
        form = await request.form()
        file = form.get("file")
        filename = form.get("filename") or "result.xlsx"
        subfolder = form.get("subfolder") or ""
        if not file:
            return JSONResponse({"error": "file 누락"}, status_code=400)

        # 안전: 파일명에서 디렉토리 traversal 방지
        safe_name = Path(str(filename)).name
        if not safe_name.lower().endswith(".xlsx"):
            safe_name += ".xlsx"
        safe_subfolder = "/".join(p for p in str(subfolder).split("/") if p and p not in ("..", "."))

        root_str = _os.environ.get("QA_RESULT_SAVE_ROOT") or str(Path.home() / "Desktop" / "QA평가결과")
        root = Path(root_str)
        target_dir = root / safe_subfolder if safe_subfolder else root
        target_dir.mkdir(parents=True, exist_ok=True)

        target_path = target_dir / safe_name
        # 동명 존재 시 접미사 추가
        if target_path.exists():
            i = 1
            stem = target_path.stem
            suf = target_path.suffix
            while True:
                cand = target_dir / f"{stem}_{i}{suf}"
                if not cand.exists():
                    target_path = cand
                    break
                i += 1
                if i > 9999:
                    return JSONResponse({"error": "너무 많은 동명 파일"}, status_code=500)

        content = await file.read()
        target_path.write_bytes(content)
        logger.info("save-xlsx: %s (%d bytes)", target_path, len(content))
        return JSONResponse({"ok": True, "path": str(target_path), "bytes": len(content)})
    except Exception as e:
        logger.exception("save-xlsx 실패")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/v2/aws-resources")
async def aws_resources(tenant_id: str = "generic") -> JSONResponse:
    """AOSS + S3 + DynamoDB + API Gateway + tenant paths 리소스 상태.

    V3 NodeDrawer AWS 패널 용 확장판:
      - aoss: 인덱스별 exists/doc_count/status/error  (qa-golden-set, qa-reasoning-index,
              qa-business-knowledge, rag-documents)
      - s3: bucket 존재 + HeadBucket 권한 상태
      - dynamodb: describe_table + 권한 상태
      - api_gateway: endpoint URL + 존재 여부
      - tenant_paths: tenant_id 별 config_dir / output_dir / golden_set_dir /
                      reasoning_index_dir 절대경로 + 존재 여부
      - identity: 현재 caller IAM (STS get-caller-identity)
      - overall_status: "ok" | "partial" | "error" (리소스 전반 요약)
    """
    import os as _os

    backend = (_os.environ.get("QA_RAG_BACKEND") or "aoss").strip().lower()
    embed_backend = (_os.environ.get("QA_RAG_EMBEDDING") or "titan").strip().lower()

    report: dict[str, Any] = {
        "tenant_id": tenant_id,
        "qa_rag_backend": backend,  # aoss | jaccard
        "qa_rag_embedding": embed_backend,  # titan | jaccard
        "region": _os.environ.get("AWS_REGION") or _os.environ.get("AWS_DEFAULT_REGION") or "us-east-1",
        "connected": False,
        "overall_status": "error",
        "aoss": None,
        "s3": None,
        "dynamodb": None,
        "api_gateway": None,
        "tenant_paths": None,
        "identity": None,
        "errors": [],
    }

    # --- tenant_paths: 로컬 디렉토리 (boto3 무관, 항상 조사) ---
    try:
        _v2_base = Path(__file__).resolve().parents[1]
        _tenants_dir = _v2_base / "tenants"
        _tenant_dir = _tenants_dir / tenant_id
        _golden_dir = _tenant_dir / "golden_set"
        _reasoning_dir = _tenant_dir / "reasoning_index"
        _knowledge_dir = _tenant_dir / "business_knowledge"
        _output_dir_env = _os.environ.get("QA_RESULT_JSON_ROOT")
        _output_dir = Path(_output_dir_env) if _output_dir_env else (Path.home() / "Desktop" / "QA평가결과" / "JSON")
        _export_dir_env = _os.environ.get("QA_HITL_EXPORT_ROOT")
        _export_dir = (
            Path(_export_dir_env) if _export_dir_env else (Path.home() / "Desktop" / "QA평가결과" / "HITL_비교")
        )

        def _path_status(p: Path) -> dict[str, Any]:
            return {"path": str(p), "exists": p.exists(), "is_dir": p.is_dir() if p.exists() else False}

        report["tenant_paths"] = {
            "tenants_root": _path_status(_tenants_dir),
            "config_dir": _path_status(_tenant_dir),
            "golden_set_dir": _path_status(_golden_dir),
            "reasoning_index_dir": _path_status(_reasoning_dir),
            "business_knowledge_dir": _path_status(_knowledge_dir),
            "output_dir": _path_status(_output_dir),
            "hitl_export_dir": _path_status(_export_dir),
        }
        if not _tenant_dir.exists():
            report["errors"].append(f"tenant 디렉토리 없음: {_tenant_dir}")
    except Exception as e:  # noqa: BLE001
        report["errors"].append(f"tenant_paths 조사 실패: {e.__class__.__name__}: {str(e)[:120]}")

    try:
        import boto3  # type: ignore
    except ImportError:
        report["errors"].append("boto3 미설치")
        return JSONResponse(report)

    ssm_region = report["region"]

    # --- identity: STS get-caller-identity — AOSS/S3 권한 디버깅에 유용 ---
    try:
        sts = boto3.client("sts", region_name=ssm_region)
        ident = sts.get_caller_identity()
        report["identity"] = {
            "account": ident.get("Account"),
            "arn": ident.get("Arn"),
            "user_id": ident.get("UserId"),
            "status": "ok",
        }
    except Exception as e:  # noqa: BLE001
        report["identity"] = {"status": "error", "error": f"{e.__class__.__name__}: {str(e)[:120]}"}
        report["errors"].append("STS get-caller-identity 실패 — AWS 자격증명 미설정 가능성")

    # --- SSM 파라미터 조회 → 엔드포인트/버킷명 확보 ---
    ssm = boto3.client("ssm", region_name=ssm_region)
    param_map: dict[str, str] = {}
    try:
        resp = ssm.get_parameters_by_path(Path="/a2a_rag/", Recursive=False)
        for p in resp.get("Parameters", []):
            param_map[p["Name"]] = p["Value"]
    except Exception as e:  # noqa: BLE001
        report["errors"].append(f"SSM 파라미터 조회 실패: {e.__class__.__name__}")
        # SSM 실패해도 tenant_paths / identity 는 응답.
        return JSONResponse(report)

    # --- 1) AOSS — collection + QA 인덱스 3개 + rag-documents ---
    try:
        from v2.rag.aoss_store import (
            GOLDEN_INDEX,
            KNOWLEDGE_INDEX,
            REASONING_INDEX,
            AossStore,
            _resolve_endpoint as _aoss_resolve_endpoint,
        )

        # env QA_AOSS_ENDPOINT 우선, 없으면 SSM 폴백 (aoss_store._resolve_endpoint 와 동일 우선순위).
        # 이 라우트가 파이프라인 실제 호출과 동일한 endpoint 를 보고하도록 통일.
        env_or_ssm_endpoint = _aoss_resolve_endpoint()
        ssm_endpoint = param_map.get("/a2a_rag/opensearch_endpoint")
        ssm_collection = param_map.get("/a2a_rag/opensearch_collection_name")
        endpoint = env_or_ssm_endpoint or ssm_endpoint

        # endpoint 가 SSM 값과 다르면 (env override) collection 이름 재해석.
        # endpoint host 의 첫 토큰이 collection ID — boto3 batch_get_collection 으로 name 조회.
        if endpoint and ssm_endpoint and endpoint != ssm_endpoint:
            from urllib.parse import urlparse

            coll_id = (urlparse(endpoint).hostname or "").split(".")[0]
            collection = coll_id  # 폴백: ID 그대로 표시
            try:
                aoss_admin = boto3.client("opensearchserverless", region_name=ssm_region)
                details = aoss_admin.batch_get_collection(ids=[coll_id]).get("collectionDetails", [])
                if details and details[0].get("name"):
                    collection = details[0]["name"]
            except Exception:  # noqa: BLE001
                pass
        else:
            collection = ssm_collection
        aoss_info: dict[str, Any] = {
            "collection": collection,
            "endpoint": endpoint,
            "status": "missing" if not endpoint else "unknown",
            "indexes": [],
        }
        if endpoint:
            # V3 NodeDrawer 가 요구하는 인덱스 전체 — QA 3개 + rag-documents
            for idx in (GOLDEN_INDEX, REASONING_INDEX, KNOWLEDGE_INDEX, "rag-documents"):
                entry: dict[str, Any] = {"name": idx, "exists": False, "doc_count": 0, "status": "missing"}
                try:
                    store = AossStore(idx, endpoint=endpoint)
                    if not store.index_exists():
                        aoss_info["indexes"].append(entry)
                        continue
                    entry["exists"] = True
                    entry["status"] = "ok"
                    tid = None if idx == "rag-documents" else tenant_id
                    entry["doc_count"] = store.count_docs(tenant_id=tid)
                except Exception as e:  # noqa: BLE001
                    msg = f"{e.__class__.__name__}: {str(e)[:200]}"
                    entry["error"] = msg
                    entry["status"] = "error"
                    if "Authorization" in e.__class__.__name__ or "Authentication" in e.__class__.__name__:
                        report["errors"].append(
                            f"AOSS 권한 거부 ({idx}) — data access policy Principals 에 현재 IAM 추가 필요"
                        )
                aoss_info["indexes"].append(entry)

            # AOSS 전체 status: 모든 인덱스 ok 면 ok, 일부 ok 면 partial, 전부 실패면 error
            _idx_statuses = [i.get("status") for i in aoss_info["indexes"]]
            if all(s == "ok" for s in _idx_statuses):
                aoss_info["status"] = "ok"
            elif any(s == "ok" for s in _idx_statuses):
                aoss_info["status"] = "partial"
            else:
                aoss_info["status"] = "error"
        report["aoss"] = aoss_info
    except Exception as e:  # noqa: BLE001
        report["errors"].append(f"AOSS 확인 실패: {e.__class__.__name__}: {str(e)[:120]}")
        report["aoss"] = {"status": "error", "error": str(e)[:120]}

    # --- 2) S3 — HeadBucket + 권한 검사 ---
    try:
        bucket = param_map.get("/a2a_rag/document_bucket")
        s3_info: dict[str, Any] = {"name": bucket, "exists": False, "status": "missing"}
        if bucket:
            s3 = boto3.client("s3", region_name=ssm_region)
            s3.head_bucket(Bucket=bucket)
            s3_info["exists"] = True
            s3_info["status"] = "ok"
            s3_info["permission"] = "head_bucket_ok"
        report["s3"] = s3_info
    except Exception as e:  # noqa: BLE001
        cls = e.__class__.__name__
        report["s3"] = {
            "name": param_map.get("/a2a_rag/document_bucket"),
            "exists": False,
            "status": "error",
            "error": f"{cls}: {str(e)[:120]}",
            "permission": "denied" if "403" in str(e) or "Forbidden" in str(e) else "unknown",
        }

    # --- 3) DynamoDB — describe_table ---
    try:
        table_name = param_map.get("/a2a_rag/metadata_table")
        ddb_info: dict[str, Any] = {"name": table_name, "exists": False, "status": "missing"}
        if table_name:
            ddb = boto3.client("dynamodb", region_name=ssm_region)
            d = ddb.describe_table(TableName=table_name)
            ddb_info["exists"] = True
            ddb_info["status"] = "ok"
            ddb_info["table_status"] = d["Table"]["TableStatus"]
            ddb_info["item_count"] = int(d["Table"].get("ItemCount", 0))
        report["dynamodb"] = ddb_info
    except Exception as e:  # noqa: BLE001
        report["dynamodb"] = {
            "name": param_map.get("/a2a_rag/metadata_table"),
            "exists": False,
            "status": "error",
            "error": str(e)[:120],
        }

    # --- 4) API Gateway — URL + reachability 표기 (HTTP 호출은 생략, URL 존재 여부만) ---
    _api_url = param_map.get("/a2a_rag/api_url")
    report["api_gateway"] = {"url": _api_url, "status": "ok" if _api_url else "missing", "exists": bool(_api_url)}

    # --- 5) overall status ---
    aoss_ok = bool(report.get("aoss") and report["aoss"].get("status") in ("ok", "partial"))
    s3_ok = bool(report.get("s3") and report["s3"].get("status") == "ok")
    ddb_ok = bool(report.get("dynamodb") and report["dynamodb"].get("status") == "ok")
    apigw_ok = bool(report.get("api_gateway") and report["api_gateway"].get("status") == "ok")

    _oks = [aoss_ok, s3_ok, ddb_ok, apigw_ok]
    if all(_oks):
        report["overall_status"] = "ok"
    elif any(_oks):
        report["overall_status"] = "partial"
    else:
        report["overall_status"] = "error"
    report["connected"] = aoss_ok or s3_ok or ddb_ok

    return JSONResponse(report)


# ---------------------------------------------------------------------------
# /evaluate — JSON 입력 → V2 graph 실행 → final_state JSON 반환
# ---------------------------------------------------------------------------


def _build_initial_state(body: dict[str, Any]) -> dict[str, Any]:
    """요청 body → QAStateV2 초기 상태.

    필수: transcript
    선택: stt_metadata, site_id / channel / department (3단계 멀티테넌트),
          session_id, customer_id, plan, consultation_type,
          llm_backend, bedrock_model_id
    레거시: tenant_id — site_id 미지정 시 fallback 으로 쓰임 (한시적).
    """
    transcript = body.get("transcript", "")
    if not transcript:
        raise ValueError("transcript 필수")

    session_id = body.get("session_id") or body.get("consultation_id") or f"v2-{int(time.time())}"

    # 3단계 멀티테넌트 — body 우선순위:
    #   site_id   : body.site_id > body.tenant_id(legacy) > "generic"
    #   channel   : body.channel > "inbound"
    #   department: body.department > "default"
    # SITE_CD → channel 매핑은 프론트엔드(EvaluateRunner.tsx) 에서 수행 후 body.channel 로 전달.
    site_id = body.get("site_id") or body.get("tenant_id") or "generic"
    channel = body.get("channel") or "inbound"
    department = body.get("department") or "default"
    tenant_key = f"{site_id}:{channel}:{department}"

    # auto_start — V3 interactive discussion 제어. 기본 True (backward compat).
    #   True  → debate 가 즉시 실행 (기존 동작).
    #   False → discussion_started 이벤트 발송 후 /v2/discussion/{id}/start 대기.
    _auto_raw = body.get("auto_start", True)
    if isinstance(_auto_raw, str):
        auto_start = _auto_raw.lower() not in ("0", "false", "no", "off")
    else:
        auto_start = bool(_auto_raw) if _auto_raw is not None else True

    return {
        "transcript": transcript,
        "consultation_id": body.get("consultation_id") or session_id,
        "session_id": session_id,
        "customer_id": body.get("customer_id") or session_id,
        "consultation_type": body.get("consultation_type", "general"),
        "site_id": site_id,
        "channel": channel,
        "department": department,
        "tenant_key": tenant_key,
        "tenant_id": site_id,  # 레거시 — site_id alias. 구 sub-agent 코드 호환.
        "llm_backend": body.get("llm_backend", "bedrock"),
        "bedrock_model_id": body.get("bedrock_model_id"),
        # GT 비교 — body.gt_sample_id 또는 JSON 샘플의 id 필드 매핑 (프론트에서 전달)
        "gt_sample_id": body.get("gt_sample_id") or body.get("sample_id") or "",
        # 프론트 토글: persona 모드 선택 ("single" | "ensemble" | None).
        # 평가 시작 직전 set_runtime_force_single() 호출에 쓰임.
        "_persona_mode_override": body.get("persona_mode") or body.get("force_single_persona"),
        # V3 interactive discussion control — debate_node 가 auto_start=False 시 gate 대기.
        "_discussion_auto_start": auto_start,
        "_discussion_gate_factory": _get_or_create_discussion_gate,
        "stt_metadata": body.get(
            "stt_metadata",
            {
                "transcription_confidence": 0.95,
                "speaker_diarization_success": True,
                "duration_sec": 60.0,
                "has_timestamps": False,
                "masking_format": {"version": "v1_symbolic"},
            },
        ),
        "plan": body.get("plan") or {},
    }


def _adapt_report_for_frontend(
    report: dict[str, Any] | None, debates: dict[int, dict[str, Any]] | None = None
) -> dict[str, Any] | None:
    """V2 report 구조를 chatbot-ui(qa_pipeline_reactflow.html) 가 기대하는 평탄 구조로 확장.

    프론트가 참조하는 필드 (V1 호환):
      - report.item_scores: [{item_number, item_name, score, max_score, evaluation_mode, deductions, evidence, ...}]
      - report.strengths / improvements / coaching_points (top-level)
      - report.deductions (top-level 집계)
    V2 원본 위치:
      - report.evaluation.categories[].items[] → item_scores 로 평탄화
      - report.summary.strengths / improvements → top-level 로 승격
      - 각 item.deductions → deductions 집계

    ``debates`` 가 주어지면 item_scores / category items 의 score 를 debate 최종값으로 교체.
    """
    if not isinstance(report, dict):
        return report

    evaluation = report.get("evaluation") or {}
    categories = evaluation.get("categories") or []

    def _patch_with_debate(it: dict[str, Any]) -> dict[str, Any]:
        if not debates:
            return it
        ino = it.get("item_number")
        if not isinstance(ino, int) or ino not in debates:
            return it
        rec = debates.get(ino) or {}
        final_score = rec.get("final_score")
        if final_score is None:
            return it
        patched = dict(it)
        patched["score"] = int(final_score)
        patched["persona_merge_rule"] = f"debate_{rec.get('merge_rule') or 'debate'}"
        patched["debate_applied"] = True
        return patched

    # categories[].items[] 도 패치 (UI 가 그대로 평탄화 참조)
    if debates:
        for cat in categories:
            cat["items"] = [_patch_with_debate(it) for it in (cat.get("items") or [])]

    if not report.get("item_scores"):
        flat_items: list[dict[str, Any]] = []
        for cat in categories:
            for it in cat.get("items") or []:
                # 프론트는 item_name 이 비어있을 때 ITEM_NAMES[num] 로 대체하므로 그대로 전달
                flat_items.append(it)
        if flat_items:
            report["item_scores"] = flat_items
    elif debates:
        report["item_scores"] = [_patch_with_debate(it) for it in report["item_scores"]]

    if not report.get("deductions"):
        all_deds: list[dict[str, Any]] = []
        for cat in categories:
            for it in cat.get("items") or []:
                for d in it.get("deductions") or []:
                    d2 = dict(d)
                    d2.setdefault("item_number", it.get("item_number"))
                    all_deds.append(d2)
        if all_deds:
            report["deductions"] = all_deds

    summary = report.get("summary") or {}
    if not report.get("strengths") and summary.get("strengths"):
        report["strengths"] = summary["strengths"]
    if not report.get("improvements") and summary.get("improvements"):
        report["improvements"] = summary["improvements"]
    if not report.get("coaching_points") and summary.get("coaching_points"):
        report["coaching_points"] = summary["coaching_points"]

    return report


def _apply_persona_mode_override(initial: dict[str, Any]) -> str | None:
    """initial state 의 `_persona_mode_override` 를 reconciler_personas 런타임에 반영.

    값은 state 에서 pop 되며 set_runtime_force_single() 이 호출된다.
    - "single" / True / "1"/"true"/"yes" → force single
    - "ensemble" / False / "0"/"false"/"no" → force ensemble (env var 무시)
    - None / 기타 → env var fallback

    반환: 적용된 모드 문자열 ("single" | "ensemble" | None).
    `/evaluate` (JSON) 과 `/evaluate/stream` (SSE) 양쪽에서 호출.
    """
    override = initial.pop("_persona_mode_override", None)
    try:
        from v2.reconciler_personas import force_single_persona, set_runtime_force_single

        if override == "single" or override is True or str(override).lower() in ("1", "true", "yes"):
            set_runtime_force_single(True)
            logger.info("persona mode = SINGLE (런타임 강제)")
            return "single"
        if override == "ensemble" or override is False or str(override).lower() in ("0", "false", "no"):
            set_runtime_force_single(False)  # env var 무시, ensemble 강제
            logger.info("persona mode = ENSEMBLE (런타임 강제, env var 무시)")
            return "ensemble"
        set_runtime_force_single(None)  # env var fallback
        logger.info("persona mode override 미지정 → env fallback (현재 SINGLE=%s)", force_single_persona())
        return None
    except Exception as _e:
        logger.warning("persona mode override 적용 실패: %s", _e)
        return None


def _apply_debate_overrides(
    evaluations: list[dict[str, Any]] | None, debates: dict[int, dict[str, Any]] | None
) -> list[dict[str, Any]] | None:
    """토론이 돌아간 항목의 score / merge_rule 을 debate 결과로 교체.

    ``evaluations`` 필드는 LangGraph ``operator.add`` 리듀서라 graph 내부에서 override 하면
    append 되어 중복된다. 대신 server 응답 직전에 ``debates[item_no].final_score`` 를
    기준으로 덮어써 UI 가 debate 결과를 "최종" 으로 본다.
    """
    if not evaluations or not debates:
        return evaluations
    out: list[dict[str, Any]] = []
    for ev in evaluations:
        if not isinstance(ev, dict):
            out.append(ev)
            continue
        inner = ev.get("evaluation") if isinstance(ev.get("evaluation"), dict) else None
        meta = inner if inner is not None else ev
        item_no = meta.get("item_number") if isinstance(meta, dict) else None
        if not isinstance(item_no, int) or item_no not in debates:
            out.append(ev)
            continue
        rec = debates.get(item_no) or {}
        final_score = rec.get("final_score")
        if final_score is None:
            out.append(ev)
            continue
        merge_rule = rec.get("merge_rule") or "debate"
        final_rationale = rec.get("final_rationale") or ""
        # 2026-04-27 개정: judge_* 필드는 rec 에서 직접 매핑. 판사 호출 성공/실패 여부에 따라
        # _invoke_post_debate_judge 가 알아서 채움 (성공: 값, 실패: None + judge_failure_reason).
        judge_score_raw = rec.get("judge_score")
        judge_reasoning = rec.get("judge_reasoning")
        judge_failure_reason = rec.get("judge_failure_reason")
        judge_deductions = rec.get("judge_deductions") or []
        judge_evidence = rec.get("judge_evidence") or []
        judge_human_cases = rec.get("judge_human_cases") or []
        # 점수가 final_score 로 갱신되면 deductions / evidence 도 정합화해야 한다.
        # 이전에는 score 만 덮어쓰고 deductions 는 페르소나 1차 + Layer3 OVERRIDE 누적분을
        # 그대로 남겨서, 판사가 만점(=감점 0) 으로 올린 케이스에 빈 reason -3 + OVERRIDE -10
        # 같은 잔재가 화면에 표시되는 모순 발생.
        # 정책:
        #   1) 판사 성공 (judge_score != None) → judge_deductions / judge_evidence 가 정답.
        #      메인 deductions / evidence 를 판사 것으로 교체.
        #   2) 판사 실패 (median fallback) → deductions 보존하되 정합성 검증:
        #      final_score == max_score 면 deductions = [] (감점 없음 = 만점 모순 방지).
        judge_succeeded = isinstance(judge_score_raw, (int, float))
        max_score_val: int | None = None
        try:
            from v2.agents.group_b.base import ITEM_MAX_SCORE  # noqa: WPS433

            max_score_val = ITEM_MAX_SCORE.get(int(item_no)) if isinstance(item_no, int) else None
        except Exception:  # pragma: no cover — base import 실패해도 fallback 으로 진행
            max_score_val = None

        def _normalized_deds(orig: list, ev_lines: list) -> tuple[list, list]:
            """판사/만점 정책에 따른 deductions / evidence 정렬."""
            if judge_succeeded:
                return list(judge_deductions), list(judge_evidence)
            if max_score_val is not None and int(final_score) >= int(max_score_val):
                # 만점 = 감점 없음 — 잔재 deductions 제거 (override / 빈 reason 등)
                return [], list(ev_lines)
            return list(orig), list(ev_lines)

        # 판사 성공 시 judgment / summary 도 판사 reasoning 으로 교체.
        # 이전에는 페르소나 1차 판정 텍스트 (예: "7점 부여") 가 그대로 남아 score=10 인데
        # judgment 는 7점 설명이라 사용자가 모순 지적. 판사가 HITL 사례로 점수를 바꿨으면
        # 그 reasoning (HITL 기반 판단) 이 화면에 노출되어야 함.
        def _judgment_text() -> str | None:
            if judge_succeeded and judge_reasoning:
                # HITL 채택 케이스를 사용자가 식별할 수 있게 헤더 prepend
                hitl_count = len(judge_human_cases or [])
                if hitl_count:
                    return f"🎭 [판사 판정 · HITL {hitl_count}건 참조] {judge_reasoning}"
                return f"🎭 [판사 판정] {judge_reasoning}"
            return None  # None = 기존 judgment 유지

        judge_text = _judgment_text()

        new_ev = dict(ev)
        if inner is not None:
            new_inner = dict(inner)
            orig_deds = new_inner.get("deductions") or []
            orig_ev = new_inner.get("evidence") or []
            cleaned_deds, cleaned_ev = _normalized_deds(orig_deds, orig_ev)
            new_inner["score"] = int(final_score)
            new_inner["deductions"] = cleaned_deds
            new_inner["evidence"] = cleaned_ev
            if judge_text is not None:
                new_inner["judgment"] = judge_text
                new_inner["summary"] = judge_text
            new_inner["persona_merge_rule"] = f"debate_{merge_rule}"
            new_inner["debate_applied"] = True
            new_inner["debate_rationale"] = final_rationale
            new_inner["judge_score"] = int(judge_score_raw) if judge_succeeded else None
            new_inner["judge_reasoning"] = judge_reasoning
            new_inner["judge_failure_reason"] = judge_failure_reason
            new_inner["judge_deductions"] = judge_deductions
            new_inner["judge_evidence"] = judge_evidence
            new_inner["judge_human_cases"] = judge_human_cases
            new_ev["evaluation"] = new_inner
        else:
            orig_deds = new_ev.get("deductions") or []
            orig_ev = new_ev.get("evidence") or []
            cleaned_deds, cleaned_ev = _normalized_deds(orig_deds, orig_ev)
            new_ev["score"] = int(final_score)
            new_ev["deductions"] = cleaned_deds
            new_ev["evidence"] = cleaned_ev
            if judge_text is not None:
                new_ev["judgment"] = judge_text
                new_ev["summary"] = judge_text
            new_ev["persona_merge_rule"] = f"debate_{merge_rule}"
            new_ev["debate_applied"] = True
            new_ev["debate_rationale"] = final_rationale
            new_ev["judge_score"] = int(judge_score_raw) if judge_succeeded else None
            new_ev["judge_reasoning"] = judge_reasoning
            new_ev["judge_failure_reason"] = judge_failure_reason
            new_ev["judge_deductions"] = judge_deductions
            new_ev["judge_evidence"] = judge_evidence
            new_ev["judge_human_cases"] = judge_human_cases
        out.append(new_ev)
    return out


def _extract_response(final_state: dict[str, Any]) -> dict[str, Any]:
    """응답 payload — V2 orchestrator / preprocessing / evaluations / report / gt_comparison."""
    debates = final_state.get("debates") or {}
    evaluations = _apply_debate_overrides(final_state.get("evaluations"), debates)
    return {
        "preprocessing": final_state.get("preprocessing"),
        "evaluations": evaluations,
        "orchestrator": final_state.get("orchestrator"),
        "debates": debates,
        "report": _adapt_report_for_frontend(final_state.get("report"), debates),
        "routing": final_state.get("routing"),
        "error": final_state.get("error"),
        "completed_nodes": final_state.get("completed_nodes"),
        "node_timings": final_state.get("node_timings"),
        "gt_comparison": final_state.get("gt_comparison"),
        "gt_evidence_comparison": final_state.get("gt_evidence_comparison"),
    }


@app.post("/evaluate")
async def evaluate(request: Request) -> JSONResponse:
    """V2 파이프라인 평가 — JSON 입력 → JSON 결과.

    Request body (JSON):
      {
        "transcript": "상담사: ...",
        "stt_metadata": {...},      // 선택
        "tenant_id": "generic",      // 선택
        "session_id": "...",         // 선택
        "plan": {"skip_phase_c_and_reporting": false}  // 선택
      }
    """
    graph = _get_graph()
    if graph is None:
        return JSONResponse({"error": "graph_build_failed", "detail": _graph_build_error}, status_code=503)

    body = await request.json()
    try:
        initial = _build_initial_state(body)
    except ValueError as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    # 프론트 토글 반영 — body.persona_mode ("single" | "ensemble") 를 런타임 주입.
    # (SSE stream 과 동일한 로직을 공통 헬퍼로 호출)
    _persona_applied = _apply_persona_mode_override(initial)

    # runtime_flags current_sample_id — 로깅/디버깅 용 (self-exclusion 은 제거됨)
    try:
        from v2.runtime_flags import set_current_sample_id

        _sid = body.get("gt_sample_id") or body.get("sample_id") or ""
        set_current_sample_id(_sid if _sid else None)
    except Exception:  # noqa: BLE001
        pass

    t0 = time.time()
    try:
        final_state = await graph.ainvoke(initial)
    except Exception as e:
        logger.exception("evaluate 실패")
        return JSONResponse({"error": "graph_invoke_failed", "detail": str(e)}, status_code=500)

    elapsed = round(time.time() - t0, 2)
    response = _extract_response(final_state)
    response["_meta"] = {
        "pipeline": "v2",
        "elapsed_sec": elapsed,
        "session_id": initial["session_id"],
        "persona_mode": _persona_applied,
    }
    return JSONResponse(response)


@app.post("/invocations")
async def invocations(request: Request) -> JSONResponse:
    """AgentCore Runtime invoke entrypoint — /evaluate 로 위임."""
    return await evaluate(request)


# ---------------------------------------------------------------------------
# /evaluate/stream — SSE 호환 엔드포인트 (실시간 progress)
# ---------------------------------------------------------------------------
# graph.astream(stream_mode="updates") 로 노드 완료 시점마다 이벤트 emit.
# 프론트엔드 (qa_pipeline_reactflow.html) 가 기대하는 이벤트:
#   - routing : {phase, phase_label, next_node, next_label}
#   - status  : {node, label, status, elapsed, scores}
#   - result  : 최종 final_state (evaluations + report + orchestrator + ...)
#   - done    : {elapsed_seconds}
#   - error   : {message, type}


_LAYER1_NODES = {"layer1"}
_LAYER2_SUB_AGENTS = (
    "greeting",
    "listening_comm",
    "language",
    "needs",
    "explanation",
    "proactiveness",
    "work_accuracy",
    "privacy",
)
# 신한 부서특화 dept 노드 — base 8 외 동적 fan-out 대상 (Layer 2 phase 분류에 포함)
try:
    from v2.agents.shinhan_dept.registry import DEPT_NODE_REGISTRY as _DEPT_REG
    _DEPT_SUB_AGENTS: tuple[str, ...] = tuple(_DEPT_REG.keys())
except Exception:
    _DEPT_SUB_AGENTS = ()
_LAYER2_BARRIER = {"layer2_barrier"}
_LAYER3_NODES = {"layer3"}
_LAYER4_NODES = {"layer4"}

# 프론트엔드 (chatbot-ui/qa_pipeline_reactflow.html) NODE_DEFS 에 맞춰 layer3/layer4 를 재매핑.
# 서버 그래프는 단일 "layer3" / "layer4" 노드지만 프론트는 세분화 (orchestrator_v2,
# confidence·tier_router·evidence_refiner·report_generator) 로 표시.
_BACKEND_TO_FRONTEND_NODES: dict[str, tuple[str, ...]] = {
    "layer3": ("orchestrator_v2",),
    "layer4": ("confidence", "tier_router", "evidence_refiner", "report_generator"),
}


def _frontend_node_names(backend_node: str) -> tuple[str, ...]:
    """백엔드 노드명을 프론트 표시용 노드명 (1개 이상) 으로 변환."""
    return _BACKEND_TO_FRONTEND_NODES.get(backend_node, (backend_node,))


def _node_phase(node: str) -> str:
    if node in _LAYER1_NODES:
        return "layer1"
    if node in _LAYER2_SUB_AGENTS or node in _DEPT_SUB_AGENTS or node in _LAYER2_BARRIER:
        return "layer2"
    if node in _LAYER3_NODES:
        return "layer3"
    if node in _LAYER4_NODES:
        return "layer4"
    return "other"


def _merge_update(accum: dict[str, Any], delta: dict[str, Any]) -> None:
    """graph.astream(updates) 델타를 누적 state 에 병합. evaluations 는 append."""
    for key, value in (delta or {}).items():
        if key == "evaluations" and isinstance(value, list):
            accum.setdefault("evaluations", []).extend(value)
        else:
            accum[key] = value


_SUB_AGENT_ITEM_MAP: dict[str, list[int]] = {
    "greeting": [1, 2],
    "listening_comm": [3, 4, 5],
    "language": [6, 7],
    "needs": [8, 9],
    "explanation": [10, 11],
    "proactiveness": [12, 13, 14],
    "work_accuracy": [15, 16],
    "privacy": [17, 18],
}


# V2 sub agent 이름 ↔ V1 agent_turn_assignments 키 매핑.
# Layer 1 의 agent_turn_assignments 는 V1 호환을 위해 V1 이름으로 키를 부여.
# Trace input 표시용으로만 사용 (sub agent 실제 동작에는 영향 없음).
_V2_TO_V1_AGENT_KEY: dict[str, str] = {
    "greeting": "greeting",
    "listening_comm": "understanding",
    "language": "courtesy",
    "needs": "mandatory",
    "explanation": "scope",
    "proactiveness": "proactiveness",
    "work_accuracy": "work_accuracy",
    "privacy": "incorrect_check",
}


def _preview_text(text: str | None, limit: int = 400) -> str | None:
    if not text:
        return text
    s = str(text)
    if len(s) <= limit:
        return s
    return s[:limit] + f"... (+{len(s) - limit} chars)"


def _sanitize_trace_output(result: Any, depth: int = 0) -> Any:
    """LangGraph delta 를 SSE payload 로 변환.

    기존 depth=5 제한이 evaluations[...].evidence[...].quote 까지 도달 (6~7 depth) 하여
    evidence quote 가 전부 "…" 로 치환되던 문제 수정 — depth 를 10 으로 확장.
    문자열 truncation 은 유지 (2000자 컷).
    """
    if depth > 10:
        return "…"
    if isinstance(result, str):
        return result[:2000] + ("…" if len(result) > 2000 else "")
    if isinstance(result, list):
        return [_sanitize_trace_output(x, depth + 1) for x in result]
    if isinstance(result, dict):
        return {k: _sanitize_trace_output(v, depth + 1) for k, v in result.items()}
    return result


def _build_trace_input(node: str, initial: dict[str, Any], accum: dict[str, Any]) -> dict[str, Any]:
    """노드별 실제 입력 컨텍스트. Trace 탭 Input 에 노출."""
    base = {"session_id": initial.get("session_id"), "tenant_id": initial.get("tenant_id"), "backend_node": node}

    if node == "layer1":
        return {
            **base,
            "transcript_preview": _preview_text(initial.get("transcript"), 600),
            "transcript_len": len(initial.get("transcript", "")),
            "stt_metadata": initial.get("stt_metadata"),
            "consultation_type": initial.get("consultation_type"),
            "llm_backend": initial.get("llm_backend"),
        }

    pre = accum.get("preprocessing") or {}

    if node in _SUB_AGENT_ITEM_MAP:
        items = _SUB_AGENT_ITEM_MAP[node]
        # agent_turn_assignments 는 V1 키로 저장됨 — V2 sub agent 명을 V1 키로 변환 후 조회
        assignment_key = _V2_TO_V1_AGENT_KEY.get(node, node)
        assignments = (pre.get("agent_turn_assignments") or {}).get(assignment_key) or {}
        rule_pv = pre.get("rule_pre_verdicts") or {}
        item_rule_pv = {f"item_{n:02d}": rule_pv.get(f"item_{n:02d}") for n in items if f"item_{n:02d}" in rule_pv}
        turns = assignments.get("turns") or []
        turns_preview = [
            {
                "turn_id": t.get("turn_id"),
                "speaker": t.get("speaker", ""),
                "segment": t.get("segment", ""),
                "text": (t.get("text", "") or "")[:120],
            }
            for t in turns[:20]
        ]
        return {
            **base,
            "sub_agent": node,
            "target_items": items,
            "assigned_turns_count": len(turns),
            "assigned_turn_ids": assignments.get("turn_ids"),
            "assigned_text_preview": _preview_text(assignments.get("text"), 800),
            "turns": turns_preview,
            "rule_pre_verdicts": item_rule_pv,
            "intent_type": pre.get("intent_type"),
            "deduction_triggers": pre.get("deduction_triggers"),
            "quality": pre.get("quality"),
        }

    if node == "layer2_barrier":
        evals = accum.get("evaluations") or []
        return {
            **base,
            "evaluations_collected": len(evals),
            "item_numbers_collected": sorted(
                {
                    (ev.get("evaluation") or {}).get("item_number")
                    for ev in evals
                    if isinstance(ev, dict) and (ev.get("evaluation") or {}).get("item_number") is not None
                }
            ),
        }

    if node == "layer3":
        evals = accum.get("evaluations") or []
        return {
            **base,
            "evaluations_count": len(evals),
            "intent_type": pre.get("intent_type"),
            "deduction_triggers": pre.get("deduction_triggers"),
            "has_all_zero_trigger": pre.get("has_all_zero_trigger"),
            "recommended_override": pre.get("recommended_override"),
        }

    if node == "layer4":
        orch = accum.get("orchestrator") or {}
        return {
            **base,
            "raw_total": (orch.get("final_score") or {}).get("raw_total"),
            "grade": (orch.get("final_score") or {}).get("grade"),
            "routing_tier_hint": orch.get("routing_tier_hint"),
            "consistency_flags_count": len(orch.get("consistency_flags") or []),
            "overrides_applied": (orch.get("overrides") or {}).get("applied"),
            "skip_phase_c_and_reporting": (accum.get("plan") or {}).get("skip_phase_c_and_reporting"),
        }

    return base


def _node_item_scores(node: str, delta: dict[str, Any]) -> list[dict[str, Any]]:  # noqa: ARG001  # node 는 호출부 트레이스 호환성 유지 위해 시그니처에 보존
    """Sub Agent 완료 시 streamingItems 에 반영할 item_number/score 리스트 추출.

    rag_evidence + persona 앙상블 메타 (persona_votes, merge_path, judge_reasoning,
    persona_details) 도 함께 반환 — 프론트가 노드 완료 시점에 실시간으로 각 평가자별
    판정과 최종 결과를 표시할 수 있게 한다.
    """
    out: list[dict[str, Any]] = []
    for ev in delta.get("evaluations") or []:
        e = ev.get("evaluation") if isinstance(ev, dict) else None
        if not isinstance(e, dict):
            continue
        # Group A 는 confidence.final, Group B 는 llm_self_confidence.score 로 다른 필드 사용
        confidence = e.get("confidence") or e.get("llm_self_confidence")
        out.append(
            {
                "item_number": e.get("item_number"),
                "item_name": e.get("item_name"),
                "score": e.get("score"),
                "max_score": e.get("max_score"),
                "evaluation_mode": e.get("evaluation_mode"),
                "confidence": confidence,
                "rag_evidence": e.get("rag_evidence"),
                # 최종 결과 표시용 (neutral 대표 / 단일 판정의 판단·감점·인용)
                "judgment": e.get("judgment"),
                "summary": e.get("summary"),
                "deductions": e.get("deductions") or [],
                "evidence": e.get("evidence") or [],
                # 3-Persona 앙상블 메타 (Phase 5, 2026-04-21)
                "persona_votes": e.get("persona_votes"),
                "persona_step_spread": e.get("persona_step_spread"),
                "persona_merge_path": e.get("persona_merge_path"),
                "persona_merge_rule": e.get("persona_merge_rule"),
                "judge_reasoning": e.get("judge_reasoning"),
                "persona_details": e.get("persona_details"),
                "mandatory_human_review": e.get("mandatory_human_review"),
            }
        )
    return out


async def _build_stream_events(body: dict[str, Any], graph: Any):
    """graph.astream(updates) 기반 실시간 SSE 이벤트.

    body / graph 는 엔드포인트에서 선처리 후 전달 — 제너레이터 내부에서 `await request.json()`
    을 호출하면 첫 `yield` 이후 본문 read 가 블록되어 SSE 스트림이 정지하는 문제가 있음.

    V3 interactive discussion:
      - 사전 바인드된 asyncio.Queue 를 통해 debate 노드 내부 (sync, thread pool) 의 실시간
        discussion_* 이벤트를 SSE 스트림으로 즉시 전파. auto_start=False 경로에서 gate
        대기 중에도 discussion_started 이벤트가 즉시 프론트에 도달.
    """
    import asyncio

    def sse(event: str, data: dict[str, Any]) -> str:
        return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

    # SSE keepalive — 아무 이벤트도 안 들어올 때 주기적으로 comment line 을 보내
    # idle-disconnect 를 방지한다. AG2 debate 는 persona 발언 1회당 수 초~수십 초
    # Bedrock 호출이라 그 사이 구간이 조용하면 프록시/브라우저가 커넥션을 끊음.
    def sse_keepalive() -> str:
        # SSE 주석 라인 (client 는 무시, 커넥션만 유지).
        return ": keepalive\n\n"

    # 3단계 멀티테넌트 해석 (start 이벤트/로그용). _build_initial_state 와 동일 fallback.
    _start_site = body.get("site_id") or body.get("tenant_id") or "generic"
    _start_channel = body.get("channel") or "inbound"
    _start_department = body.get("department") or "default"
    _start_tenant_key = f"{_start_site}:{_start_channel}:{_start_department}"

    logger.info("=" * 78)
    logger.info("🎬 SSE STREAM START — /evaluate/stream")
    logger.info(
        "  · tenant_key=%s session=%s sample=%s transcript_chars=%d",
        _start_tenant_key,
        body.get("session_id") or "?",
        body.get("gt_sample_id") or body.get("sample_id") or "?",
        len(str(body.get("transcript") or "")),
    )
    logger.info(
        "  · llm_backend=%s bedrock_model=%s persona_mode=%s",
        body.get("llm_backend"),
        body.get("bedrock_model_id"),
        body.get("persona_mode"),
    )
    logger.info("=" * 78)
    # tenant 3단계 필드를 start 이벤트에 포함 — 프론트가 tenant 노드에 표시.
    # tenant_id 필드는 레거시 호환 (site_id 와 동일 값).
    yield sse(
        "start",
        {
            "pipeline": "v2",
            "site_id": _start_site,
            "channel": _start_channel,
            "department": _start_department,
            "tenant_key": _start_tenant_key,
            "tenant_id": _start_site,
        },
    )

    try:
        initial = _build_initial_state(body)
    except ValueError as e:
        logger.warning("SSE stream: bad_request — %s", e)
        yield sse("error", {"error": "bad_request", "detail": str(e), "message": str(e)})
        return

    # 프론트 토글 반영 — body.persona_mode ("single" | "ensemble") 를 런타임 주입.
    # (JSON /evaluate 와 동일한 공통 헬퍼 호출)
    _apply_persona_mode_override(initial)

    # ── V3 interactive discussion 실시간 이벤트 큐 ────────────────────────────
    # debate_node 는 sync 노드이고 LangGraph 가 threadpool 에서 실행하므로,
    # 실시간 SSE 푸시를 위해 asyncio.Queue + call_soon_threadsafe 로 브릿지.
    _rt_loop = asyncio.get_running_loop()
    _rt_event_queue: "asyncio.Queue[tuple[str, dict[str, Any]] | None]" = asyncio.Queue()
    _allowed_realtime_events = {
        # legacy
        "debate_round_start",
        "persona_turn",
        "moderator_verdict",
        "debate_final",
        # V3 interactive
        "discussion_started",
        "persona_speaking",
        "persona_message",
        "vote_cast",
        "discussion_round_complete",
        "discussion_finalized",
        # 서버 로그 스트리밍 — 백엔드 콘솔에 찍히는 Bedrock 호출/LLM req·res 등
        "log",
    }

    # 스트림 종료 시 후속 callback 을 조용히 폐기하기 위한 플래그 (list 래퍼 — closure mutation)
    _rt_shutdown: list[bool] = [False]

    def _debate_realtime_callback(name: str, payload: dict[str, Any]) -> None:
        """debate_node → run_debate 경로에서 호출되는 sync 콜백. 실시간 큐에 푸시.

        주의: 클라이언트가 disconnect 하거나 스트림이 끝난 뒤에도 background 토론
        스레드는 계속 돌고 있을 수 있다. 이때 `_rt_loop.call_soon_threadsafe` 가
        닫힌 loop 에 접근하면 RuntimeError('Event loop is closed') 발생 — 매 이벤트
        마다 traceback 이 로그를 도배. `_rt_shutdown` 플래그 + `is_closed()` 가드로
        방어.
        """
        if _rt_shutdown[0]:
            return
        if name not in _allowed_realtime_events:
            logger.debug("SSE drop (not in allowlist): %s", name)
            return
        try:
            if _rt_loop.is_closed():
                return
            _rt_loop.call_soon_threadsafe(_rt_event_queue.put_nowait, (name, payload))
            logger.info(
                "📡 SSE queue ← %s · item=%s persona=%s round=%s",
                name,
                payload.get("item_number"),
                payload.get("persona_id"),
                payload.get("round"),
            )
        except RuntimeError as exc:
            # loop closed 류 — 한 번만 경고 찍고 이후 조용히 폐기
            msg = str(exc)
            if "Event loop is closed" in msg or "loop is closed" in msg.lower():
                # 한 번만 경고, 이후는 shutdown 플래그로 걸러짐
                pass
            else:
                logger.warning("debate realtime callback RuntimeError (%s): %s", name, msg)
        except Exception:  # pragma: no cover — 큐 실패가 토론 전체 중단시키면 안 됨
            logger.exception("debate realtime callback failed: %s", name)

    initial["_debate_on_event"] = _debate_realtime_callback

    # ── 서버 로그 → 프론트 실행 로그 스트리밍 ───────────────────────────────────
    # 백엔드 콘솔(logging)에 찍히는 메시지를 SSE `log` 이벤트로 프론트에 푸시한다.
    # 대상 로거: v2.* / nodes.* / bedrock / graph / __main__ (root 아님 — 제3자 라이브러리 제외)
    # 필터: INFO+ · 일부 장황 로거(watchfiles/asyncio/anyio/httpcore)는 WARNING 이하 컷.
    _LOG_STREAM_LOGGERS = ("v2", "nodes", "graph", "bedrock", "__main__")
    _LOG_STREAM_BLOCK_MODULES = {
        "watchfiles",
        "httpcore",
        "httpx",
        "urllib3",
        "boto3",
        "botocore",
        "asyncio",
        "anyio",
        "uvicorn.access",
    }
    _log_seen = [0]  # 전송한 레코드 수 — 과도한 폭주 방어용

    class _SSELogHandler(logging.Handler):
        """request-scoped 로그 핸들러 — 동일 _rt_event_queue 에 `log` 이벤트로 emit."""

        def emit(self, record: logging.LogRecord) -> None:
            if _rt_shutdown[0]:
                return
            # 블랙리스트 로거는 컷
            for blocked in _LOG_STREAM_BLOCK_MODULES:
                if record.name.startswith(blocked):
                    return
            try:
                msg = record.getMessage()
            except Exception:
                return
            # 너무 긴 메시지는 잘라냄 (프론트 부담 줄이기)
            if len(msg) > 2000:
                msg = msg[:2000] + "…(truncated)"
            payload = {
                "level": record.levelname,
                "logger": record.name,
                "message": msg,
                "ts": record.created,  # epoch float
            }
            try:
                if _rt_loop.is_closed():
                    return
                _rt_loop.call_soon_threadsafe(_rt_event_queue.put_nowait, ("log", payload))
                _log_seen[0] += 1
            except RuntimeError:
                # loop closed 시 조용히 폐기
                pass
            except Exception:
                pass  # 로그 스트리밍 자체가 장애 내면 안 됨

    _sse_log_handler = _SSELogHandler(level=logging.INFO)
    _sse_log_fmt = logging.Formatter("%(message)s")
    _sse_log_handler.setFormatter(_sse_log_fmt)
    _log_attached_loggers: list[logging.Logger] = []
    for _lg_name in _LOG_STREAM_LOGGERS:
        _lg = logging.getLogger(_lg_name)
        _lg.addHandler(_sse_log_handler)
        _log_attached_loggers.append(_lg)
    logger.info("📡 SSE log streaming 활성 — loggers=%s", _LOG_STREAM_LOGGERS)

    # runtime_flags 의 current_sample_id 초기화 — self-exclusion 은 제거됐으나
    # 다른 디버깅/로깅 용도로 세팅 유지 (향후 옵션 재활성화 대비).
    try:
        from v2.runtime_flags import set_current_sample_id

        _sid = body.get("gt_sample_id") or body.get("sample_id") or ""
        set_current_sample_id(_sid if _sid else None)
    except Exception:  # noqa: BLE001
        pass

    logger.info(
        "SSE stream: initial state built — session=%s tenant=%s transcript_len=%d",
        initial.get("session_id"),
        initial.get("tenant_id"),
        len(initial.get("transcript", "")),
    )

    t0 = time.time()
    last_phase: str | None = None
    node_start: dict[str, float] = {}
    accum: dict[str, Any] = dict(initial)

    phase_labels = {
        "layer1": "Layer 1 · Preprocessing",
        "layer2": "Layer 2 · Sub Agents (병렬)",
        "layer3": "Layer 3 · Orchestrator",
        "layer4": "Layer 4 · Post-processing",
    }

    # Layer 1 진입 라우팅 이벤트 (프론트가 layer1 active 로 표시)
    yield sse(
        "routing",
        {
            "phase": "layer1",
            "phase_label": phase_labels["layer1"],
            "next_node": "layer1",
            "next_label": phase_labels["layer1"],
        },
    )
    yield sse("status", {"node": "layer1", "status": "started"})
    node_start["layer1"] = time.time()
    last_phase = "layer1"

    # 이미 started 를 emit 한 노드 (중복 방지)
    started_emitted: set[str] = {"layer1"}

    def _emit_next_phase_preflight(completed_node: str):
        """완료 직후 다음 페이즈 진입을 선제적으로 알림.
        프론트가 Sub Agent 완료를 기다리지 않고 즉시 active 상태로 표시하기 위함.
        `last_phase` 를 다음 페이즈로 업데이트해 중복 routing 방지."""
        nonlocal last_phase
        events: list[str] = []

        if completed_node == "layer1":
            preprocessing = accum.get("preprocessing") or {}
            if preprocessing.get("quality", {}).get("unevaluable"):
                # unevaluable short-circuit → Layer 4 직결 (Layer 2/3 skip)
                events.append(
                    sse(
                        "routing",
                        {
                            "phase": "layer4",
                            "phase_label": phase_labels["layer4"] + " (unevaluable short-circuit)",
                            "next_node": "layer4",
                            "next_label": "Layer 4 · Short-circuit",
                        },
                    )
                )
                if "layer4" not in started_emitted:
                    events.append(sse("status", {"node": "layer4", "status": "started"}))
                    node_start["layer4"] = time.time()
                    started_emitted.add("layer4")
                last_phase = "layer4"
            else:
                # Layer 2 fan-out → 활성 Sub Agent 들 동시 started 로 표시.
                # 신한 부서일 때 work_accuracy 빠지고 dept 노드 (coll_accuracy 등) 추가됨.
                # graph_v2 의 _resolve_active_sub_agents 와 동일한 룰.
                site_id = accum.get("site_id") or accum.get("tenant_id") or "generic"
                dept = accum.get("department") or accum.get("team_id")
                active_subs: list[str] = list(_LAYER2_SUB_AGENTS)
                if site_id == "shinhan" and dept and _DEPT_SUB_AGENTS:
                    try:
                        from v2.agents.shinhan_dept.registry import get_dept_nodes_for_tenant
                        dept_nodes = get_dept_nodes_for_tenant(site_id, dept)
                    except Exception:
                        dept_nodes = []
                    if dept_nodes:
                        # work_accuracy 제외 + dept 노드 추가
                        active_subs = [n for n in active_subs if n != "work_accuracy"] + list(dept_nodes)

                events.append(
                    sse(
                        "routing",
                        {
                            "phase": "layer2",
                            "phase_label": phase_labels["layer2"],
                            "next_node": "__parallel__",
                            "next_label": f"Layer 2 · Fan-out ({len(active_subs)} Sub Agents)",
                        },
                    )
                )
                for sub_agent in active_subs:
                    if sub_agent not in started_emitted:
                        events.append(sse("status", {"node": sub_agent, "status": "started"}))
                        node_start[sub_agent] = time.time()
                        started_emitted.add(sub_agent)
                last_phase = "layer2"

        elif completed_node == "layer2_barrier":
            # Layer 2 barrier 완료 → Layer 3 진입 (프론트: orchestrator_v2)
            events.append(
                sse(
                    "routing",
                    {
                        "phase": "layer3",
                        "phase_label": phase_labels["layer3"],
                        "next_node": "orchestrator_v2",
                        "next_label": "Layer 3 · Orchestrator V2",
                    },
                )
            )
            for fe_name in _frontend_node_names("layer3"):
                if fe_name not in started_emitted:
                    events.append(sse("status", {"node": fe_name, "status": "started"}))
                    node_start[fe_name] = time.time()
                    started_emitted.add(fe_name)
            last_phase = "layer3"

        elif completed_node == "layer3":
            plan = accum.get("plan") or {}
            if not plan.get("skip_phase_c_and_reporting"):
                events.append(
                    sse(
                        "routing",
                        {
                            "phase": "layer4",
                            "phase_label": phase_labels["layer4"],
                            "next_node": "report_generator",
                            "next_label": "Layer 4 · Post-processing",
                        },
                    )
                )
                # Layer 4 는 프론트에서 4개 노드 (confidence/tier_router/evidence_refiner/report_generator)
                # 로 표시 — 모두 started 로 선제 마킹.
                for fe_name in _frontend_node_names("layer4"):
                    if fe_name not in started_emitted:
                        events.append(sse("status", {"node": fe_name, "status": "started"}))
                        node_start[fe_name] = time.time()
                        started_emitted.add(fe_name)
                last_phase = "layer4"

        return events

    # realtime 큐에서 이미 emit 한 이벤트 키 — buffer 재방출 시 중복 방지.
    # key = (event_name, discussion_id, item_number, round, persona_id) — 가능한 한 의미론적 식별.
    _rt_emitted_keys: set[tuple[str, str, int, int, str]] = set()

    def _event_dedup_key(name: str, data: dict[str, Any]) -> tuple[str, str, int, int, str]:
        try:
            d_id = str(data.get("discussion_id") or "")
            i_no = int(data.get("item_number") or 0)
            rnd = int(data.get("round") or 0)
            p_id = str(data.get("persona_id") or data.get("persona") or "")
            return (name, d_id, i_no, rnd, p_id)
        except Exception:
            return (name, "", 0, 0, "")

    def _drain_realtime_queue_nonblocking() -> list[tuple[str, dict[str, Any]]]:
        """큐에 쌓인 이벤트를 즉시 모두 뽑아 리스트로 반환 (blocking 없음)."""
        flushed: list[tuple[str, dict[str, Any]]] = []
        while True:
            try:
                item = _rt_event_queue.get_nowait()
            except asyncio.QueueEmpty:
                break
            if item is None:
                continue
            flushed.append(item)
        return flushed

    # graph.astream 을 백그라운드 task 로 돌려 실시간 큐와 race.
    _astream_iter = graph.astream(initial, stream_mode="updates").__aiter__()
    _graph_done = False
    _graph_next_task: asyncio.Task | None = None
    _rt_next_task: asyncio.Task | None = None

    try:
        while not _graph_done or not _rt_event_queue.empty():
            if _graph_next_task is None and not _graph_done:
                _graph_next_task = asyncio.ensure_future(_astream_iter.__anext__())
            if _rt_next_task is None:
                _rt_next_task = asyncio.ensure_future(_rt_event_queue.get())

            pending = [t for t in (_graph_next_task, _rt_next_task) if t is not None]
            if not pending:
                break
            # idle 시 15초 간격으로 keepalive 를 보내 프록시/브라우저 disconnect 방지.
            # debate 동안 AG2 Bedrock 호출이 길어지면 rt queue / graph iter 둘 다 조용해짐.
            done, _ = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED, timeout=15.0)
            if not done:
                # 타임아웃 — 아무 태스크도 완료 안 됨. keepalive 푸시하고 계속 대기.
                yield sse_keepalive()
                continue

            # ── realtime 이벤트 먼저 드레인 (프론트 반응성 우선) ──
            if _rt_next_task in done:
                try:
                    rt_item = _rt_next_task.result()
                except Exception:
                    rt_item = None
                _rt_next_task = None
                if rt_item is not None:
                    rt_name, rt_payload = rt_item
                    key = _event_dedup_key(rt_name, rt_payload)
                    if key not in _rt_emitted_keys:
                        _rt_emitted_keys.add(key)
                        logger.info(
                            "📤 SSE → client · %s · item=%s persona=%s round=%s",
                            rt_name,
                            rt_payload.get("item_number"),
                            rt_payload.get("persona_id"),
                            rt_payload.get("round"),
                        )
                        yield sse(rt_name, rt_payload)
                # 큐에 쌓여있는 나머지도 즉시 드레인
                for extra in _drain_realtime_queue_nonblocking():
                    rt_name, rt_payload = extra
                    key = _event_dedup_key(rt_name, rt_payload)
                    if key in _rt_emitted_keys:
                        continue
                    _rt_emitted_keys.add(key)
                    logger.info(
                        "📤 SSE → client (drained) · %s · item=%s persona=%s round=%s",
                        rt_name,
                        rt_payload.get("item_number"),
                        rt_payload.get("persona_id"),
                        rt_payload.get("round"),
                    )
                    yield sse(rt_name, rt_payload)

            # ── graph.astream 의 다음 업데이트 ──
            if _graph_next_task is not None and _graph_next_task in done:
                try:
                    event = _graph_next_task.result()
                except StopAsyncIteration:
                    _graph_done = True
                    _graph_next_task = None
                    continue
                except Exception:
                    _graph_next_task = None
                    raise
                _graph_next_task = None

                # event = {node_name: state_delta}
                if not isinstance(event, dict):
                    continue
                for node, delta in event.items():
                    if not isinstance(delta, dict):
                        delta = {}
                    _merge_update(accum, delta)

                    # Debate 이벤트 flush — _debate_events 버퍼는 realtime 큐로 이미
                    # 전달됐지만, 콜백 없이 실행된 경로 (run_direct_batch 등) 호환성을 위해
                    # 버퍼 이벤트 중 실시간 큐에 안 나온 것만 보완적으로 emit.
                    # ★ Option A (2026-04-24): sub-agent 가 inline debate 실행 시에도
                    # _debate_events 를 delta 로 반환 → debate 노드만이 아니라 모든 노드의
                    # _debate_events 를 flush.
                    if delta.get("_debate_events"):
                        debate_events = delta.get("_debate_events") or []
                        if isinstance(debate_events, list):
                            _allowed_event_names = {
                                "debate_round_start",
                                "persona_turn",
                                "moderator_verdict",
                                "debate_final",
                                "discussion_started",
                                "persona_speaking",
                                "persona_message",
                                "vote_cast",
                                "discussion_round_complete",
                                "discussion_finalized",
                            }
                            _flushed = 0
                            for ev in debate_events:
                                if not isinstance(ev, dict):
                                    continue
                                name = str(ev.get("event") or "")
                                data = ev.get("data") if isinstance(ev.get("data"), dict) else {}
                                if name in _allowed_event_names:
                                    key = _event_dedup_key(name, data)
                                    if key in _rt_emitted_keys:
                                        continue
                                    _rt_emitted_keys.add(key)
                                    yield sse(name, data)
                                    _flushed += 1
                            if _flushed:
                                logger.info("SSE stream: debate post-flush %d events (not in realtime queue)", _flushed)

                    phase = _node_phase(node)
                    _evals_count = len(accum.get("evaluations") or [])
                    _delta_keys = sorted((delta or {}).keys())
                    logger.info(
                        "🟢 NODE DONE — %s · phase=%s · %.2fs · evals=%d · delta_keys=%s",
                        node,
                        phase,
                        time.time() - t0,
                        _evals_count,
                        _delta_keys,
                    )

                    # 페이즈 전환 라우팅 이벤트 — 노드 완료로 인해 뒤늦게 발견되는 경우 보정
                    if phase != last_phase and phase != "other":
                        fe_route_node = _frontend_node_names(node)[0]
                        yield sse(
                            "routing",
                            {
                                "phase": phase,
                                "phase_label": phase_labels.get(phase, phase),
                                "next_node": fe_route_node,
                                "next_label": fe_route_node,
                            },
                        )
                        last_phase = phase

                    # 노드 완료 status 이벤트 — 프론트 노드 이름으로 확장 emit
                    scores = _node_item_scores(node, delta)
                    frontend_targets = _frontend_node_names(node)
                    for fe_name in frontend_targets:
                        started = node_start.get(fe_name, node_start.get(node, time.time()))
                        elapsed = round(time.time() - started, 2)
                        status_payload: dict[str, Any] = {
                            "node": fe_name,
                            "label": fe_name,
                            "status": "completed",
                            "elapsed": elapsed,
                        }
                        if scores and fe_name == frontend_targets[-1]:
                            # 점수는 마지막 프론트 노드에만 첨부 (중복 방지)
                            status_payload["scores"] = scores
                        yield sse("status", status_payload)

                    # node_trace 이벤트 — 프론트 Trace 탭 입력용
                    # 마지막 프론트 타겟 하나에만 원본 node 의 input/output 을 실어 중복 방지.
                    trace_target = frontend_targets[-1]
                    trace_started = node_start.get(trace_target, node_start.get(node, time.time()))
                    trace_elapsed = round(time.time() - trace_started, 2)
                    trace_payload: dict[str, Any] = {
                        "node": trace_target,
                        "label": trace_target,
                        "phase": phase,
                        "elapsed": trace_elapsed,
                        "input": _build_trace_input(node, initial, accum),
                        "output": _sanitize_trace_output(delta),
                    }
                    yield sse("node_trace", trace_payload)

                    # 선제 페이즈 진입 이벤트 (Sub Agent/Layer3/Layer4 시작 신호)
                    for preflight_event in _emit_next_phase_preflight(node):
                        yield preflight_event

    except Exception as e:
        logger.exception("💥 SSE STREAM FAIL — /evaluate/stream")
        yield sse(
            "error",
            {"error": "graph_invoke_failed", "detail": str(e), "message": str(e), "type": "graph_invoke_failed"},
        )
        # 실패 경로 — 종료 배너 + pending task 정리 후 바로 리턴
        logger.info("=" * 78)
        logger.info(
            "🏁 SSE STREAM END (error) — evals=%d debates=%d completed_nodes=%s",
            len(accum.get("evaluations") or []),
            len(accum.get("debates") or {}),
            accum.get("completed_nodes") or [],
        )
        logger.info("=" * 78)
        for _t in (_graph_next_task, _rt_next_task):
            if _t is not None and not _t.done():
                _t.cancel()
        return
    finally:
        # shutdown ON — 이후 background debate thread 에서 들어오는 callback 은 즉시 폐기.
        _rt_shutdown[0] = True
        # 정상/실패 공통 — realtime queue pending task 취소.
        if _rt_next_task is not None and not _rt_next_task.done():
            _rt_next_task.cancel()
        if _graph_next_task is not None and not _graph_next_task.done():
            _graph_next_task.cancel()
        # SSE log handler detach — 다음 요청에 끌려가지 않게 반드시 제거.
        try:
            for _lg in _log_attached_loggers:
                _lg.removeHandler(_sse_log_handler)
            logger.info("📡 SSE log streaming 종료 — 전송 %d건", _log_seen[0])
        except Exception:
            pass

    # 정상 경로에서 큐에 남은 late-arriving 이벤트를 마지막으로 flush.
    _tail_events = _drain_realtime_queue_nonblocking()
    for _rt_name, _rt_payload in _tail_events:
        _key = _event_dedup_key(_rt_name, _rt_payload)
        if _key in _rt_emitted_keys:
            continue
        _rt_emitted_keys.add(_key)
        yield sse(_rt_name, _rt_payload)

    elapsed_total = round(time.time() - t0, 2)
    response = _extract_response(accum)
    response["_meta"] = {
        "pipeline": "v2",
        "elapsed_sec": elapsed_total,
        "session_id": initial["session_id"],
        "tenant_id": initial.get("tenant_id") or "generic",
    }

    # result 이벤트 — 최종 리포트/평가 전체 payload
    logger.info("SSE stream: result emit (session=%s, elapsed=%.2fs)", initial.get("session_id"), elapsed_total)
    yield sse("result", response)

    # done 이벤트 — 프론트가 pipeline 종료 판단
    logger.info("SSE stream: done (session=%s, elapsed=%.2fs)", initial.get("session_id"), elapsed_total)
    yield sse("done", {"elapsed_seconds": elapsed_total})

    logger.info("=" * 78)
    logger.info(
        "🏁 SSE STREAM END — evals=%d debates=%d completed_nodes=%s elapsed=%.2fs",
        len(accum.get("evaluations") or []),
        len(accum.get("debates") or {}),
        accum.get("completed_nodes") or [],
        elapsed_total,
    )
    logger.info("=" * 78)


@app.post("/evaluate/stream")
async def evaluate_stream(request: Request) -> StreamingResponse:
    # body / graph 를 제너레이터 외부에서 선처리 — generator 안에서 `await request.json()`
    # 호출 시 첫 yield 이후 본문 read 가 스트림을 정지시키는 이슈 회피.
    try:
        body = await request.json()
    except Exception as exc:
        # except 블록을 벗어나면 `exc` 바인딩이 해제되므로 메시지를 로컬로 캡처해 클로저로 전달.
        err_msg = str(exc)

        async def _bad_request():
            yield (
                "event: error\n"
                f"data: {json.dumps({'error': 'bad_request', 'detail': err_msg, 'message': err_msg}, ensure_ascii=False)}\n\n"
            )

        return StreamingResponse(_bad_request(), media_type="text/event-stream")

    graph = _get_graph()
    if graph is None:

        async def _graph_error():
            yield f"event: error\ndata: {json.dumps({'error': 'graph_build_failed', 'detail': _graph_build_error, 'message': _graph_build_error or 'graph build failed'}, ensure_ascii=False)}\n\n"

        return StreamingResponse(_graph_error(), media_type="text/event-stream")

    return StreamingResponse(
        _build_stream_events(body, graph),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache, no-transform", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )


# ---------------------------------------------------------------------------
# V3 interactive discussion — auto_start=False 시 프론트가 제어하는 gate 엔드포인트
# ---------------------------------------------------------------------------
# 프론트가 /evaluate/stream 요청 시 auto_start=false 를 보냈다면:
#   1. 토론 시작 전 discussion_started 이벤트가 discussion_id 를 싣고 SSE 로 도착.
#   2. 프론트는 사용자가 "시작" 버튼 누르면 POST /v2/discussion/{discussion_id}/start 호출.
#   3. 이 엔드포인트가 threading.Event.set() 호출 → run_debate 의 wait() 해제.
#   4. 토론이 정상 진행되며 persona_speaking / persona_message / ... 이벤트 발송.


@app.post("/v2/discussion/{discussion_id}/start")
async def discussion_start(discussion_id: str) -> JSONResponse:
    """토론 시작 신호 — gate event 를 set. 토론이 이미 시작됐거나 존재 안 하면 404."""
    try:
        released = _release_discussion_gate(discussion_id)
    except Exception as exc:
        logger.exception("discussion_start 실패 — id=%s", discussion_id)
        return JSONResponse({"error": "internal", "detail": str(exc)}, status_code=500)

    if not released:
        return JSONResponse(
            {"error": "not_found", "detail": f"discussion_id={discussion_id} 이 존재하지 않음"}, status_code=404
        )
    logger.info("discussion_start: released gate for discussion_id=%s", discussion_id)
    # 메모리 누수 방지 — 200 개 초과 시 오래된 것부터 제거.
    _cleanup_old_discussion_gates()
    return JSONResponse({"ok": True, "discussion_id": discussion_id})


@app.post("/v2/discussion/{discussion_id}/next-round")
async def discussion_next_round(discussion_id: str) -> JSONResponse:
    """다음 라운드 진행 신호 — 현재 구현은 /start 와 동일한 gate 해제 (라운드 단위 pause hook).

    향후 라운드별 개별 gate 로 확장 가능 (now: 단일 gate per discussion).
    """
    try:
        released = _release_discussion_gate(discussion_id)
    except Exception as exc:
        logger.exception("discussion_next_round 실패 — id=%s", discussion_id)
        return JSONResponse({"error": "internal", "detail": str(exc)}, status_code=500)

    if not released:
        return JSONResponse(
            {"error": "not_found", "detail": f"discussion_id={discussion_id} 이 존재하지 않음"}, status_code=404
        )
    logger.info("discussion_next_round: released gate for discussion_id=%s", discussion_id)
    return JSONResponse({"ok": True, "discussion_id": discussion_id, "action": "next_round"})


@app.get("/v2/discussion/{discussion_id}/status")
async def discussion_status(discussion_id: str) -> JSONResponse:
    """디버깅 용 — gate 존재 / signaled 여부 조회."""
    with _DISCUSSION_GATES_LOCK:
        ev = _DISCUSSION_GATES.get(discussion_id)
    if ev is None:
        return JSONResponse({"discussion_id": discussion_id, "exists": False}, status_code=404)
    return JSONResponse({"discussion_id": discussion_id, "exists": True, "signaled": ev.is_set()})


# ---------------------------------------------------------------------------
# HITL (Human-in-the-loop) 엔드포인트 — human_reviews + golden-set 후보
# ---------------------------------------------------------------------------

_hitl_initialized = False


def _ensure_hitl_db() -> None:
    """첫 HITL 호출 시 lazy init (module import 시 자동 호출하지 않음)."""
    global _hitl_initialized
    if _hitl_initialized:
        return
    from v2.hitl import db as _hitl_db

    _hitl_db.init_db()
    _hitl_initialized = True


@app.post("/v2/human-review")
async def human_review_upsert(request: Request) -> JSONResponse:
    """수정 모드에서 전송된 human_score/note 저장 (consultation_id + item_number UPSERT).

    Request body (JSON):
      {
        "consultation_id": "668437",
        "item_number": 7,
        "ai_score": 5.0,
        "human_score": 3.0,
        "ai_evidence": [...] | {...} | "string",   // 선택
        "ai_judgment": "...",                       // 선택
        "human_note": "...",                        // 선택
        "ai_confidence": 0.72,                      // 선택
        "reviewer_id": "youngjoon",                 // 선택
        "reviewer_role": "senior",                  // 선택 (기본 'senior')
        "force_t3": false                           // 선택
      }
    """
    try:
        _ensure_hitl_db()
        body = await request.json()
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    consult = body.get("consultation_id")
    item_no = body.get("item_number")
    if not consult or item_no is None:
        return JSONResponse(
            {"error": "bad_request", "detail": "consultation_id, item_number required"}, status_code=400
        )

    try:
        from v2.hitl import db as _hitl_db

        review_id = _hitl_db.upsert_review(
            consultation_id=str(consult),
            item_number=int(item_no),
            ai_score=body.get("ai_score"),
            human_score=body.get("human_score"),
            ai_evidence=body.get("ai_evidence"),
            ai_judgment=body.get("ai_judgment"),
            human_note=body.get("human_note"),
            ai_confidence=body.get("ai_confidence"),
            reviewer_id=body.get("reviewer_id"),
            reviewer_role=body.get("reviewer_role"),
            force_t3=body.get("force_t3"),
            status=body.get("status"),
        )
        return JSONResponse({"ok": True, "id": review_id})
    except Exception as e:
        logger.exception("human-review upsert 실패")
        return JSONResponse({"error": "upsert_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/review/queue")
async def review_queue(status: str = "pending", force_t3_only: bool = False, limit: int = 100) -> JSONResponse:
    """검토 큐 조회. status=''(or 'all') 이면 전체."""
    try:
        _ensure_hitl_db()
        from v2.hitl import db as _hitl_db

        status_val: str | None = status or None
        if status_val and status_val.lower() == "all":
            status_val = None
        items = _hitl_db.list_reviews(status=status_val, force_t3_only=bool(force_t3_only), limit=int(limit))
        return JSONResponse({"items": items})
    except Exception as e:
        logger.exception("review queue 조회 실패")
        return JSONResponse({"error": "queue_failed", "detail": str(e)}, status_code=500)


def _resolve_hitl_edits_root():
    """사람 수정 결과 스냅샷 루트. env `QA_HITL_EDITS_ROOT` 우선, 기본 `~/Desktop/QA평가결과/HITL_수정`.
    `/v2/review/{id}/confirm` 이 매 확정 시 해당 상담의 전체 검토 행을 이 폴더에 JSON 으로 덮어씀.
    """
    import os as _os
    from pathlib import Path as _Path

    root_str = _os.environ.get("QA_HITL_EDITS_ROOT") or str(_Path.home() / "Desktop" / "QA평가결과" / "HITL_수정")
    return _Path(root_str)


def _save_consultation_edits_snapshot(consultation_id: str) -> str | None:
    """해당 consultation_id 의 모든 human_reviews 행을 JSON 스냅샷으로 폴더에 저장.

    - 파일명: `<consultation_id>.json` (경로 구분자는 _ 로 치환)
    - 내용: consultation_id, saved_at, counts(pending/confirmed/total), items[]
      items 각각: item_number, ai_score, human_score, delta, ai_confidence,
                 ai_judgment, ai_evidence, human_note, status, reviewer_id,
                 reviewer_role, force_t3, confirmed_at, created_at, updated_at
    - 실패 시 None 반환 + warning 로그 (confirm 자체는 성공).
    """
    try:
        import json as _json
        from datetime import datetime as _datetime
        from v2.hitl import db as _hitl_db  # lazy import

        with _hitl_db.get_conn() as conn:
            rows = conn.execute(
                "SELECT * FROM human_reviews WHERE consultation_id = ? ORDER BY item_number", (str(consultation_id),)
            ).fetchall()
            items_out: list[dict] = []
            pending_n = confirmed_n = 0
            for r in rows:
                d = dict(r)
                # ai_evidence 가 TEXT(json) 이면 parse
                ev_raw = d.get("ai_evidence")
                if isinstance(ev_raw, str) and ev_raw.strip():
                    try:
                        d["ai_evidence"] = _json.loads(ev_raw)
                    except Exception:
                        pass
                # delta 계산
                ai = d.get("ai_score")
                hu = d.get("human_score")
                try:
                    d["delta"] = float(ai) - float(hu) if ai is not None and hu is not None else None
                except (TypeError, ValueError):
                    d["delta"] = None
                st = d.get("status") or "pending"
                if st == "confirmed":
                    confirmed_n += 1
                else:
                    pending_n += 1
                items_out.append(d)
        root = _resolve_hitl_edits_root()
        root.mkdir(parents=True, exist_ok=True)
        safe_cid = str(consultation_id).replace("/", "_").replace("\\", "_") or "unknown"
        target = root / f"{safe_cid}.json"
        payload = {
            "consultation_id": consultation_id,
            "saved_at": _datetime.now(UTC).isoformat(),
            "counts": {"total": len(items_out), "pending": pending_n, "confirmed": confirmed_n},
            "items": items_out,
        }
        target.write_text(_json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        logger.info("hitl edits snapshot saved → %s (pending=%d, confirmed=%d)", target, pending_n, confirmed_n)
        return str(target)
    except Exception as exc:
        logger.warning("hitl edits snapshot 저장 실패 consult=%s — %s", consultation_id, exc)
        return None


# ---------------------------------------------------------------------------
# HITL → RAG 자동 export + ingest 훅 (confirm / revert 후처리)
# ---------------------------------------------------------------------------


def _resolve_result_json_root() -> Path:
    """파이프라인이 떨어뜨린 결과 JSON 루트 — env QA_RESULT_JSON_ROOT 우선."""
    root_str = os.environ.get("QA_RESULT_JSON_ROOT") or str(Path.home() / "Desktop" / "QA평가결과" / "JSON")
    return Path(root_str)


def _load_item_meta_by_number(consultation_id: str) -> dict[int, dict[str, Any]]:
    """`~/Desktop/QA평가결과/JSON/<cid>.json` 에서 item_number → {item_name, max_score} 맵 추출.

    파일이 없거나 파싱 실패 시 빈 dict — md_exporter 가 fallback (item_name=`item_<n>`).
    """
    safe_cid = str(consultation_id).replace("/", "_").replace("\\", "_") or "unknown"
    target = _resolve_result_json_root() / f"{safe_cid}.json"
    if not target.exists():
        return {}
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("item_meta_by_number 로드 실패 cid=%s — %s", consultation_id, exc)
        return {}
    report = payload.get("report") if isinstance(payload, dict) else None
    if not isinstance(report, dict):
        return {}
    evaluation = report.get("evaluation") or {}
    categories = evaluation.get("categories") or []
    out: dict[int, dict[str, Any]] = {}
    for cat in categories:
        for it in cat.get("items") or []:
            if not isinstance(it, dict):
                continue
            try:
                num = int(it.get("item_number"))
            except (TypeError, ValueError):
                continue
            out[num] = {
                "item_name": it.get("item_name") or it.get("name") or "",
                "max_score": it.get("max_score"),
            }
    return out


def _fetch_confirmed_review_rows(consultation_id: str) -> list[dict[str, Any]]:
    """해당 consultation 의 status='confirmed' 행 전체를 dict 리스트로 반환."""
    from v2.hitl import db as _hitl_db  # lazy import

    with _hitl_db.get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM human_reviews WHERE consultation_id = ? AND status = 'confirmed' "
            "ORDER BY item_number",
            (str(consultation_id),),
        ).fetchall()
    return [dict(r) for r in rows]


def _run_index_pending_safe(force: bool = False) -> None:
    """background task 진입점 — 예외 삼키고 로그. opensearch-py 미설치 등은 warning."""
    try:
        from v2.hitl import rag_ingester as _ing  # lazy import

        result = _ing.index_pending(force=force)
        logger.info("hitl rag_ingester background result: %s", result)
    except NotImplementedError as exc:
        logger.warning("hitl rag_ingester 미사용 (opensearch-py 미설치): %s", exc)
    except Exception:
        logger.exception("hitl rag_ingester background 실패")


def _export_confirmed_and_schedule_ingest(
    consultation_id: str,
    background_tasks: BackgroundTasks | None,
) -> dict[str, Any]:
    """confirm 후처리 — 확정 rows fetch → MD export → background 인덱싱 스케줄.

    반환값은 응답 body 의 `rag` 필드로 즉시 노출 (export 결과만; ingest 는 비동기).
    실패해도 confirm 자체는 성공 — 모든 예외 swallow + warning.
    """
    try:
        from v2.hitl import md_exporter as _md  # lazy import

        rows = _fetch_confirmed_review_rows(consultation_id)
        if not rows:
            return {"exported": 0, "scheduled": False, "skipped_reason": "no_confirmed_rows"}
        item_meta = _load_item_meta_by_number(consultation_id)
        written = _md.export_consultation_confirmed(
            consultation_id, review_rows=rows, item_meta_by_number=item_meta
        )
        scheduled = False
        if background_tasks is not None:
            background_tasks.add_task(_run_index_pending_safe, False)
            scheduled = True
        else:
            try:
                import asyncio as _aio

                _aio.create_task(_aio.to_thread(_run_index_pending_safe, False))
                scheduled = True
            except RuntimeError:
                scheduled = False
        return {
            "exported": len(written),
            "scheduled": scheduled,
            "files": [p.name for p in written],
        }
    except Exception as exc:
        logger.exception("hitl export+ingest 후처리 실패 cid=%s", consultation_id)
        return {"exported": 0, "scheduled": False, "error": str(exc)}


def _delete_md_and_index_for_review(consultation_id: str, item_number: int) -> dict[str, Any]:
    """revert 후처리 — 해당 (cid, item) 의 MD 파일 삭제 + AOSS doc best-effort 삭제.

    실패는 warning 만 — revert 자체에는 영향 없음.
    """
    out: dict[str, Any] = {"md_deleted": False, "index_deleted": 0}
    try:
        from v2.hitl import md_exporter as _md  # lazy import

        target = _md.md_path_for(str(consultation_id), int(item_number))
        if target.exists():
            try:
                target.unlink()
                out["md_deleted"] = True
                out["md_path"] = str(target)
            except Exception as exc:
                logger.warning("MD 파일 삭제 실패 %s — %s", target, exc)
    except Exception as exc:
        logger.warning("md_path_for 호출 실패 cid=%s item=%s — %s", consultation_id, item_number, exc)

    try:
        from v2.hitl import rag_ingester as _ing  # lazy import

        client = _ing._client()
        # external_id prefix 매칭 (signature hash 가 모든 버전을 cover) → match_phrase_prefix
        body = {
            "query": {
                "bool": {
                    "filter": [
                        {"term": {"consultation_id": str(consultation_id)}},
                        {"term": {"item_number": int(item_number)}},
                    ]
                }
            }
        }
        resp = client.delete_by_query(
            index=_ing.INDEX_NAME, body=body, refresh=True, conflicts="proceed"
        )
        out["index_deleted"] = int(resp.get("deleted") or 0)
    except NotImplementedError:
        pass
    except Exception as exc:
        logger.info("AOSS doc 삭제 실패 cid=%s item=%s (무시): %s", consultation_id, item_number, exc)
    return out


def _delete_md_and_index_for_consultation(consultation_id: str) -> dict[str, Any]:
    """revert-all 후처리 — 상담 단위 MD/index 일괄 정리. 실패는 warning."""
    out: dict[str, Any] = {"md_deleted": 0, "index_deleted": 0}
    try:
        from v2.hitl import md_exporter as _md  # lazy import

        root = _md.resolve_rag_root()
        if root.exists():
            safe_cid = str(consultation_id).replace("/", "_").replace("\\", "_") or "unknown"
            for p in root.glob(f"{safe_cid}_*.md"):
                try:
                    p.unlink()
                    out["md_deleted"] += 1
                except Exception as exc:
                    logger.warning("MD 일괄 삭제 실패 %s — %s", p, exc)
    except Exception as exc:
        logger.warning("MD 일괄 정리 실패 cid=%s — %s", consultation_id, exc)

    try:
        from v2.hitl import rag_ingester as _ing  # lazy import

        client = _ing._client()
        body = {"query": {"term": {"consultation_id": str(consultation_id)}}}
        resp = client.delete_by_query(
            index=_ing.INDEX_NAME, body=body, refresh=True, conflicts="proceed"
        )
        out["index_deleted"] = int(resp.get("deleted") or 0)
    except NotImplementedError:
        pass
    except Exception as exc:
        logger.info("AOSS doc 일괄 삭제 실패 cid=%s (무시): %s", consultation_id, exc)
    return out


@app.post("/v2/review/{review_id}/confirm")
async def review_confirm(
    review_id: int, request: Request, background_tasks: BackgroundTasks
) -> JSONResponse:
    """검토 확정. body: {reviewer_id, reviewer_role?}.

    확정 성공 시 해당 상담의 전체 검토 스냅샷을 `~/Desktop/QA평가결과/HITL_수정/<cid>.json` 에 덮어씀.
    이어서 confirmed rows → HITL_RAG MD export 후 background 로 OpenSearch 색인.
    """
    try:
        _ensure_hitl_db()
        body = await request.json()
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    reviewer_id = body.get("reviewer_id")
    if not reviewer_id:
        return JSONResponse({"error": "bad_request", "detail": "reviewer_id required"}, status_code=400)

    try:
        from v2.hitl import db as _hitl_db

        ok = _hitl_db.confirm_review(
            review_id=int(review_id), reviewer_id=str(reviewer_id), reviewer_role=body.get("reviewer_role")
        )
        if not ok:
            return JSONResponse({"error": "not_found", "id": review_id}, status_code=404)
        # 확정 성공 → 해당 상담의 전체 검토 스냅샷 폴더 저장
        row = _hitl_db.get_review(int(review_id))
        cid = (row or {}).get("consultation_id") if row else None
        snapshot_path = _save_consultation_edits_snapshot(str(cid)) if cid else None
        rag_summary = _export_confirmed_and_schedule_ingest(str(cid), background_tasks) if cid else None
        return JSONResponse(
            {
                "ok": True,
                "consultation_id": cid,
                "snapshot_path": snapshot_path,
                "snapshot_root": str(_resolve_hitl_edits_root()),
                "rag": rag_summary,
            }
        )
    except Exception as e:
        logger.exception("review confirm 실패")
        return JSONResponse({"error": "confirm_failed", "detail": str(e)}, status_code=500)


@app.post("/v2/review/{review_id}/revert")
async def review_revert(review_id: int, request: Request) -> JSONResponse:
    """검수 취소 — status='confirmed' 를 'pending' 으로 되돌림. human_score/human_note 는 유지.

    body: {reviewer_id?, reason?} — reason 은 human_note 에 `[검수취소] <reason>` 로 prepend.
    완료 후 해당 상담의 스냅샷 재저장 + 해당 (cid, item) MD/index best-effort 삭제.
    """
    try:
        _ensure_hitl_db()
        body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    reviewer_id = (body or {}).get("reviewer_id") or "ui-user"
    reason = (body or {}).get("reason") or ""

    try:
        from v2.hitl import db as _hitl_db

        with _hitl_db.get_conn() as conn:
            row = conn.execute(
                "SELECT id, consultation_id, item_number, human_note, status FROM human_reviews WHERE id = ?",
                (int(review_id),),
            ).fetchone()
            if not row:
                return JSONResponse({"error": "not_found", "id": review_id}, status_code=404)
            cid = row["consultation_id"]
            item_no = row["item_number"]
            prev_note = row["human_note"] or ""
            new_note = (f"[검수취소:{reviewer_id}] {reason}\n" + prev_note).strip() if reason else prev_note
            conn.execute(
                "UPDATE human_reviews SET status='pending', confirmed_at=NULL, human_note=? WHERE id=?",
                (new_note, int(review_id)),
            )
        snapshot_path = _save_consultation_edits_snapshot(str(cid)) if cid else None
        rag_cleanup = (
            _delete_md_and_index_for_review(str(cid), int(item_no)) if cid and item_no is not None else None
        )
        return JSONResponse(
            {
                "ok": True,
                "consultation_id": cid,
                "snapshot_path": snapshot_path,
                "rag_cleanup": rag_cleanup,
            }
        )
    except Exception as e:
        logger.exception("review revert 실패")
        return JSONResponse({"error": "revert_failed", "detail": str(e)}, status_code=500)


@app.post("/v2/review/consultation/{consultation_id}/revert-all")
async def review_revert_all(consultation_id: str, request: Request) -> JSONResponse:
    """상담 단위 일괄 취소 — 해당 상담의 모든 confirmed 행을 pending 으로 되돌림."""
    try:
        _ensure_hitl_db()
        body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    reviewer_id = (body or {}).get("reviewer_id") or "ui-user"
    reason = (body or {}).get("reason") or ""

    try:
        from v2.hitl import db as _hitl_db

        with _hitl_db.get_conn() as conn:
            rows = conn.execute(
                "SELECT id, human_note FROM human_reviews WHERE consultation_id=? AND status='confirmed'",
                (str(consultation_id),),
            ).fetchall()
            reverted = 0
            for r in rows:
                prev_note = r["human_note"] or ""
                new_note = (f"[검수취소:{reviewer_id}] {reason}\n" + prev_note).strip() if reason else prev_note
                conn.execute(
                    "UPDATE human_reviews SET status='pending', confirmed_at=NULL, human_note=? WHERE id=?",
                    (new_note, int(r["id"])),
                )
                reverted += 1
        snapshot_path = _save_consultation_edits_snapshot(str(consultation_id))
        rag_cleanup = _delete_md_and_index_for_consultation(str(consultation_id)) if reverted else None
        return JSONResponse(
            {
                "ok": True,
                "consultation_id": consultation_id,
                "reverted": reverted,
                "snapshot_path": snapshot_path,
                "rag_cleanup": rag_cleanup,
            }
        )
    except Exception as e:
        logger.exception("revert-all 실패 consult=%s", consultation_id)
        return JSONResponse({"error": "revert_all_failed", "detail": str(e)}, status_code=500)


@app.post("/v2/review/consultation/{consultation_id}/confirm-all")
async def review_confirm_all(
    consultation_id: str, request: Request, background_tasks: BackgroundTasks
) -> JSONResponse:
    """상담 단위 일괄 확정. body: {reviewer_id, reviewer_role?, accept_ai_score?=true, overwrite?=false}

    기본 동작: `status != 'confirmed'` 인 모든 항목을 확정.
    - `accept_ai_score=true` (기본): `human_score IS NULL` 인 항목은 `human_score=ai_score` 로 채움 (AI 판정 그대로 승인)
    - `overwrite=true`: 이미 확정된 항목도 재확정 (confirmed_at/reviewer 갱신)
    - 완료 후 `HITL_수정/<cid>.json` 스냅샷 1회 저장.
    """
    try:
        _ensure_hitl_db()
        body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    reviewer_id = (body or {}).get("reviewer_id") or "ui-user"
    reviewer_role = (body or {}).get("reviewer_role") or "senior"
    accept_ai_score = (body or {}).get("accept_ai_score", True)
    overwrite = (body or {}).get("overwrite", False)

    try:
        from datetime import datetime
        from v2.hitl import db as _hitl_db

        now = datetime.now(UTC).astimezone().isoformat(timespec="seconds")
        confirmed = 0
        filled_from_ai = 0
        skipped = 0

        with _hitl_db.get_conn() as conn:
            rows = conn.execute(
                "SELECT id, item_number, ai_score, human_score, status FROM human_reviews "
                "WHERE consultation_id = ? ORDER BY item_number",
                (str(consultation_id),),
            ).fetchall()
            if not rows:
                return JSONResponse(
                    {
                        "error": "not_found",
                        "consultation_id": consultation_id,
                        "detail": "no reviews for this consultation",
                    },
                    status_code=404,
                )
            for r in rows:
                rid = int(r["id"])
                cur_status = r["status"] or "pending"
                if cur_status == "confirmed" and not overwrite:
                    skipped += 1
                    continue
                hs = r["human_score"]
                if hs is None:
                    if accept_ai_score and r["ai_score"] is not None:
                        hs = r["ai_score"]
                        filled_from_ai += 1
                    else:
                        # 점수가 없으면 확정 불가 — skip
                        skipped += 1
                        continue
                conn.execute(
                    "UPDATE human_reviews SET status='confirmed', confirmed_at=?, "
                    "reviewer_id=?, reviewer_role=?, human_score=? WHERE id=?",
                    (now, str(reviewer_id), str(reviewer_role), float(hs), rid),
                )
                confirmed += 1

        snapshot_path = _save_consultation_edits_snapshot(str(consultation_id))
        rag_summary = (
            _export_confirmed_and_schedule_ingest(str(consultation_id), background_tasks)
            if confirmed
            else None
        )
        return JSONResponse(
            {
                "ok": True,
                "consultation_id": consultation_id,
                "confirmed": confirmed,
                "filled_from_ai": filled_from_ai,
                "skipped": skipped,
                "snapshot_path": snapshot_path,
                "snapshot_root": str(_resolve_hitl_edits_root()),
                "rag": rag_summary,
            }
        )
    except Exception as e:
        logger.exception("confirm-all 실패 consult=%s", consultation_id)
        return JSONResponse({"error": "confirm_all_failed", "detail": str(e)}, status_code=500)


@app.post("/v2/review/consultation/{consultation_id}/delete-all")
async def review_delete_all(consultation_id: str, request: Request) -> JSONResponse:
    """상담 단위 영구 삭제 — 해당 상담의 모든 human_reviews 행을 DB 에서 제거.

    body: ``{reviewer_id?, reason?}`` (감사 로깅용)
    """
    try:
        _ensure_hitl_db()
        body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    reviewer_id = (body or {}).get("reviewer_id") or "ui-user"
    reason = (body or {}).get("reason") or ""

    try:
        from v2.hitl import db as _hitl_db

        with _hitl_db.get_conn() as conn:
            row = conn.execute(
                "SELECT COUNT(*) AS n FROM human_reviews WHERE consultation_id=?", (str(consultation_id),)
            ).fetchone()
            count_before = int(row["n"]) if row else 0
            if count_before == 0:
                return JSONResponse(
                    {"error": "not_found", "consultation_id": consultation_id, "deleted": 0}, status_code=404
                )
            conn.execute("DELETE FROM human_reviews WHERE consultation_id=?", (str(consultation_id),))
        logger.info(
            "delete-all consult=%s deleted=%d by=%s reason=%s", consultation_id, count_before, reviewer_id, reason
        )
        return JSONResponse({"ok": True, "consultation_id": consultation_id, "deleted": count_before})
    except Exception as e:
        logger.exception("delete-all 실패 consult=%s", consultation_id)
        return JSONResponse({"error": "delete_all_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/review/edits/{consultation_id}")
async def review_edits_read(consultation_id: str) -> JSONResponse:
    """`HITL_수정` 폴더에 저장된 해당 상담의 스냅샷 JSON 반환. 없으면 404."""
    try:
        import json as _json

        safe_cid = consultation_id.replace("/", "_").replace("\\", "_")
        target = _resolve_hitl_edits_root() / f"{safe_cid}.json"
        if not target.exists():
            return JSONResponse(
                {"error": "not_found", "consultation_id": consultation_id, "path": str(target)}, status_code=404
            )
        return JSONResponse(
            {
                "ok": True,
                "consultation_id": consultation_id,
                "path": str(target),
                "data": _json.loads(target.read_text(encoding="utf-8")),
            }
        )
    except Exception as e:
        logger.exception("hitl edits read 실패")
        return JSONResponse({"error": "read_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/tuning/priorities")
async def tuning_priorities(rolling_n: int = 50) -> JSONResponse:
    """항목별 rolling-N 통계 (MAE / Bias / override_pct). sample_count < N 이면 metric=None."""
    try:
        _ensure_hitl_db()
        from v2.hitl import db as _hitl_db

        items = _hitl_db.aggregate_priorities(rolling_n=int(rolling_n))
        return JSONResponse({"items": items})
    except Exception as e:
        logger.exception("tuning priorities 집계 실패")
        return JSONResponse({"error": "priorities_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/review/export-xlsx")
async def review_export_xlsx(status: str = "all", consultation_id: str | None = None) -> JSONResponse:
    """human_reviews 전체(또는 필터)를 xlsx 로 저장.

    저장 경로: env `QA_HITL_EXPORT_ROOT` (기본 `~/Desktop/QA평가결과/HITL_비교/`).
    파일명: `<YYYY-MM-DD_HHMMSS>_reviews.xlsx`. 빈 결과도 header row 만 있는 xlsx 생성.
    """
    try:
        _ensure_hitl_db()
        from v2.hitl.export import export_reviews_to_xlsx

        path, row_count = export_reviews_to_xlsx(status=status, consultation_id=consultation_id)
        return JSONResponse({"ok": True, "path": str(path), "row_count": row_count})
    except Exception as e:
        logger.exception("review export-xlsx 실패")
        return JSONResponse({"error": "export_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/result/full/{consultation_id}")
async def result_full(consultation_id: str) -> JSONResponse:
    """검토 큐 풀뷰용 — 파이프라인이 저장해 둔 결과 JSON 반환.

    저장 위치: env `QA_RESULT_JSON_ROOT` (기본 `~/Desktop/QA평가결과/JSON/`)
    파일명: `<consultation_id>.json` — report / gt_comparison / gt_evidence_comparison 포함.
    """
    import json as _json
    import os as _os
    from pathlib import Path as _Path

    try:
        safe_cid = consultation_id.replace("/", "_").replace("\\", "_")
        if not safe_cid or safe_cid in (".", ".."):
            return JSONResponse({"error": "invalid_consultation_id"}, status_code=400)

        root_str = _os.environ.get("QA_RESULT_JSON_ROOT") or str(_Path.home() / "Desktop" / "QA평가결과" / "JSON")
        target = _Path(root_str) / f"{safe_cid}.json"
        if not target.exists():
            return JSONResponse(
                {"error": "not_found", "consultation_id": consultation_id, "path": str(target)}, status_code=404
            )
        payload = _json.loads(target.read_text(encoding="utf-8"))
        return JSONResponse({"ok": True, "consultation_id": consultation_id, "data": payload})
    except Exception as e:
        logger.exception("result full 로드 실패")
        return JSONResponse({"error": "load_failed", "detail": str(e)}, status_code=500)


# ---------------------------------------------------------------------------
# Judge LLM — 좌/우 모델 비교, 수동 평가표 비교 (Sonnet 4)
# ---------------------------------------------------------------------------


def _compact_items_for_judge(result: dict[str, Any] | None) -> list[dict[str, Any]]:
    """evaluation 결과에서 Judge LLM 프롬프트용 compact item_scores 추출.

    report.item_scores 우선, 없으면 evaluations[].evaluation 에서 재구성.
    """
    if not isinstance(result, dict):
        return []
    report = result.get("report") or {}
    items = report.get("item_scores") if isinstance(report, dict) else None
    if isinstance(items, list) and items:
        return [
            {
                "item_number": it.get("item_number"),
                "item_name": it.get("item_name"),
                "score": it.get("score"),
                "max_score": it.get("max_score"),
                "deductions": it.get("deductions"),
                "evidence": it.get("evidence"),
            }
            for it in items
            if isinstance(it, dict)
        ]
    evals = result.get("evaluations") or []
    out: list[dict[str, Any]] = []
    for ev in evals:
        if not isinstance(ev, dict):
            continue
        e = ev.get("evaluation") or {}
        out.append(
            {
                "item_number": e.get("item_number"),
                "item_name": e.get("item_name"),
                "score": e.get("score"),
                "max_score": e.get("max_score"),
                "deductions": e.get("deductions"),
                "evidence": e.get("evidence") or e.get("judgment"),
            }
        )
    return out


def _truncate_transcript(transcript: str | None, limit: int = 6000) -> str:
    if not transcript:
        return ""
    s = str(transcript)
    if len(s) <= limit:
        return s
    return s[:limit] + f"\n... [생략 — 총 {len(s)}자]"


async def _invoke_judge_llm(system_prompt: str, user_prompt: str, max_tokens: int = 4000) -> str:
    """Judge LLM 호출 — Bedrock Sonnet 4 기본. `nodes.llm.ainvoke_llm` 헬퍼 위임."""
    from langchain_core.messages import HumanMessage, SystemMessage  # type: ignore
    from nodes.llm import ainvoke_llm  # noqa: WPS433

    messages = [SystemMessage(content=system_prompt), HumanMessage(content=user_prompt)]
    return await ainvoke_llm(messages, temperature=0.2, max_tokens=max_tokens, backend="bedrock")


@app.post("/analyze-compare")
async def analyze_compare(request: Request) -> JSONResponse:
    """좌/우 두 모델 평가 결과 비교 — Sonnet judge → markdown 리포트 반환.

    Request body (JSON):
      {
        "left_result":  { ...evaluation result... },
        "right_result": { ...evaluation result... },
        "left_model":   "Sonnet 4" | "Qwen3-8B" | ...,
        "right_model":  "Haiku 3.5" | ...,
        "transcript":   "상담사: ..." (선택, 6000자 컷)
      }
    Response:
      { "status": "success", "analysis": "...markdown...", "markdown": "...", "_meta": {...} }
    """
    try:
        body = await request.json()
    except Exception as e:
        return JSONResponse({"status": "error", "error": "bad_request", "detail": str(e)}, status_code=400)

    left_result = body.get("left_result") or {}
    right_result = body.get("right_result") or {}
    left_model = body.get("left_model") or "Model A"
    right_model = body.get("right_model") or "Model B"
    transcript = _truncate_transcript(body.get("transcript"))

    if not isinstance(left_result, dict) or not isinstance(right_result, dict):
        return JSONResponse(
            {"status": "error", "error": "bad_request", "detail": "left_result/right_result dict 필수"}, status_code=400
        )
    if not left_result and not right_result:
        return JSONResponse(
            {"status": "error", "error": "bad_request", "detail": "최소 한쪽 결과 필요"}, status_code=400
        )

    left_items = _compact_items_for_judge(left_result)
    right_items = _compact_items_for_judge(right_result)
    left_total = ((left_result.get("orchestrator") or {}).get("final_score") or {}).get("raw_total")
    right_total = ((right_result.get("orchestrator") or {}).get("final_score") or {}).get("raw_total")

    import json as _json

    system_prompt = (
        "당신은 한국어 콜센터 QA 평가 비교 전문가입니다. 두 모델의 QA 평가 결과를 비교하여 "
        "마크다운 리포트를 작성하세요. 숫자 근거 없이 주장 금지. 항목(1~18) 단위로 차이 설명."
    )
    user_prompt = (
        f"## 원본 상담 스크립트 (일부)\n{transcript or '(제공되지 않음)'}\n\n"
        f"## 모델 A: {left_model}\n"
        f"총점: {left_total}\n"
        f"항목별: {_json.dumps(left_items, ensure_ascii=False)[:8000]}\n\n"
        f"## 모델 B: {right_model}\n"
        f"총점: {right_total}\n"
        f"항목별: {_json.dumps(right_items, ensure_ascii=False)[:8000]}\n\n"
        "### 작성 요구사항\n"
        "1. **총점 비교** — A/B 총점, 차이, 더 관대/엄격한 쪽.\n"
        "2. **항목별 점수 차이 Top 5** — 차이가 큰 항목 5개, 각 항목에 대해 근거 스니펫 대조.\n"
        "3. **평가 일관성** — 비슷한 근거에서 판정이 갈린 사례.\n"
        "4. **권장 — 코칭 포인트** — 이 통화에서 상담사가 개선해야 할 포인트 2~3개.\n"
        "5. **최종 판정** — 어느 모델의 평가가 실제에 더 가까워 보이는가 + 이유.\n\n"
        "출력은 순수 markdown. 머리글은 `###` 이하 사용. 테이블 활용."
    )

    try:
        t0 = time.time()
        markdown = await _invoke_judge_llm(system_prompt, user_prompt, max_tokens=4000)
        elapsed = round(time.time() - t0, 2)
    except Exception as e:  # noqa: BLE001
        logger.exception("analyze-compare judge 실패")
        return JSONResponse({"status": "error", "error": "judge_failed", "detail": str(e)}, status_code=500)

    return JSONResponse(
        {
            "status": "success",
            "analysis": markdown,
            "markdown": markdown,
            "_meta": {
                "left_model": left_model,
                "right_model": right_model,
                "left_total": left_total,
                "right_total": right_total,
                "elapsed_sec": elapsed,
            },
        }
    )


def _manual_row_lookup(manual_rows: list[dict[str, Any]] | None) -> dict[int, dict[str, Any]]:
    """manual_rows → item_number indexed dict."""
    out: dict[int, dict[str, Any]] = {}
    for r in manual_rows or []:
        if not isinstance(r, dict):
            continue
        num = r.get("item_number") or r.get("no") or r.get("number")
        try:
            num_int = int(num) if num is not None else None
        except (TypeError, ValueError):
            num_int = None
        if num_int:
            out[num_int] = r
    return out


@app.post("/analyze-manual-compare")
async def analyze_manual_compare(request: Request) -> JSONResponse:
    """사람 QA 평가표 vs 모델 결과 비교 — Sonnet judge → rows + comparison_table + markdown.

    Request body (JSON):
      {
        "models": [{"name": "Sonnet 4", "result": {...}}, ...],
        "transcript": "상담사: ...",
        "manual_evaluation": "...텍스트..." (선택),
        "manual_rows": [{"item_number": 1, "qa_score": 5, "qa_evidence": "..."} ...] (선택),
        "manual_total": 95 (선택)
      }
    Response (V2 HTML client-side fallback 포맷 호환):
      {
        "status": "success",
        "rows": [ ...행별 비교... ],
        "summary": { "manual_total": ..., "model_totals": {<name>: <num>} },
        "model_names": ["Sonnet 4", "Haiku 3.5"],
        "markdown": "...",
        "comparison_table": [ ...rows 복사... ]
      }
    """
    try:
        body = await request.json()
    except Exception as e:
        return JSONResponse({"status": "error", "error": "bad_request", "detail": str(e)}, status_code=400)

    models = body.get("models") or []
    if not isinstance(models, list) or not models:
        return JSONResponse(
            {"status": "error", "error": "bad_request", "detail": "models: Report[] 필수"}, status_code=400
        )

    transcript = _truncate_transcript(body.get("transcript"))
    manual_evaluation = body.get("manual_evaluation") or ""
    manual_rows = body.get("manual_rows")
    manual_total = body.get("manual_total")
    manual_by_num = _manual_row_lookup(manual_rows)

    model_names: list[str] = []
    model_items: list[list[dict[str, Any]]] = []
    model_totals: dict[str, Any] = {}
    for m in models:
        if not isinstance(m, dict):
            continue
        name = m.get("name") or f"Model{len(model_names) + 1}"
        model_names.append(name)
        res = m.get("result") or {}
        items = _compact_items_for_judge(res)
        model_items.append(items)
        total = ((res.get("orchestrator") or {}).get("final_score") or {}).get("raw_total")
        model_totals[name] = total

    item_lookup: list[dict[int, dict[str, Any]]] = [
        {int(it.get("item_number")): it for it in items if it.get("item_number") is not None} for items in model_items
    ]

    rows: list[dict[str, Any]] = []
    for num in range(1, 19):
        manual = manual_by_num.get(num, {})
        row: dict[str, Any] = {
            "no": num,
            "item_number": num,
            "category": manual.get("category") or "",
            "item": manual.get("item") or manual.get("item_name") or "",
            "max_score": manual.get("max_score"),
            "qa_score": manual.get("qa_score") if "qa_score" in manual else manual.get("score"),
            "qa_evidence": manual.get("qa_evidence") or manual.get("evidence") or "",
        }
        for idx, _name in enumerate(model_names, start=1):
            it = item_lookup[idx - 1].get(num, {}) if idx - 1 < len(item_lookup) else {}
            row[f"model{idx}_score"] = it.get("score")
            row[f"model{idx}_evidence"] = it.get("evidence") or ""
            row[f"model{idx}_verdict"] = (
                "일치" if row.get("qa_score") is not None and row.get("qa_score") == it.get("score") else "불일치"
            )
        rows.append(row)

    # Judge LLM 으로 diff_summary / final_verdict 채우기 (1회 호출)
    import json as _json

    elapsed = 0.0
    if rows and (manual_by_num or manual_evaluation):
        system_prompt = (
            "당신은 한국어 QA 평가 심사관입니다. 수동 평가표와 모델 평가 결과의 행별 차이를 "
            "JSON 배열로 요약하세요. 반드시 **순수 JSON** 만 출력. 마크다운 펜스 금지."
        )
        user_prompt = (
            f"## 원본 상담 스크립트\n{transcript or '(제공되지 않음)'}\n\n"
            f"## 수동 평가표 (텍스트)\n{manual_evaluation[:3000]}\n\n"
            f"## 수동 평가표 (구조화)\n{_json.dumps(manual_rows, ensure_ascii=False)[:3000]}\n\n"
            f"## 모델 결과\n"
            + "\n".join(
                f"### {n}\n{_json.dumps(model_items[i], ensure_ascii=False)[:3000]}" for i, n in enumerate(model_names)
            )
            + "\n\n"
            "### 출력 포맷 (JSON 배열, 18개 행)\n"
            '[{"item_number": 1, "diff_summary": "...", "final_verdict": "일치|부분일치|불일치"}, ...]\n'
            "각 행: 수동 평가와 모든 모델의 점수 차이를 1~2문장 요약 + 최종 판정."
        )
        try:
            t0 = time.time()
            raw = await _invoke_judge_llm(system_prompt, user_prompt, max_tokens=6000)
            elapsed = round(time.time() - t0, 2)
            try:
                from nodes.llm import parse_llm_json  # type: ignore

                diffs = parse_llm_json(raw)
            except Exception:  # noqa: BLE001
                diffs = []
            if isinstance(diffs, dict):
                diffs = diffs.get("rows") or diffs.get("items") or []
            diff_by_num = {
                int(d.get("item_number") or d.get("no") or 0): d for d in (diffs or []) if isinstance(d, dict)
            }
            for r in rows:
                d = diff_by_num.get(r["item_number"], {})
                r["diff_summary"] = d.get("diff_summary") or ""
                r["final_verdict"] = d.get("final_verdict") or r.get("model1_verdict") or ""
        except Exception:  # noqa: BLE001
            logger.exception("analyze-manual-compare judge 실패")
            for r in rows:
                r.setdefault("diff_summary", "")
                r.setdefault("final_verdict", "")
    else:
        for r in rows:
            r.setdefault("diff_summary", "")
            r.setdefault("final_verdict", "")

    # markdown 요약 테이블
    md_lines: list[str] = ["### QA 수동 평가표 비교", ""]
    md_lines.append(f"- 모델: {', '.join(model_names)}")
    if manual_total is not None:
        md_lines.append(f"- 수동 총점: {manual_total}")
    md_lines.append(f"- 모델 총점: {_json.dumps(model_totals, ensure_ascii=False)}")
    md_lines.append("")
    header = "| # | 항목 | 수동 | " + " | ".join(model_names) + " | 최종 |"
    md_lines.append(header)
    md_lines.append("|---|---|---|" + "---|" * len(model_names) + "---|")
    for r in rows:
        cells = [
            str(r["no"]),
            (r.get("item") or "")[:30],
            str(r.get("qa_score") if r.get("qa_score") is not None else "-"),
        ]
        for i in range(len(model_names)):
            s = r.get(f"model{i + 1}_score")
            cells.append(str(s) if s is not None else "-")
        cells.append(r.get("final_verdict") or "-")
        md_lines.append("| " + " | ".join(cells) + " |")
    markdown = "\n".join(md_lines)

    summary = {"manual_total": manual_total, "model_totals": model_totals}

    return JSONResponse(
        {
            "status": "success",
            "rows": rows,
            "summary": summary,
            "model_names": model_names,
            "markdown": markdown,
            "comparison_table": rows,
            "_meta": {"elapsed_sec": elapsed},
        }
    )


# ---------------------------------------------------------------------------
# HITL RAG 관리 endpoints — 프론트 RAG 관리 탭에서 사용
# ---------------------------------------------------------------------------


def _hitl_rag_md_metrics() -> dict[str, Any]:
    """HITL_RAG 폴더 스캔 → md_count / indexed_count / pending_count / last_built_at."""
    try:
        from v2.hitl import md_exporter as _md  # lazy import
    except Exception as exc:
        logger.warning("md_exporter import 실패: %s", exc)
        return {
            "rag_root": "",
            "md_count": 0,
            "indexed_count": 0,
            "pending_count": 0,
            "last_built_at": None,
        }

    root = _md.resolve_rag_root()
    out: dict[str, Any] = {
        "rag_root": str(root),
        "md_count": 0,
        "indexed_count": 0,
        "pending_count": 0,
        "last_built_at": None,
    }
    if not root.exists():
        return out

    last_built: str | None = None
    for p in root.glob("*.md"):
        out["md_count"] += 1
        try:
            parsed = _md.parse_md_file(p)
        except Exception:
            out["pending_count"] += 1
            continue
        meta = parsed.get("meta") or {}
        idx_at = meta.get("indexed_at")
        if idx_at:
            out["indexed_count"] += 1
            if isinstance(idx_at, str) and (last_built is None or idx_at > last_built):
                last_built = idx_at
        else:
            out["pending_count"] += 1
    out["last_built_at"] = last_built
    return out


def _hitl_rag_index_metrics() -> dict[str, Any]:
    """AOSS 인덱스 존재 여부 + 문서 수. opensearch-py 미설치 시 ``index_exists=False``."""
    out: dict[str, Any] = {
        "index_exists": False,
        "index_doc_count": 0,
        "tenant_id": "",
        "index_name": "",
    }
    try:
        from v2.hitl import rag_ingester as _ing  # lazy import

        out["index_name"] = _ing.INDEX_NAME
        out["tenant_id"] = _ing._resolve_tenant()
        client = _ing._client()
        if not client.indices.exists(index=_ing.INDEX_NAME):
            return out
        out["index_exists"] = True
        body = {"query": {"term": {"tenant_id": out["tenant_id"]}}}
        resp = client.count(index=_ing.INDEX_NAME, body=body)
        out["index_doc_count"] = int(resp.get("count") or 0)
    except NotImplementedError as exc:
        logger.info("hitl_rag status: opensearch-py 미설치 — %s", exc)
    except Exception as exc:
        logger.warning("hitl_rag index metrics 실패: %s", exc)
    return out


@app.get("/v2/hitl-rag/status")
async def hitl_rag_status() -> JSONResponse:
    """HITL RAG 빌드 상태 — 폴더 MD 카운트 + AOSS 인덱스 stats.

    Response:
      {
        rag_root: str, md_count: int, indexed_count: int, pending_count: int,
        index_exists: bool, index_doc_count: int, last_built_at: str | null,
        tenant_id: str, index_name: str
      }
    """
    try:
        md = _hitl_rag_md_metrics()
        idx = _hitl_rag_index_metrics()
        return JSONResponse(
            {
                **md,
                **idx,
            }
        )
    except Exception as e:
        logger.exception("hitl-rag status 실패")
        return JSONResponse({"error": "status_failed", "detail": str(e)}, status_code=500)


@app.post("/v2/hitl-rag/rebuild")
async def hitl_rag_rebuild(request: Request) -> JSONResponse:
    """HITL RAG 수동 재빌드. body: {force?: bool=false}.

    force=true 면 모든 MD 의 indexed_at 무시하고 재임베딩.
    응답: {indexed, skipped, errors}.
    """
    try:
        body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    force = bool((body or {}).get("force", False))
    try:
        from v2.hitl import rag_ingester as _ing

        result = _ing.index_pending(force=force)
        return JSONResponse({"ok": True, "force": force, **result})
    except NotImplementedError as exc:
        return JSONResponse(
            {"error": "opensearch_unavailable", "detail": str(exc)}, status_code=503
        )
    except Exception as e:
        logger.exception("hitl-rag rebuild 실패")
        return JSONResponse({"error": "rebuild_failed", "detail": str(e)}, status_code=500)


@app.post("/v2/hitl-rag/recreate-index")
async def hitl_rag_recreate_index(request: Request) -> JSONResponse:
    """인덱스 drop + create + 전체 재임베딩 (관리자 도구).

    body 강제: {confirm: "DROP-INDEX"}. 잘못 호출 방지 가드.
    """
    try:
        body = await request.json() if request.headers.get("content-length", "0") != "0" else {}
    except Exception as e:
        return JSONResponse({"error": "bad_request", "detail": str(e)}, status_code=400)

    if (body or {}).get("confirm") != "DROP-INDEX":
        return JSONResponse(
            {"error": "confirmation_required", "detail": 'body must include {"confirm": "DROP-INDEX"}'},
            status_code=400,
        )

    try:
        from v2.hitl import rag_ingester as _ing

        client = _ing._client()
        index_name = _ing.INDEX_NAME
        dropped = False
        if client.indices.exists(index=index_name):
            client.indices.delete(index=index_name)
            dropped = True
            logger.info("hitl-rag recreate: 인덱스 drop 완료 — %s", index_name)
        # MD frontmatter 의 indexed_at 도 비워야 force 없이 재임베딩 — force=True 로 전체 진행
        result = _ing.index_pending(force=True)
        return JSONResponse(
            {
                "ok": True,
                "dropped": dropped,
                "index_name": index_name,
                **result,
            }
        )
    except NotImplementedError as exc:
        return JSONResponse(
            {"error": "opensearch_unavailable", "detail": str(exc)}, status_code=503
        )
    except Exception as e:
        logger.exception("hitl-rag recreate-index 실패")
        return JSONResponse({"error": "recreate_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/hitl-rag/cases")
async def hitl_rag_cases(
    item_number: int | None = None, limit: int = 50, offset: int = 0
) -> JSONResponse:
    """HITL_RAG 폴더의 MD 메타 리스트 (frontmatter 만 파싱). 페이지네이션 지원.

    Query:
      item_number — 선택, 해당 #번 항목만
      limit       — 기본 50, 최대 500
      offset      — 기본 0
    """
    try:
        from v2.hitl import md_exporter as _md  # lazy import

        root = _md.resolve_rag_root()
        if not root.exists():
            return JSONResponse(
                {"items": [], "total": 0, "limit": int(limit), "offset": int(offset), "rag_root": str(root)}
            )

        eff_limit = max(1, min(int(limit), 500))
        eff_offset = max(0, int(offset))

        # 전체 스캔 후 metadata 추출 — frontmatter 만 파싱하므로 충분히 빠름
        all_items: list[dict[str, Any]] = []
        for p in sorted(root.glob("*.md"), key=lambda x: x.stat().st_mtime, reverse=True):
            try:
                parsed = _md.parse_md_file(p)
            except Exception:
                continue
            meta = parsed.get("meta") or {}
            if item_number is not None:
                try:
                    if int(meta.get("item_number") or 0) != int(item_number):
                        continue
                except (TypeError, ValueError):
                    continue
            # frontend HitlRagCaseListItem 인터페이스 정합 — meta nested 형식.
            # 단일 detail endpoint (`/v2/hitl-rag/case/{filename}`) 와 동일 구조.
            all_items.append(
                {
                    "filename": p.name,
                    "meta": {
                        "consultation_id": meta.get("consultation_id"),
                        "item_number": meta.get("item_number"),
                        "item_name": meta.get("item_name"),
                        "ai_score": meta.get("ai_score"),
                        "human_score": meta.get("human_score"),
                        "max_score": meta.get("max_score"),
                        "delta": meta.get("delta"),
                        "status": meta.get("status"),
                        "reviewer_id": meta.get("reviewer_id"),
                        "reviewer_role": meta.get("reviewer_role"),
                        "site_id": meta.get("site_id"),
                        "channel": meta.get("channel"),
                        "department": meta.get("department"),
                        "confirmed_at": meta.get("confirmed_at"),
                        "indexed_at": meta.get("indexed_at"),
                        "score_signature": meta.get("score_signature"),
                    },
                    "mtime": p.stat().st_mtime,
                }
            )

        total = len(all_items)
        page = all_items[eff_offset : eff_offset + eff_limit]
        return JSONResponse(
            {
                "items": page,
                "total": total,
                "limit": eff_limit,
                "offset": eff_offset,
                "rag_root": str(root),
            }
        )
    except Exception as e:
        logger.exception("hitl-rag cases 실패")
        return JSONResponse({"error": "cases_failed", "detail": str(e)}, status_code=500)


@app.get("/v2/hitl-rag/case/{filename}")
async def hitl_rag_case(filename: str) -> JSONResponse:
    """단일 MD 파일 → {meta, body}. 파일명 traversal 차단."""
    if "/" in filename or "\\" in filename or filename.startswith("..") or not filename.endswith(".md"):
        return JSONResponse({"error": "invalid_filename"}, status_code=400)
    try:
        from v2.hitl import md_exporter as _md  # lazy import

        root = _md.resolve_rag_root()
        target = root / filename
        # 실제 path 가 root 안인지 확인 (symlink 우회 방지)
        try:
            target.resolve().relative_to(root.resolve())
        except (ValueError, OSError):
            return JSONResponse({"error": "invalid_filename"}, status_code=400)
        if not target.exists():
            return JSONResponse({"error": "not_found", "filename": filename}, status_code=404)
        parsed = _md.parse_md_file(target)
        return JSONResponse(
            {
                "ok": True,
                "filename": filename,
                "path": str(target),
                "meta": parsed.get("meta") or {},
                "body": parsed.get("body") or "",
            }
        )
    except Exception as e:
        logger.exception("hitl-rag case 실패 filename=%s", filename)
        return JSONResponse({"error": "case_failed", "detail": str(e)}, status_code=500)


@app.delete("/v2/hitl-rag/case/{filename}")
async def hitl_rag_case_delete(filename: str) -> JSONResponse:
    """HITL RAG 사례 단건 삭제 — md 파일 + AOSS 인덱스 doc 모두 제거.

    안전:
      - 파일명 traversal 차단 (slash / backslash / .. / .md 외 확장자)
      - AOSS 삭제는 best-effort (실패해도 md 파일은 삭제 후 200 반환,
        warning 으로 응답 body 에 표시)
    """
    if "/" in filename or "\\" in filename or filename.startswith("..") or not filename.endswith(".md"):
        return JSONResponse({"error": "invalid_filename"}, status_code=400)

    try:
        from v2.hitl import md_exporter as _md
    except Exception as exc:
        return JSONResponse({"error": "module_load_failed", "detail": str(exc)}, status_code=500)

    root = _md.resolve_rag_root()
    target = root / filename
    try:
        target.resolve().relative_to(root.resolve())
    except (ValueError, OSError):
        return JSONResponse({"error": "invalid_filename"}, status_code=400)

    if not target.exists():
        return JSONResponse({"error": "not_found", "filename": filename}, status_code=404)

    # parse 해서 cid + item_number 확보 (AOSS dedup 키 재구성용)
    cid = None
    item_no = None
    try:
        parsed = _md.parse_md_file(target)
        meta = parsed.get("meta") or {}
        cid = meta.get("consultation_id")
        item_no = meta.get("item_number")
    except Exception as exc:
        logger.warning("hitl-rag delete: meta parse 실패 (계속 진행) — %s", exc)

    aoss_warning: str | None = None
    aoss_deleted = 0
    if cid is not None and item_no is not None:
        try:
            from v2.hitl import rag_ingester as _ing
            client = _ing._client()
            index_name = _ing.INDEX_NAME
            # external_id prefix 매칭 — md 한 파일 = AOSS doc 들 (signature 별 1개)
            ext_prefix = f"{cid}_{int(item_no):02d}_"
            resp = client.delete_by_query(
                index=index_name,
                body={"query": {"prefix": {"external_id": ext_prefix}}},
                refresh=True,
            )
            aoss_deleted = int(resp.get("deleted") or 0)
        except NotImplementedError:
            aoss_warning = "opensearch-py 미설치 — AOSS doc 미삭제, md 파일만 삭제"
        except Exception as exc:
            aoss_warning = f"AOSS delete_by_query 실패 — md 파일은 삭제됨: {exc}"
            logger.warning("hitl-rag delete: AOSS 실패 — %s", exc)

    try:
        target.unlink()
    except Exception as exc:
        logger.exception("hitl-rag delete: md unlink 실패 filename=%s", filename)
        return JSONResponse(
            {"error": "delete_failed", "detail": str(exc)}, status_code=500
        )

    logger.info(
        "hitl-rag delete: md=%s cid=%s item=%s aoss_deleted=%d warning=%s",
        filename, cid, item_no, aoss_deleted, aoss_warning,
    )
    return JSONResponse(
        {
            "ok": True,
            "filename": filename,
            "consultation_id": cid,
            "item_number": item_no,
            "aoss_deleted": aoss_deleted,
            "warning": aoss_warning,
        }
    )


# ---------------------------------------------------------------------------
# Comparison Agent — 사람 정답(human_score) vs AI 평가 차이 비교
# ---------------------------------------------------------------------------
# 평가 결과 탭의 "사람-AI 비교" 섹션에서 운영자가 한눈에 확인하기 위한 백엔드.
# 데이터 부재(JSON 없음 / confirmed 0건) 도 200 으로 일관 응답 — 프론트가
# 반복 호출하지 않도록 (사유는 `reason` 필드로 구분: "no_ai_report" | "no_confirmed_reviews").
# 지표는 MAE / RMSE / Bias / MAPE / Accuracy 만 노출 (memory: feedback_qa_metric_framing).


@app.get("/v2/result/comparison/{consultation_id}")
async def result_comparison(consultation_id: str) -> JSONResponse:
    """사람 정답 vs AI 평가 비교 결과.

    Returns
    -------
    200 — 정상  : `{available: True, summary, by_category, items, ...}`
    200 — 스킵  : `{available: False, reason: "no_ai_report" | "no_confirmed_reviews", summary: null, ...}`
    502 — 비교 계산 중 예외 발생 (`{"error": "comparison_failed", "detail": "..."}`).
    """
    try:
        from v2.agents.comparison_agent import compute_comparison  # lazy import
    except ImportError as exc:
        logger.exception("comparison_agent import 실패")
        return JSONResponse({"error": "comparison_failed", "detail": f"import error: {exc}"}, status_code=502)
    try:
        result = compute_comparison(consultation_id)
    except Exception as e:
        logger.exception("comparison 계산 실패 cid=%s", consultation_id)
        return JSONResponse({"error": "comparison_failed", "detail": str(e)}, status_code=502)
    return JSONResponse(result)
