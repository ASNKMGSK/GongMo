# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0
"""
진짜 코오롱 데이터 → 실험 corpus + test cases 변환.

입력:
    1. C:\\Users\\META M\\Desktop\\코오롱 업무 정확도 auto_qa_criteria.xlsx
       - "필수안내사항" 시트 16 rows (메인 KMS)
       - 인텐트별 시트 (회원정보/환불/교환/반품/수선/배송/취소) 의 분기 데이터
    2. C:\\Users\\META M\\Desktop\\qa 샘플\\테스트셋 (9 files) + 학습셋 (14 files)
       + 멀티테넌트 테스트 (8 files)
       - 실제 코오롱 상담 transcript

출력:
    data/kolon_kms.json — 진짜 KMS corpus (모든 분기 통합)
    data/kolon_test_cases.json — 23+8=31 transcript + multi-label GT
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

import openpyxl

XLSX_PATH = Path(r"C:\Users\META M\Desktop\코오롱 업무 정확도 auto_qa_criteria.xlsx")
SAMPLE_ROOT = Path(r"C:\Users\META M\Desktop\qa 샘플")
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "data"


# ── 1. KMS 추출 ────────────────────────────────────────────────────────


def extract_kms() -> list[dict]:
    """필수안내사항 시트 16 rows + 인텐트별 시트 보충 → 통합 corpus."""
    xl = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    rows: list[dict] = []
    seen: set[tuple[str, str]] = set()  # (intent, branch) 중복 제거

    # 1) 메인: 필수안내사항 시트
    ws = xl["필수안내사항"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # ['문의유형', '세부사항', '조건', 'condition', '키워드', 'example sentences']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        intent = (row[0] or "").strip()
        branch = (row[1] or "").strip()
        condition_ko = (row[2] or "").strip()
        condition_en = (row[3] or "").strip() if len(row) > 3 else ""
        keywords_raw = (row[4] or "").strip() if len(row) > 4 else ""
        statements = (row[5] or "").strip() if len(row) > 5 else ""
        if not intent or not branch:
            continue
        key = (intent, branch)
        if key in seen:
            continue
        seen.add(key)

        keywords = _parse_keywords(keywords_raw)
        # "평가 제외" 카드 식별
        is_skip = (
            "평가 제외" in branch
            or "평가제외" in branch
            or branch.startswith("※")
            or "평가 제외" in keywords_raw
        )

        pid = _make_pid(intent, branch)
        rows.append(
            {
                "pid": pid,
                "intent": intent,
                "branch": branch,
                "condition": condition_ko,
                "condition_en": condition_en,
                "required_keywords": keywords,
                "required_statements": _split_statements(statements),
                "is_evaluation_skip": is_skip,
                "source_sheet": "필수안내사항",
            }
        )

    # 2) 인텐트별 시트 — 추가 분기 보충 (필수안내사항에 없는 것만)
    intent_sheets = ["회원정보", "환불", "교환", "반품", "수선", "배송", "취소"]
    for sn in intent_sheets:
        if sn not in xl.sheetnames:
            continue
        ws = xl[sn]
        # 헤더는 R4 (R1-3 은 타이틀)
        for row in ws.iter_rows(min_row=5, max_row=50, values_only=True):
            if not any(row[1:6]):
                continue
            intent = (row[1] or "").strip()
            branch = (row[2] or "").strip()
            condition = (row[3] or "").strip()
            keywords_raw = (row[4] or "").strip()
            statements = (row[5] or "").strip() if len(row) > 5 else ""
            if not intent and not branch:
                continue
            # intent 비어있으면 시트명에서 추론
            if not intent:
                intent = sn
            if not branch:
                continue

            key = (intent, branch)
            if key in seen:
                continue
            seen.add(key)

            keywords = _parse_keywords(keywords_raw)
            is_skip = (
                "평가 제외" in branch
                or "평가제외" in branch
                or branch.startswith("※")
                or "평가 제외" in keywords_raw
            )
            pid = _make_pid(intent, branch)
            rows.append(
                {
                    "pid": pid,
                    "intent": intent,
                    "branch": branch,
                    "condition": condition,
                    "condition_en": "",
                    "required_keywords": keywords,
                    "required_statements": _split_statements(statements),
                    "is_evaluation_skip": is_skip,
                    "source_sheet": sn,
                }
            )

    return rows


def _parse_keywords(raw: str) -> list[str]:
    if not raw:
        return []
    parts = [p.strip() for p in re.split(r"[,/]", raw)]
    return [p for p in parts if p]


def _split_statements(raw: str) -> list[str]:
    if not raw:
        return []
    return [s.strip() for s in raw.split("\n") if s.strip()]


def _make_pid(intent: str, branch: str) -> str:
    """직관적 pid 생성. 한국어/공백/특수문자 제거 + intent_branch 형태."""
    norm = re.sub(r"[\s\-/※]+", "_", f"{intent}_{branch}").strip("_")
    norm = re.sub(r"[^\w가-힣_]", "", norm)
    return norm or "unknown"


# ── 2. Transcript test cases 변환 ─────────────────────────────────────


# 파일명 패턴: 668797_kolon_긍정_안타티카_스텐드_백_사이즈_교환_및_반품_요청(배송비_문제).json
# description 에서 인텐트 키워드 추출
INTENT_TRIGGERS = {
    "교환": ["교환", "사이즈", "색상", "맞교환"],
    "반품": ["반품", "환불"],
    "수선": ["수선", "A/S", "AS", "실밥", "수리"],
    "취소": ["취소", "주문취소"],
    "배송": ["배송", "배송지", "운송장", "주소변경"],
    "환불": ["환불", "환급", "결제수단"],
    "회원정보": ["회원정보", "탈퇴", "비밀번호", "이메일", "전화번호"],
}


def _extract_intents_from_text(text: str) -> list[str]:
    """description 또는 파일명에서 등장 인텐트 추출 (multi-label)."""
    detected = []
    for intent, keys in INTENT_TRIGGERS.items():
        for k in keys:
            if k in text:
                detected.append(intent)
                break
    return list(dict.fromkeys(detected))  # 순서 유지 dedupe


def collect_transcripts() -> list[dict]:
    cases: list[dict] = []
    for sub in ["테스트셋", "학습셋", "멀티테넌트 테스트"]:
        folder = SAMPLE_ROOT / sub
        if not folder.exists():
            continue
        for fp in sorted(folder.glob("*.json")):
            try:
                data = json.load(fp.open(encoding="utf-8"))
            except Exception as e:  # noqa: BLE001
                print(f"  skip {fp.name}: {e}")
                continue
            if not isinstance(data, dict) or "transcript" not in data:
                continue
            description = data.get("description", "")
            label = data.get("label", "")
            # GT intents — description + label + 파일명에서 추출
            gt_text = f"{description} {label} {fp.stem}"
            gt_intents = _extract_intents_from_text(gt_text)
            cases.append(
                {
                    "case_id": data.get("id") or fp.stem.split("_")[0],
                    "filename": fp.name,
                    "split": sub,
                    "description": description,
                    "label": label,
                    "tenant_id": data.get("tenant_id") or data.get("SITE_CD") or "unknown",
                    "department": data.get("department"),
                    "channel": data.get("channel"),
                    "site_id": data.get("site_id") or data.get("SITE_CD"),
                    "turns_total": data.get("turns"),
                    "transcript": data.get("transcript", ""),
                    "gt_intents": gt_intents,
                }
            )
    return cases


# ── main ──────────────────────────────────────────────────────────────


def main() -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print("Extracting KMS from xlsx...")
    kms_rows = extract_kms()
    print(f"  KMS rows: {len(kms_rows)}")
    skip_count = sum(1 for r in kms_rows if r["is_evaluation_skip"])
    print(f"  Of which 평가제외: {skip_count}")
    intents = sorted({r["intent"] for r in kms_rows})
    print(f"  Intents: {intents}")
    for it in intents:
        cnt = sum(1 for r in kms_rows if r["intent"] == it)
        print(f"    {it}: {cnt} branches")
    out_kms = DATA_DIR / "kolon_kms.json"
    out_kms.write_text(json.dumps(kms_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {out_kms}")

    print("\nCollecting transcripts...")
    cases = collect_transcripts()
    print(f"  Test cases: {len(cases)}")
    by_split: dict = {}
    for c in cases:
        by_split.setdefault(c["split"], []).append(c)
    for sp, lst in by_split.items():
        print(f"    {sp}: {len(lst)}")
    multi_intent = sum(1 for c in cases if len(c["gt_intents"]) >= 2)
    print(f"  Multi-intent (multi-KMS): {multi_intent}/{len(cases)}")
    out_cases = DATA_DIR / "kolon_test_cases.json"
    out_cases.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {out_cases}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
